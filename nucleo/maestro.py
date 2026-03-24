"""
Módulo centralizado de gestión del MAESTRO_PROVEEDORES.

Contiene:
- Proveedor: dataclass con datos de un proveedor
- MaestroProveedores: carga, indexa y busca proveedores
- normalizar_nombre_proveedor: normalización para nombres de archivo
"""

import io
import re
import unicodedata
import logging
from typing import Optional, Dict, List, Tuple
from dataclasses import dataclass, field

from openpyxl import load_workbook
from rapidfuzz import fuzz, process
import pdfplumber

try:
    from config.datos_sensibles import CIFS_PROPIOS as _CIFS_PROPIOS
except ImportError:
    _CIFS_PROPIOS = set()


# Sufijos societarios a eliminar de nombres de proveedores
SUFIJOS_ELIMINAR = [
    'S.L.L.', 'S.L.U.', 'S.COOP.', 'S.L.', 'S.A.', 'C.B.',
    'S.L.L', 'S.L.U', 'S.COOP', 'S.L', 'S.A', 'C.B'
]

UMBRAL_FUZZY_DEFAULT = 85


@dataclass
class Proveedor:
    """Datos de un proveedor del MAESTRO"""
    nombre: str
    cif: str = ""
    iban: str = ""
    forma_pago: str = ""
    email: str = ""
    alias: List[str] = field(default_factory=list)
    cuenta: str = ""
    tiene_extractor: bool = False
    archivo_extractor: str = ""


def normalizar_nombre_proveedor(nombre: str, sufijos: List[str] = None) -> str:
    """
    Normaliza el nombre de un proveedor para usar en nombre de archivo.
    Elimina acentos, sufijos societarios, caracteres especiales.
    """
    if not nombre:
        return ""

    if sufijos is None:
        sufijos = SUFIJOS_ELIMINAR

    nombre = unicodedata.normalize('NFD', nombre)
    nombre = ''.join(c for c in nombre if unicodedata.category(c) != 'Mn')
    nombre = nombre.replace("'", "").replace("\u2019", "").replace("`", "")
    nombre = nombre.replace("&", "Y")
    nombre = re.sub(r'\([^)]*\)', '', nombre)

    for sufijo in sorted(sufijos, key=len, reverse=True):
        nombre = re.sub(rf'\b{re.escape(sufijo)}\b', '', nombre, flags=re.IGNORECASE)

    # Limpiar artefactos que quedan tras eliminar sufijos
    nombre = re.sub(r'[,;.]+\s*$', '', nombre)
    nombre = re.sub(r'\s+[,;.]+\s*', ' ', nombre)
    nombre = re.sub(r'[,;.]+\s+', ' ', nombre)

    nombre = ' '.join(nombre.split())
    nombre = nombre.upper()

    return nombre


class MaestroProveedores:
    """
    Gestiona el acceso al MAESTRO_PROVEEDORES.xlsx

    Carga el Excel una vez y mantiene índices en memoria para
    búsqueda rápida por nombre, email, alias y CIF.
    """

    # CIFs propios (cliente) — excluir de identificación por PDF
    # Importado de config/datos_sensibles.py (no hardcodear aquí)
    CIFS_PROPIOS = _CIFS_PROPIOS

    def __init__(self, ruta: str, umbral_fuzzy: int = UMBRAL_FUZZY_DEFAULT):
        self.ruta = ruta
        self.umbral_fuzzy = umbral_fuzzy
        self.proveedores: Dict[str, Proveedor] = {}
        self.emails: Dict[str, str] = {}
        self.alias: Dict[str, str] = {}
        self.cifs: Dict[str, str] = {}
        self._cargar()

    def _cargar(self):
        """Carga los proveedores del Excel"""
        wb = load_workbook(self.ruta, read_only=True, data_only=True)
        ws = wb.active

        headers = {}
        for col, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[str(cell.value).upper().strip()] = col

        col_map = {
            'PROVEEDOR': headers.get('PROVEEDOR', headers.get('NOMBRE', 2)),
            'CIF': headers.get('CIF', headers.get('NIF', 4)),
            'IBAN': headers.get('IBAN', 5),
            'FORMA_PAGO': headers.get('FORMA_PAGO', headers.get('METODO_PAGO', 6)),
            'EMAIL': headers.get('EMAIL', headers.get('CORREO', 7)),
            'ALIAS': headers.get('ALIAS', 3),
            'CUENTA': headers.get('CUENTA', headers.get('CODIGO', 1)),
            'TIENE_EXTRACTOR': headers.get('TIENE_EXTRACTOR', 8),
            'ARCHIVO_EXTRACTOR': headers.get('ARCHIVO_EXTRACTOR', 9)
        }

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue

            nombre = row[col_map['PROVEEDOR'] - 1] if len(row) >= col_map['PROVEEDOR'] else None
            if not nombre:
                continue
            nombre = str(nombre).strip()

            alias_raw = row[col_map['ALIAS'] - 1] if len(row) >= col_map['ALIAS'] else ""
            alias_list = []
            if alias_raw:
                separador = "," if "," in str(alias_raw) else "|"
                alias_list = [a.strip().upper() for a in str(alias_raw).split(separador) if a.strip()]

            tiene_extractor = False
            archivo_extractor = ""
            if len(row) >= col_map['TIENE_EXTRACTOR']:
                tiene_ext = row[col_map['TIENE_EXTRACTOR'] - 1]
                tiene_extractor = str(tiene_ext).upper().strip() == 'SI' if tiene_ext else False
            if len(row) >= col_map['ARCHIVO_EXTRACTOR']:
                archivo_extractor = str(row[col_map['ARCHIVO_EXTRACTOR'] - 1] or "").strip()

            proveedor = Proveedor(
                nombre=nombre,
                cif=str(row[col_map['CIF'] - 1] or "").strip() if len(row) >= col_map['CIF'] else "",
                iban=str(row[col_map['IBAN'] - 1] or "").strip() if len(row) >= col_map['IBAN'] else "",
                forma_pago=str(row[col_map['FORMA_PAGO'] - 1] or "").strip().upper() if len(row) >= col_map['FORMA_PAGO'] else "",
                email=str(row[col_map['EMAIL'] - 1] or "").strip().lower() if len(row) >= col_map['EMAIL'] else "",
                alias=alias_list,
                cuenta=str(row[col_map['CUENTA'] - 1] or "").strip() if len(row) >= col_map['CUENTA'] else "",
                tiene_extractor=tiene_extractor,
                archivo_extractor=archivo_extractor
            )

            self.proveedores[nombre.upper()] = proveedor

            if proveedor.email:
                for em in proveedor.email.split(","):
                    em = em.strip().lower()
                    if em:
                        self.emails[em] = nombre

            for alias in alias_list:
                self.alias[alias.upper()] = nombre

            if proveedor.cif:
                cif_limpio = proveedor.cif.upper().replace(' ', '').replace('-', '')
                if cif_limpio:
                    self.cifs[cif_limpio] = nombre

        wb.close()

        # Emails hardcoded (dominio no coincide con razón social)
        _hardcoded = {
            'manuel@castrolaborda.com': 'PAGOS ALTOS DE ACERED, SL',
            'manuel@lajas.es': 'PAGOS ALTOS DE ACERED, SL',
        }
        for em, nombre_prov in _hardcoded.items():
            self.emails[em] = nombre_prov

    def buscar_por_email(self, email: str) -> Optional[Proveedor]:
        """Busca proveedor por email del remitente"""
        email = email.lower().strip()
        if "<" in email and ">" in email:
            email = email.split("<")[1].split(">")[0].strip()

        nombre = self.emails.get(email)
        if nombre:
            return self.proveedores.get(nombre.upper())

        for email_maestro, nombre_prov in self.emails.items():
            if email_maestro in email or email in email_maestro:
                return self.proveedores.get(nombre_prov.upper())

        return None

    def buscar_por_alias(self, texto: str) -> Optional[Proveedor]:
        """Busca proveedor por alias en el texto"""
        texto_upper = texto.upper().strip()
        texto_limpio = texto_upper.replace(',', ' ').replace('"', ' ').replace('(', ' ').replace(')', ' ')
        palabras_texto = set(texto_limpio.split())

        nombre = self.alias.get(texto_upper)
        if nombre:
            return self.proveedores.get(nombre.upper())

        for alias, nombre_prov in self.alias.items():
            if len(alias) >= 6 and alias in texto_upper:
                return self.proveedores.get(nombre_prov.upper())
            elif len(alias) >= 4 and alias in palabras_texto:
                return self.proveedores.get(nombre_prov.upper())

        return None

    def buscar_fuzzy(self, texto: str, umbral: int = None) -> Optional[Tuple[Proveedor, int]]:
        """Busca proveedor por coincidencia fuzzy"""
        if not texto:
            return None

        if umbral is None:
            umbral = self.umbral_fuzzy

        nombres = list(self.proveedores.keys())
        resultado = process.extractOne(
            texto.upper(),
            nombres,
            scorer=fuzz.token_sort_ratio
        )

        if resultado and resultado[1] >= umbral:
            return (self.proveedores[resultado[0]], resultado[1])
        return None

    def buscar_por_cif(self, cif: str) -> Optional[Proveedor]:
        """Busca proveedor por CIF"""
        cif = cif.upper().replace(' ', '').replace('-', '')
        nombre = self.cifs.get(cif)
        if nombre:
            return self.proveedores.get(nombre.upper())
        return None

    def identificar_por_pdf(self, contenido: bytes, logger=None, texto_pdf: str = None) -> Optional[Proveedor]:
        """
        Último recurso: busca CIF del proveedor en el texto del PDF.
        Si se pasa texto_pdf, no vuelve a abrir el PDF.
        """
        if texto_pdf is None:
            try:
                texto_pdf = ""
                with pdfplumber.open(io.BytesIO(contenido)) as pdf:
                    for page in pdf.pages[:2]:
                        t = page.extract_text()
                        if t:
                            texto_pdf += t + "\n"
            except Exception:
                return None

        if not texto_pdf:
            return None

        cifs_raw = re.findall(r'\b([A-Z])\s?(\d[\s\d]{7,11})\b', texto_pdf.upper())
        cifs_encontrados = []
        for letra, digitos in cifs_raw:
            solo_digitos = digitos.replace(' ', '')
            if len(solo_digitos) == 8:
                cifs_encontrados.append(letra + solo_digitos)
        for cif in cifs_encontrados:
            if cif in self.CIFS_PROPIOS:
                continue
            prov = self.buscar_por_cif(cif)
            if prov:
                if logger:
                    logger.info(f"  ↳ Identificado por CIF en PDF: {prov.nombre} ({cif})")
                return prov

        return None

    def identificar_proveedor(self, email: str, nombre_remitente: str, asunto: str) -> Optional[Proveedor]:
        """Cascada de identificación: email → alias → fuzzy"""
        prov = self.buscar_por_email(email)
        if prov:
            return prov

        for texto in [nombre_remitente, asunto]:
            prov = self.buscar_por_alias(texto)
            if prov:
                return prov

        resultado = self.buscar_fuzzy(nombre_remitente)
        if resultado:
            return resultado[0]

        return None
