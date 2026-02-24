#!/usr/bin/env python3
# -*- coding: utf-8 -*-
r"""
================================================================================
NORMA43.PY - Parser de ficheros bancarios Norma 43 (Sabadell)
================================================================================

Versión: 2.2
Fecha: Enero 2026
Autor: TASCA BAREA S.L.L.

DESCRIPCIÓN:
------------
Convierte ficheros bancarios Norma 43 (.n43, .txt) exportados de Banco Sabadell
en un Excel consolidado con movimientos bancarios categorizados, separados por
pestañas (una por cuenta), con deduplicación y trazabilidad completa.

CAMBIOS v2.2:
-------------
- NEW: COMISION_TPV muestra % comisión junto a remesa (ej: "6303169726 0,52%")
- El % se calcula como: (comisión / abono) × 100
- Si no encuentra abono asociado muestra "SIN_ABONO"

CAMBIOS v2.1.4:
---------------
- FIX: Detecta si el Excel está abierto y muestra mensaje claro antes de fallar

CAMBIOS v2.1.3:
---------------
- FIX: Ruta archivado cambiada a cuadre/norma43/archivados/

CAMBIOS v2.1.2:
---------------
- FIX: F.Operativa y F.Valor con formato fecha DD-MM-YY (no texto)
- FIX: Importe con formato número #,##0.00 € (no texto)

CAMBIOS v2.1.1:
---------------
- FIX: No crear tabla Excel si la hoja está vacía (evita error de reparación)

CAMBIOS v2.1:
-------------
- TRANSFERENCIA_RECIBIDA: Categoria_Detalle extrae de 2304 (CONCEPTO/OBSERVACIONES)
- TRANSFERENCIA: Categoria_Detalle busca OBSERVACIONES en 2304/2305 para nº factura
- ALQUILER_LOCAL: Nuevo tipo cuando beneficiario = "BENJAMIN ORTEGA Y JAIME FDEZ M"
- COMPRA TARJETA: Categoria_Tipo = "COMPRA TARJETA <nº>", Categoria_Detalle = vacío
- Excel: Formato tabla con filtros y estilos (TableStyleMedium4/6)

CUENTAS SOPORTADAS:
-------------------
- TASCA BAREA:       ...4495 (0081-0259-10-0001844495)
- COMESTIBLES BAREA: ...2404 (0081-0259-10-0001992404)

CATEGORÍAS DE MOVIMIENTOS:
--------------------------
| Categoria_Tipo        | Concepto Común | Descripción                          |
|-----------------------|----------------|--------------------------------------|
| TRANSFERENCIA         | 04             | Pago a proveedores                   |
| TRANSFERENCIA_RECIBIDA| 13             | Cobros recibidos                     |
| ALQUILER_LOCAL        | 04             | Alquiler (BENJAMIN ORTEGA Y JAIME)   |
| NOMINA                | 15             | Pagos a empleados                    |
| ADEUDO_CORE           | 03             | Recibos domiciliados SEPA Core       |
| ADEUDO_B2B            | 03             | Recibos domiciliados SEPA B2B        |
| COMPRA TARJETA xxxx   | 12             | Compras con tarjeta (nº en tipo)     |
| ABONO_TPV             | 12             | Ingresos por ventas TPV              |
| COMISION_TPV          | 17             | Comisiones de TPV                    |
| COMISION_DIVISA       | 06             | Comisión por pago en divisa no euro  |
| INGRESO               | 02             | Ingresos en efectivo                 |
| IMPUESTO              | 03/15          | Pagos a Hacienda, SS, Ayuntamiento   |
| DEVOLUCION            | 99             | Devoluciones                         |
| SERVICIO_TPV          | 12             | Servicio mensual de TPV              |
| OTRO                  | --             | Otros movimientos no clasificados    |

COMERCIOS TPV REGISTRADOS:
--------------------------
- 0337410674: TASCA (bar físico)
- 0354272759: WOOCOMMERCE (cursos online en COMESTIBLES)
- 0354768939: COMESTIBLES (tienda física)

COLUMNAS DEL EXCEL DE SALIDA:
-----------------------------
1.  MovimientoId         - Hash único del movimiento (20 chars)
2.  #                    - Número secuencial en la hoja
3.  F.Operativa          - Fecha de operación (DD-MM-YY)
4.  Concepto             - Descripción del movimiento
5.  Categoria_Tipo       - Tipo de categoría (ver tabla arriba)
6.  Categoria_Detalle    - Detalle: nº factura, nº remesa, etc.
7.  Proveedor_Beneficiario - Nombre del proveedor o beneficiario
8.  F.Valor              - Fecha valor (DD-MM-YY)
9.  Importe              - Importe con signo (- = cargo, + = abono)
10. Saldo                - (Reservado para futuro uso)
11. Ref1                 - Referencia 1 del banco
12. Ref2                 - Referencia 2 del banco
13. Remesa_Asociada      - Nº de remesa para vincular ABONO-COMISION TPV
14. Estado               - OK / DUPLICADO_POSIBLE
15. FicheroOrigen        - Nombre del fichero N43 procesado
16. LineaOrigen          - Línea del registro 22 en el fichero

RUTAS:
------
- Script:    C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\cuadre\norma43\norma43.py
- Salida:    C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\outputs\
- Archivado: C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\cuadre\norma43\archivados\

NOMBRE DEL ARCHIVO DE SALIDA:
-----------------------------
MovimientosSabadell_MMYY-MMYY.xlsx
Donde MMYY corresponde al mes/año inicial y final según cabecera del N43.

USO:
----
    cd C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\cuadre\norma43
    python norma43.py

Se abrirá un diálogo para seleccionar uno o varios archivos .n43 o .txt.
Los archivos procesados se moverán automáticamente a la carpeta "archivados".

REQUISITOS:
-----------
    pip install pandas openpyxl

FORMATO NORMA 43 (resumen):
---------------------------
- Registro 11: Cabecera de cuenta (fecha inicial/final, saldo inicial)
- Registro 22: Movimiento principal (fecha, importe, concepto común)
- Registro 23: Complementos de concepto (hasta 5 por movimiento)
- Registro 33: Final de cuenta
- Registro 88: Final de fichero

CONCEPTOS COMUNES NORMA 43:
---------------------------
01 = Talones/Reintegros    07 = Suscripciones
02 = Ingresos              08 = Dividendos
03 = Adeudos/Recibos       09 = Bolsa
04 = Transferencias        12 = Tarjetas
05 = Amortizaciones        14 = Devoluciones
06 = Remesas efectos       15 = Nóminas
                           17 = Comisiones/Gastos
                           99 = Varios

================================================================================
"""

from __future__ import annotations

import os
import re
import shutil
import hashlib
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import pandas as pd

# GUI Windows (tkinter viene con Python)
import tkinter as tk
from tkinter import filedialog, messagebox


# ==============================================================================
# CONFIGURACIÓN
# ==============================================================================

# Rutas del sistema
OUTPUT_DIR = Path(r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\outputs")
ARCHIVE_DIR = Path(r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\cuadre\norma43\archivados")

# Mapeo de sufijo de cuenta a nombre de pestaña
SHEET_BY_ACCOUNT_SUFFIX: Dict[str, str] = {
    "2404": "COMESTIBLES",
    "4495": "TASCA",
}

# Mapeo de número de comercio TPV a nombre
COMERCIOS_TPV: Dict[str, str] = {
    "0337410674": "TASCA",
    "0354272759": "WOOCOMMERCE",
    "0354768939": "COMESTIBLES",
}

# Prefijo del nombre de archivo de salida
EXCEL_BASENAME_PREFIX = "MovimientosSabadell"

# Formato de fecha para Excel
DATE_FORMAT_EXCEL = "%d-%m-%y"  # DD-MM-YY

# Palabras clave para detectar nóminas
KEYWORDS_NOMINA = [
    "NOMINA", "NÓMINA", "PAGA EXTRA", "SALARIO", "SUELDO", 
    "RETRIBUCION", "RETRIBUCIÓN", "MENSUALIDAD"
]

# Palabras clave para detectar impuestos
KEYWORDS_IMPUESTO = [
    "IMPUESTO", "IMP.", "TGSS", "COTIZACION", "COTIZACIÓN",
    "TAXS", "IBI", "IAE", "IVA", "IRPF", "AEAT", "AGENCIA TRIBUTARIA",
    "AYUNTAMIENTO", "HACIENDA"
]


# ==============================================================================
# MODELO DE DATOS
# ==============================================================================

@dataclass
class N43Movement:
    """
    Representa un movimiento bancario extraído de un fichero Norma 43.
    
    Attributes:
        movimiento_id: Hash único del movimiento (20 chars)
        account_suffix: Últimos 4 dígitos de la cuenta (2404 o 4495)
        oper_date: Fecha de operación
        value_date: Fecha valor
        amount: Importe con signo (negativo = cargo)
        balance: Saldo (no disponible en N43 estándar)
        concepto: Descripción del movimiento
        categoria_tipo: Tipo de categoría (TRANSFERENCIA, ADEUDO_CORE, etc.)
        categoria_detalle: Detalle (nº factura, nº remesa, etc.)
        proveedor_beneficiario: Nombre del proveedor o beneficiario
        remesa_asociada: Nº de remesa para vincular ABONO-COMISION TPV
        ref1: Referencia 1 del banco
        ref2: Referencia 2 del banco
        status: Estado (OK / DUPLICADO_POSIBLE)
        source_file: Nombre del fichero origen
        source_line_22: Número de línea del registro 22
        raw22: Contenido crudo del registro 22
        raw23: Contenido crudo de los registros 23
        raw11: Contenido crudo del registro 11 (cabecera)
    """
    movimiento_id: str
    account_suffix: str
    oper_date: Optional[date]
    value_date: Optional[date]
    amount: Optional[float]
    balance: Optional[float]
    concepto: str
    categoria_tipo: str
    categoria_detalle: str
    proveedor_beneficiario: str
    remesa_asociada: str
    ref1: str
    ref2: str
    status: str
    source_file: str
    source_line_22: int
    raw22: str
    raw23: str
    raw11: str


@dataclass
class N43Header:
    """
    Representa la cabecera de una cuenta en el fichero N43 (registro 11).
    
    Attributes:
        account_suffix: Últimos 4 dígitos de la cuenta
        fecha_inicial: Fecha inicial del período
        fecha_final: Fecha final del período
        saldo_inicial: Saldo inicial
        nombre: Nombre abreviado del titular
    """
    account_suffix: str
    fecha_inicial: Optional[date]
    fecha_final: Optional[date]
    saldo_inicial: Optional[float]
    nombre: str


# ==============================================================================
# UTILIDADES GENERALES
# ==============================================================================

def _ensure_dirs() -> None:
    """Crea los directorios de salida y archivado si no existen."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)


def _parse_yymmdd(s: str) -> Optional[date]:
    """
    Parsea una fecha en formato YYMMDD.
    
    Args:
        s: String de 6 caracteres en formato YYMMDD
        
    Returns:
        Objeto date o None si no se puede parsear
    """
    s = s.strip()
    if not re.fullmatch(r"\d{6}", s):
        return None
    try:
        return datetime.strptime(s, "%y%m%d").date()
    except ValueError:
        return None


def _parse_ddmmyy(s: str) -> Optional[date]:
    """
    Parsea una fecha en formato DDMMYY.
    
    Args:
        s: String de 6 caracteres en formato DDMMYY
        
    Returns:
        Objeto date o None si no se puede parsear
    """
    s = s.strip()
    if not re.fullmatch(r"\d{6}", s):
        return None
    try:
        return datetime.strptime(s, "%d%m%y").date()
    except ValueError:
        return None


def _parse_date6_auto(s: str) -> Optional[date]:
    """
    Parsea una fecha de 6 dígitos probando YYMMDD y DDMMYY.
    Sabadell usa YYMMDD.
    
    Args:
        s: String de 6 caracteres
        
    Returns:
        Objeto date o None si no se puede parsear
    """
    return _parse_yymmdd(s) or _parse_ddmmyy(s)


def _to_float_amount(s: str) -> Optional[float]:
    """
    Convierte un importe N43 (14 dígitos, 2 decimales implícitos) a float.
    
    Args:
        s: String numérico de hasta 14 caracteres
        
    Returns:
        Float con el importe o None si no es válido
        
    Example:
        >>> _to_float_amount("00000000100000")
        1000.00
    """
    s = str(s).strip()
    if not s or not re.fullmatch(r"\d+", s):
        return None
    return int(s) / 100.0


def _hash_movement(account_suffix: str, raw11: str, raw22: str, raw23: str) -> str:
    """
    Genera un ID único y estable para un movimiento.
    
    El hash se basa en: cuenta + cabecera + registro 22 + registros 23.
    Esto garantiza que el mismo movimiento siempre tenga el mismo ID.
    
    Args:
        account_suffix: Últimos 4 dígitos de la cuenta
        raw11: Contenido del registro 11
        raw22: Contenido del registro 22
        raw23: Contenido de los registros 23 concatenados
        
    Returns:
        String de 20 caracteres hexadecimales
    """
    h = hashlib.sha256()
    payload = f"{account_suffix}|{raw11.rstrip()}|{raw22.rstrip()}|{raw23.rstrip()}"
    h.update(payload.encode("utf-8", errors="ignore"))
    return h.hexdigest()[:20]


def _clean_text(s: str) -> str:
    """
    Limpia un texto: mayúsculas, espacios múltiples, trim.
    
    Args:
        s: Texto a limpiar
        
    Returns:
        Texto limpio
    """
    s = s.upper()
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _extract_23_text(line: str) -> str:
    """
    Extrae el texto de un registro 23, quitando el código.
    Soporta formatos '23 01 TEXTO' y '2301TEXTO'.
    
    Args:
        line: Línea completa del registro 23
        
    Returns:
        Texto del concepto sin el código
    """
    # Quitar "23" + codigo (01..05) con o sin espacios
    t = re.sub(r"^23\s*(\d{2})\s*", "", line).strip()
    return t


def _clean_prefix(text: str, prefixes: List[str]) -> str:
    """
    Elimina prefijos conocidos de un texto.
    
    Args:
        text: Texto original
        prefixes: Lista de prefijos a eliminar
        
    Returns:
        Texto sin el prefijo
    """
    t = text.strip()
    t_up = t.upper()
    for p in prefixes:
        p_up = p.upper()
        if t_up.startswith(p_up):
            return t[len(p):].strip(" :\t-")
    return t.strip()


# ==============================================================================
# SELECTOR DE ARCHIVOS (GUI WINDOWS)
# ==============================================================================

def _ask_n43_files_gui() -> List[Path]:
    """
    Abre un diálogo de Windows para seleccionar archivos N43.
    
    - Abre por defecto en la carpeta Descargas
    - Permite selección múltiple
    - Acepta extensiones .txt y .n43
    - Si se cancela, pregunta si quiere salir
    
    Returns:
        Lista de rutas de archivos seleccionados, o lista vacía si se cancela
    """
    downloads = Path(os.environ.get("USERPROFILE", str(Path.home()))) / "Downloads"

    while True:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        selected = filedialog.askopenfilenames(
            title="Selecciona uno o varios archivos Norma 43 (.txt / .n43)",
            initialdir=str(downloads),
            filetypes=[
                ("Norma 43 (*.txt, *.n43)", "*.txt *.n43"),
                ("Todos los archivos", "*.*")
            ],
        )
        root.destroy()

        if selected:
            files = [
                Path(p) for p in selected 
                if Path(p).is_file() and Path(p).suffix.lower() in (".txt", ".n43")
            ]
            if files:
                return sorted(files)

            root = tk.Tk()
            root.withdraw()
            messagebox.showwarning(
                "Selección inválida",
                "No se seleccionó ningún archivo .txt o .n43 válido.\nVuelve a intentarlo."
            )
            root.destroy()
            continue

        root = tk.Tk()
        root.withdraw()
        salir = messagebox.askyesno(
            "Cancelar selección",
            "No se seleccionaron archivos.\n¿Quieres salir?"
        )
        root.destroy()

        if salir:
            return []


# ==============================================================================
# PARSING DE REGISTROS N43
# ==============================================================================

def _extract_account_from_11(line11: str, source_name: str) -> Tuple[str, str]:
    """
    Extrae la cuenta bancaria del registro 11 (cabecera).
    
    Formato N43: posiciones 3-6 (entidad) + 7-10 (oficina) + 11-20 (cuenta)
    
    Args:
        line11: Línea completa del registro 11
        source_name: Nombre del fichero (para mensajes de error)
        
    Returns:
        Tupla (cuenta_completa, últimos_4_dígitos)
        
    Raises:
        RuntimeError: Si no se puede extraer la cuenta
    """
    # Posiciones fijas según Norma 43
    maybe = line11[2:6] + line11[6:10] + line11[10:20]
    digits = re.sub(r"\D", "", maybe)

    if len(digits) >= 18:
        account_raw = digits[:18]
        suffix = account_raw[-4:]
        return account_raw, suffix

    # Fallback: buscar secuencia de 10+ dígitos
    all_digits = re.findall(r"\d{10,}", line11)
    if not all_digits:
        raise RuntimeError(f"No se pudo extraer cuenta de cabecera 11: {source_name}")
    account_raw = max(all_digits, key=len)
    suffix = account_raw[-4:]
    return account_raw, suffix


def _parse_header_11(line11: str, source_name: str) -> N43Header:
    """
    Parsea el registro 11 (cabecera de cuenta).
    
    Estructura del registro 11:
    - Pos 1-2: Código registro (11)
    - Pos 3-6: Entidad
    - Pos 7-10: Oficina  
    - Pos 11-20: Cuenta
    - Pos 21-26: Fecha inicial (YYMMDD)
    - Pos 27-32: Fecha final (YYMMDD)
    - Pos 33: Clave D/H saldo
    - Pos 34-47: Importe saldo inicial
    - Pos 48-50: Divisa
    - Pos 51: Modalidad
    - Pos 52-77: Nombre abreviado
    
    Args:
        line11: Línea completa del registro 11
        source_name: Nombre del fichero
        
    Returns:
        Objeto N43Header con los datos de la cabecera
    """
    _, suffix = _extract_account_from_11(line11, source_name)
    
    # Extender línea a 80 caracteres
    s = (line11.rstrip("\r\n") + " " * 80)[:80]
    
    fecha_inicial = _parse_date6_auto(s[20:26])
    fecha_final = _parse_date6_auto(s[26:32])
    
    # Saldo inicial
    dh_saldo = s[32:33]
    saldo_str = s[33:47]
    saldo = _to_float_amount(saldo_str)
    if saldo is not None and dh_saldo == "1":
        saldo = -abs(saldo)  # Deudor = negativo
    
    nombre = s[51:77].strip()
    
    return N43Header(
        account_suffix=suffix,
        fecha_inicial=fecha_inicial,
        fecha_final=fecha_final,
        saldo_inicial=saldo,
        nombre=nombre
    )


def _parse_22_fixed(raw22: str) -> Tuple[
    Optional[date], Optional[date], Optional[float], 
    Optional[str], str, str, str, str
]:
    """
    Parsea el registro 22 (movimiento principal) por posiciones fijas.
    
    Estructura del registro 22:
    - Pos 1-2: Código registro (22)
    - Pos 3-6: Libre
    - Pos 7-10: Oficina origen
    - Pos 11-16: Fecha operación (YYMMDD)
    - Pos 17-22: Fecha valor (YYMMDD)
    - Pos 23-24: Concepto común
    - Pos 25-27: Concepto propio
    - Pos 28: Clave D/H
    - Pos 29-42: Importe
    - Pos 43-52: Nº documento
    - Pos 53-64: Referencia 1
    - Pos 65-80: Referencia 2
    
    Args:
        raw22: Línea completa del registro 22
        
    Returns:
        Tupla con: (fecha_oper, fecha_valor, importe, dh, ref1, ref2, 
                    concepto_comun, concepto_propio)
    """
    s = (raw22.rstrip("\r\n") + " " * 80)[:80]

    oper = _parse_date6_auto(s[10:16])
    val = _parse_date6_auto(s[16:22])

    concept_common = s[22:24]
    concept_own = s[24:27]

    # Debe/Haber y importe
    dh = s[27:28]  # '1' = Debe (cargo), '2' = Haber (abono)
    imp14 = s[28:42]
    amount = _to_float_amount(imp14)
    if amount is not None:
        if dh == "1":
            amount = -abs(amount)  # Cargo = negativo
        elif dh == "2":
            amount = abs(amount)   # Abono = positivo

    # Referencias
    ref1 = s[52:64].strip()
    ref2 = s[64:80].strip()

    return oper, val, amount, dh, ref1, ref2, concept_common, concept_own


# ==============================================================================
# CLASIFICACIÓN DE MOVIMIENTOS
# ==============================================================================

def _classify_movement(
    concept_common: str,
    block23: List[str],
    amount: Optional[float]
) -> Tuple[str, str, str, str, str]:
    """
    Clasifica un movimiento y extrae sus campos semánticos.
    
    Esta es la función principal de categorización. Analiza el concepto común
    y los registros 23 para determinar el tipo de movimiento y extraer
    la información relevante.
    
    Args:
        concept_common: Código de concepto común (2 dígitos)
        block23: Lista de líneas de registros 23
        amount: Importe del movimiento
        
    Returns:
        Tupla con: (categoria_tipo, categoria_detalle, proveedor_beneficiario,
                    concepto, remesa_asociada)
    """
    # Concatenar texto de todos los 23 para búsquedas
    all_23_text = " ".join(_extract_23_text(l) for l in block23).upper()
    
    # Extraer texto de cada registro 23 individualmente
    t23_dict: Dict[str, str] = {}
    for line in block23:
        match = re.match(r"^23\s*(\d{2})", line)
        if match:
            code = match.group(1)
            text = _extract_23_text(line)
            if code not in t23_dict:
                t23_dict[code] = text
            else:
                t23_dict[code] += " " + text
    
    # Inicializar valores por defecto
    categoria_tipo = "OTRO"
    categoria_detalle = ""
    proveedor_beneficiario = ""
    concepto = ""
    remesa_asociada = ""
    
    # =========================================================================
    # TRANSFERENCIA SALIENTE (concepto 04 con BENEFICIARIO)
    # =========================================================================
    if concept_common == "04" and "BENEFICIARIO" in all_23_text:
        categoria_tipo = "TRANSFERENCIA"
        
        # Proveedor/Beneficiario de 2301
        if "01" in t23_dict:
            proveedor_beneficiario = _clean_prefix(t23_dict["01"], [
                "BENEFICIARIO DE LA TRANSFERENCIA :",
                "BENEFICIARIO DE LA TRANSFERENCIA",
                "BENEFICIARIO"
            ])
        
        # Concepto (nº factura) - CAMBIO: buscar en 2304/2305 OBSERVACIONES primero
        # Si tiene "OBSERVACIONES :" o "OBSERVACIONES" seguido de texto, usar eso
        for code in ["04", "05"]:
            if code in t23_dict:
                texto = t23_dict[code]
                # Buscar patrón OBSERVACIONES : o OBSERVACIONES seguido de contenido
                match_obs = re.search(r"OBSERVACIONES\s*:?\s*(.+)", texto, re.IGNORECASE)
                if match_obs:
                    categoria_detalle = match_obs.group(1).strip()
                    break
                # Si es 2305, también buscar CONCEPTO DEL PAGO
                if code == "05":
                    categoria_detalle = _clean_prefix(texto, [
                        "CONCEPTO DEL PAGO",
                        "CONCEPTO"
                    ]).strip()
        
        # Concepto legible
        concepto = f"Transferencia a {proveedor_beneficiario}"
        
        # CAMBIO: Detectar ALQUILER_LOCAL si beneficiario es exactamente "BENJAMIN ORTEGA Y JAIME FDEZ M"
        beneficiario_norm = proveedor_beneficiario.strip().upper()
        if beneficiario_norm == "BENJAMIN ORTEGA Y JAIME FDEZ M":
            categoria_tipo = "ALQUILER_LOCAL"
            concepto = f"Alquiler local a {proveedor_beneficiario}"
            return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
        
        # Detectar si es NÓMINA
        for kw in KEYWORDS_NOMINA:
            if kw.upper() in all_23_text.upper():
                categoria_tipo = "NOMINA"
                concepto = f"Nómina a {proveedor_beneficiario}"
                break
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # TRANSFERENCIA RECIBIDA (concepto 13 con ORDENANTE)
    # =========================================================================
    if concept_common == "13" and "ORDENANTE" in all_23_text:
        categoria_tipo = "TRANSFERENCIA_RECIBIDA"
        
        # Ordenante de 2301
        if "01" in t23_dict:
            proveedor_beneficiario = _clean_prefix(t23_dict["01"], [
                "ORDENANTE DE LA TRANSFERENCIA :",
                "NOMBRE DEL ORDENANTE",
                "ORDENANTE"
            ])
        
        # Concepto de 2304 (CONCEPTO/OBSERVACIONES) - CAMBIO: antes era 2303
        if "04" in t23_dict:
            categoria_detalle = _clean_prefix(t23_dict["04"], [
                "CONCEPTO/OBSERVACIONES",
                "CONCEPTO",
                "OBSERVACIONES :",
                "OBSERVACIONES"
            ]).strip()
        
        concepto = f"Transferencia recibida de {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # ADEUDO B2B (concepto 03 con B2B en 2301)
    # =========================================================================
    if concept_common == "03" and "B2B" in t23_dict.get("01", "").upper():
        categoria_tipo = "ADEUDO_B2B"
        
        # Emisor de 2301 (quitar "B2B")
        if "01" in t23_dict:
            proveedor_beneficiario = _clean_prefix(t23_dict["01"], ["B2B"]).strip()
        
        # Nº factura de 2303
        if "03" in t23_dict:
            factura = t23_dict["03"].strip()
            # Extraer número de factura si tiene formato conocido
            match = re.search(r"FACTURA\s*(?:NUM\.?)?\s*(\S+)", factura, re.IGNORECASE)
            if match:
                categoria_detalle = match.group(1)
            else:
                categoria_detalle = factura[:50]
        
        concepto = f"Adeudo B2B de {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # ADEUDO CORE (concepto 03 con CORE en 2301)
    # =========================================================================
    if concept_common == "03" and "CORE" in t23_dict.get("01", "").upper():
        categoria_tipo = "ADEUDO_CORE"
        
        # Emisor de 2301 (quitar "CORE")
        if "01" in t23_dict:
            proveedor_beneficiario = _clean_prefix(t23_dict["01"], ["CORE"]).strip()
        
        # Detectar si es IMPUESTO
        for kw in KEYWORDS_IMPUESTO:
            if kw.upper() in all_23_text.upper():
                categoria_tipo = "IMPUESTO"
                break
        
        # Nº factura/referencia de 2303
        if "03" in t23_dict:
            ref = t23_dict["03"].strip()
            # Limpiar prefijos comunes
            ref = _clean_prefix(ref, [
                "FACTURA", "FRA", "NUESTRA FACTURA:",
                "SU PAGO FRA", "SU PAGO", "NTRA. FACTURA"
            ])
            # Buscar patrón de factura
            match = re.search(r"([A-Z0-9/-]+\d+[A-Z0-9/-]*)", ref, re.IGNORECASE)
            if match:
                categoria_detalle = match.group(1)[:30]
            else:
                categoria_detalle = ref[:30].strip()
        
        concepto = f"Adeudo CORE de {proveedor_beneficiario}"
        if categoria_tipo == "IMPUESTO":
            concepto = f"Impuesto/Tasa: {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # NÓMINA (concepto 15 con BENEFICIARIO = persona física)
    # =========================================================================
    if concept_common == "15":
        # Verificar si es impuesto o nómina
        is_impuesto = any(kw.upper() in all_23_text.upper() for kw in KEYWORDS_IMPUESTO)
        
        if is_impuesto:
            categoria_tipo = "IMPUESTO"
            if "01" in t23_dict:
                proveedor_beneficiario = _clean_prefix(t23_dict["01"], ["BENEFICIARIO"]).strip()
            concepto = f"Impuesto: {proveedor_beneficiario}"
        else:
            categoria_tipo = "NOMINA"
            if "01" in t23_dict:
                proveedor_beneficiario = _clean_prefix(t23_dict["01"], ["BENEFICIARIO"]).strip()
            concepto = f"Nómina a {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # TARJETA - COMPRA (concepto 12 con "COMPRA TARJ")
    # =========================================================================
    if concept_common == "12" and "COMPRA TARJ" in all_23_text:
        # CAMBIO: Categoria_Tipo = "COMPRA TARJETA <nº tarjeta>", Categoria_Detalle = vacío
        num_tarjeta = ""
        
        if "01" in t23_dict:
            texto = t23_dict["01"].replace("COMPRA TARJ.", "").strip()
            # Extraer nº tarjeta y comercio
            match = re.match(r"(\d{4}X+\d{4})\s+(.+)", texto)
            if match:
                num_tarjeta = match.group(1)  # Nº tarjeta
                proveedor_beneficiario = match.group(2).strip()  # Comercio
            else:
                proveedor_beneficiario = texto
        
        # Formato nuevo: tipo incluye nº tarjeta, detalle vacío
        categoria_tipo = f"COMPRA TARJETA {num_tarjeta}" if num_tarjeta else "COMPRA TARJETA"
        categoria_detalle = ""  # Vacío - se rellena en CUADRE
        
        concepto = f"Compra tarjeta en {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # ABONO TPV (concepto 12 con "ABONO TPV")
    # =========================================================================
    if concept_common == "12" and "ABONO TPV" in all_23_text:
        categoria_tipo = "ABONO_TPV"
        
        if "01" in t23_dict:
            texto = t23_dict["01"]
            # Patrón: ABONO TPV 0337410674 01 TASCA BAREA 6303095426
            match = re.search(r"ABONO TPV (\d{10}) (\d{2}) (.+?) (\d{10})", texto)
            if match:
                comercio_num = match.group(1)
                remesa_asociada = match.group(4)
                categoria_detalle = remesa_asociada
                proveedor_beneficiario = COMERCIOS_TPV.get(comercio_num, match.group(3).strip())
            else:
                proveedor_beneficiario = texto.replace("ABONO TPV", "").strip()
        
        concepto = f"Abono TPV {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # SERVICIO TPV (concepto 12 con "SERVICIO DE TPV")
    # =========================================================================
    if concept_common == "12" and "SERVICIO DE TPV" in all_23_text:
        categoria_tipo = "SERVICIO_TPV"
        
        if "01" in t23_dict:
            texto = t23_dict["01"]
            match = re.search(r"SERVICIO DE TPV (\d{10})", texto)
            if match:
                comercio_num = match.group(1)
                proveedor_beneficiario = COMERCIOS_TPV.get(comercio_num, comercio_num)
        
        concepto = f"Servicio TPV {proveedor_beneficiario}"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # COMISIÓN TPV (concepto 17 con "COMISIONES" y nº comercio)
    # =========================================================================
    if concept_common == "17" and "COMISIONES" in all_23_text:
        # Verificar si tiene patrón de TPV
        if "01" in t23_dict:
            texto = t23_dict["01"]
            match = re.search(r"COMISIONES (\d{10}) (\d{2}) (.+?) (\d{10})", texto)
            if match:
                categoria_tipo = "COMISION_TPV"
                comercio_num = match.group(1)
                remesa_asociada = match.group(4)
                categoria_detalle = remesa_asociada
                proveedor_beneficiario = COMERCIOS_TPV.get(comercio_num, match.group(3).strip())
                concepto = f"Comisión TPV {proveedor_beneficiario}"
                
                return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
        
        # Si no tiene patrón TPV, es comisión genérica
        categoria_tipo = "COMISION_OTRO"
        if "01" in t23_dict:
            concepto = t23_dict["01"]
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # COMISIÓN DIVISA (concepto 06)
    # =========================================================================
    if concept_common == "06" and "DIVISA" in all_23_text:
        categoria_tipo = "COMISION_DIVISA"
        concepto = "Comisión divisa no euro"
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # INGRESO EFECTIVO (concepto 02)
    # =========================================================================
    if concept_common == "02":
        categoria_tipo = "INGRESO"
        if "01" in t23_dict:
            concepto = t23_dict["01"]
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # DEVOLUCIÓN (concepto 99)
    # =========================================================================
    if concept_common == "99":
        categoria_tipo = "DEVOLUCION"
        if "01" in t23_dict:
            concepto = t23_dict["01"]
            categoria_detalle = concepto[:50]
        
        return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada
    
    # =========================================================================
    # OTROS (fallback)
    # =========================================================================
    # Construir concepto desde 2301
    if "01" in t23_dict:
        concepto = t23_dict["01"]
    elif block23:
        concepto = _extract_23_text(block23[0])
    
    return categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada


# ==============================================================================
# PARSER PRINCIPAL N43
# ==============================================================================

def parse_n43_file(path: Path) -> Tuple[List[N43Movement], List[N43Header]]:
    """
    Parsea un fichero Norma 43 completo.
    
    El fichero puede contener múltiples cuentas. Cada cuenta tiene:
    - Un registro 11 (cabecera)
    - Múltiples registros 22 (movimientos) con sus 23 (complementos)
    - Un registro 33 (cierre)
    
    Args:
        path: Ruta al fichero N43
        
    Returns:
        Tupla con:
        - Lista de movimientos (N43Movement)
        - Lista de cabeceras (N43Header)
        
    Raises:
        RuntimeError: Si el fichero está vacío o tiene errores
    """
    lines = path.read_text(encoding="latin-1", errors="ignore").splitlines()
    if not lines:
        raise RuntimeError(f"Fichero vacío: {path.name}")

    current_suffix: Optional[str] = None
    current_raw11: str = ""
    movements: List[N43Movement] = []
    headers: List[N43Header] = []

    i = 0
    while i < len(lines):
        line = lines[i]

        # =====================================================================
        # REGISTRO 11: Cabecera de cuenta
        # =====================================================================
        if line.startswith("11"):
            current_raw11 = line
            header = _parse_header_11(line, path.name)
            current_suffix = header.account_suffix
            headers.append(header)
            i += 1
            continue

        # =====================================================================
        # REGISTRO 33: Cierre de cuenta
        # =====================================================================
        if line.startswith("33"):
            i += 1
            continue

        # =====================================================================
        # REGISTRO 88: Fin de fichero
        # =====================================================================
        if line.startswith("88"):
            break

        # =====================================================================
        # REGISTRO 22: Movimiento principal
        # =====================================================================
        if line.startswith("22"):
            if not current_suffix:
                raise RuntimeError(
                    f"Registro 22 sin cabecera 11 previa en {path.name} (línea {i+1})."
                )

            raw22 = line
            line_no = i + 1

            # Recoger registros 23 asociados
            j = i + 1
            block23: List[str] = []
            while j < len(lines) and lines[j].startswith("23"):
                block23.append(lines[j])
                j += 1

            raw23 = "\n".join(block23)

            # Parsear registro 22
            oper_date, value_date, amount, dh, ref1, ref2, concept_common, concept_own = \
                _parse_22_fixed(raw22)

            # Clasificar movimiento
            categoria_tipo, categoria_detalle, proveedor_beneficiario, concepto, remesa_asociada = \
                _classify_movement(concept_common, block23, amount)

            # Generar ID único
            mov_id = _hash_movement(current_suffix, current_raw11, raw22, raw23)

            movements.append(
                N43Movement(
                    movimiento_id=mov_id,
                    account_suffix=current_suffix,
                    oper_date=oper_date,
                    value_date=value_date,
                    amount=amount,
                    balance=None,  # N43 no incluye saldo por apunte
                    concepto=concepto,
                    categoria_tipo=categoria_tipo,
                    categoria_detalle=categoria_detalle,
                    proveedor_beneficiario=proveedor_beneficiario,
                    remesa_asociada=remesa_asociada,
                    ref1=ref1,
                    ref2=ref2,
                    status="OK",
                    source_file=path.name,
                    source_line_22=line_no,
                    raw22=raw22,
                    raw23=raw23,
                    raw11=current_raw11,
                )
            )

            i = j
            continue

        i += 1

    return movements, headers


# ==============================================================================
# GESTIÓN DE EXCEL
# ==============================================================================

def standard_columns() -> List[str]:
    """
    Devuelve la lista de columnas estándar del Excel de salida.
    
    Returns:
        Lista con los nombres de las 16 columnas
    """
    return [
        "MovimientoId",
        "#",
        "F.Operativa",
        "Concepto",
        "Categoria_Tipo",
        "Categoria_Detalle",
        "Proveedor_Beneficiario",
        "F.Valor",
        "Importe",
        "Saldo",
        "Ref1",
        "Ref2",
        "Remesa_Asociada",
        "Estado",
        "FicheroOrigen",
        "LineaOrigen",
    ]


def load_existing_sheets(excel_path: Path) -> Dict[str, pd.DataFrame]:
    """
    Carga las hojas existentes de un Excel o crea DataFrames vacíos.
    
    Args:
        excel_path: Ruta al fichero Excel
        
    Returns:
        Diccionario con nombre_hoja -> DataFrame
    """
    cols = standard_columns()
    
    if not excel_path.exists():
        return {name: pd.DataFrame(columns=cols) for name in SHEET_BY_ACCOUNT_SUFFIX.values()}

    sheets = pd.read_excel(excel_path, sheet_name=None, dtype=str)
    out: Dict[str, pd.DataFrame] = {}
    
    for sheet_name in SHEET_BY_ACCOUNT_SUFFIX.values():
        df = sheets.get(sheet_name)
        if df is None:
            out[sheet_name] = pd.DataFrame(columns=cols)
        else:
            # Asegurar que existen todas las columnas
            for c in cols:
                if c not in df.columns:
                    df[c] = ""
            out[sheet_name] = df[cols].copy()
    
    return out


def _dates_close(a: Optional[date], b: Optional[date], days: int = 3) -> bool:
    """
    Comprueba si dos fechas están cerca (para detectar duplicados).
    
    Args:
        a: Primera fecha
        b: Segunda fecha
        days: Número máximo de días de diferencia
        
    Returns:
        True si las fechas están a menos de 'days' días de distancia
    """
    if a is None or b is None:
        return False
    return abs((a - b).days) <= days


def _similarity(a: str, b: str) -> float:
    """
    Calcula la similitud entre dos textos basada en tokens compartidos.
    
    Args:
        a: Primer texto
        b: Segundo texto
        
    Returns:
        Valor entre 0.0 y 1.0 (índice de Jaccard)
    """
    ta = set(_clean_text(a).split())
    tb = set(_clean_text(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _looks_like_possible_duplicate(existing_rows: List[dict], m: N43Movement) -> bool:
    """
    Detecta si un movimiento podría ser un duplicado de uno existente.
    
    Criterios:
    - Mismo importe (con tolerancia de 0.01€)
    - Fechas cercanas (±3 días)
    - Concepto similar (≥60% similitud)
    
    Args:
        existing_rows: Lista de movimientos existentes (como diccionarios)
        m: Movimiento nuevo a comprobar
        
    Returns:
        True si parece un duplicado posible
    """
    if m.amount is None:
        return False

    for r in existing_rows[-200:]:  # Solo comprobar los últimos 200
        try:
            amt = float(str(r.get("Importe", "")).replace(",", "."))
        except ValueError:
            continue

        if abs(amt - float(m.amount)) > 0.01:
            continue

        # Comparar fechas
        rop = str(r.get("F.Operativa", "")).strip()
        
        def _parse_excel_date(s: str) -> Optional[date]:
            if not s:
                return None
            try:
                return datetime.strptime(s, DATE_FORMAT_EXCEL).date()
            except ValueError:
                return None

        r_op = _parse_excel_date(rop)
        if not _dates_close(r_op, m.oper_date):
            continue

        # Comparar concepto
        sim = _similarity(str(r.get("Concepto", "")), m.concepto)
        if sim >= 0.6:
            return True

    return False


def append_movements_to_sheet(df: pd.DataFrame, movements: List[N43Movement]) -> pd.DataFrame:
    """
    Añade movimientos a un DataFrame existente con deduplicación.
    
    Args:
        df: DataFrame existente
        movements: Lista de movimientos a añadir
        
    Returns:
        DataFrame actualizado con los nuevos movimientos
    """
    if df.empty:
        existing_ids = set()
        existing_rows: List[dict] = []
    else:
        existing_ids = set(df["MovimientoId"].astype(str).tolist())
        existing_rows = df.to_dict(orient="records")

    new_rows = []
    for m in movements:
        # Saltar si ya existe por ID
        if m.movimiento_id in existing_ids:
            continue

        # Detectar posibles duplicados
        status = "OK"
        if _looks_like_possible_duplicate(existing_rows, m):
            status = "DUPLICADO_POSIBLE"

        row = {
            "MovimientoId": m.movimiento_id,
            "#": "",
            "F.Operativa": m.oper_date.strftime(DATE_FORMAT_EXCEL) if m.oper_date else "",
            "Concepto": m.concepto,
            "Categoria_Tipo": m.categoria_tipo,
            "Categoria_Detalle": m.categoria_detalle,
            "Proveedor_Beneficiario": m.proveedor_beneficiario,
            "F.Valor": m.value_date.strftime(DATE_FORMAT_EXCEL) if m.value_date else "",
            "Importe": f"{m.amount:.2f}" if isinstance(m.amount, (int, float)) else "",
            "Saldo": "",
            "Ref1": m.ref1 or "",
            "Ref2": m.ref2 or "",
            "Remesa_Asociada": m.remesa_asociada or "",
            "Estado": status,
            "FicheroOrigen": m.source_file,
            "LineaOrigen": str(m.source_line_22),
        }
        new_rows.append(row)

    if not new_rows:
        return df

    df2 = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    
    # =========================================================================
    # CALCULAR % COMISIÓN TPV
    # Para cada COMISION_TPV, buscar el ABONO_TPV con la misma remesa
    # y calcular el porcentaje: (comisión / abono) × 100
    # =========================================================================
    df2 = _calculate_tpv_commission_percentage(df2)
    
    # Renumerar
    df2["#"] = [str(i + 1) for i in range(len(df2))]
    
    return df2


def _calculate_tpv_commission_percentage(df: pd.DataFrame) -> pd.DataFrame:
    """
    Calcula el % de comisión TPV y lo añade a Categoria_Detalle.
    
    Para cada fila COMISION_TPV, busca el ABONO_TPV con la misma remesa
    y calcula: % = (|comisión| / abono) × 100
    
    El resultado se muestra como: "6303169726 0,52%"
    Si no encuentra abono: "6303169726 SIN_ABONO"
    
    Args:
        df: DataFrame con los movimientos
        
    Returns:
        DataFrame con Categoria_Detalle actualizado para COMISION_TPV
    """
    if df.empty:
        return df
    
    # Crear diccionario de ABONO_TPV: remesa -> importe
    abonos_tpv = {}
    for idx, row in df.iterrows():
        if row.get("Categoria_Tipo") == "ABONO_TPV":
            remesa = str(row.get("Categoria_Detalle", "")).strip()
            if remesa:
                try:
                    importe = abs(float(str(row.get("Importe", "0")).replace(",", ".")))
                    abonos_tpv[remesa] = importe
                except (ValueError, TypeError):
                    pass
    
    # Actualizar COMISION_TPV con el %
    for idx, row in df.iterrows():
        if row.get("Categoria_Tipo") == "COMISION_TPV":
            detalle = str(row.get("Categoria_Detalle", "")).strip()
            
            # Extraer nº de remesa (puede tener ya texto adicional)
            remesa = detalle.split()[0] if detalle else ""
            
            if remesa and remesa in abonos_tpv:
                abono = abonos_tpv[remesa]
                if abono > 0:
                    try:
                        comision = abs(float(str(row.get("Importe", "0")).replace(",", ".")))
                        porcentaje = (comision / abono) * 100
                        # Formato: "6303169726 0,52%"
                        df.at[idx, "Categoria_Detalle"] = f"{remesa} {porcentaje:.2f}%".replace(".", ",")
                    except (ValueError, TypeError, ZeroDivisionError):
                        df.at[idx, "Categoria_Detalle"] = f"{remesa} SIN_ABONO"
                else:
                    df.at[idx, "Categoria_Detalle"] = f"{remesa} SIN_ABONO"
            elif remesa:
                df.at[idx, "Categoria_Detalle"] = f"{remesa} SIN_ABONO"
    
    return df


def _check_file_writable(path: Path) -> bool:
    """
    Verifica si un archivo se puede escribir (no está abierto por otro programa).
    
    Args:
        path: Ruta del archivo a verificar
        
    Returns:
        True si se puede escribir, False si está bloqueado
    """
    if not path.exists():
        return True
    
    try:
        # Intentar abrir en modo append para verificar si está bloqueado
        with open(path, 'a'):
            pass
        return True
    except PermissionError:
        return False


def save_excel(excel_path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    """
    Guarda el Excel con todas las hojas aplicando formato de tabla.
    
    Args:
        excel_path: Ruta de destino
        sheets: Diccionario nombre_hoja -> DataFrame
        
    Raises:
        PermissionError: Si el archivo está abierto por otro programa
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import numbers
    
    # Verificar si el archivo está bloqueado (abierto en Excel)
    if not _check_file_writable(excel_path):
        raise PermissionError(
            f"\n❌ ERROR: El archivo está abierto en otro programa.\n"
            f"   Cierra el Excel antes de continuar:\n"
            f"   {excel_path}\n"
        )
    
    # Estilos de tabla por hoja
    TABLE_STYLES = {
        "COMESTIBLES": "TableStyleMedium4",
        "TASCA": "TableStyleMedium6",
    }
    
    # Índices de columnas (1-based) para aplicar formatos
    # Basado en: MovimientoId(1), #(2), F.Operativa(3), Concepto(4), Categoria_Tipo(5),
    #            Categoria_Detalle(6), Proveedor_Beneficiario(7), F.Valor(8), Importe(9), ...
    COL_F_OPERATIVA = 3
    COL_F_VALOR = 8
    COL_IMPORTE = 9
    
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        table_counter = 0
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Obtener la hoja
            ws = writer.sheets[sheet_name]
            
            # Solo crear tabla si hay datos (al menos 1 fila además de cabecera)
            if len(df) == 0:
                print(f"   ⚠️  {sheet_name}: Sin datos, no se crea tabla")
                continue
            
            table_counter += 1
            
            # Calcular rango de la tabla
            num_rows = len(df) + 1  # +1 para cabecera
            num_cols = len(df.columns)
            end_col = get_column_letter(num_cols)
            table_range = f"A1:{end_col}{num_rows}"
            
            # =========================================================
            # APLICAR FORMATOS A LAS COLUMNAS
            # =========================================================
            
            # Formato fecha DD-MM-YY para F.Operativa y F.Valor
            for col_idx in [COL_F_OPERATIVA, COL_F_VALOR]:
                col_letter = get_column_letter(col_idx)
                for row in range(2, num_rows + 1):  # Desde fila 2 (después de cabecera)
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value:
                        # Convertir string "DD-MM-YY" a fecha Excel
                        try:
                            dt = datetime.strptime(str(cell.value), "%d-%m-%y")
                            cell.value = dt
                            cell.number_format = "DD-MM-YY"
                        except ValueError:
                            pass  # Mantener valor original si no es fecha válida
            
            # Formato número con 2 decimales y símbolo € para Importe
            col_letter = get_column_letter(COL_IMPORTE)
            for row in range(2, num_rows + 1):
                cell = ws.cell(row=row, column=COL_IMPORTE)
                if cell.value:
                    try:
                        # Convertir a número
                        val = float(str(cell.value).replace(",", "."))
                        cell.value = val
                        cell.number_format = '#,##0.00 €'
                    except ValueError:
                        pass
            
            # =========================================================
            # CREAR TABLA CON ESTILO
            # =========================================================
            table_name = f"Tabla{table_counter}"
            style_name = TABLE_STYLES.get(sheet_name, "TableStyleMedium2")
            
            table = Table(displayName=table_name, ref=table_range)
            style = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            
            # Ajustar ancho de columnas
            for col_idx, column in enumerate(df.columns, start=1):
                col_letter = get_column_letter(col_idx)
                # Calcular ancho basado en contenido
                max_length = len(str(column))
                for cell in ws[col_letter][1:num_rows]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                # Limitar ancho máximo
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)


def archive_file(src: Path) -> None:
    """
    Mueve un fichero procesado a la carpeta de archivados.
    Si ya existe, añade un timestamp al nombre.
    
    Args:
        src: Ruta del fichero a archivar
    """
    dest = ARCHIVE_DIR / src.name
    if dest.exists():
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = ARCHIVE_DIR / f"{src.stem}_{stamp}{src.suffix}"
    shutil.move(str(src), str(dest))


def _generate_excel_filename(headers: List[N43Header]) -> str:
    """
    Genera el nombre del archivo Excel basado en las fechas de las cabeceras.
    
    Formato: MovimientosSabadell_MMYY-MMYY.xlsx
    
    Args:
        headers: Lista de cabeceras N43
        
    Returns:
        Nombre del archivo
    """
    fecha_min = None
    fecha_max = None
    
    for h in headers:
        if h.fecha_inicial and (fecha_min is None or h.fecha_inicial < fecha_min):
            fecha_min = h.fecha_inicial
        if h.fecha_final and (fecha_max is None or h.fecha_final > fecha_max):
            fecha_max = h.fecha_final
    
    if fecha_min and fecha_max:
        mm1 = fecha_min.strftime("%m%y")
        mm2 = fecha_max.strftime("%m%y")
        return f"{EXCEL_BASENAME_PREFIX}_{mm1}-{mm2}.xlsx"
    elif fecha_max:
        mm2 = fecha_max.strftime("%m%y")
        return f"{EXCEL_BASENAME_PREFIX}_{mm2}.xlsx"
    else:
        # Fallback al año actual
        yy = datetime.now().strftime("%y")
        return f"{EXCEL_BASENAME_PREFIX}{yy}.xlsx"


# ==============================================================================
# FUNCIÓN PRINCIPAL
# ==============================================================================

def main() -> None:
    """
    Punto de entrada principal del script.
    
    1. Crea directorios necesarios
    2. Muestra diálogo para seleccionar archivos N43
    3. Parsea cada archivo
    4. Clasifica y deduplica movimientos
    5. Guarda Excel consolidado
    6. Archiva ficheros procesados
    """
    _ensure_dirs()
    
    print("=" * 70)
    print("NORMA43.PY v2.2 - Parser N43 Sabadell → Excel")
    print("=" * 70)
    print()

    # Seleccionar archivos
    files = _ask_n43_files_gui()
    if not files:
        print("❌ No se seleccionaron archivos. Saliendo.")
        return

    print(f"📁 Archivos seleccionados: {len(files)}")
    for f in files:
        print(f"   • {f.name}")
    print()

    # Procesar todos los archivos primero para obtener las fechas
    all_movements: List[N43Movement] = []
    all_headers: List[N43Header] = []
    
    for f in files:
        print(f"🔄 Procesando: {f.name}")
        try:
            movements, headers = parse_n43_file(f)
            all_movements.extend(movements)
            all_headers.extend(headers)
            print(f"   ✅ {len(movements)} movimientos, {len(headers)} cuenta(s)")
        except Exception as e:
            print(f"   ❌ Error: {e}")
            continue

    if not all_movements:
        print("\n❌ No se encontraron movimientos. Saliendo.")
        return

    # Generar nombre del archivo de salida
    excel_filename = _generate_excel_filename(all_headers)
    excel_path = OUTPUT_DIR / excel_filename
    
    print(f"\n📊 Archivo de salida: {excel_path}")

    # Cargar o crear hojas
    sheets = load_existing_sheets(excel_path)

    # Separar movimientos por cuenta
    by_suffix: Dict[str, List[N43Movement]] = {}
    for m in all_movements:
        by_suffix.setdefault(m.account_suffix, []).append(m)

    # Validar cuentas
    unknown = [sfx for sfx in by_suffix.keys() if sfx not in SHEET_BY_ACCOUNT_SUFFIX]
    if unknown:
        print(f"\n⚠️  Cuentas no reconocidas: {unknown}")
        print(f"   Solo se permiten: {list(SHEET_BY_ACCOUNT_SUFFIX.keys())}")

    # Añadir movimientos a cada hoja
    added_total = 0
    for suffix, movs in by_suffix.items():
        if suffix not in SHEET_BY_ACCOUNT_SUFFIX:
            continue
            
        sheet_name = SHEET_BY_ACCOUNT_SUFFIX[suffix]
        before = len(sheets[sheet_name])
        sheets[sheet_name] = append_movements_to_sheet(sheets[sheet_name], movs)
        after = len(sheets[sheet_name])
        added = after - before
        added_total += added
        
        print(f"   📋 {sheet_name}: +{added} nuevos (total: {after})")

    # Guardar Excel
    save_excel(excel_path, sheets)
    print(f"\n💾 Excel guardado: {excel_path}")

    # Archivar ficheros procesados
    print("\n🗄️  Archivando ficheros procesados...")
    for f in files:
        try:
            archive_file(f)
            print(f"   • {f.name} → archivados/")
        except Exception as e:
            print(f"   ⚠️  No se pudo archivar {f.name}: {e}")

    # Resumen final
    print()
    print("=" * 70)
    print("✅ PROCESO COMPLETADO")
    print("=" * 70)
    print(f"   📦 Ficheros procesados: {len(files)}")
    print(f"   ➕ Movimientos añadidos: {added_total}")
    print(f"   📄 Excel: {excel_path}")
    print()


# ==============================================================================
# PUNTO DE ENTRADA
# ==============================================================================

if __name__ == "__main__":
    main()
