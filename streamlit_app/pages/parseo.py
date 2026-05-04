"""
Página Parseo — Extracción de facturas PDF usando motor Parseo real.
"""

import io
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

from utils.auth import require_role
from utils.entorno import ruta_existe_seguro

require_role(["admin"])

# ── Constantes ────────────────────────────────────────────────────────────────

DROPBOX_BASE = Path(
    r"C:\Users\jaime\Dropbox\File inviati\TASCA BAREA S.L.L\CONTABILIDAD"
)
DICCIONARIO_PATH = Path(
    r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\datos"
    r"\DiccionarioProveedoresCategoria.xlsx"
)
SNAPSHOTS_DIR = Path(r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\datos\snapshots")

# ── Importar motor Parseo (modo local) ───────────────────────────────────────

PARSEO_DISPONIBLE = False
try:
    if r"C:\_ARCHIVOS\TRABAJO\Facturas\Parseo" not in sys.path:
        sys.path.insert(0, r"C:\_ARCHIVOS\TRABAJO\Facturas\Parseo")
    from main import procesar_factura, cargar_diccionario  # type: ignore
    from extractores import listar_extractores  # type: ignore

    PARSEO_DISPONIBLE = True
except ImportError:
    pass


# ── Helpers ───────────────────────────────────────────────────────────────────

def _fmt_eur(valor) -> str:
    """Formatea un número como moneda española: 1.234,56 €"""
    if valor is None:
        return "—"
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "—"
    txt = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{txt} €"


@st.cache_data(ttl=60)
def _detectar_trimestres() -> list[dict]:
    """Escanea carpetas de Dropbox y devuelve trimestres disponibles."""
    if not ruta_existe_seguro(DROPBOX_BASE):
        return []
    trimestres: list[dict] = []
    for year_dir in sorted(DROPBOX_BASE.glob("FACTURAS *"), reverse=True):
        m = re.search(r"FACTURAS\s+(\d{4})", year_dir.name)
        if not m:
            continue
        year = m.group(1)
        recibidas = year_dir / "FACTURAS RECIBIDAS"
        if not ruta_existe_seguro(recibidas):
            continue
        # Buscar ambos patrones: "N TRI YYYY" (2025) y "N TRIMESTRE YYYY" (2026+)
        tri_dirs = sorted(recibidas.glob("* TRI*"), reverse=True)
        for tri_dir in tri_dirs:
            if not tri_dir.is_dir():
                continue
            tm = re.search(r"(\d)\s*TRI", tri_dir.name)
            if not tm:
                continue
            num = tm.group(1)
            label = f"{num}T{year[2:]}"
            atrasadas = tri_dir / "ATRASADAS"
            pdfs = list(tri_dir.glob("*.pdf"))
            atrasadas_existe = ruta_existe_seguro(atrasadas)
            pdfs_a = list(atrasadas.glob("*.pdf")) if atrasadas_existe else []
            if not pdfs and not pdfs_a:
                continue
            trimestres.append(
                {
                    "year": year,
                    "label": label,
                    "display": tri_dir.name,
                    "path": str(tri_dir),
                    "tiene_atrasadas": atrasadas_existe and len(pdfs_a) > 0,
                    "path_atrasadas": str(atrasadas) if atrasadas_existe else None,
                    "num_pdfs": len(pdfs),
                    "num_atrasadas": len(pdfs_a),
                }
            )
    return trimestres


@st.cache_data(ttl=300)
def _cargar_indice() -> dict:
    """Carga el diccionario de categorías para el parseo."""
    if not ruta_existe_seguro(DICCIONARIO_PATH):
        return {}
    _, _, indice = cargar_diccionario(DICCIONARIO_PATH)
    return indice


def _icono_cuadre(cuadre: str) -> str:
    if cuadre == "OK":
        return "✅"
    try:
        val = abs(float(cuadre.replace("DESCUADRE_", "").replace(",", ".")))
        return "⚠️" if val < 1.0 else "❌"
    except (ValueError, AttributeError):
        return "❌"


def _extraer_texto_pdf(ruta_pdf: Path) -> str:
    """Extrae texto de un PDF con pdfplumber (solo modo local)."""
    try:
        import pdfplumber
        with pdfplumber.open(str(ruta_pdf)) as pdf:
            return '\n'.join(p.extract_text() or '' for p in pdf.pages)
    except Exception as e:
        return f"Error extrayendo texto: {e}"


def _resolver_ruta_pdf(archivo: str) -> Path | None:
    """Busca la ruta completa de un PDF en las carpetas del trimestre activo."""
    tri_data = st.session_state.get("parseo_tri")
    if not tri_data:
        return None
    ruta = Path(tri_data["path"]) / archivo
    if ruta_existe_seguro(ruta):
        return ruta
    # Buscar en ATRASADAS
    if tri_data.get("path_atrasadas"):
        ruta_a = Path(tri_data["path_atrasadas"]) / archivo
        if ruta_existe_seguro(ruta_a):
            return ruta_a
    return None


# ── UI ────────────────────────────────────────────────────────────────────────

st.header("🔍 Parseo de Facturas")

if not PARSEO_DISPONIBLE:
    st.warning(
        "⚠️ Motor de parseo local no disponible — requiere acceso al PC "
        "con la carpeta Parseo."
    )
    uploaded = st.file_uploader(
        "Sube PDFs para parseo básico", type=["pdf"], accept_multiple_files=True
    )
    if uploaded:
        st.info(
            "El parseo con extractores completos solo funciona en modo local. "
            "Los archivos subidos se guardarían para procesamiento posterior."
        )
    st.stop()

# ── Controles ─────────────────────────────────────────────────────────────────

trimestres = _detectar_trimestres()

if not trimestres:
    st.error("No se encontraron carpetas de facturas en Dropbox.")
    st.stop()

# Dropdown 1: Año
anyos = sorted({t["year"] for t in trimestres}, reverse=True)
anyo = st.selectbox("Año", anyos)

# Dropdown 2: Trimestre (filtrado por año)
tris_anyo = [t for t in trimestres if t["year"] == anyo]
opc_tri = [f"{t['label']}  ({t['num_pdfs']} facturas)" for t in tris_anyo]
idx = st.selectbox("Trimestre", range(len(opc_tri)), format_func=lambda i: opc_tri[i])
tri = tris_anyo[idx]

# Selección atrasadas
opc_atrasadas = ["Solo trimestre"]
if tri["tiene_atrasadas"]:
    opc_atrasadas += [
        f"Solo ATRASADAS ({tri['num_atrasadas']})",
        f"Trimestre + ATRASADAS ({tri['num_pdfs']} + {tri['num_atrasadas']})",
    ]
sel_atrasadas = st.radio("Carpeta", opc_atrasadas, horizontal=True) if len(opc_atrasadas) > 1 else opc_atrasadas[0]

# Recopilar PDFs según selección
archivos: list[Path] = []
if "Solo ATRASADAS" not in sel_atrasadas:
    archivos += sorted(Path(tri["path"]).glob("*.pdf"))
if "ATRASADAS" in sel_atrasadas and "Solo trimestre" not in sel_atrasadas and tri["path_atrasadas"]:
    archivos += sorted(Path(tri["path_atrasadas"]).glob("*.pdf"))

modo = st.radio(
    "Facturas a procesar",
    ["Todas", "Factura concreta"],
    horizontal=True,
)

if modo == "Factura concreta":
    nombres = [p.name for p in archivos]
    sel = st.selectbox("Seleccionar PDF", nombres)
    archivos = [p for p in archivos if p.name == sel]

st.caption(f"{len(archivos)} archivo(s) seleccionado(s)")

# ── Ejecutar parseo ──────────────────────────────────────────────────────────

if st.button("▶️ Ejecutar Parseo", type="primary", use_container_width=True):
    if not archivos:
        st.warning("No hay PDFs para procesar.")
        st.stop()

    indice = _cargar_indice()
    resultados = []
    barra = st.progress(0, text="Parseando facturas...")

    for i, pdf in enumerate(archivos):
        try:
            factura = procesar_factura(pdf, indice)
        except Exception as exc:
            # Crear factura de error sin romper el lote
            from nucleo.factura import Factura  # type: ignore

            factura = Factura(archivo=pdf.name, numero="")
            factura.errores.append(str(exc))
            factura.cuadre = "ERROR"
        resultados.append(factura)
        barra.progress(
            (i + 1) / len(archivos),
            text=f"[{i + 1}/{len(archivos)}] {pdf.name[:50]}",
        )

    barra.empty()
    st.session_state["parseo_resultados"] = resultados
    st.session_state["parseo_trimestre"] = tri["label"]
    st.session_state["parseo_tri"] = tri

    # Guardar snapshot del parseo original
    try:
        from nucleo.aprendizaje import guardar_snapshot
        snapshot_path = guardar_snapshot(resultados, tri["label"])
        st.caption(f"Snapshot guardado: {snapshot_path.name}")
    except Exception:
        pass  # No bloquear el parseo si falla el snapshot

# ── Resultados ────────────────────────────────────────────────────────────────

resultados = st.session_state.get("parseo_resultados")
if not resultados:
    st.stop()

trimestre_label = st.session_state.get("parseo_trimestre", "")

st.divider()
st.subheader(f"Resultados — {trimestre_label}")

# Métricas resumen
total_ok = sum(1 for f in resultados if f.cuadre == "OK")
total_desc = sum(1 for f in resultados if f.cuadre.startswith("DESCUADRE"))
total_err = len(resultados) - total_ok - total_desc
total_lineas = sum(f.num_lineas for f in resultados)

c1, c2, c3, c4 = st.columns(4)
c1.metric("Total", len(resultados))
c2.metric("OK", total_ok)
c3.metric("Descuadre", total_desc)
c4.metric("Errores / Sin extractor", total_err)
st.metric("Líneas extraídas", total_lineas)

# Panel de sugerencias de correcciones anteriores
try:
    from nucleo.aprendizaje import obtener_sugerencias

    todas_sugerencias = []
    for r in resultados:
        sugs = obtener_sugerencias(r.lineas, r.proveedor)
        todas_sugerencias.extend(sugs)

    if todas_sugerencias:
        st.markdown("---")
        st.subheader(f"💡 {len(todas_sugerencias)} sugerencias de correcciones anteriores")

        df_sug = pd.DataFrame(todas_sugerencias)
        df_sug_display = df_sug[['proveedor', 'articulo', 'campo', 'valor_original', 'valor_sugerido', 'confianza']]
        df_sug_display.columns = ['Proveedor', 'Artículo', 'Campo', 'Antes', 'Sugerido', 'Confianza']

        st.dataframe(df_sug_display, use_container_width=True, hide_index=True)

        col1, col2 = st.columns(2)
        if col1.button("✅ Aplicar sugerencias", key="btn_aplicar_sug"):
            for sug in todas_sugerencias:
                idx = sug['indice']
                campo = sug['campo']
                for r in resultados:
                    if r.proveedor.upper() == sug['proveedor'].upper():
                        if idx < len(r.lineas):
                            setattr(r.lineas[idx], campo, sug['valor_sugerido'])

            from nucleo.aprendizaje import confirmar_correcciones
            proveedores_arts = {}
            for sug in todas_sugerencias:
                prov = sug['proveedor']
                if prov not in proveedores_arts:
                    proveedores_arts[prov] = []
                proveedores_arts[prov].append(sug['articulo'])
            for prov, arts in proveedores_arts.items():
                confirmar_correcciones(prov, arts)

            st.success(f"Aplicadas {len(todas_sugerencias)} correcciones")
            st.rerun()

        if col2.button("❌ Ignorar", key="btn_ignorar_sug"):
            pass

except ImportError:
    pass  # Parseo no disponible (modo cloud)
except Exception as e:
    st.caption(f"Sugerencias no disponibles: {e}")

# Tabla resumen
filas_resumen = []
for f in resultados:
    filas_resumen.append(
        {
            "Archivo": f.archivo,
            "Proveedor": f.proveedor,
            "Líneas": f.num_lineas,
            "Total factura": _fmt_eur(f.total),
            "Total calculado": _fmt_eur(f.total_calculado),
            "Cuadre": f"{_icono_cuadre(f.cuadre)} {f.cuadre}",
            "Extractor": getattr(f, "extractor_nombre", "—"),
        }
    )

df_resumen = pd.DataFrame(filas_resumen)
st.dataframe(df_resumen, use_container_width=True, hide_index=True)

# Detalle expandible por factura
st.subheader("Detalle por factura")
for f in resultados:
    label = f"{_icono_cuadre(f.cuadre)} {f.archivo} — {f.proveedor}"
    with st.expander(label):
        mc1, mc2, mc3 = st.columns(3)
        mc1.write(f"**Fecha:** {f.fecha or '—'}")
        mc2.write(f"**Ref:** {f.referencia or '—'}")
        mc3.write(f"**CIF:** {f.cif or '—'}")

        if f.tiene_lineas:
            lineas_data = []
            for ln in f.lineas:
                lineas_data.append(
                    {
                        "Artículo": ln.articulo,
                        "Categoría": ln.categoria,
                        "Cant.": ln.cantidad if ln.cantidad else "",
                        "Precio ud.": _fmt_eur(ln.precio_ud),
                        "Base": _fmt_eur(ln.base),
                        "IVA %": ln.iva,
                        "Cuota IVA": _fmt_eur(ln.cuota_iva),
                        "Total": _fmt_eur(ln.total),
                    }
                )
            st.dataframe(
                pd.DataFrame(lineas_data),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.warning("Sin líneas extraídas")

        if f.errores:
            for err in f.errores:
                st.error(err)

        if f.texto_raw:
            with st.expander("📄 Texto raw"):
                st.code(f.texto_raw[:3000], language=None)

# ── Requiere atención (panel de revisión interactivo) ───────────────────────

st.divider()

# Inicializar ignorados
if 'ignorados' not in st.session_state:
    st.session_state.ignorados = set()
_ignorados = st.session_state.ignorados

# Clasificar problemas — guardar objetos Factura completos
_desc_mayor: list = []
_desc_menor: list = []
_sin_total: list = []
_pdf_vacio: list = []
_sin_lineas: list = []

for f in resultados:
    if f.archivo in _ignorados:
        continue
    cuadre = f.cuadre

    if cuadre.startswith("DESCUADRE"):
        try:
            dv = abs(float(cuadre.replace("DESCUADRE_", "").replace(",", ".")))
        except ValueError:
            dv = 999
        if dv >= 1.0:
            _desc_mayor.append(f)
        else:
            _desc_menor.append(f)
    if f.total is None:
        _sin_total.append(f)
    if not f.texto_raw:
        _pdf_vacio.append(f)
    elif not f.tiene_lineas and cuadre != "OK":
        _sin_lineas.append(f)

_grupos_check = [_desc_mayor, _desc_menor, _sin_total, _pdf_vacio, _sin_lineas]
_hay_problemas = any(_grupos_check)

if _hay_problemas:
    st.subheader("⚠️ Requiere atención")

    # ── DESCUADRE > 1 € ─────────────────────────────────────────────────
    if _desc_mayor:
        with st.expander(f"🔴 DESCUADRE > 1 € ({len(_desc_mayor)})", expanded=True):
            for f in _desc_mayor:
                archivo_key = f.archivo.replace(".", "_").replace(" ", "_")
                suma_lineas = f.total_calculado
                diff = abs((f.total or 0) - suma_lineas)

                with st.expander(f"❌ {f.archivo} — {f.proveedor} — diff {_fmt_eur(diff)}"):
                    # Métricas
                    mc1, mc2, mc3 = st.columns(3)
                    mc1.metric("Total factura", _fmt_eur(f.total))
                    mc2.metric("Total calculado", _fmt_eur(suma_lineas))
                    mc3.metric("Diferencia", _fmt_eur(diff))

                    # Tabla de líneas
                    if f.tiene_lineas:
                        lineas_data = [{
                            "Artículo": ln.articulo,
                            "Cant.": ln.cantidad or "",
                            "Precio ud.": _fmt_eur(ln.precio_ud),
                            "Base": _fmt_eur(ln.base),
                            "IVA %": ln.iva,
                            "Total": _fmt_eur(ln.total),
                        } for ln in f.lineas]
                        st.dataframe(pd.DataFrame(lineas_data), use_container_width=True, hide_index=True)

                    # Texto crudo del PDF
                    ruta_pdf = _resolver_ruta_pdf(f.archivo)
                    texto_pdf = ""
                    if ruta_pdf:
                        texto_pdf = f.texto_raw or _extraer_texto_pdf(ruta_pdf)
                        if texto_pdf:
                            with st.expander("📄 Texto crudo del PDF"):
                                st.code(texto_pdf[:3000], language=None)

                    # Botones de acción
                    bc1, bc2, bc3 = st.columns(3)
                    if bc1.button("🔧 Generar fix", key=f"fix_{archivo_key}"):
                        prompt = (
                            f"Debug del extractor de {f.proveedor}.\n"
                            f"Archivo: {f.archivo}\n"
                            f"Total factura: {f.total}\n"
                            f"Total calculado: {suma_lineas}\n"
                            f"Diferencia: {diff:.2f}\n"
                            f"Líneas extraídas: {f.num_lineas}\n"
                            f"Texto del PDF:\n{(texto_pdf or f.texto_raw or '')[:3000]}\n\n"
                            f"Investiga por qué hay descuadre y corrige el extractor."
                        )
                        ruta_fix = Path(f"tasks/fix_{f.proveedor.replace(' ', '_')}.md")
                        ruta_fix.parent.mkdir(exist_ok=True)
                        ruta_fix.write_text(prompt, encoding='utf-8')
                        st.success(f"Guardado: {ruta_fix}")
                        st.code(f'claude "Lee {ruta_fix} y ejecuta"', language='bash')

                    if bc2.button("📄 Abrir PDF", key=f"open_{archivo_key}"):
                        if ruta_pdf:
                            st.code(str(ruta_pdf), language='text')
                        else:
                            st.warning("Ruta PDF no disponible")

                    if bc3.button("⏭️ Ignorar", key=f"skip_{archivo_key}"):
                        st.session_state.ignorados.add(f.archivo)
                        st.rerun()

    # ── DESCUADRE < 1 € ─────────────────────────────────────────────────
    if _desc_menor:
        with st.expander(f"🟡 DESCUADRE < 1 € ({len(_desc_menor)})"):
            filas = [{
                "Archivo": f.archivo,
                "Proveedor": f.proveedor,
                "Total factura": _fmt_eur(f.total),
                "Total calculado": _fmt_eur(f.total_calculado),
                "Diferencia": _fmt_eur(abs((f.total or 0) - f.total_calculado)),
            } for f in _desc_menor]
            st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    # ── SIN TOTAL ────────────────────────────────────────────────────────
    if _sin_total:
        with st.expander(f"🔴 SIN TOTAL ({len(_sin_total)})", expanded=True):
            for f in _sin_total:
                archivo_key = f.archivo.replace(".", "_").replace(" ", "_")

                with st.expander(f"❌ {f.archivo} — {f.proveedor}"):
                    # Mostrar líneas si hay
                    if f.tiene_lineas:
                        st.caption(f"{f.num_lineas} líneas extraídas — total calculado: {_fmt_eur(f.total_calculado)}")
                        lineas_data = [{
                            "Artículo": ln.articulo,
                            "Base": _fmt_eur(ln.base),
                            "IVA %": ln.iva,
                            "Total": _fmt_eur(ln.total),
                        } for ln in f.lineas]
                        st.dataframe(pd.DataFrame(lineas_data), use_container_width=True, hide_index=True)

                    # Texto crudo
                    ruta_pdf = _resolver_ruta_pdf(f.archivo)
                    texto_pdf = f.texto_raw or ""
                    if ruta_pdf and not texto_pdf:
                        texto_pdf = _extraer_texto_pdf(ruta_pdf)
                    if texto_pdf:
                        with st.expander("📄 Texto crudo del PDF"):
                            st.code(texto_pdf[:2000], language=None)

                    # Botones
                    bc1, bc2, bc3 = st.columns(3)
                    if bc1.button("🔧 Generar fix", key=f"fixtot_{archivo_key}"):
                        prompt = (
                            f"El extractor de {f.proveedor} no extrae el total de la factura.\n"
                            f"Archivo: {f.archivo}\n"
                            f"Líneas extraídas: {f.num_lineas}\n"
                            f"Texto del PDF:\n{(texto_pdf)[:3000]}\n\n"
                            f"Corrige extraer_total() en el extractor."
                        )
                        ruta_fix = Path(f"tasks/fix_total_{f.proveedor.replace(' ', '_')}.md")
                        ruta_fix.parent.mkdir(exist_ok=True)
                        ruta_fix.write_text(prompt, encoding='utf-8')
                        st.success(f"Guardado: {ruta_fix}")

                    if bc2.button("📝 Rellenar total", key=f"filltot_{archivo_key}"):
                        st.session_state[f"show_fill_total_{archivo_key}"] = True

                    if bc3.button("⏭️ Ignorar", key=f"skiptot_{archivo_key}"):
                        st.session_state.ignorados.add(f.archivo)
                        st.rerun()

                    # Input manual de total
                    if st.session_state.get(f"show_fill_total_{archivo_key}"):
                        total_manual = st.number_input(
                            "Total (€)", min_value=0.0, step=0.01,
                            key=f"input_total_{archivo_key}"
                        )
                        if st.button("💾 Guardar total", key=f"savetot_{archivo_key}"):
                            f.total = total_manual
                            st.success(f"Total asignado: {_fmt_eur(total_manual)}")

    # ── PDF VACÍO ────────────────────────────────────────────────────────
    if _pdf_vacio:
        with st.expander(f"🔴 PDF VACÍO ({len(_pdf_vacio)})"):
            filas = [{"Archivo": f.archivo, "Proveedor": f.proveedor} for f in _pdf_vacio]
            st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)

    # ── SIN LÍNEAS ───────────────────────────────────────────────────────
    if _sin_lineas:
        with st.expander(f"🟡 SIN LÍNEAS ({len(_sin_lineas)})", expanded=True):
            for f in _sin_lineas:
                archivo_key = f.archivo.replace(".", "_").replace(" ", "_")
                ext_nombre = getattr(f, "extractor_nombre", "")
                tiene_ext = ext_nombre and ext_nombre != "—" and "genérico" not in ext_nombre.lower()

                # Detectar si tiene extractor via módulo Parseo
                try:
                    from extractores import tiene_extractor as _tiene_ext_fn
                    tiene_ext = _tiene_ext_fn(f.proveedor)
                except ImportError:
                    pass

                tag = "🔵 Tiene extractor" if tiene_ext else "⚪ Sin extractor"
                with st.expander(f"{tag} | {f.archivo} — {f.proveedor}"):
                    ruta_pdf = _resolver_ruta_pdf(f.archivo)
                    texto_pdf = f.texto_raw or ""
                    if ruta_pdf and not texto_pdf:
                        texto_pdf = _extraer_texto_pdf(ruta_pdf)

                    if tiene_ext:
                        # CON extractor pero sin líneas → debug
                        bc1, bc2, bc3 = st.columns(3)
                        if bc1.button("🔧 Debug extractor", key=f"dbg_{archivo_key}"):
                            prompt = (
                                f"El extractor de {f.proveedor} ({ext_nombre}) no extrae líneas.\n"
                                f"Archivo: {f.archivo}\n"
                                f"Texto del PDF:\n{texto_pdf[:3000]}\n\n"
                                f"Investiga por qué extraer_lineas() devuelve vacío y corrige."
                            )
                            ruta_fix = Path(f"tasks/fix_{f.proveedor.replace(' ', '_')}.md")
                            ruta_fix.parent.mkdir(exist_ok=True)
                            ruta_fix.write_text(prompt, encoding='utf-8')
                            st.success(f"Guardado: {ruta_fix}")

                        if bc2.button("📝 Rellenar mínimo", key=f"fillmin_{archivo_key}"):
                            st.write(f"**Proveedor:** {f.proveedor}")
                            st.write(f"**Fecha:** {f.fecha or '—'}")
                            st.write(f"**Total:** {_fmt_eur(f.total)}")
                            st.write(f"**Ref:** {f.referencia or '—'}")

                        if bc3.button("⏭️ Ignorar", key=f"skiplin_{archivo_key}"):
                            st.session_state.ignorados.add(f.archivo)
                            st.rerun()
                    else:
                        # SIN extractor → buscar/crear/rellenar
                        bc1, bc2, bc3, bc4 = st.columns(4)

                        if bc1.button("🔍 Buscar extractor", key=f"search_{archivo_key}"):
                            st.session_state[f"show_search_{archivo_key}"] = True

                        if bc2.button("➕ Crear extractor", key=f"create_{archivo_key}"):
                            prompt = (
                                f"Crear extractor nuevo para {f.proveedor}.\n"
                                f"Archivo ejemplo: {f.archivo}\n"
                                f"Texto del PDF:\n{texto_pdf[:3000]}\n\n"
                                f"Sigue /extractor para crear el extractor completo."
                            )
                            ruta_fix = Path(f"tasks/fix_nuevo_{f.proveedor.replace(' ', '_')}.md")
                            ruta_fix.parent.mkdir(exist_ok=True)
                            ruta_fix.write_text(prompt, encoding='utf-8')
                            st.success(f"Guardado: {ruta_fix}")
                            st.code(f'claude "Lee {ruta_fix} y ejecuta"', language='bash')

                        if bc3.button("📝 Rellenar mínimo", key=f"fillmin2_{archivo_key}"):
                            st.write(f"**Proveedor:** {f.proveedor}")
                            st.write(f"**Fecha:** {f.fecha or '—'}")
                            st.write(f"**Total:** {_fmt_eur(f.total)}")
                            st.write(f"**Ref:** {f.referencia or '—'}")

                        if bc4.button("⏭️ Ignorar", key=f"skipnew_{archivo_key}"):
                            st.session_state.ignorados.add(f.archivo)
                            st.rerun()

                        # Panel de búsqueda de extractor (TAREA 3)
                        if st.session_state.get(f"show_search_{archivo_key}"):
                            try:
                                from extractores import listar_extractores as _listar, obtener_extractor as _obtener

                                exts_disponibles = _listar()
                                nombres_ext = sorted(set(
                                    cls.nombre for cls in exts_disponibles.values()
                                    if hasattr(cls, 'nombre') and cls.nombre
                                ))

                                extractor_elegido = st.selectbox(
                                    "Elige un extractor para probar:",
                                    options=["(seleccionar)"] + nombres_ext,
                                    key=f"select_ext_{archivo_key}"
                                )

                                if extractor_elegido and extractor_elegido != "(seleccionar)":
                                    if st.button("▶️ Probar extractor", key=f"try_{archivo_key}"):
                                        ext_obj = _obtener(extractor_elegido)
                                        if ext_obj and texto_pdf:
                                            lineas_test = ext_obj.extraer_lineas(texto_pdf)
                                            total_test = ext_obj.extraer_total(texto_pdf)
                                            fecha_test = ext_obj.extraer_fecha(texto_pdf)
                                            ref_test = ext_obj.extraer_referencia(texto_pdf)

                                            if lineas_test:
                                                suma_test = sum(
                                                    l.get('base', 0) * (1 + l.get('iva', 21) / 100)
                                                    for l in lineas_test
                                                )
                                                cuadre_test = abs((total_test or 0) - suma_test)

                                                if cuadre_test < 0.05:
                                                    st.success(f"✅ Cuadre OK — {len(lineas_test)} líneas, total {_fmt_eur(suma_test)}")
                                                elif cuadre_test < 1:
                                                    st.warning(f"⚠️ Descuadre menor — {len(lineas_test)} líneas, diff {_fmt_eur(cuadre_test)}")
                                                else:
                                                    st.error(f"❌ Descuadre — {len(lineas_test)} líneas, diff {_fmt_eur(cuadre_test)}")

                                                df_test = pd.DataFrame(lineas_test)
                                                st.dataframe(df_test, use_container_width=True, hide_index=True)
                                                st.caption(f"Fecha: {fecha_test} | Ref: {ref_test} | Total: {_fmt_eur(total_test)}")

                                                st.session_state[f"test_ok_{archivo_key}"] = {
                                                    'extractor': extractor_elegido,
                                                    'lineas': len(lineas_test),
                                                    'cuadre': cuadre_test,
                                                    'archivo_nombre': f.proveedor,
                                                }
                                            else:
                                                st.error(f"❌ El extractor {extractor_elegido} no extrajo líneas de este PDF")
                                        elif not texto_pdf:
                                            st.error("No hay texto disponible del PDF")

                                # Botón añadir alias (si test fue OK)
                                test_key = f"test_ok_{archivo_key}"
                                if test_key in st.session_state:
                                    test = st.session_state[test_key]
                                    st.success(f"Extractor {test['extractor']} funciona: {test['lineas']} líneas, cuadre {_fmt_eur(test['cuadre'])}")

                                    if st.button("✅ Añadir alias al extractor", key=f"alias_{archivo_key}"):
                                        prompt = (
                                            f"En el extractor de {test['extractor']} en "
                                            f"C:\\_ARCHIVOS\\TRABAJO\\Facturas\\Parseo\\extractores\\,\n"
                                            f"añade el alias '{test['archivo_nombre']}' al decorador @registrar.\n"
                                            f"También añade '{test['archivo_nombre']}' a ALIAS_DICCIONARIO en main.py si es necesario.\n\n"
                                            f"Contexto: El archivo '{f.archivo}' no fue reconocido porque el nombre\n"
                                            f"'{test['archivo_nombre']}' no matcheaba con el extractor '{test['extractor']}'.\n"
                                            f"Al probar manualmente, el extractor funciona perfectamente "
                                            f"({test['lineas']} líneas, cuadre {test['cuadre']:.2f}).\n"
                                            f"Solo falta el alias."
                                        )
                                        ruta_fix = Path(f"tasks/fix_alias_{test['archivo_nombre'].replace(' ', '_')}.md")
                                        ruta_fix.parent.mkdir(exist_ok=True)
                                        ruta_fix.write_text(prompt, encoding='utf-8')
                                        st.success(f"Guardado: {ruta_fix}")
                                        st.code(f'claude "Lee {ruta_fix} y ejecuta"', language='bash')

                            except ImportError:
                                st.warning("Módulo extractores no disponible (modo cloud)")

                    # Texto crudo compartido
                    if texto_pdf:
                        with st.expander("📄 Texto crudo del PDF"):
                            st.code(texto_pdf[:2000], language=None)

    # ── Archivos ignorados ───────────────────────────────────────────────
    if _ignorados:
        with st.expander(f"⏭️ Ignorados ({len(_ignorados)})"):
            for arch in sorted(_ignorados):
                st.text(arch)
            if st.button("🔄 Restaurar todos", key="btn_restaurar_ignorados"):
                st.session_state.ignorados.clear()
                st.rerun()

else:
    st.success("✅ Todas las facturas parseadas correctamente")

# ── Export Excel ──────────────────────────────────────────────────────────────

st.divider()

filas_lineas = []
filas_facturas = []
for i, f in enumerate(resultados, 1):
    filas_facturas.append(
        {
            "#": i,
            "ARCHIVO": f.archivo,
            "PROVEEDOR": f.proveedor,
            "FECHA": f.fecha,
            "REF": f.referencia,
            "CIF": f.cif,
            "TOTAL_FACTURA": f.total,
            "TOTAL_CALCULADO": round(f.total_calculado, 2),
            "CUADRE": f.cuadre,
            "EXTRACTOR": getattr(f, "extractor_nombre", ""),
            "LÍNEAS": f.num_lineas,
            "ERRORES": "; ".join(f.errores) if f.errores else "",
        }
    )
    for ln in f.lineas:
        filas_lineas.append(
            {
                "#": i,
                "ARCHIVO": f.archivo,
                "PROVEEDOR": f.proveedor,
                "CÓDIGO": ln.codigo,
                "ARTÍCULO": ln.articulo,
                "CATEGORÍA": ln.categoria,
                "CANTIDAD": ln.cantidad,
                "PRECIO_UD": ln.precio_ud,
                "TIPO_IVA": ln.iva,
                "BASE": round(ln.base, 2),
                "CUOTA_IVA": round(ln.cuota_iva, 2),
                "TOTAL": round(ln.total, 2),
                "MATCH": getattr(ln, "match_info", ""),
            }
        )

df_lineas = pd.DataFrame(filas_lineas)
df_facturas = pd.DataFrame(filas_facturas)


def _formatear_excel(buffer: io.BytesIO, df_lin: pd.DataFrame, df_fac: pd.DataFrame):
    """Genera Excel profesional con formato corporativo Tasca Barea."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    FONT_BASE = Font(name="Arial", size=11)
    FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    FILL_HEADER = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    FILL_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_WARN = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FILL_ERR = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    BORDER_THIN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    FMT_EUR = '#,##0.00\\ "€"'
    FMT_PCT = '0%'

    # Columnas de moneda por hoja
    COLS_EUR_LINEAS = {"PRECIO_UD", "BASE", "CUOTA_IVA", "TOTAL"}
    COLS_EUR_FACTURAS = {"TOTAL_FACTURA", "TOTAL_CALCULADO"}
    COLS_PCT = {"TIPO_IVA"}

    wb = Workbook()

    def _escribir_hoja(ws, df: pd.DataFrame, nombre_tabla: str, cols_eur: set):
        # Cabeceras
        headers = list(df.columns)
        for c, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER_THIN

        # Datos
        for r, row in enumerate(df.itertuples(index=False), 2):
            for c, (col_name, val) in enumerate(zip(headers, row), 1):
                cell = ws.cell(row=r, column=c)
                # Convertir NaN/None
                if pd.isna(val) if not isinstance(val, str) else False:
                    cell.value = None
                else:
                    cell.value = val
                cell.font = FONT_BASE
                cell.border = BORDER_THIN

                # Formato moneda
                if col_name in cols_eur and isinstance(val, (int, float)):
                    cell.number_format = FMT_EUR
                # Formato porcentaje (IVA viene como entero 4/10/21 → dividir entre 100)
                elif col_name in COLS_PCT and isinstance(val, (int, float)):
                    cell.value = val / 100
                    cell.number_format = FMT_PCT

                # Condicional CUADRE
                if col_name == "CUADRE" and isinstance(val, str):
                    if val == "OK":
                        cell.fill = FILL_OK
                    elif val.startswith("DESCUADRE"):
                        try:
                            dv = abs(float(val.replace("DESCUADRE_", "").replace(",", ".")))
                            cell.fill = FILL_WARN if dv < 1.0 else FILL_ERR
                        except ValueError:
                            cell.fill = FILL_ERR
                    elif val:
                        cell.fill = FILL_ERR

        # Tabla Excel
        if len(df) > 0:
            last_col = get_column_letter(len(headers))
            ref = f"A1:{last_col}{len(df) + 1}"
            tabla = Table(displayName=nombre_tabla, ref=ref)
            tabla.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            ws.add_table(tabla)

        # Anchos automáticos
        for c, col_name in enumerate(headers, 1):
            max_len = len(str(col_name))
            for r in range(2, min(len(df) + 2, 102)):  # muestreo hasta 100 filas
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(c)].width = max(10, min(max_len + 3, 40))

        # Freeze panes
        ws.freeze_panes = "A2"

    # Hoja Lineas
    ws_lin = wb.active
    ws_lin.title = "Lineas"
    _escribir_hoja(ws_lin, df_lin, "Lineas", COLS_EUR_LINEAS)

    # Hoja Facturas
    ws_fac = wb.create_sheet("Facturas")
    _escribir_hoja(ws_fac, df_fac, "Facturas", COLS_EUR_FACTURAS)

    wb.save(buffer)


buffer = io.BytesIO()
_formatear_excel(buffer, df_lineas, df_facturas)

st.download_button(
    "📥 Descargar Excel",
    data=buffer.getvalue(),
    file_name=f"Facturas_{trimestre_label}_{datetime.now().strftime('%Y%m%d')}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

# ── Aprender de correcciones ────────────────────────────────────────────────

st.markdown("---")
st.subheader("📚 Aprender de correcciones")
st.caption("Sube el Excel COMPRAS que hayas corregido manualmente. El sistema detectará los cambios y los aplicará en futuros parseos.")

excel_corregido = st.file_uploader("Excel corregido", type=["xlsx"], key="uploader_aprender")

if excel_corregido:
    try:
        from nucleo.aprendizaje import detectar_correcciones, guardar_correcciones_nuevas

        # Buscar snapshot más reciente
        snapshots = sorted(SNAPSHOTS_DIR.glob("parseo_*.json"), reverse=True) if ruta_existe_seguro(SNAPSHOTS_DIR) else []

        if not snapshots:
            st.warning("No hay snapshots de parseos anteriores. Ejecuta un parseo primero.")
        else:
            snapshot_path = snapshots[0]
            st.caption(f"Comparando con: {snapshot_path.name}")

            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(excel_corregido.read())
                tmp_path = Path(tmp.name)

            correcciones = detectar_correcciones(snapshot_path, tmp_path)
            tmp_path.unlink()

            if correcciones:
                st.success(f"Detectadas {len(correcciones)} correcciones")

                for corr in correcciones:
                    cambios_str = ", ".join(f"{k}: {v}" for k, v in corr['cambios'].items())
                    st.markdown(f"- **{corr['proveedor']}** | {corr['articulo_original']} → {cambios_str}")

                if st.button("💾 Guardar correcciones", key="btn_guardar_corr"):
                    guardar_correcciones_nuevas(correcciones)
                    st.success(f"Guardadas {len(correcciones)} correcciones en correcciones.json")
                    st.balloons()
            else:
                st.info("No se detectaron diferencias entre el parseo original y el Excel corregido.")

    except ImportError:
        st.error("Módulo de aprendizaje no disponible (requiere acceso local a Parseo)")
    except Exception as e:
        st.error(f"Error: {e}")
