"""
Página Mov. Banco — Actualizar movimientos bancarios desde archivos Sabadell .xls
"""

import os
import sys
import tempfile
from pathlib import Path

import streamlit as st

from utils.auth import require_role
from utils.entorno import ruta_existe_seguro

require_role(["admin"])

# ── Importar script de actualización ─────────────────────────────────────────

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'scripts'))
from actualizar_movimientos import actualizar, leer_xls_sabadell, mostrar_info, detectar_cuenta  # type: ignore

CONSOLIDADO = Path(os.path.join(os.path.dirname(__file__), '..', '..', 'datos', 'Movimientos_Cuenta_26.xlsx'))
AÑO_CONSOLIDADO = 2026  # Extraído del nombre del archivo (_26)


# ── Helpers ──────────────────────────────────────────────────────────────────

def _fmt_eur(valor) -> str:
    """Formatea un número como moneda española: 1.234,56 €"""
    if valor is None:
        return "—"
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return "—"
    txt = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{txt} €"


def _info_consolidado() -> dict:
    """Lee estado actual del consolidado."""
    info = {"existe": False, "pestañas": []}
    if not ruta_existe_seguro(CONSOLIDADO):
        return info
    info["existe"] = True
    from openpyxl import load_workbook
    wb = load_workbook(str(CONSOLIDADO), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n = ws.max_row - 1
        if n < 1:
            continue
        primera = ws.cell(2, 2).value
        ultima = ws.cell(ws.max_row, 2).value
        saldo_actual = ws.cell(ws.max_row, 6).value
        info["pestañas"].append({
            "nombre": sheet_name,
            "filas": n,
            "primera": primera,
            "ultima": ultima,
            "saldo": saldo_actual,
        })
    return info


def _verificar_continuidad_saldos() -> dict:
    """Verifica que saldo_anterior + importe = saldo_actual en cada fila."""
    from openpyxl import load_workbook
    wb = load_workbook(str(CONSOLIDADO), data_only=True)
    resultado = {}
    for s in wb.sheetnames:
        ws = wb[s]
        errores = 0
        for r in range(3, ws.max_row + 1):
            sa = float(ws.cell(r - 1, 6).value or 0)
            imp = float(ws.cell(r, 5).value or 0)
            sc = float(ws.cell(r, 6).value or 0)
            if abs(round(sa + imp, 2) - sc) > 0.01:
                errores += 1
        resultado[s] = {"filas": ws.max_row - 1, "errores": errores}
    return resultado


# ── UI ───────────────────────────────────────────────────────────────────────

st.header("🏦 Actualizar Movimientos Banco")

# ── Estado actual del consolidado ────────────────────────────────────────────

info = _info_consolidado()

if info["existe"]:
    st.subheader("Estado actual")
    cols = st.columns(len(info["pestañas"]))
    for col, p in zip(cols, info["pestañas"]):
        col.metric(p["nombre"], f"{p['filas']} movimientos")
        col.caption(f"{p['primera']:%d/%m/%Y} → {p['ultima']:%d/%m/%Y}")
        col.caption(f"Saldo: {_fmt_eur(p['saldo'])}")
else:
    st.info("El consolidado no existe todavía. Se creará al subir los primeros archivos.")

st.divider()

# ── Subir archivos ───────────────────────────────────────────────────────────

st.subheader("Subir archivos del banco")
uploaded_files = st.file_uploader(
    "Archivos .xls de Banco Sabadell",
    type=["xls"],
    accept_multiple_files=True,
    key="upload_xls"
)

if uploaded_files:
    # Guardar temporalmente y analizar
    tmp_paths = []
    analisis = []
    hay_errores = False
    for uf in uploaded_files:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xls')
        tmp.write(uf.read())
        tmp.close()
        tmp_paths.append(tmp.name)

        try:
            cuenta, filas = leer_xls_sabadell(tmp.name, año_esperado=AÑO_CONSOLIDADO)
            primera = filas[0]["F. Operativa"] if filas else None
            ultima = filas[-1]["F. Operativa"] if filas else None
            analisis.append({
                "archivo": uf.name,
                "cuenta": cuenta,
                "movimientos": len(filas),
                "primera": primera,
                "ultima": ultima,
                "tmp": tmp.name,
            })
        except ValueError as e:
            hay_errores = True
            analisis.append({
                "archivo": uf.name,
                "cuenta": "ERROR",
                "movimientos": 0,
                "primera": None,
                "ultima": None,
                "tmp": tmp.name,
                "error": str(e),
            })
        except Exception as e:
            hay_errores = True
            analisis.append({
                "archivo": uf.name,
                "cuenta": "ERROR",
                "movimientos": 0,
                "primera": None,
                "ultima": None,
                "tmp": tmp.name,
                "error": str(e),
            })

    # Mostrar preview
    st.subheader("Preview")
    for a in analisis:
        if a["cuenta"] == "ERROR":
            st.error(f"❌ {a['archivo']}: {a.get('error', 'Error desconocido')}")
        else:
            fecha_rango = ""
            if a["primera"] and a["ultima"]:
                fecha_rango = f" | {a['primera']:%d/%m/%Y} → {a['ultima']:%d/%m/%Y}"
            st.success(f"✅ {a['archivo']} → **{a['cuenta']}** ({a['movimientos']} movimientos{fecha_rango})")

    if hay_errores:
        st.error("Corrige los archivos con errores antes de actualizar.")
        st.button("Actualizar consolidado", disabled=True, use_container_width=True)

    # Dry-run
    if not hay_errores:
        st.markdown("---")
        st.subheader("Simulación (dry-run)")

    archivos_validos = [a["tmp"] for a in analisis if a["cuenta"] != "ERROR"]

    if archivos_validos and not hay_errores:
        res_dry = actualizar(archivos_validos, str(CONSOLIDADO), dry_run=True)

        c1, c2, c3 = st.columns(3)
        c1.metric("Nuevos Tasca", res_dry["nuevos_tasca"])
        c2.metric("Nuevos Comestibles", res_dry["nuevos_comestibles"])
        c3.metric("Duplicados ignorados", res_dry["duplicados_ignorados"])

        for aviso in res_dry["avisos"]:
            st.warning(aviso)

        # Botón actualizar
        st.markdown("---")
        total_nuevos = res_dry["nuevos_tasca"] + res_dry["nuevos_comestibles"]

        if total_nuevos == 0:
            st.info("No hay movimientos nuevos que incorporar (todos son duplicados).")
        else:
            if st.button(f"✅ Actualizar consolidado ({total_nuevos} nuevos)", type="primary", use_container_width=True):
                res = actualizar(archivos_validos, str(CONSOLIDADO), dry_run=False)

                st.success(
                    f"Actualización completada: "
                    f"Tasca +{res['nuevos_tasca']}, "
                    f"Comestibles +{res['nuevos_comestibles']}, "
                    f"{res['duplicados_ignorados']} duplicados ignorados"
                )

                if res["backup_path"]:
                    st.caption(f"Backup: {res['backup_path']}")

                for aviso in res["avisos"]:
                    st.warning(aviso)

                # Sync a Drive
                try:
                    from nucleo.sync_drive import sync_datos
                    synced = sync_datos()
                    st.success(f"{len(synced)} archivos sincronizados con Drive")
                except Exception as e:
                    st.warning(f"Sync Drive falló: {e}")

                # Verificar continuidad de saldos
                st.markdown("---")
                st.subheader("Verificación de integridad")
                verif = _verificar_continuidad_saldos()
                for nombre, datos in verif.items():
                    if datos["errores"] == 0:
                        st.success(f"✅ {nombre}: {datos['filas']} filas — continuidad de saldos OK")
                    else:
                        st.error(f"❌ {nombre}: {datos['filas']} filas — {datos['errores']} errores de continuidad")

    # Limpiar temporales
    for tmp in tmp_paths:
        try:
            os.unlink(tmp)
        except OSError:
            pass
