"""
Módulo de generación de Excel.

Genera archivos Excel con las facturas procesadas.

CAMBIOS v5.16 (07/01/2026):
- NUEVO: Columna ARCHIVO en hoja Facturas (después de #)
- NUEVO: Columna TOTAL FACTURA (total extraído del PDF)
- NUEVO: Columna Total ahora es el total calculado de líneas (base × IVA)
- Orden columnas Facturas: #, ARCHIVO, CUENTA, TITULO, Fec.Fac., REF, TOTAL FACTURA, Total, OBSERVACIONES

CAMBIOS v5.9 (02/01/2026):
- FIX: Sanitización de caracteres ilegales para Excel (IllegalCharacterError)
- Función sanitizar_para_excel() elimina caracteres de control

CAMBIOS v5.8 (02/01/2026):
- Nueva hoja "Facturas" con cabeceras (una fila por factura)
- Hoja "Lineas" renombrada (antes "Facturas")
- Integración con DiccionarioEmisorTitulo.xlsx para CUENTA y TITULO
- Auto-numeración TMP001, TMP002... para facturas sin número de gestoría
"""
import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional, TYPE_CHECKING
from datetime import datetime
from difflib import SequenceMatcher
import re

# Imports para formateo Excel
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from nucleo.factura import Factura


# ==============================================================================
# FORMATEO DE TABLAS EXCEL
# ==============================================================================

def formatear_hoja_como_tabla(worksheet, nombre_tabla: str, estilo: str = 'TableStyleMedium9'):
    """
    Formatea una hoja Excel como tabla con estilo y fuente Aptos Narrow 9.
    
    Args:
        worksheet: Hoja de openpyxl
        nombre_tabla: Nombre único para la tabla
        estilo: Estilo de tabla Excel
               - TableStyleMedium9 = Gris
               - TableStyleMedium16 = Morado claro
    """
    if worksheet.max_row < 1:
        return
    
    # Definir rango de la tabla
    max_col = worksheet.max_column
    max_row = worksheet.max_row
    
    # Crear referencia de rango (A1:X100)
    rango = f"A1:{get_column_letter(max_col)}{max_row}"
    
    # Crear tabla
    tabla = Table(displayName=nombre_tabla, ref=rango)
    
    # Estilo de tabla
    style = TableStyleInfo(
        name=estilo,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = style
    
    # Añadir tabla a la hoja
    worksheet.add_table(tabla)
    
    # Aplicar fuente Aptos Narrow 9 a todas las celdas
    fuente = Font(name='Aptos Narrow', size=9)
    fuente_cabecera = Font(name='Aptos Narrow', size=9, bold=True)
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col), 1):
        for cell in row:
            if row_idx == 1:
                cell.font = fuente_cabecera
            else:
                cell.font = fuente
    
    # Ajustar ancho de columnas (aproximado)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        # Ancho basado en el contenido de la cabecera + margen
        header_value = worksheet.cell(row=1, column=col_idx).value
        if header_value:
            width = max(len(str(header_value)) + 2, 8)
            worksheet.column_dimensions[col_letter].width = min(width, 50)


# ==============================================================================
# SANITIZACIÓN DE CARACTERES PARA EXCEL
# ==============================================================================

def sanitizar_para_excel(valor):
    """
    Elimina caracteres ilegales para Excel (caracteres de control).
    
    Excel no acepta caracteres de control (0x00-0x1F) excepto:
    - 0x09 (tab)
    - 0x0A (newline)
    - 0x0D (carriage return)
    
    Args:
        valor: Cualquier valor a sanitizar
        
    Returns:
        Valor sanitizado (string sin caracteres ilegales)
    """
    if valor is None:
        return valor
    if isinstance(valor, str):
        # Elimina caracteres de control excepto tab, newline, carriage return
        return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', valor)
    return valor


def sanitizar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Sanitiza todas las columnas de texto de un DataFrame.
    
    Args:
        df: DataFrame a sanitizar
        
    Returns:
        DataFrame con textos sanitizados
    """
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].apply(sanitizar_para_excel)
    return df


# ==============================================================================
# CONFIGURACIÓN DEL DICCIONARIO DE CUENTAS
# ==============================================================================

# Ruta por defecto al diccionario (se puede sobreescribir)
DICCIONARIO_CUENTAS_DEFAULT = r"C:\_ARCHIVOS\TRABAJO\Facturas\ParsearFacturas-main\datos\DiccionarioEmisorTitulo.xlsx"

# Caché global para evitar recargar el diccionario en cada llamada
_CACHE_CUENTAS: Dict[str, str] = {}  # TITULO -> CUENTA
_CACHE_ALIAS: Dict[str, str] = {}     # NOMBRE_EN_CONCEPTO -> TITULO_FACTURA
_CACHE_CARGADO: bool = False


def cargar_diccionario_cuentas(ruta: Optional[Path] = None) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Carga el diccionario de cuentas desde el Excel.
    
    Args:
        ruta: Ruta al archivo DiccionarioEmisorTitulo.xlsx
        
    Returns:
        Tupla (dict_cuentas, dict_alias)
        - dict_cuentas: {TITULO_NORMALIZADO: CUENTA}
        - dict_alias: {ALIAS_NORMALIZADO: TITULO}
    """
    global _CACHE_CUENTAS, _CACHE_ALIAS, _CACHE_CARGADO
    
    if _CACHE_CARGADO:
        return _CACHE_CUENTAS, _CACHE_ALIAS
    
    if ruta is None:
        ruta = Path(DICCIONARIO_CUENTAS_DEFAULT)
    
    dict_cuentas = {}
    dict_alias = {}
    
    try:
        # Cargar hoja de cuentas (Hoja1)
        df_cuentas = pd.read_excel(ruta, sheet_name='Hoja1')
        for _, row in df_cuentas.iterrows():
            cuenta = str(row.get('CUENTA ', row.get('CUENTA', ''))).strip()
            cliente = str(row.get('CLIENTE', '')).strip().upper()
            if cuenta and cliente:
                dict_cuentas[cliente] = cuenta
                # También añadir versiones sin puntuación
                cliente_limpio = re.sub(r'[.,]', '', cliente)
                if cliente_limpio != cliente:
                    dict_cuentas[cliente_limpio] = cuenta
        
        # Cargar hoja de alias (Sheet1)
        df_alias = pd.read_excel(ruta, sheet_name='Sheet1')
        for _, row in df_alias.iterrows():
            nombre = str(row.get('NOMBRE_EN_CONCEPTO', '')).strip().upper()
            titulo = str(row.get('TITULO_FACTURA', '')).strip().upper()
            if nombre and titulo and titulo != 'CASO ESPECIAL':
                dict_alias[nombre] = titulo
        
        _CACHE_CUENTAS = dict_cuentas
        _CACHE_ALIAS = dict_alias
        _CACHE_CARGADO = True
        
    except Exception as e:
        print(f"[AVISO] No se pudo cargar DiccionarioEmisorTitulo.xlsx: {e}")
    
    return dict_cuentas, dict_alias


def normalizar_para_busqueda(texto) -> str:
    """
    Normaliza un texto para búsqueda flexible.
    """
    if texto is None or (isinstance(texto, float) and pd.isna(texto)):
        return ''
    texto = str(texto).upper().strip()
    if texto == 'NAN':
        return ''
    # Quitar prefijos comunes de archivos
    texto = re.sub(r'^\d+T\d*\s+\d+\s+', '', texto)  # Quitar "4T25 1001 "
    texto = re.sub(r'^ATRASADA\s+', '', texto)
    # Quitar sufijos de tipo de factura
    texto = re.sub(r'\s+(TF|TJ|RC)\.?PDF$', '', texto, flags=re.IGNORECASE)
    texto = re.sub(r'\s+(TF|TJ|RC)$', '', texto, flags=re.IGNORECASE)
    # Normalizar espacios
    texto = ' '.join(texto.split())
    return texto


def buscar_cuenta_titulo(proveedor: str, ruta_diccionario: Optional[Path] = None) -> Tuple[str, str]:
    """
    Busca la CUENTA y TITULO oficial de un proveedor.
    
    Flujo de búsqueda:
    1. Buscar en alias (Sheet1) → obtener TITULO_FACTURA
    2. Buscar TITULO en cuentas (Hoja1) → obtener CUENTA
    3. Si no hay alias, buscar directamente en cuentas por similitud
    
    Args:
        proveedor: Nombre del proveedor (del extractor o archivo)
        ruta_diccionario: Ruta al diccionario (opcional)
        
    Returns:
        Tupla (CUENTA, TITULO) o ('PENDIENTE', proveedor_original) si no encuentra
    """
    dict_cuentas, dict_alias = cargar_diccionario_cuentas(ruta_diccionario)
    
    if not proveedor or (isinstance(proveedor, float) and pd.isna(proveedor)):
        return ('PENDIENTE', '')
    
    proveedor = str(proveedor)
    proveedor_norm = normalizar_para_busqueda(proveedor)
    proveedor_upper = proveedor.upper().strip()
    
    # 1. Búsqueda exacta en alias
    if proveedor_upper in dict_alias:
        titulo = dict_alias[proveedor_upper]
        if titulo in dict_cuentas:
            return (dict_cuentas[titulo], titulo)
    
    if proveedor_norm in dict_alias:
        titulo = dict_alias[proveedor_norm]
        if titulo in dict_cuentas:
            return (dict_cuentas[titulo], titulo)
    
    # 2. Búsqueda exacta en cuentas
    if proveedor_upper in dict_cuentas:
        return (dict_cuentas[proveedor_upper], proveedor_upper)
    
    if proveedor_norm in dict_cuentas:
        return (dict_cuentas[proveedor_norm], proveedor_norm)
    
    # 3. Búsqueda parcial en alias (contiene)
    for alias, titulo in dict_alias.items():
        if alias in proveedor_upper or proveedor_upper in alias:
            if titulo in dict_cuentas:
                return (dict_cuentas[titulo], titulo)
    
    # 4. Búsqueda parcial en cuentas (contiene)
    for cliente, cuenta in dict_cuentas.items():
        if cliente in proveedor_upper or proveedor_upper in cliente:
            return (cuenta, cliente)
    
    # 5. Búsqueda por similitud (último recurso)
    mejor_match = None
    mejor_ratio = 0.0
    
    # Primero en alias
    for alias, titulo in dict_alias.items():
        ratio = SequenceMatcher(None, proveedor_norm, alias).ratio()
        if ratio > mejor_ratio and ratio > 0.6:
            mejor_ratio = ratio
            mejor_match = ('alias', alias, titulo)
    
    # Luego en cuentas
    for cliente, cuenta in dict_cuentas.items():
        ratio = SequenceMatcher(None, proveedor_norm, cliente).ratio()
        if ratio > mejor_ratio and ratio > 0.6:
            mejor_ratio = ratio
            mejor_match = ('cuenta', cliente, cuenta)
    
    if mejor_match:
        if mejor_match[0] == 'alias':
            titulo = mejor_match[2]
            if titulo in dict_cuentas:
                return (dict_cuentas[titulo], titulo)
        else:
            return (mejor_match[2], mejor_match[1])
    
    # No encontrado
    return ('PENDIENTE', proveedor)


def extraer_numero_gestoria(archivo: str, numero_factura) -> Tuple[str, bool]:
    """
    Extrae el número de gestoría del nombre de archivo.
    
    Patrones válidos:
    - "3005 3T25 0704 CERES TF.pdf" → 3005
    - "946 T25 0619 DE LUIS..." → 946
    - "4001 4T 1108 ANTHROPIC..." → 4001
    
    Args:
        archivo: Nombre del archivo PDF
        numero_factura: Número asignado por el sistema (f.numero)
        
    Returns:
        Tupla (numero, es_temporal)
        - numero: El número de gestoría o TMPxxx si no tiene
        - es_temporal: True si es número temporal generado
    """
    if not archivo:
        return ('', True)
    
    # Patrón: número al inicio (3-4 dígitos) seguido de espacio
    match = re.match(r'^(\d{3,4})\s+', archivo)
    if match:
        return (match.group(1), False)
    
    # Si el número de factura del sistema es válido (no empieza por 9 y tiene 4 dígitos)
    if numero_factura:
        num_str = str(numero_factura)
        if re.match(r'^\d{3,4}$', num_str):
            # Verificar que el archivo empieza con este número
            if archivo.startswith(num_str):
                return (num_str, False)
    
    # No tiene número de gestoría
    return ('', True)


def formatear_fecha_factura(fecha) -> str:
    """
    Formatea la fecha al formato DD-MM-YY.
    
    Args:
        fecha: Fecha en varios formatos posibles
        
    Returns:
        Fecha formateada o cadena vacía
    """
    if not fecha:
        return ''
    
    fecha_str = str(fecha).strip()
    
    # Ya está en formato DD-MM-YY
    if re.match(r'^\d{2}-\d{2}-\d{2}$', fecha_str):
        return fecha_str
    
    # Formato DD/MM/YYYY
    match = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', fecha_str)
    if match:
        dd, mm, yyyy = match.groups()
        return f"{dd}-{mm}-{yyyy[2:]}"
    
    # Formato DD/MM/YY
    match = re.match(r'^(\d{2})/(\d{2})/(\d{2})$', fecha_str)
    if match:
        dd, mm, yy = match.groups()
        return f"{dd}-{mm}-{yy}"
    
    # Formato YYYY-MM-DD
    match = re.match(r'^(\d{4})-(\d{2})-(\d{2})', fecha_str)
    if match:
        yyyy, mm, dd = match.groups()
        return f"{dd}-{mm}-{yyyy[2:]}"
    
    # Devolver tal cual si no reconoce el formato
    return fecha_str


def generar_excel(facturas: List['Factura'], ruta: Path, nombre_hoja: str = 'Lineas',
                  ruta_diccionario: Optional[Path] = None) -> int:
    """
    Genera el Excel con las facturas procesadas.
    
    Crea dos hojas:
    - "Lineas": Detalle de todas las líneas de factura (antes "Facturas")
    - "Facturas": Cabeceras con una fila por factura
    
    Args:
        facturas: Lista de facturas procesadas
        ruta: Ruta donde guardar el archivo
        nombre_hoja: Nombre de la hoja de líneas (por compatibilidad)
        ruta_diccionario: Ruta al DiccionarioEmisorTitulo.xlsx
        
    Returns:
        Número de filas de líneas generadas
    """
    # Asegurar que el directorio existe
    ruta.parent.mkdir(parents=True, exist_ok=True)
    
    # Contador para números temporales
    contador_tmp = 1
    
    # =========================================================================
    # HOJA 1: LINEAS (detalle de artículos)
    # =========================================================================
    filas_lineas = []
    
    for f in facturas:
        # Obtener nombre del extractor (nombre de la clase)
        nombre_extractor = ''
        if hasattr(f, 'extractor_nombre') and f.extractor_nombre:
            nombre_extractor = f.extractor_nombre
        elif hasattr(f, 'extractor') and f.extractor:
            nombre_extractor = f.extractor.__class__.__name__
        
        if f.lineas:
            for linea in f.lineas:
                filas_lineas.append({
                    '#': f.numero,
                    'FECHA': f.fecha or '',
                    'REF': f.referencia or '',
                    'PROVEEDOR': f.proveedor,
                    'ARTICULO': linea.articulo,
                    'CATEGORIA': linea.categoria or 'PENDIENTE',
                    'CANTIDAD': linea.cantidad if linea.cantidad else '',
                    'PRECIO_UD': linea.precio_ud if linea.precio_ud else '',
                    'TIPO IVA': linea.iva,
                    'BASE (€)': linea.base,
                    'CUOTA IVA': linea.cuota_iva,
                    'TOTAL FAC': f.total or '',
                    'CUADRE': f.cuadre,
                    'ARCHIVO': f.archivo,
                    'EXTRACTOR': nombre_extractor
                })
        else:
            # Factura sin líneas extraídas
            filas_lineas.append({
                '#': f.numero,
                'FECHA': f.fecha or '',
                'REF': f.referencia or '',
                'PROVEEDOR': f.proveedor,
                'ARTICULO': 'VER FACTURA',
                'CATEGORIA': 'PENDIENTE',
                'CANTIDAD': '',
                'PRECIO_UD': '',
                'TIPO IVA': '',
                'BASE (€)': f.total or '',
                'CUOTA IVA': '',
                'TOTAL FAC': f.total or '',
                'CUADRE': f.cuadre,
                'ARCHIVO': f.archivo,
                'EXTRACTOR': nombre_extractor
            })
    
    # =========================================================================
    # HOJA 2: FACTURAS (cabeceras, una fila por factura)
    # Columnas: #, ARCHIVO, CUENTA, TITULO, Fec.Fac., REF, TOTAL FACTURA, Total, OBSERVACIONES
    # - TOTAL FACTURA = total extraído del PDF por el extractor
    # - Total = suma calculada de líneas (base × IVA)
    # =========================================================================
    filas_facturas = []
    
    for f in facturas:
        # Extraer número de gestoría
        num_gestoria, es_temporal = extraer_numero_gestoria(f.archivo, f.numero)
        
        if es_temporal or not num_gestoria:
            num_gestoria = f"TMP{contador_tmp:03d}"
            contador_tmp += 1
        
        # Buscar CUENTA y TITULO
        cuenta, titulo = buscar_cuenta_titulo(f.proveedor, ruta_diccionario)
        
        # Formatear fecha
        fecha_formateada = formatear_fecha_factura(f.fecha)
        
        # Calcular total desde líneas (suma de base * (1 + iva/100))
        total_calculado = ''
        if f.lineas:
            total_calc = sum(l.base * (1 + l.iva/100) for l in f.lineas)
            total_calculado = round(total_calc, 2)
        
        # Construir observaciones
        observaciones = f.cuadre or ''
        if es_temporal:
            if observaciones:
                observaciones += ', SIN_NUM_GESTORIA'
            else:
                observaciones = 'SIN_NUM_GESTORIA'
        
        # Nombre del archivo sin extensión para la columna ARCHIVO
        archivo_nombre = f.archivo if f.archivo else ''
        if archivo_nombre.lower().endswith('.pdf'):
            archivo_nombre = archivo_nombre[:-4]
        elif archivo_nombre.lower().endswith(('.jpg', '.png')):
            archivo_nombre = archivo_nombre[:-4]
        elif archivo_nombre.lower().endswith('.jpeg'):
            archivo_nombre = archivo_nombre[:-5]
        
        filas_facturas.append({
            '#': num_gestoria,
            'ARCHIVO': archivo_nombre,
            'CUENTA': cuenta,
            'TITULO': titulo,
            'Fec.Fac.': fecha_formateada,
            'REF': f.referencia or '',
            'TOTAL FACTURA': f.total or '',  # Total extraído del PDF
            'Total': total_calculado,         # Total calculado de líneas
            'OBSERVACIONES': observaciones
        })
    
    # =========================================================================
    # GUARDAR EXCEL CON AMBAS HOJAS Y FORMATO TABLA
    # =========================================================================
    df_lineas = pd.DataFrame(filas_lineas)
    df_facturas = pd.DataFrame(filas_facturas)
    
    # SANITIZAR antes de escribir (evita IllegalCharacterError)
    df_lineas = sanitizar_dataframe(df_lineas)
    df_facturas = sanitizar_dataframe(df_facturas)
    
    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
        # Orden: Lineas primero, Facturas después
        df_lineas.to_excel(writer, index=False, sheet_name='Lineas')
        df_facturas.to_excel(writer, index=False, sheet_name='Facturas')
        
        # Obtener workbook para formatear
        workbook = writer.book
        
        # Formatear hoja Lineas (tabla gris)
        ws_lineas = workbook['Lineas']
        formatear_hoja_como_tabla(ws_lineas, 'TablaLineas', 'TableStyleMedium9')
        
        # Formatear hoja Facturas (tabla morado claro)
        ws_facturas = workbook['Facturas']
        formatear_hoja_como_tabla(ws_facturas, 'TablaFacturas', 'TableStyleMedium16')
    
    return len(filas_lineas)


def generar_excel_resumen(facturas: List['Factura'], ruta: Path) -> int:
    """
    Genera un Excel resumen con totales por proveedor.
    
    Args:
        facturas: Lista de facturas procesadas
        ruta: Ruta donde guardar el archivo
        
    Returns:
        Número de filas generadas
    """
    # Agrupar por proveedor
    resumen = {}
    for f in facturas:
        prov = f.proveedor or 'DESCONOCIDO'
        if prov not in resumen:
            resumen[prov] = {
                'facturas': 0,
                'total': 0.0,
                'lineas': 0,
                'ok': 0,
                'errores': 0
            }
        resumen[prov]['facturas'] += 1
        resumen[prov]['total'] += f.total or 0
        resumen[prov]['lineas'] += len(f.lineas)
        if f.cuadre == 'OK':
            resumen[prov]['ok'] += 1
        else:
            resumen[prov]['errores'] += 1
    
    # Convertir a filas
    filas = []
    for prov, datos in sorted(resumen.items()):
        filas.append({
            'PROVEEDOR': prov,
            'FACTURAS': datos['facturas'],
            'TOTAL €': round(datos['total'], 2),
            'LINEAS': datos['lineas'],
            'OK': datos['ok'],
            'ERRORES': datos['errores'],
            '% ÉXITO': f"{100 * datos['ok'] / datos['facturas']:.1f}%" if datos['facturas'] > 0 else '0%'
        })
    
    df = pd.DataFrame(filas)
    df = sanitizar_dataframe(df)  # SANITIZAR
    ruta.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(ruta, index=False, sheet_name='Resumen')
    
    return len(filas)


def generar_excel_errores(facturas: List['Factura'], ruta: Path) -> int:
    """
    Genera un Excel con las facturas que tienen errores.
    
    Args:
        facturas: Lista de facturas procesadas
        ruta: Ruta donde guardar el archivo
        
    Returns:
        Número de filas generadas
    """
    filas = []
    
    for f in facturas:
        if f.tiene_errores or f.cuadre != 'OK':
            filas.append({
                '#': f.numero,
                'ARCHIVO': f.archivo,
                'PROVEEDOR': f.proveedor,
                'FECHA': f.fecha or '',
                'TOTAL': f.total or '',
                'CUADRE': f.cuadre,
                'ERRORES': ', '.join(f.errores) if f.errores else '',
                'LINEAS': len(f.lineas),
                'RUTA': str(f.ruta) if f.ruta else ''
            })
    
    if filas:
        df = pd.DataFrame(filas)
        df = sanitizar_dataframe(df)  # SANITIZAR
        ruta.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(ruta, index=False, sheet_name='Errores')
    
    return len(filas)


def generar_excel_multihoja(facturas: List['Factura'], ruta: Path) -> dict:
    """
    Genera un Excel con múltiples hojas:
    - Lineas: todas las líneas (detalle)
    - Facturas: cabeceras (una por factura)
    - Resumen: totales por proveedor
    - Errores: facturas con problemas
    
    Args:
        facturas: Lista de facturas procesadas
        ruta: Ruta donde guardar el archivo
        
    Returns:
        Dict con conteo de filas por hoja
    """
    ruta.parent.mkdir(parents=True, exist_ok=True)
    
    contador_tmp = 1
    
    with pd.ExcelWriter(ruta, engine='openpyxl') as writer:
        # Hoja 1: Lineas (detalle)
        filas_lineas = []
        for f in facturas:
            if f.lineas:
                for linea in f.lineas:
                    filas_lineas.append({
                        '#': f.numero,
                        'FECHA': f.fecha or '',
                        'PROVEEDOR': f.proveedor,
                        'ARTICULO': linea.articulo,
                        'CATEGORIA': linea.categoria or 'PENDIENTE',
                        'CANTIDAD': linea.cantidad if linea.cantidad else '',
                        'PRECIO_UD': linea.precio_ud if linea.precio_ud else '',
                        'IVA': linea.iva,
                        'BASE': linea.base,
                        'TOTAL FAC': f.total or '',
                        'CUADRE': f.cuadre
                    })
            else:
                filas_lineas.append({
                    '#': f.numero,
                    'FECHA': f.fecha or '',
                    'PROVEEDOR': f.proveedor,
                    'ARTICULO': 'VER FACTURA',
                    'CATEGORIA': 'PENDIENTE',
                    'CANTIDAD': '',
                    'PRECIO_UD': '',
                    'IVA': '',
                    'BASE': f.total or '',
                    'TOTAL FAC': f.total or '',
                    'CUADRE': f.cuadre
                })
        
        df_lineas = pd.DataFrame(filas_lineas)
        df_lineas = sanitizar_dataframe(df_lineas)  # SANITIZAR
        df_lineas.to_excel(writer, index=False, sheet_name='Lineas')
        
        # Hoja 2: Facturas (cabeceras)
        filas_facturas = []
        for f in facturas:
            num_gestoria, es_temporal = extraer_numero_gestoria(f.archivo, f.numero)
            if es_temporal or not num_gestoria:
                num_gestoria = f"TMP{contador_tmp:03d}"
                contador_tmp += 1
            
            cuenta, titulo = buscar_cuenta_titulo(f.proveedor)
            fecha_formateada = formatear_fecha_factura(f.fecha)
            
            observaciones = f.cuadre or ''
            if es_temporal:
                observaciones += ', SIN_NUM_GESTORIA' if observaciones else 'SIN_NUM_GESTORIA'
            
            filas_facturas.append({
                '#': num_gestoria,
                'CUENTA': cuenta,
                'TITULO': titulo,
                'Fec.Fac.': fecha_formateada,
                'REF': f.referencia or '',
                'Total': f.total or '',
                'OBSERVACIONES': observaciones
            })
        
        df_facturas = pd.DataFrame(filas_facturas)
        df_facturas = sanitizar_dataframe(df_facturas)  # SANITIZAR
        df_facturas.to_excel(writer, index=False, sheet_name='Facturas')
        
        # Hoja 3: Resumen
        resumen = {}
        for f in facturas:
            prov = f.proveedor or 'DESCONOCIDO'
            if prov not in resumen:
                resumen[prov] = {'facturas': 0, 'total': 0.0, 'ok': 0}
            resumen[prov]['facturas'] += 1
            resumen[prov]['total'] += f.total or 0
            if f.cuadre == 'OK':
                resumen[prov]['ok'] += 1
        
        filas_resumen = [
            {'PROVEEDOR': p, 'FACTURAS': d['facturas'], 'TOTAL': round(d['total'], 2), 
             'OK': d['ok'], '%': f"{100*d['ok']/d['facturas']:.0f}%"}
            for p, d in sorted(resumen.items())
        ]
        df_resumen = pd.DataFrame(filas_resumen)
        df_resumen = sanitizar_dataframe(df_resumen)  # SANITIZAR
        df_resumen.to_excel(writer, index=False, sheet_name='Resumen')
        
        # Hoja 4: Errores
        filas_errores = [
            {'ARCHIVO': f.archivo, 'PROVEEDOR': f.proveedor, 'CUADRE': f.cuadre, 
             'ERRORES': ', '.join(f.errores)}
            for f in facturas if f.tiene_errores or f.cuadre != 'OK'
        ]
        df_errores = pd.DataFrame(filas_errores)
        df_errores = sanitizar_dataframe(df_errores)  # SANITIZAR
        df_errores.to_excel(writer, index=False, sheet_name='Errores')
    
    return {
        'Lineas': len(filas_lineas),
        'Facturas': len(filas_facturas),
        'Resumen': len(filas_resumen),
        'Errores': len(filas_errores)
    }
