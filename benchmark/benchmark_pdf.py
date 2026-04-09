#!/usr/bin/env python3
"""
BENCHMARK PDF EXTRACTION - Tasca Barea / Comestibles Barea
===========================================================
Compara pdfplumber, PyMuPDF, pypdf, camelot y tabula sobre facturas reales.
Mide éxito en extracción de: TOTAL, FECHA, NIF, PROVEEDOR, LÍNEAS de producto.

Uso:
    python benchmark_pdf.py                          # Escanea rutas por defecto
    python benchmark_pdf.py --pdf-dir "C:\ruta"      # Carpeta específica
    python benchmark_pdf.py --sample 50              # Solo 50 PDFs aleatorios
    python benchmark_pdf.py --ground-truth compras.xlsx  # Ground truth manual

Resultados en: gestion-facturas/benchmark/results/
"""

import argparse
import csv
import glob
import json
import os
import re
import sys
import time
import traceback
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

DROPBOX_BASE = r"C:\Users\jaime\Dropbox\File inviati\TASCA BAREA S.L.L\CONTABILIDAD"
YEARS = [2025, 2026]
TRIMESTRES = ["1 TRIMESTRE", "2 TRIMESTRE", "3 TRIMESTRE", "4 TRIMESTRE"]

# Carpeta de resultados (relativa al script, que ya está en benchmark/)
SCRIPT_DIR = Path(__file__).parent
RESULTS_DIR = SCRIPT_DIR / "results"
RESULTS_DIR.mkdir(parents=True, exist_ok=True)

# Engines disponibles
ENGINES = ["pdfplumber", "pymupdf", "pypdf", "camelot", "tabula"]

# ============================================================================
# ENGINE WRAPPERS
# ============================================================================

def extract_pdfplumber(pdf_path):
    """Extrae texto y tablas con pdfplumber."""
    import pdfplumber
    result = {"text": "", "tables": [], "pages": 0, "error": None}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            result["pages"] = len(pdf.pages)
            texts = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                texts.append(t)
                tables = page.extract_tables() or []
                for table in tables:
                    result["tables"].append(table)
            result["text"] = "\n".join(texts)
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_pymupdf(pdf_path):
    """Extrae texto y bloques con PyMuPDF (fitz)."""
    import fitz
    result = {"text": "", "blocks": [], "pages": 0, "error": None}
    try:
        doc = fitz.open(pdf_path)
        result["pages"] = len(doc)
        texts = []
        for page in doc:
            texts.append(page.get_text())
            blocks = page.get_text("blocks")
            for b in blocks:
                if b[6] == 0:  # text block
                    result["blocks"].append({
                        "bbox": [b[0], b[1], b[2], b[3]],
                        "text": b[4].strip()
                    })
        result["text"] = "\n".join(texts)
        doc.close()
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_pypdf(pdf_path):
    """Extrae texto con pypdf."""
    import pypdf
    result = {"text": "", "pages": 0, "error": None}
    try:
        reader = pypdf.PdfReader(pdf_path)
        result["pages"] = len(reader.pages)
        texts = []
        for page in reader.pages:
            t = page.extract_text() or ""
            texts.append(t)
        result["text"] = "\n".join(texts)
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_camelot(pdf_path):
    """Extrae tablas con camelot."""
    result = {"text": "", "tables": [], "pages": 0, "error": None}
    try:
        import camelot
        # Lattice mode (for bordered tables)
        tables_lattice = camelot.read_pdf(str(pdf_path), pages="all", flavor="lattice")
        # Stream mode (for borderless tables)
        tables_stream = camelot.read_pdf(str(pdf_path), pages="all", flavor="stream")

        all_tables = []
        texts = []
        for t in tables_lattice:
            df = t.df
            all_tables.append({"flavor": "lattice", "accuracy": t.accuracy,
                               "rows": df.values.tolist(), "shape": list(df.shape)})
            texts.append(df.to_string(index=False))
        for t in tables_stream:
            df = t.df
            all_tables.append({"flavor": "stream", "accuracy": t.accuracy,
                               "rows": df.values.tolist(), "shape": list(df.shape)})
            texts.append(df.to_string(index=False))

        result["tables"] = all_tables
        result["text"] = "\n".join(texts)
    except Exception as e:
        result["error"] = str(e)
    return result


def extract_tabula(pdf_path):
    """Extrae tablas con tabula-py."""
    result = {"text": "", "tables": [], "pages": 0, "error": None}
    try:
        import tabula
        # Lattice
        dfs_lattice = tabula.read_pdf(str(pdf_path), pages="all",
                                       lattice=True, silent=True)
        # Stream
        dfs_stream = tabula.read_pdf(str(pdf_path), pages="all",
                                      stream=True, silent=True)

        texts = []
        all_tables = []
        for df in (dfs_lattice or []):
            all_tables.append({"flavor": "lattice", "rows": df.values.tolist(),
                               "shape": list(df.shape)})
            texts.append(df.to_string(index=False))
        for df in (dfs_stream or []):
            all_tables.append({"flavor": "stream", "rows": df.values.tolist(),
                               "shape": list(df.shape)})
            texts.append(df.to_string(index=False))

        result["tables"] = all_tables
        result["text"] = "\n".join(texts)
    except Exception as e:
        result["error"] = str(e)
    return result


ENGINE_FUNCS = {
    "pdfplumber": extract_pdfplumber,
    "pymupdf": extract_pymupdf,
    "pypdf": extract_pypdf,
    "camelot": extract_camelot,
    "tabula": extract_tabula,
}

# ============================================================================
# EXTRACTORES DE CAMPOS (genéricos, regex-based)
# ============================================================================

def parse_spanish_float(text):
    """Convierte '1.234,56' o '1234,56' o '1234.56' a float."""
    if not text:
        return None
    text = text.strip().replace(" ", "")
    # Formato europeo: 1.234,56
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def extract_total(text):
    """Busca el TOTAL de la factura en el texto."""
    patterns = [
        r'TOTAL\s*(?:FACTURA)?\s*:?\s*([\d.,]+)',
        r'TOTAL\s+([\d.,]+)\s*€?',
        r'Total\s*(?:factura)?\s*:?\s*([\d.,]+)',
        r'IMPORTE\s+TOTAL\s*:?\s*([\d.,]+)',
        r'TOTAL\s*A\s*PAGAR\s*:?\s*([\d.,]+)',
        r'T\s*O\s*T\s*A\s*L\s*:?\s*([\d.,]+)',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            val = parse_spanish_float(m.group(1))
            if val and val > 0:
                return val
    # Fallback: buscar el mayor importe en la última parte del texto
    last_quarter = text[len(text)*3//4:]
    amounts = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', last_quarter)
    if amounts:
        vals = [parse_spanish_float(a) for a in amounts]
        vals = [v for v in vals if v and v > 0]
        if vals:
            return max(vals)
    return None


def extract_fecha(text):
    """Busca la fecha de factura."""
    patterns = [
        r'(?:Fecha|FECHA)\s*(?:factura|FACTURA|Fac\.?)?\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
        r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})',
        r'(\d{1,2}[/-]\d{1,2}[/-]\d{2})\b',
        r'(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
    ]
    for p in patterns:
        m = re.search(p, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return None


def extract_nif(text):
    """Busca NIF/CIF del proveedor."""
    patterns = [
        r'(?:N\.?I\.?F\.?|C\.?I\.?F\.?)\s*:?\s*([A-Z]\d{7,8}[A-Z0-9]?)',
        r'(?:N\.?I\.?F\.?|C\.?I\.?F\.?)\s*:?\s*(\d{8}[A-Z])',
        r'\b([A-Z]\d{8})\b',
        r'\b([A-Z]\d{7}[A-Z])\b',
    ]
    for p in patterns:
        m = re.search(p, text)
        if m:
            return m.group(1).strip()
    return None


def extract_iva(text):
    """Busca tipo/importe IVA."""
    results = []
    # Buscar pares tipo% + importe
    patterns = [
        r'(\d{1,2})[,%]\s*(?:I\.?V\.?A\.?)\s*:?\s*([\d.,]+)',
        r'I\.?V\.?A\.?\s*(\d{1,2})\s*%?\s*:?\s*([\d.,]+)',
        r'(\d{1,2})\s*%\s*([\d.,]+)',
    ]
    for p in patterns:
        for m in re.finditer(p, text, re.IGNORECASE):
            tipo = parse_spanish_float(m.group(1))
            importe = parse_spanish_float(m.group(2))
            if tipo and importe:
                results.append({"tipo": tipo, "importe": importe})
    return results


def extract_lineas(text):
    """Intenta extraer líneas de producto (descripción + importes)."""
    lines = text.split("\n")
    product_lines = []
    for line in lines:
        # Buscar líneas con al menos un importe y algo de texto
        amounts = re.findall(r'(\d{1,3}(?:\.\d{3})*,\d{2})', line)
        # Filtrar líneas que son solo totales/headers
        skip_words = ['TOTAL', 'SUBTOTAL', 'BASE', 'I.V.A', 'IMPORTE',
                       'OBSERV', 'IBAN', 'CUENTA', 'PORTES', 'DESCUENTO',
                       'TIPO', 'R.E.', 'RECARGO']
        line_upper = line.upper().strip()
        if any(line_upper.startswith(w) for w in skip_words):
            continue
        if len(amounts) >= 1 and len(line.strip()) > 10:
            # Intentar separar descripción de números
            desc_match = re.match(r'^(.*?)(\d{1,3}(?:\.\d{3})*,\d{2})', line)
            if desc_match:
                desc = desc_match.group(1).strip()
                # Limpiar códigos de artículo del inicio
                desc = re.sub(r'^[\d-]+\s+', '', desc).strip()
                if desc and len(desc) > 2:
                    product_lines.append({
                        "descripcion": desc,
                        "importes": [parse_spanish_float(a) for a in amounts],
                        "raw": line.strip()
                    })
    return product_lines


def extract_all_fields(text):
    """Extrae todos los campos de un texto."""
    return {
        "total": extract_total(text),
        "fecha": extract_fecha(text),
        "nif": extract_nif(text),
        "iva": extract_iva(text),
        "lineas": extract_lineas(text),
        "num_lineas": len(extract_lineas(text)),
        "text_length": len(text),
    }


# ============================================================================
# GROUND TRUTH - Carga de COMPRAS xlsx
# ============================================================================

def load_ground_truth(xlsx_paths):
    """Carga ground truth desde uno o varios COMPRAS xlsx.
    
    Soporta múltiples formatos:
    - Formato PARSEO (tabs Lineas + Facturas con ARCHIVO)
    - Formato Provisional (tab Facturas con NOMBRE + Total)
    - Formato Kinema histórico (tabs AÑO XX con #, PROVEEDOR, BASE por línea)
    """
    import openpyxl
    gt = {}

    for xlsx_path in xlsx_paths:
        if not os.path.exists(xlsx_path):
            print(f"  [WARN] Ground truth no encontrado: {xlsx_path}")
            continue

        print(f"  Cargando: {xlsx_path}")
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        loaded = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            col_map = {h.upper(): i for i, h in enumerate(headers)}

            # ---- Formato Provisional: NOMBRE, PROVEEDOR, Fec.Fac., Total ----
            if "NOMBRE" in col_map:
                ci_nombre = col_map["NOMBRE"]
                ci_total = col_map.get("TOTAL", col_map.get("TOTAL FACTURA", -1))
                ci_fecha = col_map.get("FEC.FAC.", col_map.get("FEC.FAC", -1))
                ci_prov = col_map.get("PROVEEDOR", -1)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row = list(row)
                    nombre = str(row[ci_nombre] or "").strip()
                    if not nombre:
                        continue
                    total = row[ci_total] if ci_total >= 0 else None
                    fecha = row[ci_fecha] if ci_fecha >= 0 else None
                    prov = str(row[ci_prov] or "") if ci_prov >= 0 else ""

                    norm_key = normalize_filename(nombre)
                    gt[norm_key] = {
                        "archivo": nombre,
                        "total": float(total) if total else None,
                        "fecha": str(fecha) if fecha else None,
                        "titulo": prov,
                        "source": f"{Path(xlsx_path).name}:{sheet_name}",
                    }
                    loaded += 1

            # ---- Formato PARSEO: tab Facturas con ARCHIVO ----
            elif "ARCHIVO" in col_map:
                ci_arch = col_map["ARCHIVO"]
                ci_total = col_map.get("TOTAL FACTURA", col_map.get("TOTAL", -1))
                ci_fecha = col_map.get("FEC.FAC", col_map.get("FEC.FAC.", -1))
                ci_titulo = col_map.get("TITULO", -1)
                ci_cuenta = col_map.get("CUENTA", -1)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row = list(row)
                    archivo = str(row[ci_arch] or "").strip()
                    if not archivo:
                        continue
                    total = row[ci_total] if ci_total >= 0 else None
                    fecha = row[ci_fecha] if ci_fecha >= 0 else None

                    norm_key = normalize_filename(archivo)
                    gt[norm_key] = {
                        "archivo": archivo,
                        "total": float(total) if total else None,
                        "fecha": str(fecha) if fecha else None,
                        "titulo": str(row[ci_titulo] or "") if ci_titulo >= 0 else "",
                        "cuenta": str(row[ci_cuenta] or "") if ci_cuenta >= 0 else "",
                        "source": f"{Path(xlsx_path).name}:{sheet_name}",
                    }
                    loaded += 1

            # ---- Formato Kinema contable: CÓD., TÍTULO, FEC.FAC., TOTAL ----
            elif "CÓD." in col_map and "TOTAL" in col_map:
                ci_cod = col_map["CÓD."]
                ci_total = col_map["TOTAL"]
                ci_fecha = col_map.get("FEC.FAC.", col_map.get("FEC.FAC", -1))
                ci_titulo = col_map.get("TÍTULO", col_map.get("TITULO", -1))

                for row in ws.iter_rows(min_row=2, values_only=True):
                    row = list(row)
                    cod = row[ci_cod]
                    if not cod:
                        continue
                    cod_str = str(int(cod)) if isinstance(cod, (int, float)) else str(cod).strip()
                    total = row[ci_total] if ci_total >= 0 else None
                    fecha = row[ci_fecha] if ci_fecha >= 0 else None
                    titulo = str(row[ci_titulo] or "") if ci_titulo >= 0 else ""

                    # Clave: KINEMA_{código} — el PDF empieza con este número
                    gt_key = f"KINEMA_{cod_str}"
                    gt[gt_key] = {
                        "archivo": f"#{cod_str}",
                        "total": float(total) if total else None,
                        "fecha": str(fecha) if fecha else None,
                        "titulo": titulo,
                        "source": f"{Path(xlsx_path).name}:{sheet_name}",
                    }
                    loaded += 1

            # ---- Formato Kinema histórico: #, FECHA, PROVEEDOR, BASE ----
            elif "#" in col_map and "BASE" in col_map:
                ci_num = col_map["#"]
                ci_fecha = col_map.get("FECHA", -1)
                ci_prov = col_map.get("PROVEEDOR", -1)
                ci_base = col_map.get("BASE", -1)
                ci_iva_tipo = col_map.get("TIPO IVA", -1)

                # Agregar líneas por factura (#)
                facturas_agg = defaultdict(lambda: {
                    "total_base": 0.0, "fecha": None, "proveedor": "",
                    "num_lineas": 0, "iva_tipos": set()
                })
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row = list(row)
                    num = str(row[ci_num] or "").strip()
                    if not num:
                        continue
                    base = row[ci_base] if ci_base >= 0 else None
                    fa = facturas_agg[num]
                    if base:
                        try:
                            fa["total_base"] += float(base)
                        except (ValueError, TypeError):
                            pass
                    fa["num_lineas"] += 1
                    if ci_fecha >= 0 and row[ci_fecha] and not fa["fecha"]:
                        fa["fecha"] = str(row[ci_fecha])
                    if ci_prov >= 0 and row[ci_prov]:
                        fa["proveedor"] = str(row[ci_prov])
                    if ci_iva_tipo >= 0 and row[ci_iva_tipo]:
                        try:
                            fa["iva_tipos"].add(float(row[ci_iva_tipo]))
                        except (ValueError, TypeError):
                            pass

                # Crear GT entries — clave por nº Kinema (aparece al inicio del PDF)
                for num, fa in facturas_agg.items():
                    # El PDF se llama "1023 1T25 0307 CERES RC.pdf"
                    # El # Kinema es "1023" — guardamos por nº para matching
                    norm_key = f"KINEMA_{num}"

                    # Calcular total con IVA (base * (1 + tipo/100))
                    # Si hay múltiples tipos, no podemos calcular exacto, guardamos base
                    gt[norm_key] = {
                        "archivo": f"#{num}",
                        "total_base": round(fa["total_base"], 2),
                        "total": None,  # No podemos saber total con IVA exacto
                        "fecha": fa["fecha"],
                        "titulo": fa["proveedor"],
                        "num_lineas": fa["num_lineas"],
                        "iva_tipos": list(fa["iva_tipos"]),
                        "source": f"{Path(xlsx_path).name}:{sheet_name}",
                    }
                    loaded += 1

        print(f"    → {loaded} entradas de {Path(xlsx_path).name}")
        wb.close()

    print(f"  Ground truth total: {len(gt)} facturas")
    return gt


def normalize_filename(name):
    """Normaliza nombre de archivo para matching flexible."""
    name = str(name).strip()
    # Quitar extensión
    name = re.sub(r'\.(pdf|PDF)$', '', name)
    # Quitar número secuencial de Kinema al inicio (ej: "1023 ")
    name = re.sub(r'^\d{1,4}\s+', '', name)
    # Underscores → espacios
    name = name.replace("_", " ")
    # Normalizar espacios
    name = re.sub(r'\s+', ' ', name).strip().upper()
    return name


# ============================================================================
# DESCUBRIMIENTO DE PDFs
# ============================================================================

def discover_pdfs(base_dir=None, years=None):
    """Busca PDFs en la estructura de Dropbox."""
    if base_dir:
        # Carpeta manual
        pdfs = list(Path(base_dir).rglob("*.pdf")) + list(Path(base_dir).rglob("*.PDF"))
        print(f"  Encontrados {len(pdfs)} PDFs en {base_dir}")
        return pdfs

    if years is None:
        years = YEARS

    pdfs = []
    for year in years:
        facturas_dir = Path(DROPBOX_BASE) / f"FACTURAS {year}" / "FACTURAS RECIBIDAS"
        if not facturas_dir.exists():
            print(f"  [WARN] No existe: {facturas_dir}")
            continue

        for trim in TRIMESTRES:
            trim_dir = facturas_dir / f"{trim} {year}"
            if not trim_dir.exists():
                continue

            # PDFs en carpeta principal del trimestre
            found = list(trim_dir.glob("*.pdf")) + list(trim_dir.glob("*.PDF"))
            pdfs.extend(found)

            # PDFs en ATRASADAS
            atrasadas = trim_dir / "ATRASADAS"
            if atrasadas.exists():
                found_atr = list(atrasadas.glob("*.pdf")) + list(atrasadas.glob("*.PDF"))
                pdfs.extend(found_atr)

    # Deduplicar por path
    pdfs = list(set(pdfs))
    print(f"  Encontrados {len(pdfs)} PDFs en total")
    return pdfs


# ============================================================================
# SCORING
# ============================================================================

def score_extraction(extracted, ground_truth_entry):
    """Puntúa una extracción contra el ground truth."""
    scores = {}
    gt = ground_truth_entry

    # TOTAL - tolerancia 0.02€
    if gt.get("total"):
        ext_total = extracted.get("total")
        if ext_total and abs(ext_total - gt["total"]) < 0.02:
            scores["total"] = 1.0
        elif ext_total and abs(ext_total - gt["total"]) < 1.0:
            scores["total"] = 0.5  # Cercano
        else:
            scores["total"] = 0.0
    else:
        scores["total"] = None  # Sin referencia

    # FECHA
    if gt.get("fecha"):
        ext_fecha = extracted.get("fecha")
        if ext_fecha:
            # Normalizar ambas fechas para comparar
            scores["fecha"] = 1.0 if dates_match(ext_fecha, gt["fecha"]) else 0.0
        else:
            scores["fecha"] = 0.0
    else:
        scores["fecha"] = None

    # NIF
    if extracted.get("nif"):
        scores["nif"] = 1.0  # Encontró algo (sin ref para validar)
    else:
        scores["nif"] = 0.0

    # LINEAS
    if gt.get("num_lineas"):
        ext_lineas = extracted.get("num_lineas", 0)
        if ext_lineas == gt["num_lineas"]:
            scores["lineas"] = 1.0
        elif ext_lineas > 0:
            ratio = min(ext_lineas, gt["num_lineas"]) / max(ext_lineas, gt["num_lineas"])
            scores["lineas"] = ratio
        else:
            scores["lineas"] = 0.0
    else:
        scores["lineas"] = None

    # IVA
    if extracted.get("iva"):
        scores["iva"] = 1.0
    else:
        scores["iva"] = 0.0

    # Texto extraído
    scores["has_text"] = 1.0 if extracted.get("text_length", 0) > 50 else 0.0

    return scores


def dates_match(date_str1, date_str2):
    """Compara dos fechas en formatos posiblemente distintos."""
    def parse_date(s):
        s = str(s).strip()
        # Handle Excel datetime strings like "2026-03-23 00:00:00"
        s = re.sub(r'\s+\d{2}:\d{2}:\d{2}$', '', s)
        for fmt in ["%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
                     "%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y", "%d.%m.%y"]:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        # Intentar extraer con regex
        m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', s)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if y < 100:
                y += 2000
            try:
                from datetime import date
                return date(y, mo, d)
            except ValueError:
                pass
        return None

    d1 = parse_date(date_str1)
    d2 = parse_date(date_str2)
    if d1 and d2:
        return d1 == d2
    return False


# ============================================================================
# INFORME
# ============================================================================

def generate_report(all_results, gt, output_dir):
    """Genera informe CSV + HTML con resultados."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # --- CSV detallado ---
    csv_path = output_dir / f"benchmark_detail_{timestamp}.csv"
    fieldnames = ["pdf_file", "engine", "time_ms", "error",
                  "total_found", "total_gt", "total_score",
                  "fecha_found", "fecha_gt", "fecha_score",
                  "nif_found", "nif_score",
                  "num_lineas_found", "num_lineas_gt", "lineas_score",
                  "iva_found", "iva_score",
                  "has_text", "text_length"]

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()

        for r in all_results:
            for eng, eng_data in r["engines"].items():
                row = {
                    "pdf_file": r["filename"],
                    "engine": eng,
                    "time_ms": eng_data.get("time_ms", 0),
                    "error": eng_data.get("error", ""),
                    "total_found": eng_data["fields"].get("total", ""),
                    "total_gt": r.get("gt_total", ""),
                    "total_score": eng_data["scores"].get("total", ""),
                    "fecha_found": eng_data["fields"].get("fecha", ""),
                    "fecha_gt": r.get("gt_fecha", ""),
                    "fecha_score": eng_data["scores"].get("fecha", ""),
                    "nif_found": eng_data["fields"].get("nif", ""),
                    "nif_score": eng_data["scores"].get("nif", ""),
                    "num_lineas_found": eng_data["fields"].get("num_lineas", ""),
                    "num_lineas_gt": r.get("gt_num_lineas", ""),
                    "lineas_score": eng_data["scores"].get("lineas", ""),
                    "iva_found": len(eng_data["fields"].get("iva", [])),
                    "iva_score": eng_data["scores"].get("iva", ""),
                    "has_text": eng_data["scores"].get("has_text", ""),
                    "text_length": eng_data["fields"].get("text_length", 0),
                }
                writer.writerow(row)

    print(f"  CSV detallado: {csv_path}")

    # --- Resumen por engine ---
    summary = defaultdict(lambda: defaultdict(list))
    for r in all_results:
        for eng, eng_data in r["engines"].items():
            for field, score in eng_data["scores"].items():
                if score is not None:
                    summary[eng][field].append(score)

    csv_summary_path = output_dir / f"benchmark_summary_{timestamp}.csv"
    with open(csv_summary_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        fields = ["total", "fecha", "nif", "lineas", "iva", "has_text"]
        writer.writerow(["engine", "n_pdfs"] + [f"{fl}_avg" for fl in fields] +
                        [f"{fl}_n" for fl in fields] + ["avg_time_ms", "errors"])

        for eng in ENGINES:
            if eng not in summary:
                continue
            n_pdfs = max(len(v) for v in summary[eng].values()) if summary[eng] else 0
            row = [eng, n_pdfs]
            for fl in fields:
                vals = summary[eng].get(fl, [])
                row.append(f"{sum(vals)/len(vals):.3f}" if vals else "N/A")
            for fl in fields:
                row.append(len(summary[eng].get(fl, [])))
            # Tiempo medio
            times = []
            errors = 0
            for r in all_results:
                if eng in r["engines"]:
                    times.append(r["engines"][eng].get("time_ms", 0))
                    if r["engines"][eng].get("error"):
                        errors += 1
            row.append(f"{sum(times)/len(times):.0f}" if times else "N/A")
            row.append(errors)
            writer.writerow(row)

    print(f"  CSV resumen: {csv_summary_path}")

    # --- HTML visual ---
    html_path = output_dir / f"benchmark_report_{timestamp}.html"
    generate_html_report(all_results, summary, html_path, timestamp)
    print(f"  HTML report: {html_path}")

    return csv_path, csv_summary_path, html_path


def generate_html_report(all_results, summary, html_path, timestamp):
    """Genera un informe HTML visual."""
    fields = ["total", "fecha", "nif", "lineas", "iva", "has_text"]
    field_labels = {"total": "TOTAL €", "fecha": "FECHA", "nif": "NIF",
                    "lineas": "LÍNEAS", "iva": "IVA", "has_text": "TEXTO"}

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Benchmark PDF Extraction - {timestamp}</title>
<style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 20px; background: #f5f5f5; }}
h1 {{ color: #333; border-bottom: 3px solid #CA3026; padding-bottom: 10px; }}
h2 {{ color: #555; margin-top: 30px; }}
table {{ border-collapse: collapse; margin: 15px 0; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }}
th {{ background: #CA3026; color: white; padding: 10px 15px; text-align: left; }}
td {{ padding: 8px 15px; border-bottom: 1px solid #eee; }}
tr:hover {{ background: #f9f9f9; }}
.score-1 {{ background: #d4edda; color: #155724; font-weight: bold; }}
.score-05 {{ background: #fff3cd; color: #856404; }}
.score-0 {{ background: #f8d7da; color: #721c24; }}
.error {{ color: #dc3545; }}
.winner {{ background: #d4edda; font-weight: bold; }}
.stats {{ display: flex; gap: 20px; flex-wrap: wrap; margin: 20px 0; }}
.stat-card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 1px 3px rgba(0,0,0,0.1);
              min-width: 150px; text-align: center; }}
.stat-card .number {{ font-size: 2em; font-weight: bold; color: #CA3026; }}
.stat-card .label {{ color: #666; margin-top: 5px; }}
.bar {{ display: inline-block; height: 20px; background: #CA3026; border-radius: 3px; }}
</style></head><body>
<h1>🔍 Benchmark PDF Extraction</h1>
<p>Generado: {timestamp} | PDFs analizados: {len(all_results)}</p>
"""

    # Stats cards
    html += '<div class="stats">'
    html += f'<div class="stat-card"><div class="number">{len(all_results)}</div><div class="label">PDFs</div></div>'
    for eng in ENGINES:
        if eng in summary:
            vals = summary[eng].get("total", [])
            avg = sum(vals)/len(vals) if vals else 0
            html += f'<div class="stat-card"><div class="number">{avg:.0%}</div><div class="label">{eng}<br>TOTAL accuracy</div></div>'
    html += '</div>'

    # Summary table
    html += '<h2>Resumen por Motor</h2><table><tr><th>Motor</th>'
    for fl in fields:
        html += f'<th>{field_labels[fl]}</th>'
    html += '<th>Tiempo medio</th><th>Errores</th></tr>'

    # Find winners per field
    best_per_field = {}
    for fl in fields:
        best_score = -1
        best_eng = None
        for eng in ENGINES:
            if eng in summary:
                vals = summary[eng].get(fl, [])
                avg = sum(vals)/len(vals) if vals else 0
                if avg > best_score:
                    best_score = avg
                    best_eng = eng
        best_per_field[fl] = best_eng

    for eng in ENGINES:
        if eng not in summary:
            continue
        html += f'<tr><td><strong>{eng}</strong></td>'
        for fl in fields:
            vals = summary[eng].get(fl, [])
            if vals:
                avg = sum(vals)/len(vals)
                cls = "winner" if best_per_field[fl] == eng else ""
                html += f'<td class="{cls}">{avg:.1%} <small>({len(vals)})</small></td>'
            else:
                html += '<td>-</td>'
        # Time
        times = [r["engines"][eng].get("time_ms", 0) for r in all_results if eng in r["engines"]]
        avg_t = sum(times)/len(times) if times else 0
        errors = sum(1 for r in all_results if eng in r["engines"] and r["engines"][eng].get("error"))
        html += f'<td>{avg_t:.0f}ms</td><td>{errors}</td></tr>'
    html += '</table>'

    # Detailed table (first 100)
    html += '<h2>Detalle por PDF (primeros 100)</h2>'
    html += '<table><tr><th>PDF</th><th>GT Total</th>'
    for eng in ENGINES:
        html += f'<th>{eng}</th>'
    html += '</tr>'

    for r in all_results[:100]:
        html += f'<tr><td title="{r["filename"]}">{r["filename"][:40]}</td>'
        html += f'<td>{r.get("gt_total", "-")}</td>'
        for eng in ENGINES:
            if eng in r["engines"]:
                ed = r["engines"][eng]
                total = ed["fields"].get("total", "")
                score = ed["scores"].get("total")
                if ed.get("error"):
                    html += f'<td class="error" title="{ed["error"]}">ERROR</td>'
                elif score == 1.0:
                    html += f'<td class="score-1">{total}</td>'
                elif score == 0.5:
                    html += f'<td class="score-05">{total}</td>'
                elif score == 0.0:
                    html += f'<td class="score-0">{total or "-"}</td>'
                else:
                    html += f'<td>{total or "-"}</td>'
            else:
                html += '<td>-</td>'
        html += '</tr>'

    html += '</table></body></html>'

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)


# ============================================================================
# MAIN
# ============================================================================

def run_benchmark(pdf_dir=None, ground_truth_paths=None, sample=None,
                  engines=None, years=None):
    """Ejecuta el benchmark completo."""
    print("=" * 70)
    print("  BENCHMARK PDF EXTRACTION")
    print("=" * 70)

    # 1. Descubrir PDFs
    print("\n[1/4] Buscando PDFs...")
    pdfs = discover_pdfs(pdf_dir, years)
    if not pdfs:
        print("  ERROR: No se encontraron PDFs")
        return

    if sample and sample < len(pdfs):
        import random
        random.seed(42)
        pdfs = random.sample(pdfs, sample)
        print(f"  Muestra aleatoria: {sample} PDFs")

    # 2. Cargar ground truth
    print("\n[2/4] Cargando ground truth...")
    gt = {}
    if ground_truth_paths:
        gt = load_ground_truth(ground_truth_paths)
    else:
        # Buscar COMPRAS xlsx automáticamente en gestion-facturas
        auto_paths = list(SCRIPT_DIR.glob("**/Facturas_XT*.xlsx"))
        auto_paths += list(SCRIPT_DIR.glob("**/COMPRAS*.xlsx"))
        if auto_paths:
            print(f"  Auto-detectados: {[str(p) for p in auto_paths]}")
            gt = load_ground_truth(auto_paths)
        else:
            print("  [WARN] Sin ground truth - solo se medirá extracción bruta")

    # 3. Ejecutar engines
    active_engines = engines or ENGINES
    print(f"\n[3/4] Ejecutando {len(active_engines)} engines sobre {len(pdfs)} PDFs...")
    print(f"  Engines: {', '.join(active_engines)}")

    all_results = []
    for idx, pdf_path in enumerate(pdfs):
        filename = pdf_path.name
        norm_key = normalize_filename(filename)

        if (idx + 1) % 10 == 0 or idx == 0:
            print(f"  [{idx+1}/{len(pdfs)}] {filename[:50]}...")

        result = {
            "filename": filename,
            "path": str(pdf_path),
            "norm_key": norm_key,
            "engines": {},
        }

        # Ground truth match
        gt_entry = gt.get(norm_key)
        # Fallback: try Kinema number (prefix digits in filename)
        if not gt_entry:
            kinema_match = re.match(r'^(\d{1,4})\s+', filename)
            if kinema_match:
                kinema_key = f"KINEMA_{kinema_match.group(1)}"
                gt_entry = gt.get(kinema_key)
        if gt_entry:
            result["gt_total"] = gt_entry.get("total")
            result["gt_fecha"] = gt_entry.get("fecha")
            result["gt_num_lineas"] = gt_entry.get("num_lineas")
            result["gt_matched"] = True
        else:
            result["gt_matched"] = False

        # Run each engine
        for eng in active_engines:
            eng_result = {"fields": {}, "scores": {}, "error": None, "time_ms": 0}

            t0 = time.time()
            try:
                raw = ENGINE_FUNCS[eng](str(pdf_path))
                eng_result["time_ms"] = int((time.time() - t0) * 1000)

                if raw.get("error"):
                    eng_result["error"] = raw["error"]
                    eng_result["fields"] = {"text_length": 0}
                else:
                    eng_result["fields"] = extract_all_fields(raw.get("text", ""))

                # Score against GT
                if gt_entry:
                    eng_result["scores"] = score_extraction(eng_result["fields"], gt_entry)
                else:
                    # Sin GT, solo medir si extrajo algo
                    eng_result["scores"] = {
                        "total": 1.0 if eng_result["fields"].get("total") else 0.0,
                        "fecha": 1.0 if eng_result["fields"].get("fecha") else 0.0,
                        "nif": 1.0 if eng_result["fields"].get("nif") else 0.0,
                        "lineas": 1.0 if eng_result["fields"].get("num_lineas", 0) > 0 else 0.0,
                        "iva": 1.0 if eng_result["fields"].get("iva") else 0.0,
                        "has_text": 1.0 if eng_result["fields"].get("text_length", 0) > 50 else 0.0,
                    }

            except Exception as e:
                eng_result["time_ms"] = int((time.time() - t0) * 1000)
                eng_result["error"] = traceback.format_exc()
                eng_result["fields"] = {"text_length": 0}
                eng_result["scores"] = {k: 0.0 for k in ["total", "fecha", "nif", "lineas", "iva", "has_text"]}

            result["engines"][eng] = eng_result

        all_results.append(result)

    # 4. Generar informe
    print(f"\n[4/4] Generando informe...")
    csv_detail, csv_summary, html_report = generate_report(all_results, gt, RESULTS_DIR)

    # Print summary to console
    print("\n" + "=" * 70)
    print("  RESULTADOS")
    print("=" * 70)

    summary = defaultdict(lambda: defaultdict(list))
    for r in all_results:
        for eng, eng_data in r["engines"].items():
            for field, score in eng_data["scores"].items():
                if score is not None:
                    summary[eng][field].append(score)

    print(f"\n{'Engine':<15} {'TOTAL':>8} {'FECHA':>8} {'NIF':>8} {'LÍNEAS':>8} {'IVA':>8} {'TEXTO':>8} {'ms':>8} {'Err':>5}")
    print("-" * 85)
    for eng in ENGINES:
        if eng not in summary:
            continue
        cols = []
        for fl in ["total", "fecha", "nif", "lineas", "iva", "has_text"]:
            vals = summary[eng].get(fl, [])
            cols.append(f"{sum(vals)/len(vals):.1%}" if vals else "N/A")
        times = [r["engines"][eng].get("time_ms", 0) for r in all_results if eng in r["engines"]]
        avg_t = sum(times)/len(times) if times else 0
        errors = sum(1 for r in all_results if eng in r["engines"] and r["engines"][eng].get("error"))
        print(f"{eng:<15} {cols[0]:>8} {cols[1]:>8} {cols[2]:>8} {cols[3]:>8} {cols[4]:>8} {cols[5]:>8} {avg_t:>7.0f} {errors:>5}")

    gt_matched = sum(1 for r in all_results if r.get("gt_matched"))
    print(f"\nPDFs con ground truth: {gt_matched}/{len(all_results)}")
    print(f"\nInformes guardados en: {RESULTS_DIR}")

    # Save raw JSON for further analysis
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = RESULTS_DIR / f"benchmark_raw_{ts}.json"
    def safe_json(obj):
        if isinstance(obj, Path):
            return str(obj)
        return str(obj)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=safe_json)
    print(f"  JSON raw: {json_path}")


def main():
    parser = argparse.ArgumentParser(description="Benchmark PDF Extraction")
    parser.add_argument("--pdf-dir", help="Carpeta con PDFs (en vez de Dropbox)")
    parser.add_argument("--ground-truth", nargs="+", help="COMPRAS xlsx path(s)")
    parser.add_argument("--sample", type=int, help="Número de PDFs aleatorios a probar")
    parser.add_argument("--engines", nargs="+", choices=ENGINES, help="Solo estos engines")
    parser.add_argument("--years", nargs="+", type=int, default=YEARS, help="Años a escanear")
    args = parser.parse_args()

    run_benchmark(
        pdf_dir=args.pdf_dir,
        ground_truth_paths=args.ground_truth,
        sample=args.sample,
        engines=args.engines,
        years=args.years,
    )


if __name__ == "__main__":
    main()
