#!/usr/bin/env python3
"""
historico_eventos.py — Genera Excel con histórico de eventos desde WooCommerce.

Descarga todos los productos con fecha en el nombre (>= fecha_desde),
parsea las descripciones HTML y genera un Excel estructurado.

Uso:
    python historico_eventos.py                    # desde 01/11/2025
    python historico_eventos.py --desde 01/01/2026
    python historico_eventos.py --output mi_archivo.xlsx
"""

import argparse
import os
import re
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("ERROR: Falta beautifulsoup4. Ejecuta: pip install beautifulsoup4")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)

from dotenv import load_dotenv
from woocommerce import API

# ── Config ───────────────────────────────────────────────────────────────────

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_SCRIPT_DIR)
_OUTPUT_DIR = os.path.join(_PROJECT_ROOT, "outputs")

load_dotenv(os.path.join(_SCRIPT_DIR, ".env"))

_FECHA_RE = re.compile(r"(\d{1,2}/\d{2}/\d{2})\s*$")

# Campos estructurados que aparecen en <strong>CAMPO:</strong>
_CAMPOS_STRUCT = {
    "HORARIO": "Horario",
    "DIRIGIDO A": "Dirigido a",
    "DURACIÓN": "Duración",
    "NÚMERO ASISTENTES": "Asistentes",
    "NÚMERO de ASISTENTES": "Asistentes",
    "LUGAR": "Lugar",
}


# ── WooCommerce ──────────────────────────────────────────────────────────────

def _get_wc():
    return API(
        url=os.getenv("WC_URL"),
        consumer_key=os.getenv("WC_KEY"),
        consumer_secret=os.getenv("WC_SECRET"),
        version="wc/v3",
        timeout=30,
    )


def descargar_eventos(wc, fecha_desde):
    """Descarga productos con fecha en nombre >= fecha_desde."""
    eventos = []
    page = 1
    while True:
        resp = wc.get("products", params={
            "per_page": 100, "page": page, "status": "publish",
        }).json()
        if not isinstance(resp, list) or not resp:
            break
        for p in resp:
            nombre_raw = p.get("name", "")
            nombre_limpio = re.sub(r"<[^>]+>", "", nombre_raw).strip()
            m = _FECHA_RE.search(nombre_limpio)
            if m:
                try:
                    dt = datetime.strptime(m.group(1), "%d/%m/%y").date()
                    if dt >= fecha_desde:
                        eventos.append({
                            "id": p["id"],
                            "nombre_raw": nombre_limpio,
                            "fecha": dt,
                            "description": p.get("description", ""),
                            "total_sales": p.get("total_sales", 0),
                            "stock_quantity": p.get("stock_quantity"),
                        })
                except ValueError:
                    pass
        page += 1
        if len(resp) < 100:
            break

    # También buscar private (CERRADOS antiguos y tests)
    page = 1
    while True:
        resp = wc.get("products", params={
            "per_page": 100, "page": page, "status": "private",
        }).json()
        if not isinstance(resp, list) or not resp:
            break
        for p in resp:
            nombre_raw = p.get("name", "")
            nombre_limpio = re.sub(r"<[^>]+>", "", nombre_raw).strip()
            # Excluir tests
            if nombre_limpio.startswith("[TEST]"):
                continue
            m = _FECHA_RE.search(nombre_limpio)
            if m:
                try:
                    dt = datetime.strptime(m.group(1), "%d/%m/%y").date()
                    if dt >= fecha_desde:
                        eventos.append({
                            "id": p["id"],
                            "nombre_raw": nombre_limpio,
                            "fecha": dt,
                            "description": p.get("description", ""),
                            "total_sales": p.get("total_sales", 0),
                            "stock_quantity": p.get("stock_quantity"),
                        })
                except ValueError:
                    pass
        page += 1
        if len(resp) < 100:
            break

    return eventos


# ── Parseo de descripciones ──────────────────────────────────────────────────

def parsear_descripcion(html):
    """Extrae campos estructurados y descripción libre del HTML."""
    resultado = {v: "" for v in _CAMPOS_STRUCT.values()}
    resultado["Descripción"] = ""

    if not html or not html.strip():
        return resultado

    soup = BeautifulSoup(html, "html.parser")
    texto_completo = soup.get_text(separator="\n").strip()

    # Buscar campos <strong>CAMPO:</strong> valor
    for strong in soup.find_all("strong"):
        texto_strong = strong.get_text().strip()
        for campo_html, col_excel in _CAMPOS_STRUCT.items():
            if texto_strong.upper().startswith(campo_html.upper()):
                # El valor está después del <strong> hasta el siguiente <br> o <strong>
                valor = ""
                for sibling in strong.next_siblings:
                    if hasattr(sibling, "name") and sibling.name in ("strong", "br"):
                        break
                    text = sibling.get_text() if hasattr(sibling, "get_text") else str(sibling)
                    valor += text
                valor = valor.strip().lstrip(":").strip()
                if valor:
                    resultado[col_excel] = valor
                break

    # Horario: también buscar en texto libre "HORARIO: de X a Y"
    if not resultado["Horario"]:
        m = re.search(r"HORARIO:\s*de\s*(\d{1,2}:\d{2})\s*a?\s*(\d{1,2}:\d{2})?",
                       texto_completo, re.IGNORECASE)
        if m:
            h = m.group(1)
            if m.group(2):
                h += f" a {m.group(2)}"
            resultado["Horario"] = h

    # Descripción: el texto que no son los campos estructurados
    # Quitar la primera línea si es repetición del nombre ("CATA-TALLER Comestibles Barea")
    lineas = texto_completo.split("\n")
    desc_lineas = []
    campos_encontrados = set()
    for linea in lineas:
        linea = linea.strip()
        if not linea or linea == "\xa0":
            continue
        # Es un campo estructurado?
        es_campo = False
        for campo_html in _CAMPOS_STRUCT:
            if linea.upper().startswith(campo_html.upper()):
                es_campo = True
                campos_encontrados.add(campo_html)
                break
        if linea.upper().startswith("FECHA:"):
            es_campo = True
        if not es_campo:
            desc_lineas.append(linea)

    resultado["Descripción"] = "\n".join(desc_lineas).strip()
    return resultado


def limpiar_nombre(nombre_raw):
    """Quita fecha, CERRADO, y limpia el nombre."""
    nombre = _FECHA_RE.sub("", nombre_raw).strip()
    # Quitar prefijos CERRADO con variantes
    nombre = re.sub(r"^CERRADO[-_\s]*", "", nombre, flags=re.IGNORECASE).strip()
    return nombre


# ── Generación Excel ─────────────────────────────────────────────────────────

COLUMNAS = [
    "Nombre", "Fecha", "Tipo", "Horario", "Dirigido a",
    "Asistentes", "Duración", "Lugar", "Descripción",
    "Vendidas", "Stock",
]

ANCHOS = {
    "Nombre": 40, "Fecha": 14, "Tipo": 12, "Horario": 20,
    "Dirigido a": 50, "Asistentes": 15, "Duración": 15,
    "Lugar": 50, "Descripción": 50, "Vendidas": 10, "Stock": 10,
}

HEADER_FILL = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
HEADER_FONT = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(name="Aptos", size=10)
BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def generar_excel(eventos, output_path):
    """Genera el Excel histórico."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Eventos"

    # Headers
    for c, col in enumerate(COLUMNAS, 1):
        cell = ws.cell(1, c, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    # Datos (ordenados por fecha descendente)
    eventos.sort(key=lambda e: e["fecha"], reverse=True)

    for r, ev in enumerate(eventos, 2):
        nombre_raw = ev["nombre_raw"]
        es_cerrado = nombre_raw.upper().startswith("CERRADO")
        nombre_limpio = limpiar_nombre(nombre_raw)
        campos = parsear_descripcion(ev["description"])

        fila = {
            "Nombre": nombre_limpio,
            "Fecha": ev["fecha"],
            "Tipo": "CERRADO" if es_cerrado else "Abierto",
            "Horario": campos.get("Horario", ""),
            "Dirigido a": campos.get("Dirigido a", ""),
            "Asistentes": campos.get("Asistentes", ""),
            "Duración": campos.get("Duración", ""),
            "Lugar": campos.get("Lugar", ""),
            "Descripción": campos.get("Descripción", ""),
            "Vendidas": ev.get("total_sales", 0),
            "Stock": ev.get("stock_quantity") if ev.get("stock_quantity") is not None else "",
        }

        for c, col in enumerate(COLUMNAS, 1):
            valor = fila[col]
            cell = ws.cell(r, c, valor)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if col == "Fecha" and isinstance(valor, (datetime,)):
                cell.number_format = "DD/MM/YYYY"
            elif col == "Fecha":
                cell.number_format = "DD/MM/YYYY"
            if col in ("Descripción", "Dirigido a", "Lugar"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col in ("Vendidas", "Stock"):
                cell.alignment = Alignment(horizontal="center")

    # Anchos
    for c, col in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(c)].width = ANCHOS.get(col, 12)

    ws.freeze_panes = "A2"
    wb.save(output_path)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Histórico de eventos — Comestibles Barea")
    parser.add_argument("--desde", default="01/11/2025", help="Fecha desde (DD/MM/YYYY)")
    parser.add_argument("--output", "-o", default=None, help="Archivo de salida")
    args = parser.parse_args()

    fecha_desde = datetime.strptime(args.desde, "%d/%m/%Y").date()
    if args.output:
        output_path = args.output
    else:
        os.makedirs(_OUTPUT_DIR, exist_ok=True)
        output_path = os.path.join(
            _OUTPUT_DIR,
            f"Historico_Eventos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

    print("=" * 50)
    print("HISTÓRICO DE EVENTOS — COMESTIBLES BAREA")
    print("=" * 50)
    print(f"  Desde: {fecha_desde.strftime('%d/%m/%Y')}")

    wc = _get_wc()
    eventos = descargar_eventos(wc, fecha_desde)
    print(f"  Eventos encontrados: {len(eventos)}")

    if not eventos:
        print("  Sin eventos para exportar.")
        return

    # Resumen
    cerrados = sum(1 for e in eventos if e["nombre_raw"].upper().startswith("CERRADO"))
    abiertos = len(eventos) - cerrados
    print(f"  Abiertos: {abiertos} | Cerrados: {cerrados}")

    generar_excel(eventos, output_path)
    print(f"\n  Guardado: {output_path}")
    print("=" * 50)


if __name__ == "__main__":
    main()
