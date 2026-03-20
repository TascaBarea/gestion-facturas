#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ASISTENTES TALLER v1.0 — ventas_semana/asistentes_taller.py
============================================================
Envía por email la lista de asistentes al taller del día.
El inventario de talleres lo genera script_barea.py cada lunes en talleres_programados.json.

Uso:
  python asistentes_taller.py                    # envía email si hay taller hoy
  python asistentes_taller.py --test             # muestra asistentes en pantalla, no envía
  python asistentes_taller.py --fecha 28/03/26   # simula esa fecha
  python asistentes_taller.py --fecha 28/03/26 --test
"""

import os
import sys
import json
import re
import base64
import argparse
import logging
from datetime import datetime

from dotenv import load_dotenv

# === Paths ===
_script_dir  = os.path.dirname(os.path.abspath(__file__))
_project_root = os.path.dirname(_script_dir)
_env_path    = os.path.join(_script_dir, ".env")
_JSON_PATH   = os.path.join(_script_dir, "talleres_programados.json")
_GMAIL_TOKEN = os.path.join(_project_root, "gmail", "token.json")
_EMAIL_DESTINO = "hola@comestiblesbarea.com"

load_dotenv(_env_path)

# === Logging ===
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("asistentes")


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _parse_fecha(fecha_str: str) -> datetime:
    """Parsea DD/MM/YY o D/MM/YY a datetime."""
    for fmt in ("%d/%m/%y", "%-d/%m/%y"):
        try:
            return datetime.strptime(fecha_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Formato de fecha no reconocido: {fecha_str!r}")


def _fecha_hoy() -> datetime:
    return datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)


# ---------------------------------------------------------------------------
# Inventario y búsqueda de taller
# ---------------------------------------------------------------------------

def cargar_inventario() -> dict:
    if not os.path.exists(_JSON_PATH):
        log.error("No existe %s — ejecuta primero script_barea.py para generar el inventario.", _JSON_PATH)
        sys.exit(1)
    with open(_JSON_PATH, encoding="utf-8") as f:
        return json.load(f)


def buscar_taller_hoy(inventario: dict, fecha_override: str | None = None) -> dict | None:
    """Devuelve el taller de hoy (o de la fecha indicada), o None si no hay."""
    if fecha_override:
        ref = _parse_fecha(fecha_override)
    else:
        ref = _fecha_hoy()

    for t in inventario.get("talleres", []):
        try:
            dt = _parse_fecha(t["fecha"])
            if dt.date() == ref.date():
                return t
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# WooCommerce: pedidos del taller
# ---------------------------------------------------------------------------

def obtener_pedidos(product_id: int) -> list[dict]:
    """Descarga todos los pedidos del taller filtrando por product_id.
    Excluye estados cancelled y refunded."""
    from woocommerce import API

    wc = API(
        url=os.getenv("WC_URL"),
        consumer_key=os.getenv("WC_KEY"),
        consumer_secret=os.getenv("WC_SECRET"),
        version="wc/v3",
        timeout=30,
    )

    ESTADOS_EXCLUIR = {"cancelled", "refunded"}
    pedidos = []
    page = 1

    while True:
        batch = wc.get("orders", params={
            "per_page": 100,
            "page": page,
            "product": product_id,
        }).json()

        if not isinstance(batch, list) or not batch:
            break

        for order in batch:
            if order.get("status") not in ESTADOS_EXCLUIR:
                pedidos.append(order)

        page += 1
        if len(batch) < 100:
            break

    return pedidos


def procesar_asistentes(pedidos: list[dict]) -> list[dict]:
    """Extrae campos relevantes de cada pedido."""
    asistentes = []
    for order in pedidos:
        billing = order.get("billing", {})
        nombre  = f"{billing.get('first_name', '')} {billing.get('last_name', '')}".strip()
        entradas = sum(item.get("quantity", 0) for item in order.get("line_items", []))
        asistentes.append({
            "Nombre":          nombre,
            "Email":           billing.get("email", ""),
            "Teléfono":        billing.get("phone", ""),
            "Entradas":        entradas,
            "Importe (€)":     float(order.get("total", 0)),
            "Fecha compra":    order.get("date_created", "")[:10],
            "Método de pago":  order.get("payment_method_title", ""),
        })
    return asistentes


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generar_excel(asistentes: list[dict], taller: dict) -> str:
    """Genera Excel con la lista de asistentes. Devuelve la ruta del archivo."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    COLS = ["Nombre", "Email", "Teléfono", "Entradas", "Importe (€)", "Fecha compra", "Método de pago"]
    WIDTHS = [25, 32, 15, 10, 13, 14, 22]
    COLOR_HEADER = "2E4057"

    fecha_dt  = _parse_fecha(taller["fecha"])
    fecha_fmt = fecha_dt.strftime("%d/%m/%Y")
    fecha_safe = taller["fecha"].replace("/", "")
    nombre_safe = re.sub(r'[\\/:*?"<>|]', '', taller["nombre"])[:40].replace(" ", "_")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistentes"

    # Fila 1: título
    ws.merge_cells(f"A1:{chr(64 + len(COLS))}1")
    c = ws["A1"]
    c.value = f"Asistentes — {taller['nombre']} — {fecha_fmt}"
    c.font = Font(bold=True, size=13)
    c.alignment = Alignment(horizontal="center")

    # Fila 2: cabeceras
    for col_i, col_name in enumerate(COLS, 1):
        cell = ws.cell(row=2, column=col_i, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center")

    # Datos
    for row_i, asistente in enumerate(asistentes, 3):
        for col_i, col_name in enumerate(COLS, 1):
            ws.cell(row=row_i, column=col_i, value=asistente[col_name])

    # Fila totales
    fila_total = len(asistentes) + 3
    ws.cell(row=fila_total, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila_total, column=4, value=sum(a["Entradas"]    for a in asistentes)).font = Font(bold=True)
    ws.cell(row=fila_total, column=5, value=round(sum(a["Importe (€)"] for a in asistentes), 2)).font = Font(bold=True)

    # Anchos de columna
    for col_i, width in enumerate(WIDTHS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = width

    nombre_archivo = f"Asistentes_{nombre_safe}_{fecha_safe}.xlsx"
    path = os.path.join(_script_dir, nombre_archivo)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Gmail OAuth2
# ---------------------------------------------------------------------------

def _get_gmail_service():
    """Devuelve un servicio Gmail autenticado, o None si falla."""
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except ImportError:
        log.error("Faltan paquetes: pip install google-auth google-api-python-client")
        return None

    if not os.path.exists(_GMAIL_TOKEN):
        log.error("No existe token.json en %s", _GMAIL_TOKEN)
        return None

    scopes = [
        "https://www.googleapis.com/auth/gmail.send",
        "https://www.googleapis.com/auth/gmail.modify",
    ]
    creds = Credentials.from_authorized_user_file(_GMAIL_TOKEN, scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(_GMAIL_TOKEN, "w") as f:
                f.write(creds.to_json())
        else:
            log.error("Credenciales Gmail expiradas — ejecuta renovar_token_business.py")
            return None

    return build("gmail", "v1", credentials=creds)


def enviar_email(taller: dict, asistentes: list[dict], excel_path: str) -> bool:
    """Envía email con resumen HTML y Excel adjunto."""
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    service = _get_gmail_service()
    if not service:
        return False

    fecha_dt   = _parse_fecha(taller["fecha"])
    fecha_fmt  = fecha_dt.strftime("%d/%m/%Y")
    horario    = taller.get("hora_inicio") or ""
    hora_txt   = f" · {horario}" if horario else ""
    n_personas = sum(a["Entradas"] for a in asistentes)
    total_eur  = round(sum(a["Importe (€)"] for a in asistentes), 2)

    asunto = f"🎟️ Asistentes HOY — {taller['nombre']} — {n_personas} personas"

    filas_html = "".join(
        f"<tr>"
        f"<td>{a['Nombre']}</td>"
        f"<td>{a['Email']}</td>"
        f"<td>{a['Teléfono']}</td>"
        f"<td style='text-align:center'>{a['Entradas']}</td>"
        f"<td style='text-align:right'>{a['Importe (€)']:.2f} €</td>"
        f"</tr>"
        for a in asistentes
    )

    html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#333;max-width:800px">
      <h2 style="color:#2E4057">🎟️ {taller['nombre']}</h2>
      <p>
        <strong>Fecha:</strong> {fecha_fmt}{hora_txt}<br>
        <strong>Total asistentes:</strong> {n_personas} personas &nbsp;|&nbsp;
        <strong>Recaudación:</strong> {total_eur:.2f} €
      </p>
      <table border="1" cellpadding="6" cellspacing="0"
             style="border-collapse:collapse;width:100%;font-size:13px">
        <thead style="background:#2E4057;color:white">
          <tr>
            <th>Nombre</th><th>Email</th><th>Teléfono</th>
            <th>Entradas</th><th>Importe</th>
          </tr>
        </thead>
        <tbody>{filas_html}</tbody>
        <tfoot style="font-weight:bold;background:#f0f0f0">
          <tr>
            <td colspan="3">TOTAL</td>
            <td style="text-align:center">{n_personas}</td>
            <td style="text-align:right">{total_eur:.2f} €</td>
          </tr>
        </tfoot>
      </table>
      <p style="margin-top:20px;color:#aaa;font-size:11px">
        Generado automáticamente · asistentes_taller.py
      </p>
    </body></html>
    """

    msg = MIMEMultipart("mixed")
    msg["To"] = _EMAIL_DESTINO
    msg["Subject"] = asunto
    msg.attach(MIMEText(html, "html"))

    with open(excel_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition",
                        f'attachment; filename="{os.path.basename(excel_path)}"')
        msg.attach(part)

    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    log.info("Email enviado a %s", _EMAIL_DESTINO)
    return True


# ---------------------------------------------------------------------------
# Modo test: mostrar en pantalla
# ---------------------------------------------------------------------------

def mostrar_test(taller: dict, asistentes: list[dict]) -> None:
    n_personas = sum(a["Entradas"] for a in asistentes)
    total_eur  = round(sum(a["Importe (€)"] for a in asistentes), 2)
    sep = "=" * 70
    print(f"\n{sep}")
    print(f"[TEST] {taller['nombre']} — {taller['fecha']}  {taller.get('hora_inicio') or ''}")
    print(sep)
    print(f"{'Nombre':<26} {'Email':<30} {'Tel':<13} Ent  Importe")
    print("-" * 70)
    for a in asistentes:
        print(f"{a['Nombre']:<26} {a['Email']:<30} {a['Teléfono']:<13} "
              f" x{a['Entradas']}  {a['Importe (€)']:.2f} €")
    print("-" * 70)
    print(f"{'TOTAL':<26} {'':<30} {'':<13}  x{n_personas}  {total_eur:.2f} €")
    print(f"{sep}\n")
    log.info("Modo --test: email NO enviado.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Asistentes Taller — envío automático de lista por email el día del taller"
    )
    parser.add_argument("--test",  action="store_true",
                        help="Mostrar asistentes en pantalla sin enviar email")
    parser.add_argument("--fecha", metavar="DD/MM/YY",
                        help="Simular una fecha concreta (ej: 28/03/26)")
    args = parser.parse_args()

    inventario = cargar_inventario()
    taller = buscar_taller_hoy(inventario, args.fecha)

    if not taller:
        ref = args.fecha or datetime.now().strftime("%d/%m/%y")
        log.info("Sin taller programado para %s — nada que hacer.", ref)
        sys.exit(0)

    log.info("Taller: %s (%s %s)",
             taller["nombre"], taller["fecha"], taller.get("hora_inicio") or "")

    pedidos = obtener_pedidos(taller["id"])
    log.info("%d pedido(s) encontrado(s)", len(pedidos))

    if not pedidos:
        log.info("Sin pedidos para este taller — no se envía email.")
        sys.exit(0)

    asistentes = procesar_asistentes(pedidos)

    if args.test:
        mostrar_test(taller, asistentes)
        return

    excel_path = generar_excel(asistentes, taller)
    log.info("Excel generado: %s", excel_path)

    if not enviar_email(taller, asistentes, excel_path):
        log.error("Fallo al enviar email.")
        sys.exit(1)

    os.remove(excel_path)
    log.info("Proceso completado.")


if __name__ == "__main__":
    main()
