"""
api/maestro.py — Lógica de negocio para MAESTRO_PROVEEDORES.

Funciones puras para leer, escribir y validar el MAESTRO.
Usado por los endpoints en api/server.py.
"""

import os
import shutil
import logging
from datetime import datetime

from openpyxl import load_workbook, Workbook
from pydantic import BaseModel

from api.config import PROJECT_ROOT

logger = logging.getLogger(__name__)

MAESTRO_PATH = os.path.join(PROJECT_ROOT, "datos", "MAESTRO_PROVEEDORES.xlsx")
BACKUP_DIR = os.path.join(PROJECT_ROOT, "datos", "backups")

FORMAS_PAGO_VALIDAS = {"TF", "TJ", "RC", "EF", ""}


# ── Pydantic models ─────────────────────────────────────────────────────────

class ProveedorUpdate(BaseModel):
    CUENTA: str | None = None
    CLASE: str | None = None
    CIF: str | None = None
    IBAN: str | None = None
    FORMA_PAGO: str | None = None
    EMAIL: str | None = None
    ALIAS: list[str] | None = None
    TIENE_EXTRACTOR: str | None = None
    ARCHIVO_EXTRACTOR: str | None = None
    TIPO_CATEGORIA: str | None = None
    CATEGORIA_FIJA: str | None = None
    METODO_PDF: str | None = None
    ACTIVO: str | None = None
    NOTAS: str | None = None


class ProveedorCreate(ProveedorUpdate):
    PROVEEDOR: str


# ── Funciones de lectura ─────────────────────────────────────────────────────

def leer_maestro() -> tuple[list[str], list[dict]]:
    """Lee MAESTRO_PROVEEDORES.xlsx y devuelve (cabeceras, lista de dicts).

    Lee TODAS las columnas dinámicamente.
    ALIAS se devuelve como lista (split por coma).
    """
    wb = load_workbook(MAESTRO_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Cabeceras originales
    cabeceras = []
    for cell in ws[1]:
        if cell.value:
            cabeceras.append(str(cell.value).strip())
        else:
            cabeceras.append("")

    proveedores = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        prov = {}
        for i, col_name in enumerate(cabeceras):
            if not col_name:
                continue
            val = row[i] if i < len(row) else None
            prov[col_name] = str(val).strip() if val is not None else ""

        if not prov.get("PROVEEDOR"):
            continue

        # ALIAS → lista
        alias_raw = prov.get("ALIAS", "")
        if alias_raw:
            sep = "," if "," in alias_raw else "|"
            prov["ALIAS"] = [a.strip() for a in alias_raw.split(sep) if a.strip()]
        else:
            prov["ALIAS"] = []

        proveedores.append(prov)

    wb.close()
    return cabeceras, proveedores


def leer_maestro_simple() -> list[dict]:
    """Wrapper que devuelve solo la lista de proveedores."""
    _, proveedores = leer_maestro()
    return proveedores


def verificar_no_abierto():
    """Comprueba que el Excel no esté abierto.

    Raises:
        PermissionError: si el archivo está bloqueado (abierto en Excel).
    """
    try:
        wb = load_workbook(MAESTRO_PATH)
        wb.close()
    except PermissionError:
        raise PermissionError(
            "El archivo MAESTRO_PROVEEDORES.xlsx está abierto en Excel. "
            "Ciérralo antes de guardar cambios."
        )


def backup_maestro() -> str:
    """Crea backup del MAESTRO antes de escribir."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(BACKUP_DIR, f"MAESTRO_PROVEEDORES_{ts}.xlsx")
    shutil.copy2(MAESTRO_PATH, backup_path)
    logger.info(f"Backup MAESTRO: {backup_path}")
    return backup_path


# ── Validación ───────────────────────────────────────────────────────────────

def validar_proveedor(data: dict, es_nuevo: bool = False) -> list[str]:
    """Valida campos de un proveedor. Devuelve lista de errores (vacía = OK)."""
    errores = []

    if es_nuevo:
        if not data.get("PROVEEDOR", "").strip():
            errores.append("PROVEEDOR es obligatorio.")

    forma = data.get("FORMA_PAGO")
    if forma is not None and forma.upper() not in FORMAS_PAGO_VALIDAS:
        errores.append(
            f"FORMA_PAGO inválida: '{forma}'. "
            f"Permitidos: {', '.join(sorted(FORMAS_PAGO_VALIDAS - {''}))} o vacío."
        )

    return errores


# ── Escritura ────────────────────────────────────────────────────────────────

def guardar_maestro(cabeceras: list[str], proveedores: list[dict]):
    """Escribe lista de proveedores al Excel preservando todas las columnas.

    ALIAS (lista) se convierte a string separado por comas.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "MAESTRO"

    # Cabeceras
    for col_idx, col_name in enumerate(cabeceras, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Datos
    for row_idx, prov in enumerate(proveedores, 2):
        for col_idx, col_name in enumerate(cabeceras, 1):
            if not col_name:
                continue
            valor = prov.get(col_name, "")
            if col_name == "ALIAS" and isinstance(valor, list):
                valor = ", ".join(valor)
            # Intentar preservar números para CUENTA
            if col_name == "CUENTA" and valor and valor.isdigit():
                valor = int(valor)
            ws.cell(row=row_idx, column=col_idx, value=valor)

    wb.save(MAESTRO_PATH)
    wb.close()
    logger.info(f"MAESTRO guardado: {len(proveedores)} proveedores")


def actualizar_proveedor(nombre: str, cambios: dict) -> dict:
    """Actualiza un proveedor existente.

    Flow: verificar_no_abierto → leer → buscar → validar → backup → guardar.
    """
    verificar_no_abierto()
    cabeceras, proveedores = leer_maestro()

    # Buscar por nombre (case-insensitive)
    idx = None
    for i, p in enumerate(proveedores):
        if p["PROVEEDOR"].upper() == nombre.upper():
            idx = i
            break

    if idx is None:
        raise KeyError(f"Proveedor '{nombre}' no encontrado.")

    # Aplicar cambios
    prov = proveedores[idx]
    for campo, valor in cambios.items():
        if valor is not None:
            prov[campo] = valor

    # Validar
    errores = validar_proveedor(prov)
    if errores:
        raise ValueError("; ".join(errores))

    backup_maestro()
    guardar_maestro(cabeceras, proveedores)
    return prov


def crear_proveedor(data: dict) -> dict:
    """Crea un proveedor nuevo."""
    verificar_no_abierto()

    errores = validar_proveedor(data, es_nuevo=True)
    if errores:
        raise ValueError("; ".join(errores))

    cabeceras, proveedores = leer_maestro()

    # Comprobar duplicados
    nombre_nuevo = data["PROVEEDOR"].strip().upper()
    for p in proveedores:
        if p["PROVEEDOR"].upper() == nombre_nuevo:
            raise ValueError(f"Ya existe un proveedor con nombre '{data['PROVEEDOR']}'.")

    # Construir nuevo registro con todas las cabeceras
    nuevo = {col: "" for col in cabeceras if col}
    nuevo["ALIAS"] = []
    nuevo["PROVEEDOR"] = data["PROVEEDOR"].strip()
    for campo, valor in data.items():
        if valor is not None:
            nuevo[campo] = valor

    proveedores.append(nuevo)
    backup_maestro()
    guardar_maestro(cabeceras, proveedores)
    return nuevo
