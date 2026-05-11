"""
auditoria_iva_mixto.py
======================
Script read-only que identifica proveedores candidatos al refactor de IVA mixto.

Lee Excels con hoja "Lineas" (formato canonico Facturas_<trim>.xlsx), agrupa
por proveedor, y reporta cuales tienen facturas con mas de un TIPO IVA distinto
entre sus lineas. Las keywords de transporte en ARTICULO se reportan como
senal complementaria (TIENE_PORTES), no como condicion necesaria — la regla
del repo manda prorratear portes en lineas de producto, asi que casi nunca
aparecen como linea separada en historico.

Output:
- Consola: tabla por NIVEL_SOSPECHA descendente (filtra SIN por defecto).
- Excel:   outputs/auditoria_iva_mixto_<YYYYMMDD_HHMM>.xlsx con todos.

Uso:
    python scripts/auditoria_iva_mixto.py
    python scripts/auditoria_iva_mixto.py --trimestres 4T25,1T26
    python scripts/auditoria_iva_mixto.py --extra-files <path|glob>
    python scripts/auditoria_iva_mixto.py --no-excel
"""
from __future__ import annotations

import argparse
import glob
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple

import openpyxl

KEYWORDS_DEFAULT: List[str] = [
    "PORTES", "TRANSPORTE", "ENVIO", "ENVIO", "GASTOS DE ENVIO",
    "GASTOS ENV", "LOGISTICA", "EXPEDICION", "FLETE", "MENSAJERIA",
    "COURIER", "SHIPPING", "EMBALAJE", "PORT.",
]

CUADRE_OK_VALUES = {"OK", "SIN_LINEAS", "", None}

# Patrones que indican que el "proveedor" detectado es realmente un trozo de
# filename que el extractor de identificacion no resolvio a un nombre real.
ID_DUDOSA_PATTERNS = [
    re.compile(r"^\d?T\d{2}\s"),       # "1T25 ", "T26 ", "2T25 "
    re.compile(r"ATRASADA", re.IGNORECASE),
    re.compile(r"PAGADA", re.IGNORECASE),
    re.compile(r"DIFERENCIA", re.IGNORECASE),
    re.compile(r"\bRET\b", re.IGNORECASE),
    re.compile(r"\bOJO\b", re.IGNORECASE),
    re.compile(r"\d{6,}"),             # 6+ digitos seguidos
]


def es_id_dudosa(proveedor: str) -> bool:
    """Marca True si el nombre tiene pinta de ser filename mal extraido."""
    if not proveedor:
        return False
    return any(p.search(proveedor) for p in ID_DUDOSA_PATTERNS)


PROJECT_ROOT = Path(__file__).resolve().parent.parent

TRIMESTRES_CANONICOS: Dict[str, Path] = {
    "1T25": PROJECT_ROOT / "outputs" / "Facturas_1T25.xlsx",
    "2T25": PROJECT_ROOT / "outputs" / "Facturas_2T25v1.xlsx",
    "3T25": PROJECT_ROOT / "outputs" / "Facturas_3T25.xlsx",
    "4T25": PROJECT_ROOT / "outputs" / "Facturas_4T25.xlsx",
}


@dataclass
class ProveedorStats:
    archivos: Set[str] = field(default_factory=set)
    ivas_por_archivo: Dict[str, Set[float]] = field(default_factory=lambda: defaultdict(set))
    descuadres_archivos: Set[str] = field(default_factory=set)
    keywords_detectadas: Set[str] = field(default_factory=set)
    trimestres: Set[str] = field(default_factory=set)


def descubrir_trimestres_disponibles() -> Dict[str, Path]:
    """Devuelve dict trimestre -> path para los Excels con hoja 'Lineas'."""
    out: Dict[str, Path] = {}
    for trim, path in TRIMESTRES_CANONICOS.items():
        if not path.exists():
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            if "Lineas" in wb.sheetnames:
                out[trim] = path
            wb.close()
        except Exception as exc:
            print(f"  WARN no se pudo abrir {path}: {exc}", file=sys.stderr)
    return out


def analizar_excel(path: Path, trimestre: str, pattern: re.Pattern) -> Dict[str, ProveedorStats]:
    """Lee un Excel y devuelve estadisticas por proveedor para ese trimestre."""
    result: Dict[str, ProveedorStats] = defaultdict(ProveedorStats)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        print(f"  WARN [{trimestre}] no se pudo abrir {path}: {exc}", file=sys.stderr)
        return dict(result)

    if "Lineas" not in wb.sheetnames:
        print(f"  WARN [{trimestre}] sin hoja Lineas: {path}", file=sys.stderr)
        wb.close()
        return dict(result)

    ws = wb["Lineas"]
    headers = [c.value for c in ws[1]]
    try:
        idx_prov = headers.index("PROVEEDOR")
        idx_art = headers.index("ARTICULO")
        idx_iva = headers.index("TIPO IVA")
        idx_cua = headers.index("CUADRE")
        idx_arc = headers.index("ARCHIVO")
    except ValueError as exc:
        print(f"  WARN [{trimestre}] columnas faltantes en {path}: {exc}", file=sys.stderr)
        wb.close()
        return dict(result)

    for row in ws.iter_rows(min_row=2, values_only=True):
        proveedor = row[idx_prov]
        archivo = row[idx_arc]
        if not proveedor or not archivo:
            continue
        rec = result[str(proveedor).strip()]
        rec.archivos.add(str(archivo))
        rec.trimestres.add(trimestre)

        iva = row[idx_iva]
        if iva is not None:
            try:
                rec.ivas_por_archivo[str(archivo)].add(float(iva))
            except (TypeError, ValueError):
                pass

        cuadre = row[idx_cua]
        cuadre_norm = str(cuadre).strip().upper() if cuadre is not None else ""
        if cuadre_norm not in {"OK", "SIN_LINEAS", ""}:
            rec.descuadres_archivos.add(str(archivo))

        articulo = row[idx_art]
        if articulo:
            match = pattern.search(str(articulo))
            if match:
                rec.keywords_detectadas.add(match.group(0).upper())

    wb.close()
    return dict(result)


def merge_resultados(resultados: List[Tuple[str, Dict[str, ProveedorStats]]]) -> Dict[str, ProveedorStats]:
    """Combina resultados de varios trimestres en un solo dict por proveedor."""
    out: Dict[str, ProveedorStats] = defaultdict(ProveedorStats)
    for _trim, by_prov in resultados:
        for proveedor, rec in by_prov.items():
            dst = out[proveedor]
            dst.archivos |= rec.archivos
            for archivo, ivas in rec.ivas_por_archivo.items():
                dst.ivas_por_archivo[archivo] |= ivas
            dst.descuadres_archivos |= rec.descuadres_archivos
            dst.keywords_detectadas |= rec.keywords_detectadas
            dst.trimestres |= rec.trimestres
    return dict(out)


def calcular_nivel(rec: ProveedorStats) -> str:
    """
    NIVEL_SOSPECHA (iva_mixto como senal primaria):
      ALTO       : iva_mixto AND descuadres > 0
      MEDIO-ALTO : iva_mixto AND keyword detectada (sin descuadre)
      MEDIO      : iva_mixto sin keyword (sin descuadre)
      BAJO       : keyword detectada pero IVA uniforme
      SIN        : ninguna

    iva_mixto = intra-factura (varios IVAs en la misma factura) OR
                inter-factura (distintos IVAs entre facturas del mismo proveedor).
    El caso inter cubre extractores que extraen 1 sola linea por factura pero
    el regimen real es mixto — sintoma tipico PANRUJE 2T26.
    """
    facturas_mixtas = [a for a, ivas in rec.ivas_por_archivo.items() if len(ivas) > 1]
    has_iva_mixto_intra = bool(facturas_mixtas)
    ivas_globales: Set[float] = set()
    for ivas in rec.ivas_por_archivo.values():
        ivas_globales |= ivas
    has_iva_mixto_inter = len(ivas_globales) > 1
    has_iva_mixto = has_iva_mixto_intra or has_iva_mixto_inter
    tiene_keyword = bool(rec.keywords_detectadas)
    descuadres = len(rec.descuadres_archivos)

    if has_iva_mixto and descuadres > 0:
        return "ALTO"
    if has_iva_mixto and tiene_keyword:
        return "MEDIO-ALTO"
    if has_iva_mixto:
        return "MEDIO"
    if tiene_keyword:
        return "BAJO"
    return "SIN"


def construir_filas(merged: Dict[str, ProveedorStats]) -> List[Dict]:
    """Devuelve lista ordenada por NIVEL_SOSPECHA / N_FACTURAS."""
    filas = []
    for proveedor, rec in merged.items():
        ivas_globales: Set[float] = set()
        for ivas in rec.ivas_por_archivo.values():
            ivas_globales |= ivas
        facturas_mixtas = [a for a, ivas in rec.ivas_por_archivo.items() if len(ivas) > 1]
        filas.append({
            "PROVEEDOR": proveedor,
            "ID_DUDOSA": es_id_dudosa(proveedor),
            "N_FACTURAS": len(rec.archivos),
            "TIENE_PORTES": bool(rec.keywords_detectadas),
            "KEYWORDS_DETECTADAS": ", ".join(sorted(rec.keywords_detectadas)),
            "IVAS_EN_LINEAS": str(sorted(ivas_globales)),
            "HAS_FACTURA_IVA_MIXTO": bool(facturas_mixtas),
            "N_FACTURAS_IVA_MIXTO": len(facturas_mixtas),
            "HAS_IVAS_DISTINTOS_ENTRE_FACTURAS": len(ivas_globales) > 1,
            "N_FACTURAS_CON_DESCUADRE": len(rec.descuadres_archivos),
            "TRIMESTRES_ANALIZADOS": ", ".join(sorted(rec.trimestres)),
            "NIVEL_SOSPECHA": calcular_nivel(rec),
        })
    orden = {"ALTO": 0, "MEDIO-ALTO": 1, "MEDIO": 2, "BAJO": 3, "SIN": 4}
    filas.sort(key=lambda f: (
        orden[f["NIVEL_SOSPECHA"]],
        f["ID_DUDOSA"],          # False (0) primero, True (1) despues
        -f["N_FACTURAS"],
    ))
    return filas


def imprimir_tabla(filas: List[Dict], incluir_sin: bool = False) -> None:
    print(f"{'PROVEEDOR':35s}  {'DUD':3s}  {'N_FAC':>5s}  {'N_MIX':>5s}  {'N_DES':>5s}  "
          f"{'IVAS':30s}  {'KEYW':25s}  NIVEL")
    print("-" * 125)
    for f in filas:
        if not incluir_sin and f["NIVEL_SOSPECHA"] == "SIN":
            continue
        print(f"{f['PROVEEDOR'][:34]:35s}  "
              f"{'!' if f['ID_DUDOSA'] else ' ':3s}  "
              f"{f['N_FACTURAS']:5d}  "
              f"{f['N_FACTURAS_IVA_MIXTO']:5d}  "
              f"{f['N_FACTURAS_CON_DESCUADRE']:5d}  "
              f"{f['IVAS_EN_LINEAS'][:30]:30s}  "
              f"{f['KEYWORDS_DETECTADAS'][:25]:25s}  "
              f"{f['NIVEL_SOSPECHA']}")
    print()


def escribir_excel(filas: List[Dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Auditoria"
    if filas:
        headers = list(filas[0].keys())
        ws.append(headers)
        for f in filas:
            ws.append([f[h] for h in headers])
    wb.save(path)


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Auditoria de IVA mixto en Facturas_<trim>.xlsx")
    parser.add_argument("--trimestres", default=None,
                        help="Lista separada por comas, ej '4T25,1T26'. Default: todos disponibles.")
    parser.add_argument("--extra-files", nargs="*", default=[],
                        help="Paths adicionales (acepta glob).")
    parser.add_argument("--keywords-extra", default="",
                        help="Keywords extra separadas por comas.")
    parser.add_argument("--no-excel", action="store_true",
                        help="Solo salida por consola.")
    parser.add_argument("--output", default=None,
                        help="Path del Excel de salida.")
    parser.add_argument("--incluir-sin", action="store_true",
                        help="Incluir nivel SIN en la tabla de consola.")
    args = parser.parse_args(argv)

    keywords = list(KEYWORDS_DEFAULT)
    if args.keywords_extra:
        keywords.extend(k.strip() for k in args.keywords_extra.split(",") if k.strip())
    pattern = re.compile("|".join(re.escape(k) for k in keywords), re.IGNORECASE)

    disponibles = descubrir_trimestres_disponibles()
    if args.trimestres:
        seleccionados = [t.strip() for t in args.trimestres.split(",")]
        files: List[Tuple[str, Path]] = [(t, disponibles[t]) for t in seleccionados if t in disponibles]
        missing = [t for t in seleccionados if t not in disponibles]
        if missing:
            print(f"WARN trimestres no disponibles: {missing}", file=sys.stderr)
    else:
        files = list(disponibles.items())

    for spec in args.extra_files:
        for path_str in sorted(glob.glob(spec)):
            files.append((Path(path_str).stem, Path(path_str)))

    if not files:
        print("ERROR: ningun archivo a analizar", file=sys.stderr)
        return 1

    print(f"Analizando {len(files)} archivo(s):")
    for label, path in files:
        print(f"  [{label}] {path}")
    print()

    resultados: List[Tuple[str, Dict[str, ProveedorStats]]] = []
    for label, path in files:
        resultados.append((label, analizar_excel(path, label, pattern)))

    merged = merge_resultados(resultados)
    filas = construir_filas(merged)

    imprimir_tabla(filas, incluir_sin=args.incluir_sin)

    if not args.no_excel:
        out_path = Path(args.output) if args.output else (
            PROJECT_ROOT / "outputs" / f"auditoria_iva_mixto_{datetime.now():%Y%m%d_%H%M}.xlsx"
        )
        out_path.parent.mkdir(parents=True, exist_ok=True)
        escribir_excel(filas, out_path)
        print(f"Excel: {out_path}")

    by_level: Dict[str, int] = defaultdict(int)
    for f in filas:
        by_level[f["NIVEL_SOSPECHA"]] += 1
    print(f"\nTotal proveedores: {len(filas)}")
    for level in ("ALTO", "MEDIO-ALTO", "MEDIO", "BAJO", "SIN"):
        print(f"  {level}: {by_level[level]}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
