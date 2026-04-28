#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
resucitar_zombis.py — Reprocesa filas zombi en PAGOS_Gmail.

Una "fila zombi" es una entrada en PAGOS_Gmail_<periodo>.xlsx que quedó
con TOTAL vacío y/o "ALERTA ROJA" en OBS porque la versión de gmail.py
que la insertó tenía bugs ya corregidos. Los extractores actuales sí
parsean correctamente esos PDFs.

Este script:
  1. Lee el Excel y detecta filas zombi (TOTAL vacío o ALERTA ROJA).
  2. Localiza el PDF original en Dropbox según el nombre del archivo.
  3. Resuelve el extractor adecuado vía MAESTRO (campo `archivo_extractor`),
     con fallback a un mapeo manual.
  4. Invoca el extractor (paridad con gmail.py: dual A/B según firma).
  5. Muestra diff actual vs propuesto y pide confirmación fila por fila.
  6. Si --apply: backup + escritura del Excel.

Uso:
    python scripts\\resucitar_zombis.py                       # dry-run, Excel default (Drive)
    python scripts\\resucitar_zombis.py --apply               # apply real (interactivo)
    python scripts\\resucitar_zombis.py --excel <ruta>        # override Excel
    python scripts\\resucitar_zombis.py --carpeta <dir>       # buscar PDFs solo aquí

TASCA BAREA S.L.L. — Abril 2026
Versión: 1.0
"""
from __future__ import annotations

import argparse
import importlib.util
import logging
import re
import shutil
import sys
import traceback
from datetime import datetime
from pathlib import Path
from typing import Optional

# Asegurar que el root del proyecto esté en sys.path antes de importar core/nucleo
_THIS_DIR = Path(__file__).resolve().parent
_PROJECT_ROOT = _THIS_DIR.parent  # gestion-facturas/
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from core.config import get_config
from nucleo.maestro import MaestroProveedores
from nucleo.pdf import extraer_texto_pdf

# Stdout en UTF-8 para emojis (Windows cp1252 los rompe)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, OSError):
    pass

# Garantizar Parseo/ y Parseo/extractores/ en sys.path
# (replicado de gmail/gmail.py:2346-2351 para que funcione `from extractores import ...`)
_CFG_INIT = get_config()
for _p in (str(_CFG_INIT.parseo_dir), str(_CFG_INIT.extractores_dir)):
    if _p not in sys.path:
        sys.path.insert(0, _p)


# ============================================================================
# CONFIGURACIÓN
# ============================================================================

EXCEL_DEFAULT = Path(
    r"G:\Mi unidad\Barea - Datos Compartidos\Compras\Año en curso"
    r"\PAGOS_Gmail_2T26.xlsx"
)

# Mapeo MANUAL fallback: solo se usa si MAESTRO no tiene `archivo_extractor`
# poblado para un proveedor. Fuente primaria es el MAESTRO (paridad con gmail.py).
MAPEO_EXTRACTORES_FALLBACK = {
    "SABORES DE PATERNA SCA": "sabores_paterna.py",
    "WEBEMPRESA EUROPA SL":   "webempresa.py",
    "MIGUEZ CAL SL":          "miguez_cal.py",
    "CERES CERVEZA SL":       "ceres.py",
    "DEBORA GARCIA TOLEDANO": "debora_garcia.py",
}

COLUMNAS_RESUCITABLES = (
    "PROVEEDOR", "CIF", "FECHA_FACTURA", "REF", "TOTAL",
    "IBAN", "FORMA_PAGO", "CUENTA",
)

# Columnas cuya fuente canónica es el MAESTRO (no el extractor): si el Excel
# ya tiene valor, no se sobreescribe — solo se logea WARN si el extractor
# discrepa, porque puede indicar incidencia (factura mal asociada, datos
# del proveedor desactualizados en MAESTRO, etc.).
COLUMNAS_PROTEGIDAS_MAESTRO = ("PROVEEDOR", "CIF", "IBAN")

PATRON_ALERTA_OBS = "ALERTA ROJA"

# Formato canónico de fecha en el Excel (.claude/rules/excel.md)
FORMATO_FECHA_EXCEL = "%d/%m/%y"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("resucitar_zombis")


# ============================================================================
# BÚSQUEDA DE PDFs (réplica simplificada de la lógica de gmail.py)
# ============================================================================

def deducir_trimestre_year(nombre_archivo: str) -> Optional[tuple[int, int]]:
    """Deduce (trimestre, año) del nombre, ej. '2T26 ...' → (2, 2026)."""
    m = re.search(r"(\d)T(\d{2})", nombre_archivo)
    if not m:
        return None
    return (int(m.group(1)), 2000 + int(m.group(2)))


def carpetas_busqueda_pdf(nombre_archivo: str, dropbox_base: Path,
                          excel_trim_year: Optional[tuple[int, int]] = None
                          ) -> list[Path]:
    """Lista carpetas candidatas donde puede estar el PDF, en orden de prioridad.

    Si nombre_archivo empieza por 'ATRASADA' y se conoce el trimestre del
    Excel destino (excel_trim_year), `<excel_trim>/ATRASADAS` se añade como
    PRIMERA candidata: las facturas atrasadas se contabilizan en el trimestre
    actual, no en el de su fecha original.
    """
    candidatas: list[Path] = []
    if not dropbox_base or not dropbox_base.exists():
        return candidatas

    es_atrasada = nombre_archivo.upper().startswith("ATRASADA")

    # PRIORIDAD 1: si es ATRASADA, carpeta ATRASADAS del trimestre del Excel
    if es_atrasada and excel_trim_year:
        t_exc, y_exc = excel_trim_year
        candidatas.append(
            dropbox_base / f"FACTURAS {y_exc}" / "FACTURAS RECIBIDAS"
            / f"{t_exc} TRIMESTRE {y_exc}" / "ATRASADAS"
        )

    info = deducir_trimestre_year(nombre_archivo)
    if not info:
        return candidatas

    trimestre, year = info
    raiz_year = dropbox_base / f"FACTURAS {year}" / "FACTURAS RECIBIDAS"
    carpeta_trim = raiz_year / f"{trimestre} TRIMESTRE {year}"
    if carpeta_trim not in candidatas:
        candidatas.append(carpeta_trim)
    if (carpeta_trim / "ATRASADAS") not in candidatas:
        candidatas.append(carpeta_trim / "ATRASADAS")

    if es_atrasada:
        for t in range(1, 5):
            if t != trimestre:
                otra = raiz_year / f"{t} TRIMESTRE {year}" / "ATRASADAS"
                if otra not in candidatas:
                    candidatas.append(otra)

    for t in range(1, 5):
        if t != trimestre:
            otra = raiz_year / f"{t} TRIMESTRE {year}"
            if otra not in candidatas:
                candidatas.append(otra)

    # Año anterior (ATRASADAS de 4T25 archivadas en estructura del 26)
    if year >= 2025:
        for t in range(1, 5):
            otra = (dropbox_base / f"FACTURAS {year - 1}" / "FACTURAS RECIBIDAS"
                    / f"{t} TRIMESTRE {year - 1}")
            if otra not in candidatas:
                candidatas.append(otra)

    return candidatas


def buscar_pdf(nombre_archivo: str, dropbox_base: Path,
               carpeta_override: Optional[Path],
               excel_trim_year: Optional[tuple[int, int]] = None
               ) -> Optional[Path]:
    """Busca el PDF en disco; soporta variante Kinema (prefijo numérico)."""
    if not nombre_archivo.lower().endswith(".pdf"):
        nombre_archivo = nombre_archivo + ".pdf"

    carpetas: list[Path] = []
    if carpeta_override:
        carpetas.append(carpeta_override)
    carpetas.extend(carpetas_busqueda_pdf(nombre_archivo, dropbox_base, excel_trim_year))

    patron_kinema = re.compile(rf"^\d+ {re.escape(nombre_archivo)}$")

    for carpeta in carpetas:
        if not carpeta.exists():
            continue
        candidato = carpeta / nombre_archivo
        if candidato.is_file():
            return candidato
        try:
            for f in carpeta.iterdir():
                if f.is_file() and patron_kinema.match(f.name):
                    return f
        except (PermissionError, OSError):
            continue

    return None


# ============================================================================
# CARGA E INVOCACIÓN DE EXTRACTORES
# ============================================================================

def resolver_archivo_extractor(proveedor: str,
                               maestro: MaestroProveedores) -> Optional[str]:
    """Resuelve nombre del .py del extractor.

    Primario: MAESTRO_PROVEEDORES.archivo_extractor (paridad con gmail.py).
    Fallback: dict manual MAPEO_EXTRACTORES_FALLBACK.
    """
    nombre_upper = proveedor.upper().strip()
    prov_obj = maestro.proveedores.get(nombre_upper)
    if prov_obj and prov_obj.archivo_extractor:
        return prov_obj.archivo_extractor
    return MAPEO_EXTRACTORES_FALLBACK.get(nombre_upper)


def importar_extractor_modulo(nombre_archivo: str, dir_extractores: Path):
    """Carga el módulo .py del extractor por nombre de archivo."""
    if not nombre_archivo.endswith(".py"):
        nombre_archivo += ".py"

    ruta = dir_extractores / nombre_archivo
    if not ruta.is_file():
        log.error(f"  ↳ ❌ Extractor no encontrado: {ruta}")
        return None

    nombre_modulo = f"extractor_{nombre_archivo[:-3]}"
    spec = importlib.util.spec_from_file_location(nombre_modulo, ruta)
    if not spec or not spec.loader:
        log.error(f"  ↳ ❌ Spec inválido para {ruta}")
        return None

    modulo = importlib.util.module_from_spec(spec)
    sys.modules[nombre_modulo] = modulo
    try:
        spec.loader.exec_module(modulo)
    except Exception as e:
        log.error(f"  ↳ ❌ Error cargando {nombre_archivo}: {e}")
        return None
    return modulo


def aplicar_extractor(modulo, pdf_path: Path) -> Optional[dict]:
    """Invoca el extractor sobre un PDF y normaliza el retorno.

    ──────────────────────────────────────────────────────────────────
    Lógica replicada de gmail/gmail.py:2376-2453.
    Mantener sincronizada con esa función. Si gmail.py cambia el flujo
    dual A/B, este script debe actualizarse también.
    ──────────────────────────────────────────────────────────────────

    Patrón A (CERES y similares): clase con método extraer(pdf_path) que
        abre el PDF y devuelve dict con todas las claves.
    Patrón B (la mayoría): clase solo con extraer_lineas/total/fecha/referencia
        que reciben TEXTO. Aquí extraemos el texto con pdfplumber + fallback OCR
        (vía nucleo.pdf.extraer_texto_pdf) y luego invocamos cada hook.
    """
    clase_extractor = None
    for nombre in dir(modulo):
        obj = getattr(modulo, nombre)
        if (isinstance(obj, type)
                and nombre.startswith("Extractor")
                and nombre != "ExtractorBase"):
            clase_extractor = obj
            break

    if not clase_extractor:
        log.error("  ↳ ❌ Módulo sin clase Extractor*")
        return None

    try:
        instancia = clase_extractor()
    except Exception as e:
        log.error(f"  ↳ ❌ No se puede instanciar {clase_extractor.__name__}: {e}")
        return None

    datos: Optional[dict] = None

    # Patrón A: instancia.extraer(pdf_path) devuelve dict
    if hasattr(instancia, "extraer") and callable(getattr(instancia, "extraer")):
        try:
            datos = instancia.extraer(str(pdf_path))
        except Exception as e:
            log.error(f"  ↳ ❌ extraer() lanzó excepción: {e}")
            traceback.print_exc()
            return None

    # Patrón B: hooks individuales sobre TEXTO
    elif hasattr(instancia, "extraer_total") or hasattr(instancia, "extraer_fecha"):
        try:
            texto_pdf = extraer_texto_pdf(pdf_path, metodo="pdfplumber", fallback=True)
        except Exception as e:
            log.error(f"  ↳ ❌ Error extrayendo texto del PDF: {e}")
            return None

        if not texto_pdf:
            log.error("  ↳ ❌ Texto vacío tras pdfplumber+OCR")
            return None

        datos = {}
        for hook, key in (
            ("extraer_fecha",      "fecha"),
            ("extraer_total",      "total"),
            ("extraer_referencia", "referencia"),
        ):
            if hasattr(instancia, hook):
                try:
                    datos[key] = getattr(instancia, hook)(texto_pdf)
                except Exception as e:
                    log.warning(f"  ↳ ⚠️ {hook} falló: {e}")

    else:
        log.error("  ↳ ❌ Clase sin métodos extractor reconocibles")
        return None

    if not datos:
        log.error("  ↳ ❌ Extractor no devolvió datos")
        return None

    cif_fijo = getattr(instancia, "cif", "") or ""
    iban_fijo = getattr(instancia, "iban", "") or ""

    def get_first(d, *keys):
        for k in keys:
            if k in d and d[k] not in (None, "", 0, 0.0):
                return d[k]
        return None

    return {
        "PROVEEDOR":     get_first(datos, "PROVEEDOR", "proveedor", "nombre"),
        "CIF":           get_first(datos, "CIF", "cif", "nif") or (cif_fijo or None),
        "FECHA_FACTURA": get_first(datos, "FECHA_FACTURA", "fecha_factura", "fecha"),
        "REF":           get_first(datos, "REF", "referencia", "ref",
                                   "numero", "numero_factura"),
        "TOTAL":         get_first(datos, "TOTAL", "total", "importe", "total_factura"),
        "IBAN":          get_first(datos, "IBAN", "iban") or (iban_fijo or None),
        "FORMA_PAGO":    get_first(datos, "FORMA_PAGO", "forma_pago", "metodo_pago"),
        "CUENTA":        get_first(datos, "CUENTA", "cuenta"),
    }


# ============================================================================
# DETECCIÓN DE ZOMBIS
# ============================================================================

def es_fila_zombi(fila: dict) -> tuple[bool, str]:
    """Decide si una fila debe reprocesarse."""
    obs = str(fila.get("OBS") or "")
    total = fila.get("TOTAL")
    archivo = fila.get("ARCHIVO")
    proveedor = fila.get("PROVEEDOR")

    if not archivo:
        return (False, "sin nombre de archivo")
    if not proveedor:
        return (False, "sin proveedor identificado (requiere alta MAESTRO)")
    if PATRON_ALERTA_OBS in obs:
        return (True, "OBS contiene ALERTA ROJA")
    if total in (None, "", 0, 0.0):
        return (True, "TOTAL vacío")
    return (False, "fila OK")


# ============================================================================
# UI INTERACTIVA
# ============================================================================

def normalizar_fecha_dd_mm_yy(valor) -> Optional[str]:
    """Devuelve la fecha en formato DD/MM/YY (estándar del Excel) o None.

    Acepta datetime, o string en formatos: DD/MM/YY, DD/MM/YYYY, DD-MM-YY,
    DD-MM-YYYY, YYYY-MM-DD. Si no se reconoce, devuelve la cadena original.
    """
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.strftime(FORMATO_FECHA_EXCEL)
    s = str(valor).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime(FORMATO_FECHA_EXCEL)
        except ValueError:
            continue
    return s


def normalizar_para_diff(col: str, valor) -> str:
    """Normaliza un valor para comparación de diff."""
    if valor is None:
        return ""
    if col == "FECHA_FACTURA":
        norm = normalizar_fecha_dd_mm_yy(valor)
        return norm if norm is not None else ""
    return str(valor).strip()


def normalizar_propuesto(propuesto: dict) -> dict:
    """Normaliza valores propuestos antes de diff/escritura.

    En particular: FECHA_FACTURA siempre como DD/MM/YY (estándar del proyecto).
    """
    if propuesto.get("FECHA_FACTURA"):
        propuesto["FECHA_FACTURA"] = normalizar_fecha_dd_mm_yy(
            propuesto["FECHA_FACTURA"]
        ) or propuesto["FECHA_FACTURA"]
    return propuesto


def mostrar_diff(actual: dict, propuesto: dict) -> list[str]:
    """Imprime diff y devuelve lista de columnas que cambian.

    Reglas:
    - Si actual y propuesto son equivalentes (incluido fechas mismo día),
      no se cuenta como cambio.
    - Para columnas en COLUMNAS_PROTEGIDAS_MAESTRO, si Excel ya tiene valor
      y discrepa del extractor: se loga WARN pero NO se sobreescribe (la
      fuente canónica es el MAESTRO, no el extractor).
    """
    cambios = []
    for col in COLUMNAS_RESUCITABLES:
        v_a = actual.get(col)
        v_p = propuesto.get(col)
        if v_p is None:
            continue
        if normalizar_para_diff(col, v_a) == normalizar_para_diff(col, v_p):
            continue

        # Protección MAESTRO: NO sobreescribir si Excel tiene valor.
        if col in COLUMNAS_PROTEGIDAS_MAESTRO and v_a not in (None, ""):
            log.warning(
                f"  ⚠️ {col}: extractor devolvió {v_p!r}, "
                f"Excel/MAESTRO tiene {v_a!r} → se mantiene Excel"
            )
            continue

        cambios.append(col)
        simbolo = "+" if v_a in (None, "") else "Δ"
        print(f"    {simbolo} {col:18s}: {repr(v_a):<32} -> {repr(v_p)}")
    return cambios


def pedir_confirmacion(prompt: str) -> str:
    while True:
        r = input(f"  {prompt} [s/n/e/q]: ").strip().lower()
        if r in ("s", "n", "e", "q"):
            return r
        print("    Respuestas válidas: s (sí), n (no), e (editar), q (salir)")


def editar_manual(propuesto: dict) -> dict:
    """Permite override manual de campos clave + nota OBS."""
    print("    --- EDICIÓN MANUAL (Enter para mantener valor) ---")
    editado = dict(propuesto)
    for col in ("FECHA_FACTURA", "REF", "TOTAL", "FORMA_PAGO"):
        actual = editado.get(col, "")
        nuevo = input(f"    {col} [{actual}]: ").strip()
        if nuevo:
            editado[col] = nuevo
    nota = input("    Nota OBS (opcional): ").strip()
    if nota:
        editado["__OBS_EXTRA__"] = nota
    return editado


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Reprocesa filas zombi en PAGOS_Gmail."
    )
    parser.add_argument("--excel", type=Path, default=EXCEL_DEFAULT,
                        help=f"Ruta al Excel (default: {EXCEL_DEFAULT})")
    parser.add_argument("--carpeta", type=Path, default=None,
                        help="Override carpeta donde buscar PDFs")
    parser.add_argument("--apply", action="store_true",
                        help="Aplicar cambios. Sin esto, modo dry-run.")
    parser.add_argument("--hoja", default="FACTURAS",
                        help="Nombre de la hoja (default: FACTURAS)")
    args = parser.parse_args()

    if not args.excel.exists():
        log.error(f"❌ Excel no encontrado: {args.excel}")
        sys.exit(1)
    if args.carpeta and not args.carpeta.exists():
        log.error(f"❌ --carpeta no existe: {args.carpeta}")
        sys.exit(1)

    cfg = get_config()
    modo = "APPLY (escribirá)" if args.apply else "DRY-RUN (no escribe)"
    log.info("=" * 70)
    log.info(f"resucitar_zombis.py v1.0 — MODO: {modo}")
    log.info(f"  Excel:        {args.excel}")
    log.info(f"  Extractores:  {cfg.extractores_dir}")
    log.info(f"  Dropbox base: {cfg.dropbox_base}")
    log.info(f"  MAESTRO:      {cfg.maestro_path}")
    log.info("=" * 70)

    if not cfg.maestro_path.exists():
        log.error(f"❌ MAESTRO no existe: {cfg.maestro_path}")
        sys.exit(1)
    if not cfg.extractores_dir.exists():
        log.error(f"❌ Carpeta de extractores no existe: {cfg.extractores_dir}")
        sys.exit(1)

    maestro = MaestroProveedores(str(cfg.maestro_path))
    log.info(f"📚 MAESTRO cargado: {len(maestro.proveedores)} proveedores")

    # Backup ANTES de cargar el Excel para escritura
    if args.apply:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        backup = args.excel.with_name(
            f"{args.excel.stem}_backup_{ts}{args.excel.suffix}"
        )
        try:
            shutil.copy2(args.excel, backup)
            log.info(f"📦 Backup creado: {backup.name}")
        except Exception as e:
            log.error(f"❌ Error creando backup: {e}")
            sys.exit(1)

    wb: Workbook = openpyxl.load_workbook(args.excel)
    if args.hoja not in wb.sheetnames:
        log.error(f"❌ Hoja '{args.hoja}' no existe. Disponibles: {wb.sheetnames}")
        sys.exit(1)
    ws: Worksheet = wb[args.hoja]
    columnas = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {c: i + 1 for i, c in enumerate(columnas) if c}

    # Deducir trimestre/año del nombre del Excel para localizar ATRASADAS
    # (PAGOS_Gmail_2T26.xlsx → (2, 2026))
    excel_trim_year = deducir_trimestre_year(args.excel.stem)
    if excel_trim_year:
        log.info(f"📅 Trimestre del Excel: {excel_trim_year[0]}T{excel_trim_year[1]}")

    zombis = []
    for r in range(2, ws.max_row + 1):
        fila = {c: ws.cell(r, col_idx[c]).value for c in col_idx}
        es_z, motivo = es_fila_zombi(fila)
        if es_z:
            zombis.append((r, fila, motivo))

    log.info(f"\n🔍 Filas zombi detectadas: {len(zombis)}\n")
    if not zombis:
        log.info("Nada que hacer. Excel ya está limpio.")
        return

    resumen = {"resucitadas": 0, "saltadas": 0, "editadas": 0, "fallidas": 0}

    for idx, (fila_excel, fila, motivo) in enumerate(zombis, 1):
        print("\n" + "─" * 70)
        log.info(f"[{idx}/{len(zombis)}] Fila Excel {fila_excel} — {motivo}")
        log.info(f"  Archivo:   {fila.get('ARCHIVO')}")
        log.info(f"  Proveedor: {fila.get('PROVEEDOR')}")

        archivo = str(fila.get("ARCHIVO") or "")
        pdf = buscar_pdf(archivo, cfg.dropbox_base, args.carpeta, excel_trim_year)
        if not pdf:
            log.error("  ❌ PDF no encontrado en ninguna carpeta")
            resumen["fallidas"] += 1
            continue
        log.info(f"  📄 PDF localizado: {pdf}")

        proveedor = str(fila.get("PROVEEDOR") or "")
        archivo_ext = resolver_archivo_extractor(proveedor, maestro)
        if not archivo_ext:
            log.warning(f"  ⚠️ Sin extractor para '{proveedor}' "
                        f"(ni MAESTRO ni fallback manual)")
            resumen["fallidas"] += 1
            continue
        log.info(f"  🧩 Extractor: {archivo_ext}")

        modulo = importar_extractor_modulo(archivo_ext, cfg.extractores_dir)
        if not modulo:
            resumen["fallidas"] += 1
            continue

        propuesto = aplicar_extractor(modulo, pdf)
        if not propuesto:
            resumen["fallidas"] += 1
            continue
        propuesto = normalizar_propuesto(propuesto)

        print("\n  📋 Cambios propuestos:")
        cambios = mostrar_diff(fila, propuesto)
        if not cambios:
            log.info("  ℹ️ Sin cambios reales")
            resumen["saltadas"] += 1
            continue

        # Dry-run: solo reportamos. Apply: preguntamos.
        if not args.apply:
            log.info(f"  ✅ (dry-run) Fila {fila_excel} se actualizaría "
                     f"con {len(cambios)} cambios")
            resumen["resucitadas"] += 1
            continue

        accion = pedir_confirmacion(f"¿Aplicar los {len(cambios)} cambios?")
        if accion == "q":
            log.info("⛔ Salida solicitada por el usuario.")
            break
        if accion == "n":
            log.info("  ⏭️  Saltada por el usuario")
            resumen["saltadas"] += 1
            continue
        if accion == "e":
            propuesto = editar_manual(propuesto)
            propuesto = normalizar_propuesto(propuesto)
            # Recalcular cambios tras edición
            cambios = mostrar_diff(fila, propuesto)
            resumen["editadas"] += 1

        # Escribir
        for col in cambios:
            if col in col_idx:
                ws.cell(fila_excel, col_idx[col]).value = propuesto.get(col)

        # Actualizar OBS
        if "OBS" in col_idx:
            obs_actual = str(fila.get("OBS") or "")
            nota_extra = propuesto.get("__OBS_EXTRA__", "")
            if PATRON_ALERTA_OBS in obs_actual:
                nuevo_obs = f"resucitada {datetime.now():%Y-%m-%d}"
                if nota_extra:
                    nuevo_obs += f" | {nota_extra}"
                ws.cell(fila_excel, col_idx["OBS"]).value = nuevo_obs
            elif nota_extra:
                ws.cell(fila_excel, col_idx["OBS"]).value = (
                    obs_actual + " | " + nota_extra
                ).strip(" |")

        log.info(f"  ✅ Fila {fila_excel} actualizada")
        resumen["resucitadas"] += 1

    # Guardar
    if args.apply and resumen["resucitadas"] > 0:
        try:
            wb.save(args.excel)
            log.info(f"\n💾 Excel guardado: {args.excel}")
        except PermissionError:
            log.error("\n❌ Excel ABIERTO en otra app. NO se ha guardado.")
            log.error("   Cierra el Excel y reintenta. El backup está intacto.")
            sys.exit(2)

    print("\n" + "=" * 70)
    log.info("RESUMEN")
    log.info(f"  Resucitadas: {resumen['resucitadas']}")
    log.info(f"  Saltadas:    {resumen['saltadas']}")
    log.info(f"  Editadas:    {resumen['editadas']}")
    log.info(f"  Fallidas:    {resumen['fallidas']}")
    if not args.apply:
        log.info("ℹ️  Era DRY-RUN. Para aplicar, repite con --apply")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
