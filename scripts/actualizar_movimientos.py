#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
actualizar_movimientos.py — Actualiza Movimientos_Cuenta_XX.xlsx
con nuevas descargas de Banco Sabadell (.xls)

TASCA BAREA S.L.L. — Abril 2026

Uso:
    python actualizar_movimientos.py descarga1.xls descarga2.xls
    python actualizar_movimientos.py --carpeta C:\\Users\\jaime\\Downloads
    python actualizar_movimientos.py --info

Requisitos: pip install xlrd openpyxl
"""

import argparse
import os
import sys
import re
import shutil
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import xlrd
except ImportError:
    sys.exit("ERROR: pip install xlrd")
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    sys.exit("ERROR: pip install openpyxl")


# ── Configuración ──────────────────────────────────────────────
CUENTAS = {
    "1844495": "Tasca",
    "1992404": "Comestibles",
}
COLS = ["#", "F. Operativa", "Concepto", "F. Valor", "Importe", "Saldo", "Referencia 1", "Referencia 2"]
HEADER_ROW_XLS = 8   # fila 8 (0-indexed) en los .xls de Sabadell
CUENTA_ROW = 3        # fila 3 col 1 contiene el número de cuenta

# Estilos Excel
HEADER_FONT = Font(name="Aptos", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
DATE_FMT = "DD/MM/YYYY"
MONEY_FMT = '#,##0.00'


# ── Lectura de .xls de Sabadell ────────────────────────────────
def detectar_cuenta(ws):
    """Detecta Tasca/Comestibles a partir de la fila de cuenta."""
    val = str(ws.cell_value(CUENTA_ROW, 1))
    for num, nombre in CUENTAS.items():
        if num in val:
            return nombre
    return None


def leer_xls_sabadell(filepath, año_esperado=None):
    """Lee un .xls de Sabadell y devuelve (cuenta, [filas]).
    Cada fila es un dict con las claves de COLS (sin #).

    Args:
        filepath: ruta al .xls
        año_esperado: si se proporciona, valida que los datos sean de ese año

    Raises:
        ValueError: si la cuenta no se reconoce o el año no coincide
    """
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    cuenta = detectar_cuenta(ws)
    if not cuenta:
        raise ValueError(f"No se reconoce la cuenta en {filepath}")

    # Validar año del rango de fechas (fila 6: "Desde DD / MM / YYYY hasta DD / MM / YYYY.")
    if año_esperado:
        try:
            seleccion = str(ws.cell_value(6, 1)).strip()
            años_encontrados = re.findall(r'(\d{4})', seleccion)
            if años_encontrados:
                año_archivo = int(años_encontrados[0])
                if año_archivo != año_esperado:
                    raise ValueError(
                        f"AÑO INCORRECTO: {os.path.basename(filepath)} "
                        f"contiene datos de {año_archivo} pero el consolidado es de {año_esperado}. "
                        f"Fila 'Selección': {seleccion}. "
                        f"¿Descargaste el rango correcto en Sabadell?"
                    )
        except (IndexError, TypeError):
            pass  # Si no se puede leer fila 6, continuar sin validar

    filas = []
    for r in range(HEADER_ROW_XLS + 1, ws.nrows):
        fecha_op_raw = ws.cell_value(r, 0)
        if not fecha_op_raw:
            continue

        # Parsear fecha (texto "DD/MM/YYYY")
        try:
            fecha_op = datetime.strptime(str(fecha_op_raw).strip(), "%d/%m/%Y")
        except ValueError:
            continue

        concepto = str(ws.cell_value(r, 1)).strip()
        fecha_val_raw = ws.cell_value(r, 2)
        try:
            fecha_val = datetime.strptime(str(fecha_val_raw).strip(), "%d/%m/%Y")
        except ValueError:
            fecha_val = fecha_op

        importe = float(ws.cell_value(r, 3)) if ws.cell_value(r, 3) != "" else 0.0
        saldo = float(ws.cell_value(r, 4)) if ws.cell_value(r, 4) != "" else 0.0
        ref1 = str(ws.cell_value(r, 5)).strip() if ws.cell_value(r, 5) else None
        ref2 = str(ws.cell_value(r, 6)).strip() if ws.cell_value(r, 6) else None

        filas.append({
            "F. Operativa": fecha_op,
            "Concepto": concepto,
            "F. Valor": fecha_val,
            "Importe": importe,
            "Saldo": saldo,
            "Referencia 1": ref1,
            "Referencia 2": ref2,
        })

    # Sabadell entrega en orden cronológico inverso → invertir
    filas.reverse()
    return cuenta, filas


# ── Clave de deduplicación ─────────────────────────────────────
def clave(row):
    """Clave compuesta: F.Operativa + Concepto + Importe + Saldo."""
    fop = row["F. Operativa"]
    if isinstance(fop, datetime):
        fop = fop.strftime("%Y-%m-%d")
    return (str(fop), row["Concepto"], round(row["Importe"], 2), round(row["Saldo"], 2))


# ── Lectura del consolidado existente ──────────────────────────
def leer_consolidado(filepath, sheet_name):
    """Lee una pestaña del consolidado y devuelve lista de dicts."""
    wb = load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    filas = []
    for r in range(2, ws.max_row + 1):
        fop = ws.cell(r, 2).value
        if fop is None:
            continue
        filas.append({
            "F. Operativa": fop if isinstance(fop, datetime) else datetime.strptime(str(fop), "%d/%m/%Y"),
            "Concepto": str(ws.cell(r, 3).value or ""),
            "F. Valor": ws.cell(r, 4).value or fop,
            "Importe": float(ws.cell(r, 5).value or 0),
            "Saldo": float(ws.cell(r, 6).value or 0),
            "Referencia 1": ws.cell(r, 7).value,
            "Referencia 2": ws.cell(r, 8).value,
        })
    return filas


# ── Escritura del consolidado ──────────────────────────────────
def escribir_pestaña(wb, sheet_name, filas):
    """Escribe/reescribe una pestaña. Las filas ya vienen en orden correcto."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Headers
    for c, col_name in enumerate(COLS, 1):
        cell = ws.cell(1, c, col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    # Escribir datos (ya vienen en orden correcto — NO reordenar)
    for i, row in enumerate(filas, 1):
        r = i + 1
        ws.cell(r, 1, i).border = BORDER  # #
        ws.cell(r, 1).alignment = Alignment(horizontal="center")

        for col_idx, key in [(2, "F. Operativa"), (4, "F. Valor")]:
            c = ws.cell(r, col_idx, row[key])
            c.number_format = DATE_FMT
            c.border = BORDER

        ws.cell(r, 3, row["Concepto"]).border = BORDER

        for col_idx, key in [(5, "Importe"), (6, "Saldo")]:
            c = ws.cell(r, col_idx, row[key])
            c.number_format = MONEY_FMT
            c.border = BORDER

        ws.cell(r, 7, row.get("Referencia 1") or "").border = BORDER
        ws.cell(r, 8, row.get("Referencia 2") or "").border = BORDER

    # Anchos de columna
    anchos = [6, 14, 65, 14, 13, 14, 18, 22]
    for i, w in enumerate(anchos, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Inmovilizar fila de cabecera
    ws.freeze_panes = "A2"

    return len(filas)


# ── Detectar año del consolidado ───────────────────────────────
def detectar_archivo_consolidado(directorio):
    """Busca Movimientos_Cuenta_XX.xlsx en el directorio."""
    patron = re.compile(r"Movimientos_Cuenta_(\d{2})\.xlsx", re.IGNORECASE)
    encontrados = []
    for f in os.listdir(directorio):
        m = patron.match(f)
        if m:
            encontrados.append((f, int(m.group(1))))
    return encontrados


# ── Main ───────────────────────────────────────────────────────
def actualizar(archivos_xls, consolidado_path, dry_run=False):
    """Proceso principal. Retorna dict con resumen de cambios."""
    resultado = {
        "archivos_procesados": [],
        "nuevos_tasca": 0,
        "nuevos_comestibles": 0,
        "duplicados_ignorados": 0,
        "avisos": [],
        "backup_path": None,
        "total_tasca": 0,
        "total_comestibles": 0,
    }

    # Detectar año del consolidado
    año_consolidado = None
    match = re.search(r"_(\d{2})\.", os.path.basename(consolidado_path))
    if match:
        año_consolidado = 2000 + int(match.group(1))

    # Leer todos los .xls nuevos (con validación de año)
    nuevos = {"Tasca": [], "Comestibles": []}
    for f in archivos_xls:
        try:
            cuenta, filas = leer_xls_sabadell(f, año_esperado=año_consolidado)
            nuevos[cuenta].extend(filas)
            resultado["archivos_procesados"].append({
                "archivo": os.path.basename(f), "cuenta": cuenta,
                "movimientos": len(filas),
            })
        except ValueError as e:
            resultado["archivos_procesados"].append({
                "archivo": os.path.basename(f), "cuenta": None,
                "movimientos": 0, "error": str(e), "rechazado": True,
            })
            resultado["avisos"].append(f"❌ {os.path.basename(f)}: {e}")
        except Exception as e:
            resultado["avisos"].append(f"Error leyendo {f}: {e}")

    if not any(nuevos.values()):
        resultado["avisos"].append("No se encontraron movimientos nuevos en los archivos proporcionados.")
        return resultado

    # Filtrar movimientos de año incorrecto (resumen compacto)
    if año_consolidado:
        for cuenta in nuevos:
            descartados = [f for f in nuevos[cuenta] if f["F. Operativa"].year != año_consolidado]
            if descartados:
                años_malos = set(f["F. Operativa"].year for f in descartados)
                resultado["avisos"].append(
                    f"❌ {cuenta}: {len(descartados)} movimientos de "
                    f"{', '.join(str(a) for a in sorted(años_malos))} descartados "
                    f"(consolidado es de {año_consolidado})"
                )
            nuevos[cuenta] = [f for f in nuevos[cuenta] if f["F. Operativa"].year == año_consolidado]

    # Backup
    if os.path.exists(consolidado_path) and not dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(consolidado_path)
        backup = f"{base}_backup_{ts}{ext}"
        shutil.copy2(consolidado_path, backup)
        resultado["backup_path"] = backup

    # Cargar consolidado existente
    if os.path.exists(consolidado_path):
        wb = load_workbook(consolidado_path)
    else:
        from openpyxl import Workbook
        wb = Workbook()
        # Eliminar sheet por defecto
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for cuenta in ["Tasca", "Comestibles"]:
        existentes = leer_consolidado(consolidado_path, cuenta) if os.path.exists(consolidado_path) else []
        claves_existentes = set(clave(r) for r in existentes)

        añadidos = 0
        for fila in nuevos[cuenta]:
            if clave(fila) not in claves_existentes:
                existentes.append(fila)
                claves_existentes.add(clave(fila))
                añadidos += 1
            else:
                resultado["duplicados_ignorados"] += 1

        total = escribir_pestaña(wb, cuenta, existentes)

        if cuenta == "Tasca":
            resultado["nuevos_tasca"] = añadidos
            resultado["total_tasca"] = total
        else:
            resultado["nuevos_comestibles"] = añadidos
            resultado["total_comestibles"] = total

    # Asegurar orden de pestañas
    orden = ["Tasca", "Comestibles"]
    for i, name in enumerate(orden):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    if not dry_run:
        wb.save(consolidado_path)

    return resultado


def mostrar_info(consolidado_path):
    """Muestra estado actual del consolidado."""
    if not os.path.exists(consolidado_path):
        print(f"No existe: {consolidado_path}")
        return

    wb = load_workbook(consolidado_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        n = ws.max_row - 1
        if n < 1:
            continue
        primera = ws.cell(2, 2).value
        ultima = ws.cell(ws.max_row, 2).value
        print(f"  {sheet_name}: {n} movimientos | {primera:%d/%m/%Y} → {ultima:%d/%m/%Y}")


# ── CLI ────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Actualizar movimientos bancarios consolidados")
    parser.add_argument("archivos", nargs="*", help="Archivos .xls de Sabadell")
    parser.add_argument("--carpeta", help="Carpeta con archivos .xls")
    parser.add_argument("--consolidado", default=None, help="Ruta del archivo consolidado")
    parser.add_argument("--info", action="store_true", help="Mostrar estado actual")
    parser.add_argument("--dry-run", action="store_true", help="Simular sin escribir")
    args = parser.parse_args()

    # Encontrar consolidado (R.4: ahora vive en Drive Desktop)
    DIR_MOV_BANCO = r"G:\Mi unidad\Barea - Datos Compartidos\Movimientos Banco\Año en curso"
    if args.consolidado:
        consolidado = args.consolidado
    else:
        # Buscar en la carpeta Drive del año en curso
        try:
            encontrados = detectar_archivo_consolidado(DIR_MOV_BANCO)
        except (FileNotFoundError, OSError):
            encontrados = []
        if encontrados:
            consolidado = os.path.join(DIR_MOV_BANCO, encontrados[0][0])
        else:
            consolidado = os.path.join(DIR_MOV_BANCO, "Movimientos_Cuenta_26.xlsx")

    if args.info:
        print(f"Consolidado: {consolidado}")
        mostrar_info(consolidado)
        return

    # Recoger archivos
    archivos = list(args.archivos)
    if args.carpeta:
        for f in os.listdir(args.carpeta):
            if f.lower().endswith(".xls") and not f.lower().endswith(".xlsx"):
                archivos.append(os.path.join(args.carpeta, f))

    if not archivos:
        parser.print_help()
        return

    print(f"Consolidado: {consolidado}")
    print(f"Archivos a procesar: {len(archivos)}")

    res = actualizar(archivos, consolidado, dry_run=args.dry_run)

    for a in res["archivos_procesados"]:
        print(f"  ✓ {a['archivo']} → {a['cuenta']} ({a['movimientos']} movimientos)")

    print(f"\nNuevos incorporados: Tasca={res['nuevos_tasca']}, Comestibles={res['nuevos_comestibles']}")
    print(f"Duplicados ignorados: {res['duplicados_ignorados']}")
    print(f"Total: Tasca={res['total_tasca']}, Comestibles={res['total_comestibles']}")

    if res["backup_path"]:
        print(f"Backup: {res['backup_path']}")

    for aviso in res["avisos"]:
        print(aviso)

    if args.dry_run:
        print("\n[DRY-RUN] No se ha escrito nada.")
    else:
        # Sync a Google Drive (Movimientos Banco/Año en curso/) — best-effort
        try:
            import sys as _sys
            _sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            from nucleo.sync_drive import sync_archivos
            sync_archivos([consolidado], carpeta=["Movimientos Banco", "Año en curso"])
            print(f"[DRIVE OK] {consolidado} → Movimientos Banco/Año en curso/")
        except Exception as e:
            print(f"[DRIVE FALLO] {e}")


if __name__ == "__main__":
    main()
