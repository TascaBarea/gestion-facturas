#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
mov_banco.py — Actualiza el archivo de movimientos bancarios
TASCA BAREA S.L.L. — Marzo 2026

Lee archivos .xls descargados de Banco Sabadell y los añade al archivo
Movimientos Cuenta 26.xlsx (pestañas Tasca y Comestibles).

Uso:
    python mov_banco.py archivo1.xls archivo2.xls
    python mov_banco.py --carpeta C:\\Descargas
    python mov_banco.py --info
    python mov_banco.py --dry-run archivo.xls

Requisitos: pip install xlrd openpyxl
"""

import argparse
import os
import sys
import re
from datetime import datetime
from collections import defaultdict
from typing import Optional

# Windows cp1252 no soporta caracteres especiales
if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

try:
    import xlrd
except ImportError:
    print("ERROR: Falta xlrd. Ejecuta: pip install xlrd")
    sys.exit(1)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Falta openpyxl. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ============================================================================
# CONFIGURACIÓN
# ============================================================================

# Archivo de trabajo
ARCHIVO_MOV = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "datos", "Movimientos Cuenta 26.xlsx"
)

# Mapeo cuenta bancaria → pestaña
CUENTAS = {
    "1844495": "Tasca",
    "1992404": "Comestibles",
}

# Columnas del archivo
COLUMNAS = ["#", "F. Operativa", "Concepto", "F. Valor", "Importe", "Saldo",
            "Referencia 1", "Referencia 2"]

# Estilos
ESTILOS = {
    "Tasca": {
        "header_fill": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        "header_font": Font(name="Aptos", bold=True, color="FFFFFF", size=11),
        "alt_fill": PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid"),
    },
    "Comestibles": {
        "header_fill": PatternFill(start_color="375623", end_color="375623", fill_type="solid"),
        "header_font": Font(name="Aptos", bold=True, color="FFFFFF", size=11),
        "alt_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    },
}
FONT_NORMAL = Font(name="Aptos", size=11)
BORDER_THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ANCHOS = {"#": 6, "F. Operativa": 14, "Concepto": 55, "F. Valor": 14,
          "Importe": 12, "Saldo": 14, "Referencia 1": 18, "Referencia 2": 20}


# ============================================================================
# PARSEO DE FORMATOS MIXTOS
# ============================================================================

def parsear_fecha(valor) -> Optional[datetime]:
    """Convierte texto o datetime a datetime. Acepta múltiples formatos."""
    if isinstance(valor, datetime):
        return valor
    if not valor:
        return None
    texto = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(texto.split(" ")[0] if " " in texto else texto, fmt)
        except ValueError:
            continue
    return None


def parsear_numero(valor) -> float:
    """Convierte texto con formato español (1.234,56) o float a float."""
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    if not valor:
        return 0.0
    texto = str(valor).strip()
    # Formato español: 17.428,53 → quitar puntos de miles, coma → punto
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return round(float(texto), 2)
    except (ValueError, TypeError):
        return 0.0


# ============================================================================
# LECTURA DE ARCHIVOS SABADELL (.xls)
# ============================================================================

def detectar_cuenta(ws) -> Optional[str]:
    """Detecta la cuenta desde fila 3 del .xls de Sabadell."""
    try:
        valor = str(ws.cell_value(3, 1)).replace("-", "")
        for clave, nombre in CUENTAS.items():
            if clave in valor:
                return nombre
    except Exception:
        pass
    return None


def leer_xls_sabadell(ruta: str) -> dict:
    """Lee un .xls de Sabadell. Devuelve {cuenta, periodo, movimientos, errores}."""
    resultado = {"cuenta": None, "periodo": "", "movimientos": [], "errores": []}
    nombre = os.path.basename(ruta)

    if not os.path.exists(ruta):
        resultado["errores"].append(f"No encontrado: {ruta}")
        return resultado

    try:
        wb = xlrd.open_workbook(ruta)
    except Exception as e:
        resultado["errores"].append(f"{nombre}: Error abriendo — {e}")
        return resultado

    ws = wb.sheet_by_index(0)

    cuenta = detectar_cuenta(ws)
    if not cuenta:
        resultado["errores"].append(f"{nombre}: No se detectó la cuenta bancaria")
        return resultado
    resultado["cuenta"] = cuenta

    try:
        resultado["periodo"] = str(ws.cell_value(6, 1)).strip()
    except Exception:
        pass

    # Buscar fila de headers
    fila_headers = None
    for r in range(7, min(12, ws.nrows)):
        if "operativa" in str(ws.cell_value(r, 0)).lower():
            fila_headers = r
            break

    if fila_headers is None:
        resultado["errores"].append(f"{nombre}: No se encontró fila de encabezados")
        return resultado

    for r in range(fila_headers + 1, ws.nrows):
        fecha_op_raw = ws.cell_value(r, 0)
        concepto = str(ws.cell_value(r, 1)).strip()
        fecha_val_raw = ws.cell_value(r, 2)
        importe = ws.cell_value(r, 3)
        saldo = ws.cell_value(r, 4)
        ref1 = str(ws.cell_value(r, 5)).strip()
        ref2 = str(ws.cell_value(r, 6)).strip()

        if not fecha_op_raw and not concepto:
            continue

        fecha_op = parsear_fecha(fecha_op_raw)
        if not fecha_op:
            continue

        resultado["movimientos"].append({
            "F. Operativa": fecha_op,
            "Concepto": concepto,
            "F. Valor": parsear_fecha(fecha_val_raw) or fecha_op,
            "Importe": parsear_numero(importe),
            "Saldo": parsear_numero(saldo),
            "Referencia 1": ref1,
            "Referencia 2": ref2,
        })

    # Sabadell entrega más reciente primero → invertir
    resultado["movimientos"].reverse()
    return resultado


# ============================================================================
# LECTURA DEL ARCHIVO EXISTENTE
# ============================================================================

def leer_pestana_existente(ruta: str, pestana: str) -> list[dict]:
    """Lee movimientos existentes de una pestaña, normalizando formatos mixtos."""
    if not os.path.exists(ruta):
        return []

    try:
        wb = load_workbook(ruta, data_only=True)
    except Exception:
        return []

    if pestana not in wb.sheetnames:
        return []

    ws = wb[pestana]
    movimientos = []

    for r in range(2, ws.max_row + 1):
        fecha_op = parsear_fecha(ws.cell(r, 2).value)
        if not fecha_op:
            continue

        movimientos.append({
            "F. Operativa": fecha_op,
            "Concepto": str(ws.cell(r, 3).value or "").strip(),
            "F. Valor": parsear_fecha(ws.cell(r, 4).value) or fecha_op,
            "Importe": parsear_numero(ws.cell(r, 5).value),
            "Saldo": parsear_numero(ws.cell(r, 6).value),
            "Referencia 1": str(ws.cell(r, 7).value or "").strip(),
            "Referencia 2": str(ws.cell(r, 8).value or "").strip(),
        })

    return movimientos


# ============================================================================
# DEDUPLICACIÓN Y MERGE
# ============================================================================

def clave_dedup(mov: dict) -> tuple:
    """F.Operativa + Concepto + Importe + Saldo."""
    fecha = mov["F. Operativa"]
    return (
        fecha.strftime("%Y%m%d") if isinstance(fecha, datetime) else str(fecha),
        mov.get("Concepto", ""),
        mov.get("Importe", 0),
        mov.get("Saldo", 0),
    )


def merge_y_dedup(existentes: list[dict], nuevos: list[dict]) -> tuple[list[dict], int]:
    """Merge existentes + nuevos eliminando duplicados. Existentes tienen prioridad."""
    claves = set()
    resultado = []

    for mov in existentes:
        c = clave_dedup(mov)
        if c not in claves:
            claves.add(c)
            resultado.append(mov)

    dupes = 0
    for mov in nuevos:
        c = clave_dedup(mov)
        if c in claves:
            dupes += 1
        else:
            claves.add(c)
            resultado.append(mov)

    return resultado, dupes


def ordenar_y_numerar(movimientos: list[dict]) -> list[dict]:
    """Ordena cronológicamente y renumera # desde 1."""
    movimientos.sort(key=lambda m: (
        m.get("F. Operativa") or datetime(2000, 1, 1),
        m.get("Saldo", 0),
    ))
    for i, mov in enumerate(movimientos, 1):
        mov["#"] = i
    return movimientos


# ============================================================================
# ESCRITURA
# ============================================================================

def escribir_pestana(wb, pestana: str, movimientos: list[dict]):
    """Escribe (reemplaza) una pestaña en el workbook."""
    if pestana in wb.sheetnames:
        del wb[pestana]

    ws = wb.create_sheet(pestana)
    estilo = ESTILOS.get(pestana, ESTILOS["Tasca"])

    # Headers
    for c, col in enumerate(COLUMNAS, 1):
        cell = ws.cell(1, c, col)
        cell.fill = estilo["header_fill"]
        cell.font = estilo["header_font"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_THIN

    # Datos
    for r, mov in enumerate(movimientos, 2):
        es_par = (r % 2 == 0)
        fill = estilo["alt_fill"] if es_par else None

        def escribir(col, valor, fmt=None, align=None):
            cell = ws.cell(r, col, valor)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if fmt:
                cell.number_format = fmt
            if align:
                cell.alignment = Alignment(horizontal=align)
            if fill:
                cell.fill = fill

        escribir(1, mov["#"], align="center")
        escribir(2, mov["F. Operativa"], fmt="DD/MM/YYYY", align="center")
        escribir(3, mov["Concepto"])
        escribir(4, mov["F. Valor"], fmt="DD/MM/YYYY", align="center")
        escribir(5, mov["Importe"], fmt='#,##0.00 €', align="right")
        escribir(6, mov["Saldo"], fmt='#,##0.00 €', align="right")
        escribir(7, mov["Referencia 1"])
        escribir(8, mov["Referencia 2"])

    # Anchos
    for c, col in enumerate(COLUMNAS, 1):
        ws.column_dimensions[get_column_letter(c)].width = ANCHOS.get(col, 12)

    ws.freeze_panes = "A2"


def guardar_archivo(ruta: str, datos_por_pestana: dict):
    """Guarda el archivo completo preservando el orden de pestañas."""
    # Crear workbook nuevo (limpio, sin formatos heredados rotos)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Escribir en orden: Tasca primero, Comestibles segundo
    for pestana in ["Tasca", "Comestibles"]:
        if pestana in datos_por_pestana:
            escribir_pestana(wb, pestana, datos_por_pestana[pestana])

    wb.save(ruta)


# ============================================================================
# LÓGICA PRINCIPAL
# ============================================================================

def procesar(archivos_xls: list[str], archivo_mov: str, dry_run: bool = False):
    """Lee archivos Sabadell, merge con existente, dedup, renumera, guarda."""

    errores = []
    avisos = []
    nuevos_por_pestana = defaultdict(list)

    # 1. Leer archivos Sabadell
    for ruta in archivos_xls:
        resultado = leer_xls_sabadell(ruta)
        if resultado["errores"]:
            errores.extend(resultado["errores"])
            continue

        cuenta = resultado["cuenta"]
        movs = resultado["movimientos"]
        nombre = os.path.basename(ruta)

        if not movs:
            avisos.append(f"{nombre}: sin movimientos")
            continue

        nuevos_por_pestana[cuenta].extend(movs)
        print(f"  {nombre}: {cuenta} · {len(movs)} movimientos · {resultado['periodo']}")

    if errores:
        print(f"\n  ERRORES:")
        for e in errores:
            print(f"    {e}")

    if avisos:
        print(f"\n  AVISOS:")
        for a in avisos:
            print(f"    {a}")

    if not nuevos_por_pestana:
        print("\n  Sin movimientos nuevos para procesar.")
        return

    # 2. Para cada pestaña: merge con existente
    print()
    datos_final = {}

    for pestana in ["Tasca", "Comestibles"]:
        existentes = leer_pestana_existente(archivo_mov, pestana)
        nuevos = nuevos_por_pestana.get(pestana, [])

        if not nuevos and not existentes:
            continue

        if nuevos:
            mergeados, n_dupes = merge_y_dedup(existentes, nuevos)
            mergeados = ordenar_y_numerar(mergeados)
            n_nuevos = len(mergeados) - len(existentes)

            print(f"  {pestana}:")
            print(f"    Existentes: {len(existentes)} · Nuevos: {n_nuevos} · Duplicados: {n_dupes} · Total: {len(mergeados)}")

            if n_nuevos > 0:
                fechas_nuevas = [m["F. Operativa"] for m in mergeados[len(existentes) - n_dupes:]]
                if fechas_nuevas:
                    print(f"    Rango nuevos: {min(fechas_nuevas).strftime('%d/%m/%Y')} → {max(fechas_nuevas).strftime('%d/%m/%Y')}")

            datos_final[pestana] = mergeados
        else:
            # Sin nuevos pero hay existentes → renumerar y limpiar formato
            existentes = ordenar_y_numerar(existentes)
            datos_final[pestana] = existentes
            print(f"  {pestana}: sin nuevos (se mantienen {len(existentes)} existentes)")

    if not datos_final:
        print("\n  Sin cambios.")
        return

    # 3. Guardar
    if dry_run:
        print(f"\n  [DRY-RUN] No se escribe nada")
    else:
        guardar_archivo(archivo_mov, datos_final)
        print(f"\n  Guardado: {archivo_mov}")


def mostrar_info(archivo_mov: str):
    """Muestra estado actual del archivo."""
    if not os.path.exists(archivo_mov):
        print(f"  No existe: {archivo_mov}")
        return

    wb = load_workbook(archivo_mov, data_only=True)
    print(f"\n  {os.path.basename(archivo_mov)}")
    for pestana in wb.sheetnames:
        ws = wb[pestana]
        n = ws.max_row - 1 if ws.max_row > 1 else 0
        if n > 0:
            f1 = parsear_fecha(ws.cell(2, 2).value)
            fn = parsear_fecha(ws.cell(ws.max_row, 2).value)
            f1s = f1.strftime("%d/%m/%Y") if f1 else "?"
            fns = fn.strftime("%d/%m/%Y") if fn else "?"
            print(f"    {pestana}: {n} movimientos ({f1s} → {fns})")
        else:
            print(f"    {pestana}: vacía")


# ============================================================================
# CLI
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Movimientos bancarios — TASCA BAREA S.L.L.",
        epilog="Ejemplo: python mov_banco.py descarga_tasca.xls descarga_comes.xls"
    )
    parser.add_argument("archivos", nargs="*", help="Archivos .xls de Sabadell")
    parser.add_argument("--carpeta", "-c", help="Carpeta con archivos .xls")
    parser.add_argument("--archivo", "-a", default=ARCHIVO_MOV, help="Archivo de movimientos (default: %(default)s)")
    parser.add_argument("--info", "-i", action="store_true", help="Mostrar estado actual")
    parser.add_argument("--dry-run", "-d", action="store_true", help="Simular sin escribir")
    args = parser.parse_args()

    archivo_mov = args.archivo

    print("=" * 60)
    print("MOVIMIENTOS CUENTA — TASCA BAREA S.L.L.")
    print("=" * 60)

    if args.info:
        mostrar_info(archivo_mov)
        return

    # Recopilar archivos
    archivos = list(args.archivos or [])
    if args.carpeta:
        if not os.path.isdir(args.carpeta):
            print(f"  ERROR: No existe: {args.carpeta}")
            sys.exit(1)
        archivos.extend(
            os.path.join(args.carpeta, f)
            for f in os.listdir(args.carpeta)
            if f.lower().endswith(".xls") and not f.startswith("~")
        )

    if not archivos:
        print("\n  Sin archivos para procesar.")
        print("  Uso: python mov_banco.py archivo.xls [archivo2.xls ...]")
        print("       python mov_banco.py --carpeta C:\\Descargas")
        print("       python mov_banco.py --info")
        return

    print(f"\n  Archivo: {os.path.basename(archivo_mov)}")
    print(f"  Procesando {len(archivos)} archivo(s)...\n")

    procesar(archivos, archivo_mov, dry_run=args.dry_run)
    print(f"\n{'=' * 60}")


if __name__ == "__main__":
    main()
