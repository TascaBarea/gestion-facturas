#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GENERADOR DE FICHEROS SEPA XML (pain.001.001.09)
================================================

Lee la pestaña SEPA del Excel y genera fichero XML para subir al banco.
Solo procesa las filas marcadas con ✓ en columna INCLUIR.

USO:
    python generar_sepa.py [ruta_excel]
    
    Si no se especifica ruta, busca el más reciente en outputs/

FORMATO:
    ISO 20022 pain.001.001.09 (Transferencias SEPA)
    Adaptado a la guía AEB/CECA/UNACC de España (versión 2025)

Creado: 02/02/2026
"""

import os
import sys
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List, Optional
import re
import glob


# ============================================================================
# CONFIGURACIÓN
# ============================================================================

@dataclass
class ConfigSEPA:
    """Configuración para generación SEPA"""
    # Datos del ordenante (cargados de config/datos_sensibles.py)
    NOMBRE_ORDENANTE: str = ""
    NIF_SUFIJO: str = ""
    BIC_ORDENANTE: str = ""

    # IBANs disponibles
    IBAN_TASCA: str = ""
    IBAN_COMESTIBLES: str = ""

    def __post_init__(self):
        try:
            from config.datos_sensibles import (
                NOMBRE_ORDENANTE, NIF_SUFIJO, BIC_ORDENANTE,
                IBAN_TASCA, IBAN_COMESTIBLES,
            )
            self.NOMBRE_ORDENANTE = NOMBRE_ORDENANTE
            self.NIF_SUFIJO = NIF_SUFIJO
            self.BIC_ORDENANTE = BIC_ORDENANTE
            self.IBAN_TASCA = IBAN_TASCA
            self.IBAN_COMESTIBLES = IBAN_COMESTIBLES
        except ImportError:
            print("ERROR: config/datos_sensibles.py no encontrado. SEPA no funcionara.")
    
    # Rutas
    OUTPUT_PATH: str = r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas\outputs"
    
    # Formato
    NIVEL_SERVICIO: str = "SEPA"  # SEPA para transferencias normales
    CLAUSULA_GASTOS: str = "SLEV"  # Según nivel de servicio


CONFIG = ConfigSEPA()


# ============================================================================
# MODELO DE DATOS
# ============================================================================

@dataclass
class TransferenciaSEPA:
    """Datos de una transferencia SEPA"""
    numero: int
    proveedor: str
    cif: str
    iban_beneficiario: str
    importe: Decimal
    referencia: str
    concepto: str
    fecha_ejecucion: datetime
    iban_ordenante: str
    archivo: str = ""


# ============================================================================
# LECTOR DE EXCEL
# ============================================================================

def leer_transferencias_excel(ruta_excel: str) -> List[TransferenciaSEPA]:
    """
    Lee las transferencias marcadas con ✓ de la pestaña SEPA.
    
    Columnas esperadas (fila 7 = cabecera, datos desde fila 8):
    A: #
    B: INCLUIR (✓ para incluir)
    C: PROVEEDOR
    D: CIF
    E: IBAN_BENEFICIARIO
    F: IMPORTE
    G: REF_FACTURA
    H: CONCEPTO
    I: FECHA_EJECUCION
    J: IBAN_ORDENANTE
    K: ARCHIVO
    """
    wb = load_workbook(ruta_excel, data_only=True)
    
    if "SEPA" not in wb.sheetnames:
        raise ValueError("El Excel no tiene pestaña SEPA")
    
    ws = wb["SEPA"]
    transferencias = []
    
    # Datos empiezan en fila 8
    for row in ws.iter_rows(min_row=8, values_only=True):
        # Verificar si está marcada para incluir
        incluir = row[1] if len(row) > 1 else None
        if not incluir or incluir != "✓":
            continue
        
        # Verificar datos mínimos
        proveedor = row[2] if len(row) > 2 else None
        iban_beneficiario = row[4] if len(row) > 4 else None
        importe = row[5] if len(row) > 5 else None
        iban_ordenante = row[9] if len(row) > 9 else None
        
        if not proveedor or not iban_beneficiario or not importe or not iban_ordenante:
            print(f"⚠️ Fila incompleta ignorada: {proveedor or 'sin nombre'}")
            continue
        
        # Verificar que no sea "FALTA IBAN"
        if "FALTA" in str(iban_beneficiario):
            print(f"⚠️ IBAN faltante ignorado: {proveedor}")
            continue
        
        # Parsear fecha
        fecha_str = row[8] if len(row) > 8 else None
        if isinstance(fecha_str, datetime):
            fecha = fecha_str
        elif fecha_str:
            try:
                fecha = datetime.strptime(str(fecha_str), "%d/%m/%y")
            except (ValueError, TypeError):
                fecha = datetime.now()
        else:
            fecha = datetime.now()
        
        # Convertir importe a Decimal
        try:
            if isinstance(importe, str):
                importe = importe.replace("€", "").replace(" ", "").replace(",", ".")
            importe_decimal = Decimal(str(importe)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except (ValueError, TypeError, InvalidOperation):
            print(f"⚠️ Importe inválido ignorado: {proveedor} - {importe}")
            continue
        
        transferencia = TransferenciaSEPA(
            numero=row[0] if row[0] else 0,
            proveedor=str(proveedor).strip(),
            cif=str(row[3] or "").strip() if len(row) > 3 else "",
            iban_beneficiario=limpiar_iban(str(iban_beneficiario)),
            importe=importe_decimal,
            referencia=str(row[6] or "").strip() if len(row) > 6 else "",
            concepto=str(row[7] or "").strip() if len(row) > 7 else "",
            fecha_ejecucion=fecha,
            iban_ordenante=limpiar_iban(str(iban_ordenante)),
            archivo=str(row[10] or "").strip() if len(row) > 10 else ""
        )
        
        transferencias.append(transferencia)
    
    wb.close()
    return transferencias


def limpiar_iban(iban: str) -> str:
    """Limpia IBAN quitando espacios"""
    return iban.replace(" ", "").upper()


def limpiar_texto_sepa(texto: str) -> str:
    """
    Limpia texto para cumplir con caracteres SEPA permitidos.
    Solo: A-Z, a-z, 0-9, /, -, ?, :, (, ), ., ,, ', +, espacio
    """
    # Mapeo de caracteres especiales
    texto = texto.replace("Ñ", "N").replace("ñ", "n")
    texto = texto.replace("Ç", "C").replace("ç", "c")
    
    # Eliminar acentos (simplificado)
    acentos = {
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'ü': 'u', 'Ü': 'U'
    }
    for acento, sin_acento in acentos.items():
        texto = texto.replace(acento, sin_acento)
    
    # Solo caracteres permitidos
    permitidos = set("ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789/-?:().,'+  ")
    texto = "".join(c if c in permitidos else " " for c in texto)
    
    # Limpiar espacios múltiples
    texto = " ".join(texto.split())
    
    return texto[:140]  # Máximo 140 caracteres para concepto


# ============================================================================
# GENERADOR XML SEPA
# ============================================================================

def generar_xml_sepa(transferencias: List[TransferenciaSEPA], iban_ordenante: str = None) -> str:
    """
    Genera fichero XML SEPA pain.001.001.09.
    
    Estructura:
    - CstmrCdtTrfInitn (raíz)
      - GrpHdr (cabecera)
      - PmtInf (información de pago - una por IBAN ordenante)
        - CdtTrfTxInf (transferencias individuales)
    """
    # Agrupar por IBAN ordenante si hay varios
    por_iban = {}
    for t in transferencias:
        iban = t.iban_ordenante
        if iban not in por_iban:
            por_iban[iban] = []
        por_iban[iban].append(t)
    
    # Si se especifica un IBAN, filtrar solo ese
    if iban_ordenante:
        iban_ordenante = limpiar_iban(iban_ordenante)
        if iban_ordenante in por_iban:
            por_iban = {iban_ordenante: por_iban[iban_ordenante]}
        else:
            raise ValueError(f"No hay transferencias para IBAN {iban_ordenante}")
    
    # Namespace
    ns = "urn:iso:std:iso:20022:tech:xsd:pain.001.001.09"
    
    # Crear documento
    root = ET.Element("Document", xmlns=ns)
    cstmr = ET.SubElement(root, "CstmrCdtTrfInitn")
    
    # === CABECERA (GrpHdr) ===
    ahora = datetime.now()
    msg_id = f"TASCABAREA-{ahora.strftime('%Y%m%d%H%M%S')}"
    
    total_operaciones = len(transferencias)
    total_importe = sum(t.importe for t in transferencias)
    
    grp_hdr = ET.SubElement(cstmr, "GrpHdr")
    ET.SubElement(grp_hdr, "MsgId").text = msg_id
    ET.SubElement(grp_hdr, "CreDtTm").text = ahora.strftime("%Y-%m-%dT%H:%M:%S")
    ET.SubElement(grp_hdr, "NbOfTxs").text = str(total_operaciones)
    ET.SubElement(grp_hdr, "CtrlSum").text = str(total_importe)
    
    # Parte iniciadora
    initg_pty = ET.SubElement(grp_hdr, "InitgPty")
    ET.SubElement(initg_pty, "Nm").text = limpiar_texto_sepa(CONFIG.NOMBRE_ORDENANTE)
    initg_id = ET.SubElement(initg_pty, "Id")
    org_id = ET.SubElement(initg_id, "OrgId")
    othr = ET.SubElement(org_id, "Othr")
    ET.SubElement(othr, "Id").text = CONFIG.NIF_SUFIJO
    
    # === INFORMACIÓN DE PAGO (PmtInf) - Una por IBAN ordenante ===
    for idx, (iban_ord, transferencias_iban) in enumerate(por_iban.items(), 1):
        pmt_inf = ET.SubElement(cstmr, "PmtInf")
        
        # ID de información de pago
        pmt_inf_id = f"PAGO-{ahora.strftime('%Y%m%d')}-{idx:03d}"
        ET.SubElement(pmt_inf, "PmtInfId").text = pmt_inf_id
        ET.SubElement(pmt_inf, "PmtMtd").text = "TRF"
        ET.SubElement(pmt_inf, "NbOfTxs").text = str(len(transferencias_iban))
        ET.SubElement(pmt_inf, "CtrlSum").text = str(sum(t.importe for t in transferencias_iban))
        
        # Tipo de pago
        pmt_tp_inf = ET.SubElement(pmt_inf, "PmtTpInf")
        svc_lvl = ET.SubElement(pmt_tp_inf, "SvcLvl")
        ET.SubElement(svc_lvl, "Cd").text = CONFIG.NIVEL_SERVICIO
        
        # Fecha de ejecución (la primera del grupo)
        fecha_ejec = transferencias_iban[0].fecha_ejecucion
        reqd_exctn_dt = ET.SubElement(pmt_inf, "ReqdExctnDt")
        ET.SubElement(reqd_exctn_dt, "Dt").text = fecha_ejec.strftime("%Y-%m-%d")
        
        # Ordenante
        dbtr = ET.SubElement(pmt_inf, "Dbtr")
        ET.SubElement(dbtr, "Nm").text = limpiar_texto_sepa(CONFIG.NOMBRE_ORDENANTE)
        
        # Cuenta del ordenante
        dbtr_acct = ET.SubElement(pmt_inf, "DbtrAcct")
        dbtr_acct_id = ET.SubElement(dbtr_acct, "Id")
        ET.SubElement(dbtr_acct_id, "IBAN").text = iban_ord
        
        # Entidad del ordenante
        dbtr_agt = ET.SubElement(pmt_inf, "DbtrAgt")
        fin_instn_id = ET.SubElement(dbtr_agt, "FinInstnId")
        ET.SubElement(fin_instn_id, "BICFI").text = CONFIG.BIC_ORDENANTE
        
        # Cláusula de gastos
        ET.SubElement(pmt_inf, "ChrgBr").text = CONFIG.CLAUSULA_GASTOS
        
        # === TRANSFERENCIAS INDIVIDUALES (CdtTrfTxInf) ===
        for t in transferencias_iban:
            cdt_trf = ET.SubElement(pmt_inf, "CdtTrfTxInf")
            
            # Identificación del pago
            pmt_id = ET.SubElement(cdt_trf, "PmtId")
            end_to_end_id = f"{t.numero:04d}-{t.referencia[:20]}" if t.referencia else f"{t.numero:04d}"
            ET.SubElement(pmt_id, "EndToEndId").text = limpiar_texto_sepa(end_to_end_id)[:35]
            
            # Importe
            amt = ET.SubElement(cdt_trf, "Amt")
            instd_amt = ET.SubElement(amt, "InstdAmt", Ccy="EUR")
            instd_amt.text = str(t.importe)
            
            # Beneficiario
            cdtr = ET.SubElement(cdt_trf, "Cdtr")
            ET.SubElement(cdtr, "Nm").text = limpiar_texto_sepa(t.proveedor)[:70]
            
            # Cuenta del beneficiario
            cdtr_acct = ET.SubElement(cdt_trf, "CdtrAcct")
            cdtr_acct_id = ET.SubElement(cdtr_acct, "Id")
            ET.SubElement(cdtr_acct_id, "IBAN").text = t.iban_beneficiario
            
            # Concepto
            if t.concepto:
                rmt_inf = ET.SubElement(cdt_trf, "RmtInf")
                ET.SubElement(rmt_inf, "Ustrd").text = limpiar_texto_sepa(t.concepto)
    
    # Convertir a string con formato
    xml_str = ET.tostring(root, encoding="unicode")
    
    # Añadir declaración XML y formatear
    dom = minidom.parseString(xml_str)
    xml_formatted = dom.toprettyxml(indent="  ", encoding="UTF-8")
    
    return xml_formatted.decode("utf-8")


# ============================================================================
# MAIN
# ============================================================================

def encontrar_excel_mas_reciente() -> str:
    """Encuentra el Excel de pagos más reciente"""
    patron = os.path.join(CONFIG.OUTPUT_PATH, "PAGOS_Gmail_*.xlsx")
    archivos = glob.glob(patron)
    
    if not archivos:
        raise FileNotFoundError(f"No se encontraron archivos PAGOS_Gmail_*.xlsx en {CONFIG.OUTPUT_PATH}")
    
    # Ordenar por fecha de modificación (más reciente primero)
    archivos.sort(key=os.path.getmtime, reverse=True)
    return archivos[0]


def main():
    """Función principal"""
    print("=" * 60)
    print("GENERADOR DE FICHEROS SEPA XML")
    print("=" * 60)
    print()
    
    # Obtener ruta del Excel
    if len(sys.argv) > 1:
        ruta_excel = sys.argv[1]
    else:
        try:
            ruta_excel = encontrar_excel_mas_reciente()
            print(f"📄 Usando Excel más reciente: {os.path.basename(ruta_excel)}")
        except FileNotFoundError as e:
            print(f"❌ Error: {e}")
            return 1
    
    if not os.path.exists(ruta_excel):
        print(f"❌ Error: No existe el archivo {ruta_excel}")
        return 1
    
    # Leer transferencias marcadas
    print(f"\n📖 Leyendo pestaña SEPA...")
    try:
        transferencias = leer_transferencias_excel(ruta_excel)
    except Exception as e:
        print(f"❌ Error leyendo Excel: {e}")
        return 1
    
    if not transferencias:
        print("⚠️ No hay transferencias marcadas con ✓ en la pestaña SEPA")
        return 0
    
    # Resumen
    print(f"\n📊 Transferencias a procesar: {len(transferencias)}")
    total = sum(t.importe for t in transferencias)
    print(f"💰 Importe total: {total:.2f} €")
    
    # Agrupar por IBAN ordenante
    por_iban = {}
    for t in transferencias:
        if t.iban_ordenante not in por_iban:
            por_iban[t.iban_ordenante] = []
        por_iban[t.iban_ordenante].append(t)
    
    print(f"\n📋 Desglose por cuenta ordenante:")
    for iban, trans in por_iban.items():
        subtotal = sum(t.importe for t in trans)
        cuenta = "TASCA" if "4495" in iban else "COMESTIBLES" if "2404" in iban else "OTRA"
        print(f"   {cuenta}: {len(trans)} transferencias, {subtotal:.2f} €")
    
    # Generar XML
    print(f"\n🔧 Generando XML SEPA...")
    try:
        xml_content = generar_xml_sepa(transferencias)
    except Exception as e:
        print(f"❌ Error generando XML: {e}")
        return 1
    
    # Guardar archivo
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_xml = f"SEPA_{fecha_str}.xml"
    ruta_xml = os.path.join(CONFIG.OUTPUT_PATH, nombre_xml)
    
    with open(ruta_xml, "w", encoding="utf-8") as f:
        f.write(xml_content)
    
    print(f"\n✅ Fichero SEPA generado: {ruta_xml}")
    print(f"\n📤 Sube este fichero al banco para ejecutar las transferencias.")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
