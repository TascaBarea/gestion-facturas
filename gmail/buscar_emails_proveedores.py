#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BUSCADOR DE EMAILS DE PROVEEDORES
Ejecutar en tu máquina local para encontrar emails de proveedores sin registrar

Uso: python buscar_emails_proveedores.py

Requisitos:
- token.json válido en la misma carpeta
- MAESTRO_PROVEEDORES.xlsx en ../datos/
"""

import os
import json
import re
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Optional, Dict, List, Set
import base64

# Google API
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

# Excel
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# Fuzzy matching
from rapidfuzz import fuzz

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

SCOPES = [
    'https://www.googleapis.com/auth/gmail.readonly',
]

# Carpetas
BASE_PATH = os.path.dirname(os.path.abspath(__file__))
GMAIL_PATH = os.path.dirname(BASE_PATH)  # Subir un nivel
DATOS_PATH = os.path.join(GMAIL_PATH, "datos")
MAESTRO_PATH = os.path.join(DATOS_PATH, "MAESTRO_PROVEEDORES.xlsx")

# Búsqueda
MESES_ATRAS = 12  # Buscar en los últimos 12 meses
ETIQUETA = "FACTURAS"  # Etiqueta donde buscar
MAX_EMAILS = 500  # Máximo de emails a analizar
UMBRAL_FUZZY = 80  # % mínimo de coincidencia

# ============================================================================
# GMAIL CONNECTION
# ============================================================================

def conectar_gmail():
    """Conecta a Gmail API"""
    creds = None
    token_path = os.path.join(BASE_PATH, "token.json")
    creds_path = os.path.join(BASE_PATH, "credentials.json")
    
    if os.path.exists(token_path):
        creds = Credentials.from_authorized_user_file(token_path, SCOPES)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("Renovando token...")
            creds.refresh(Request())
            # Guardar token renovado
            with open(token_path, 'w') as token:
                token.write(creds.to_json())
            print("✓ Token renovado")
        else:
            print("❌ Token inválido. Ejecuta primero gmail.py para autenticarte.")
            return None
    
    return build('gmail', 'v1', credentials=creds)

# ============================================================================
# CARGAR PROVEEDORES
# ============================================================================

def cargar_proveedores_sin_email() -> List[Dict]:
    """Carga proveedores del MAESTRO que no tienen email"""
    proveedores = []
    
    wb = load_workbook(MAESTRO_PATH, read_only=True, data_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 7:
            continue
        
        nombre = str(row[1] or "").strip()
        if not nombre:
            continue
        
        email = str(row[6] or "").strip()
        if email:  # Ya tiene email, saltar
            continue
        
        alias_raw = str(row[2] or "")
        alias_list = []
        if alias_raw:
            sep = "," if "," in alias_raw else "|"
            alias_list = [a.strip().upper() for a in alias_raw.split(sep) if a.strip()]
        
        proveedores.append({
            'cuenta': str(row[0] or ""),
            'nombre': nombre,
            'alias': alias_list,
            'cif': str(row[3] or ""),
            'tiene_extractor': str(row[7] or "").upper() == "SI" if len(row) > 7 else False
        })
    
    wb.close()
    return proveedores

# ============================================================================
# BUSCAR EN GMAIL
# ============================================================================

def obtener_emails_facturas(service, meses: int = 12) -> List[Dict]:
    """Obtiene emails de la etiqueta FACTURAS de los últimos N meses"""
    
    # Obtener ID de la etiqueta
    labels = service.users().labels().list(userId='me').execute()
    label_id = None
    for label in labels.get('labels', []):
        if label['name'].upper() == ETIQUETA:
            label_id = label['id']
            break
    
    if not label_id:
        print(f"❌ Etiqueta '{ETIQUETA}' no encontrada")
        return []
    
    # Calcular fecha límite
    fecha_limite = datetime.now() - timedelta(days=meses * 30)
    query = f"after:{fecha_limite.strftime('%Y/%m/%d')}"
    
    print(f"Buscando emails en '{ETIQUETA}' desde {fecha_limite.strftime('%d/%m/%Y')}...")
    
    # Obtener lista de emails
    emails = []
    page_token = None
    
    while len(emails) < MAX_EMAILS:
        results = service.users().messages().list(
            userId='me',
            labelIds=[label_id],
            q=query,
            maxResults=min(100, MAX_EMAILS - len(emails)),
            pageToken=page_token
        ).execute()
        
        messages = results.get('messages', [])
        if not messages:
            break
        
        for msg in messages:
            # Obtener headers del mensaje
            msg_data = service.users().messages().get(
                userId='me',
                id=msg['id'],
                format='metadata',
                metadataHeaders=['From', 'Subject', 'Date']
            ).execute()
            
            headers = {h['name']: h['value'] for h in msg_data.get('payload', {}).get('headers', [])}
            
            emails.append({
                'id': msg['id'],
                'from': headers.get('From', ''),
                'subject': headers.get('Subject', ''),
                'date': headers.get('Date', '')
            })
        
        page_token = results.get('nextPageToken')
        if not page_token:
            break
        
        print(f"  ... {len(emails)} emails cargados")
    
    print(f"✓ {len(emails)} emails encontrados")
    return emails

def extraer_email_limpio(remitente: str) -> str:
    """Extrae solo el email del campo From"""
    match = re.search(r'<([^>]+)>', remitente)
    if match:
        return match.group(1).lower()
    # Si no hay <>, asumir que es solo el email
    return remitente.lower().strip()

def extraer_nombre_remitente(remitente: str) -> str:
    """Extrae el nombre del remitente"""
    if '<' in remitente:
        return remitente.split('<')[0].strip().strip('"')
    return remitente

# ============================================================================
# MATCHING
# ============================================================================

def buscar_coincidencias(proveedores: List[Dict], emails: List[Dict]) -> Dict:
    """
    Busca coincidencias entre proveedores sin email y emails recibidos.
    Retorna: {proveedor_nombre: [(email, score, metodo, asunto), ...]}
    """
    coincidencias = defaultdict(list)
    emails_vistos = set()  # Para no duplicar
    
    for prov in proveedores:
        nombre_upper = prov['nombre'].upper()
        alias_list = prov['alias']
        
        for email in emails:
            email_addr = extraer_email_limpio(email['from'])
            nombre_remitente = extraer_nombre_remitente(email['from']).upper()
            asunto = email['subject'].upper()
            
            # Evitar duplicados de email para este proveedor
            key = (prov['nombre'], email_addr)
            if key in emails_vistos:
                continue
            
            # 1. Buscar por ALIAS en nombre del remitente o asunto
            for alias in alias_list:
                if len(alias) >= 6:
                    # Alias largos: puede estar contenido
                    if alias in nombre_remitente or alias in asunto:
                        coincidencias[prov['nombre']].append((email_addr, 100, f"ALIAS ({alias})", email['subject']))
                        emails_vistos.add(key)
                        break
                elif len(alias) >= 4:
                    # Alias cortos: palabra completa
                    palabras_remitente = set(nombre_remitente.replace(',', ' ').split())
                    palabras_asunto = set(asunto.replace(',', ' ').split())
                    if alias in palabras_remitente or alias in palabras_asunto:
                        coincidencias[prov['nombre']].append((email_addr, 95, f"ALIAS ({alias})", email['subject']))
                        emails_vistos.add(key)
                        break
            
            if key in emails_vistos:
                continue
            
            # 2. Fuzzy matching con nombre del proveedor
            score_remitente = fuzz.partial_ratio(nombre_upper, nombre_remitente)
            score_asunto = fuzz.partial_ratio(nombre_upper, asunto)
            
            max_score = max(score_remitente, score_asunto)
            if max_score >= UMBRAL_FUZZY:
                metodo = "FUZZY (remitente)" if score_remitente > score_asunto else "FUZZY (asunto)"
                coincidencias[prov['nombre']].append((email_addr, max_score, metodo, email['subject']))
                emails_vistos.add(key)
    
    return coincidencias

# ============================================================================
# GENERAR REPORTE
# ============================================================================

def generar_reporte(proveedores: List[Dict], coincidencias: Dict, output_path: str):
    """Genera Excel con resultados"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "EMAILS_ENCONTRADOS"
    
    # Estilos
    font_header = Font(name='Aptos Display', size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_encontrado = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_multiple = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_no_encontrado = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeceras
    headers = ["PROVEEDOR", "CUENTA", "TIENE_EXTRACTOR", "EMAIL_SUGERIDO", "SCORE", "MÉTODO", "ASUNTO_EJEMPLO", "OTROS_EMAILS"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
    
    # Datos
    row_num = 2
    encontrados = 0
    
    # Ordenar: primero los que tienen extractor, luego por nombre
    proveedores_sorted = sorted(proveedores, key=lambda x: (not x['tiene_extractor'], x['nombre']))
    
    for prov in proveedores_sorted:
        nombre = prov['nombre']
        emails_encontrados = coincidencias.get(nombre, [])
        
        # Ordenar por score
        emails_encontrados.sort(key=lambda x: -x[1])
        
        ws.cell(row=row_num, column=1, value=nombre).border = border
        ws.cell(row=row_num, column=2, value=prov['cuenta']).border = border
        ws.cell(row=row_num, column=3, value="SI" if prov['tiene_extractor'] else "NO").border = border
        
        if emails_encontrados:
            encontrados += 1
            mejor = emails_encontrados[0]
            ws.cell(row=row_num, column=4, value=mejor[0]).border = border
            ws.cell(row=row_num, column=5, value=mejor[1]).border = border
            ws.cell(row=row_num, column=6, value=mejor[2]).border = border
            ws.cell(row=row_num, column=7, value=mejor[3][:50] + "..." if len(mejor[3]) > 50 else mejor[3]).border = border
            
            # Otros emails si hay más de uno
            otros = [e[0] for e in emails_encontrados[1:3]]  # Máximo 2 más
            ws.cell(row=row_num, column=8, value=", ".join(otros)).border = border
            
            # Color según cantidad
            fill = fill_encontrado if len(emails_encontrados) == 1 else fill_multiple
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = fill
        else:
            for col in range(4, 9):
                ws.cell(row=row_num, column=col, value="").border = border
            # Color rojo para no encontrados
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = fill_no_encontrado
        
        row_num += 1
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 50
    ws.column_dimensions['H'].width = 40
    
    # Hoja de resumen
    ws_resumen = wb.create_sheet("RESUMEN")
    ws_resumen['A1'] = "RESUMEN DE BÚSQUEDA"
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_resumen['A4'] = f"Proveedores sin email: {len(proveedores)}"
    ws_resumen['A5'] = f"Emails encontrados: {encontrados} ({encontrados*100//len(proveedores)}%)"
    ws_resumen['A6'] = f"Sin coincidencias: {len(proveedores) - encontrados}"
    ws_resumen['A8'] = "LEYENDA:"
    ws_resumen['A9'] = "Verde = Email encontrado (único)"
    ws_resumen['A10'] = "Amarillo = Múltiples emails posibles (revisar)"
    ws_resumen['A11'] = "Rojo = No encontrado"
    
    wb.save(output_path)
    return encontrados

# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 60)
    print("BUSCADOR DE EMAILS DE PROVEEDORES")
    print(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 60)
    print()
    
    # Conectar a Gmail
    print("Conectando a Gmail...")
    service = conectar_gmail()
    if not service:
        return
    print("✓ Conectado")
    print()
    
    # Cargar proveedores sin email
    print("Cargando proveedores sin email del MAESTRO...")
    proveedores = cargar_proveedores_sin_email()
    print(f"✓ {len(proveedores)} proveedores sin email")
    con_extractor = sum(1 for p in proveedores if p['tiene_extractor'])
    print(f"  - Con extractor (prioritarios): {con_extractor}")
    print()
    
    # Obtener emails
    emails = obtener_emails_facturas(service, MESES_ATRAS)
    if not emails:
        print("No se encontraron emails")
        return
    print()
    
    # Buscar coincidencias
    print("Buscando coincidencias...")
    coincidencias = buscar_coincidencias(proveedores, emails)
    encontrados = sum(1 for c in coincidencias.values() if c)
    print(f"✓ {encontrados} proveedores con email encontrado")
    print()
    
    # Generar reporte
    output_path = os.path.join(BASE_PATH, f"EMAILS_ENCONTRADOS_{datetime.now().strftime('%Y%m%d')}.xlsx")
    total_encontrados = generar_reporte(proveedores, coincidencias, output_path)
    
    print("=" * 60)
    print("RESUMEN")
    print("=" * 60)
    print(f"  Proveedores analizados: {len(proveedores)}")
    print(f"  Emails encontrados:     {total_encontrados} ({total_encontrados*100//len(proveedores)}%)")
    print(f"  Sin coincidencias:      {len(proveedores) - total_encontrados}")
    print()
    print(f"✅ Reporte guardado: {output_path}")
    print()
    print("PRÓXIMOS PASOS:")
    print("1. Revisa el Excel generado")
    print("2. Verifica los emails sugeridos (especialmente los amarillos)")
    print("3. Actualiza MAESTRO_PROVEEDORES.xlsx con los emails confirmados")
    print("=" * 60)

if __name__ == "__main__":
    main()
