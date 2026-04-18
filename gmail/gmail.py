#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GMAIL MODULE v1.18
Sistema Automatizado de Procesamiento de Facturas
TASCA BAREA S.L.L. - Abril 2026

MEJORAS v1.18:
- VENTANA DE GRACIA TRIMESTRAL: determinar_destino_factura() con 4 destinos
  Días 1-11 → GRACIA (carpeta trimestre anterior), 12-20 → PENDIENTE, 21+ → ATRASADA
- Cola JSON datos/facturas_pendientes.json para pendientes en modo automático
- Pregunta interactiva en terminal en modo manual
- Sección Streamlit en Log Gmail para resolver pendientes
- LocalDropboxClient y DropboxAPIClient: parámetro destino en subir_archivo()

MEJORAS v1.17:
- FIX: Errores de extractores dedicados ahora logean a WARNING (antes DEBUG invisible)
- FIX: Validación REF anti-basura reforzada (min 3 chars, requiere dígito en genérico)
- FIX: factura.total ya no se sobreescribe a 0.00 cuando falta (queda None → celda vacía)
- Constante VERSION centralizada

MEJORAS v1.15:
- DEDUP DROPBOX: Escanea carpeta destino por hash antes de subir; si contenido
  idéntico ya existe (en cualquier archivo), salta subida Y escritura Excel.
  Dos facturas distintas del mismo proveedor/fecha se suben con sufijo " 2".
- FAIL-FAST EXCEL: Detecta Excel bloqueado al inicio, antes de procesar emails.

MEJORAS v1.14:
- FALLBACK PARCIAL: Si extractor dedicado obtiene fecha pero no total (o viceversa),
  complementa con extractor genérico en vez de marcar REVISAR directamente

MEJORAS v1.13:
- CURSOR TEMPORAL: Solo procesa emails posteriores a la última ejecución (after:YYYY/MM/DD)
- Fecha de última ejecución guardada en emails_procesados.json

Ejecuta: python gmail.py --produccion
         python gmail.py --test (modo prueba sin modificar archivos)
"""

VERSION = "1.18"

import os
import sys
import io
import json
import hashlib
import logging
import argparse
import platform
import re
import time
import tempfile
import traceback
import importlib.util
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from typing import Optional, Dict, List, Tuple, Any
from dataclasses import dataclass, field
import shutil

# Google API
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import base64

# Procesamiento PDF
import pdfplumber
try:
    import pytesseract
    from PIL import Image
    OCR_DISPONIBLE = True
except ImportError:
    OCR_DISPONIBLE = False

# Maestro proveedores (centralizado en nucleo/)
from nucleo.maestro import Proveedor, MaestroProveedores, normalizar_nombre_proveedor
from nucleo.parser import extraer_fecha, extraer_total, extraer_referencia, extraer_iban

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

@dataclass
class Config:
    """Configuración del sistema v1.17"""
    # Rutas base — detección automática Windows/Linux
    _ES_WINDOWS: bool = field(default_factory=lambda: platform.system() == "Windows", repr=False)

    BASE_PATH: str = field(default_factory=lambda: os.environ.get(
        "GESTION_FACTURAS_DIR",
        r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas" if platform.system() == "Windows"
        else "/opt/gestion-facturas"))
    GMAIL_PATH: str = field(default="")
    DROPBOX_BASE: str = field(default_factory=lambda: os.environ.get(
        "DROPBOX_BASE",
        r"C:\Users\jaime\Dropbox\File inviati\TASCA BAREA S.L.L\CONTABILIDAD" if platform.system() == "Windows"
        else ""))

    # Dropbox API (para VPS — refresh token no expira)
    DROPBOX_REFRESH_TOKEN: str = ""
    DROPBOX_APP_KEY: str = ""
    DROPBOX_APP_SECRET: str = ""
    DROPBOX_TOKEN: str = ""  # legacy, fallback
    DROPBOX_API_BASE: str = "/File inviati/TASCA BAREA S.L.L/CONTABILIDAD"

    # RUTA EXTRACTORES v1.4 — detección automática Windows/Linux
    EXTRACTORES_PATH: str = field(default_factory=lambda: os.path.join(
        os.environ.get("PARSEO_DIR",
                       r"C:\_ARCHIVOS\TRABAJO\Facturas\Parseo" if platform.system() == "Windows"
                       else "/opt/Parseo"),
        "extractores"))
    
    # IBANs propios (cargados de config/datos_sensibles.py)
    IBAN_TASCA: str = ""
    IBAN_COMESTIBLES: str = ""
    BIC_ORDENANTE: str = ""
    NIF_SUFIJO: str = ""
    NOMBRE_ORDENANTE: str = ""
    
    # Archivos
    MAESTRO_PATH: str = field(default="")
    JSON_PATH: str = field(default="")
    OUTPUT_PATH: str = field(default="")
    LOGS_PATH: str = field(default="")
    BACKUPS_PATH: str = field(default="")
    
    # Gmail
    GMAIL_SCOPES: List[str] = field(default_factory=lambda: [
        'https://www.googleapis.com/auth/gmail.readonly',
        'https://www.googleapis.com/auth/gmail.modify',
        'https://www.googleapis.com/auth/webmasters.readonly',
        'https://www.googleapis.com/auth/drive',
    ])
    LABEL_ORIGEN: str = "FACTURAS"
    LABEL_DESTINO: str = "FACTURAS_PROCESADAS"
    MAX_EMAILS: int = 200
    
    # Emails a ignorar (reenvíos)
    EMAILS_IGNORAR: List[str] = field(default_factory=lambda: [
        'tascabarea@gmail.com',
        'comunidadrodas2@gmail.com',
        'hola@comestiblesbarea.com'
    ])
    
    # Procesamiento
    UMBRAL_FUZZY: int = 85
    REINTENTOS: int = 3
    
    # Red / Reconexión (v1.6)
    TIMEOUT_CONEXION: int = 30  # segundos máx por operación API
    ESPERA_RECONEXION: int = 10  # segundos entre reintentos de reconexión
    
    # Notificaciones
    EMAIL_NOTIFICACION: str = "tascabarea@gmail.com"
    
    # Formatos
    FECHA_EXCEL: str = "%d/%m/%y"
    
    # Validación de fechas
    MARGEN_FECHA_FUTURA_DIAS: int = 3
    
    # Sufijos a eliminar de nombres de proveedores
    SUFIJOS_ELIMINAR: List[str] = field(default_factory=lambda: [
        'S.L.L.', 'S.L.U.', 'S.COOP.', 'S.L.', 'S.A.', 'C.B.', 
        'S.L.L', 'S.L.U', 'S.COOP', 'S.L', 'S.A', 'C.B'
    ])
    
    def __post_init__(self):
        # Cargar datos sensibles (IBANs, NIF, BIC)
        try:
            from config.datos_sensibles import (
                IBAN_TASCA, IBAN_COMESTIBLES, BIC_ORDENANTE,
                NIF_SUFIJO, NOMBRE_ORDENANTE,
            )
            self.IBAN_TASCA = IBAN_TASCA
            self.IBAN_COMESTIBLES = IBAN_COMESTIBLES
            self.BIC_ORDENANTE = BIC_ORDENANTE
            self.NIF_SUFIJO = NIF_SUFIJO
            self.NOMBRE_ORDENANTE = NOMBRE_ORDENANTE
        except ImportError:
            print("AVISO: config/datos_sensibles.py no encontrado. IBANs no disponibles.")
        # Cargar Dropbox config si existe (para VPS)
        try:
            from config import datos_sensibles as _ds
            for attr in ('DROPBOX_REFRESH_TOKEN', 'DROPBOX_APP_KEY', 'DROPBOX_APP_SECRET',
                         'DROPBOX_TOKEN', 'DROPBOX_API_BASE'):
                if hasattr(_ds, attr):
                    setattr(self, attr, getattr(_ds, attr))
        except (ImportError, AttributeError):
            pass
        if not self.GMAIL_PATH:
            self.GMAIL_PATH = os.path.join(self.BASE_PATH, "gmail")
        self.MAESTRO_PATH = os.path.join(self.BASE_PATH, "datos", "MAESTRO_PROVEEDORES.xlsx")
        self.JSON_PATH = os.path.join(self.BASE_PATH, "datos", "emails_procesados.json")
        self.OUTPUT_PATH = os.path.join(self.BASE_PATH, "outputs")
        self.LOGS_PATH = os.path.join(self.BASE_PATH, "outputs", "logs_gmail")
        self.BACKUPS_PATH = os.path.join(self.BASE_PATH, "outputs", "backups")


# Instancia global de configuración
CONFIG = Config()

# ============================================================================
# MODELOS DE DATOS
# ============================================================================

# Proveedor: importado de nucleo.maestro

@dataclass
class FacturaExtraida:
    """Datos extraídos de una factura PDF"""
    fecha: Optional[datetime] = None
    total: Optional[float] = None
    referencia: str = ""
    iban_detectado: str = ""
    exito: bool = False
    metodo: str = ""
    fecha_futura: bool = False
    es_proforma: bool = False


@dataclass
class ResultadoProcesamiento:
    """Resultado del procesamiento de un email"""
    email_id: str
    message_id: str = ""
    remitente: str = ""
    asunto: str = ""
    proveedor: Optional[Proveedor] = None
    proveedor_identificado: bool = False
    factura: Optional[FacturaExtraida] = None
    archivo_generado: str = ""
    dropbox_path: str = ""
    hash_pdf: str = ""
    error: str = ""
    requiere_revision: bool = False
    motivo_revision: str = ""
    iban_sugerido: str = ""
    alerta_roja: bool = False
    es_duplicado: bool = False


# ============================================================================
# LOGGING
# ============================================================================

def configurar_logging(modo_test: bool = False) -> logging.Logger:
    """Configura el sistema de logging usando nucleo.logging_config."""
    from nucleo.logging_config import setup_logging
    sufijo = "_test" if modo_test else ""
    return setup_logging(
        "gmail_module",
        log_subdir="logs_gmail",
        sufijo=sufijo,
    )


# ============================================================================
# UTILIDADES
# ============================================================================

# normalizar_nombre_proveedor: importado de nucleo.maestro

def obtener_trimestre(fecha: datetime) -> str:
    """Devuelve el trimestre en formato XTyy (ej: 1T26)"""
    trimestre = (fecha.month - 1) // 3 + 1
    año = fecha.year % 100
    return f"{trimestre}T{año:02d}"


def es_atrasada(fecha_factura: datetime, fecha_ejecucion: datetime) -> bool:
    """
    Determina si una factura es ATRASADA (legacy, mantener para compatibilidad).
    Usa determinar_destino_factura() para lógica completa con ventana de gracia.
    """
    return obtener_trimestre(fecha_factura) != obtener_trimestre(fecha_ejecucion)


# Constantes ventana de gracia (NO configurables)
GRACIA_HASTA_DIA = 11      # Días 1-11: ventana de gracia
PENDIENTE_HASTA_DIA = 20   # Días 12-20: zona gris / pendiente
MESES_INICIO_TRIMESTRE = {1, 4, 7, 10}  # enero, abril, julio, octubre

# Validaciones de negocio (v1.18.2)
TOTAL_MIN_SOSPECHOSO = 0.50      # Facturas < 0.50€ → revisar
TOTAL_MAX_SOSPECHOSO = 50_000    # Facturas > 50.000€ → revisar
FECHA_MAX_ANTIGUEDAD_DIAS = 730  # Facturas de hace > 2 años → revisar


def trimestre_de_fecha(fecha: datetime) -> Tuple[int, int]:
    """Devuelve (trimestre, año). Ej: date(2026,3,28) → (1, 2026)"""
    trimestre = (fecha.month - 1) // 3 + 1
    return trimestre, fecha.year


def es_trimestre_inmediatamente_anterior(trim_fac: int, year_fac: int,
                                          trim_hoy: int, year_hoy: int) -> bool:
    """True si el trimestre de la factura es exactamente el anterior al actual."""
    if trim_hoy == 1:
        return trim_fac == 4 and year_fac == year_hoy - 1
    else:
        return trim_fac == trim_hoy - 1 and year_fac == year_hoy


def determinar_destino_factura(fecha_factura: datetime, fecha_proceso: datetime = None) -> str:
    """
    Determina dónde debe ir una factura según la ventana de gracia trimestral.

    Kinema acepta facturas del trimestre anterior durante los primeros días
    del nuevo trimestre. Esta función implementa 3 zonas:

    Returns:
        'NORMAL'              → factura del trimestre actual, carpeta normal
        'GRACIA'              → trimestre anterior, dentro de ventana (días 1-11)
        'PENDIENTE_UBICACION' → zona gris (días 12-20), requiere decisión manual
        'ATRASADA'            → fuera de ventana, carpeta ATRASADAS/
    """
    if fecha_proceso is None:
        fecha_proceso = datetime.now()

    trim_fac, year_fac = trimestre_de_fecha(fecha_factura)
    trim_hoy, year_hoy = trimestre_de_fecha(fecha_proceso)

    # Caso 1: La factura es del trimestre actual → NORMAL
    if trim_fac == trim_hoy and year_fac == year_hoy:
        return 'NORMAL'

    # Caso 2: La factura es del trimestre inmediatamente anterior
    if es_trimestre_inmediatamente_anterior(trim_fac, year_fac, trim_hoy, year_hoy):
        # ¿Estamos en el primer mes del trimestre?
        if fecha_proceso.month in MESES_INICIO_TRIMESTRE:
            dia = fecha_proceso.day
            if dia <= GRACIA_HASTA_DIA:
                return 'GRACIA'
            elif dia <= PENDIENTE_HASTA_DIA:
                return 'PENDIENTE_UBICACION'
            else:
                return 'ATRASADA'
        else:
            # 2º o 3er mes del trimestre → siempre ATRASADA
            return 'ATRASADA'

    # Caso 3: Trimestre más antiguo → siempre ATRASADA
    return 'ATRASADA'


def generar_nombre_archivo(
    proveedor: Optional[Proveedor],
    factura: FacturaExtraida,
    fecha_proceso: datetime,
    remitente: str = "",
    destino: str = None
) -> str:
    """
    Genera el nombre del archivo PDF según nomenclatura.
    v1.6: Sanitización de caracteres inválidos en Windows
    v1.18: Ventana de gracia — destino controla si se añade prefijo ATRASADA
    """
    fecha_factura = factura.fecha if factura.fecha else fecha_proceso
    trimestre = obtener_trimestre(fecha_factura)
    mmdd = fecha_factura.strftime("%m%d")

    if proveedor and proveedor.nombre:
        nombre_norm = normalizar_nombre_proveedor(proveedor.nombre)
        forma_pago = proveedor.forma_pago if proveedor.forma_pago else ""
    else:
        # v1.6: Sanitizar remitente - extraer solo email, sin comillas ni nombre display
        email_limpio = remitente
        if "<" in email_limpio and ">" in email_limpio:
            email_limpio = email_limpio.split("<")[1].split(">")[0].strip()
        # Quitar caracteres prohibidos en Windows: " * : < > ? | / \
        email_limpio = re.sub(r'[\"*:<>?|/\\]', '', email_limpio)
        nombre_norm = f"({email_limpio})"
        forma_pago = ""

    partes = []

    # v1.18: Usar destino si se proporciona, si no calcular con lógica legacy
    if destino is None:
        destino = determinar_destino_factura(fecha_factura, fecha_proceso) if factura.fecha else 'NORMAL'

    if destino == 'ATRASADA':
        partes.append("ATRASADA")
    # GRACIA y NORMAL: sin prefijo ATRASADA

    if not proveedor or not proveedor.nombre:
        partes.insert(0, "REVISAR")

    partes.append(trimestre)
    partes.append(mmdd)
    partes.append(nombre_norm)

    if forma_pago:
        partes.append(forma_pago)

    nombre = " ".join(partes) + ".pdf"

    # v1.6: Limpieza final - quitar caracteres Windows prohibidos que hayan sobrevivido
    nombre = re.sub(r'[\"*:<>?|]', '', nombre)

    return nombre


def calcular_hash_archivo(contenido: bytes) -> str:
    """Calcula SHA-256 del contenido"""
    return hashlib.sha256(contenido).hexdigest()


# ============================================================================
# VENTANA DE GRACIA — Cola de pendientes y pregunta manual (v1.18)
# ============================================================================

COLA_PENDIENTES_PATH = os.path.join(
    os.environ.get("GESTION_FACTURAS_DIR",
                   r"C:\_ARCHIVOS\TRABAJO\Facturas\gestion-facturas" if platform.system() == "Windows"
                   else "/opt/gestion-facturas"),
    "datos", "facturas_pendientes.json"
)
PENDIENTES_DIR = os.path.join(os.path.dirname(COLA_PENDIENTES_PATH), "pendientes")


def guardar_en_cola_pendientes(factura_info: dict, pdf_bytes: bytes, logger: logging.Logger):
    """Guarda factura en cola JSON + PDF en carpeta temporal (modo automático)."""
    import uuid

    os.makedirs(PENDIENTES_DIR, exist_ok=True)

    # Anti-duplicado: no añadir si ya existe por archivo_renombrado
    cola = []
    if os.path.exists(COLA_PENDIENTES_PATH):
        with open(COLA_PENDIENTES_PATH, 'r', encoding='utf-8') as f:
            cola = json.load(f)
    for entrada in cola:
        if entrada.get('archivo_renombrado') == factura_info.get('archivo_renombrado') \
                and entrada.get('estado') == 'pendiente':
            logger.info(f"  ↳ Ya está en cola de pendientes, no duplicar")
            return

    # Guardar PDF temporal
    nombre_temp = factura_info['archivo_renombrado'].replace(' ', '_')
    ruta_pdf = os.path.join(PENDIENTES_DIR, nombre_temp)
    with open(ruta_pdf, 'wb') as f:
        f.write(pdf_bytes)

    entrada = {
        "id": str(uuid.uuid4()),
        **factura_info,
        "ruta_pdf_temporal": ruta_pdf,
        "estado": "pendiente",
        "resuelto_por": None,
        "fecha_resolucion": None,
    }

    cola.append(entrada)

    # Escribir atómicamente
    tmp_path = COLA_PENDIENTES_PATH + ".tmp"
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(cola, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, COLA_PENDIENTES_PATH)

    logger.warning(
        f"⚠️ PENDIENTE: {factura_info['archivo_renombrado']} "
        f"— zona gris día {factura_info['dia_proceso']}, guardada en cola"
    )


def preguntar_destino_manual(factura_info: dict) -> str:
    """Pausa el proceso y pregunta al usuario en terminal (modo manual)."""
    print("\n" + "=" * 60)
    print("⚠️  FACTURA EN ZONA GRIS — REQUIERE DECISIÓN")
    print("=" * 60)
    print(f"  Archivo:    {factura_info['archivo_renombrado']}")
    print(f"  Proveedor:  {factura_info['proveedor']}")
    print(f"  Fecha fac.: {factura_info['fecha_factura']}")
    print(f"  Total:      {factura_info['total']}€")
    print(f"  Trimestre:  {factura_info['trimestre_factura']} "
          f"(procesada día {factura_info['dia_proceso']} de {factura_info['trimestre_actual']})")
    print()
    print(f"  [1] → Subir a {factura_info['opciones']['trimestre_anterior']['carpeta']}")
    print(f"        (como factura normal del trimestre anterior)")
    print(f"  [2] → Subir a {factura_info['opciones']['atrasada']['carpeta']}")
    print(f"        (como ATRASADA del trimestre actual)")
    print()

    while True:
        respuesta = input("  Elige [1/2]: ").strip()
        if respuesta == '1':
            return 'GRACIA'
        elif respuesta == '2':
            return 'ATRASADA'
        else:
            print("  → Respuesta no válida. Escribe 1 o 2.")


def construir_info_pendiente(
    resultado: 'ResultadoProcesamiento',
    factura: 'FacturaExtraida',
    nombre_generado: str,
    fecha_proceso: datetime
) -> dict:
    """Construye el dict de info para cola pendientes / pregunta manual."""
    fecha_factura = factura.fecha if factura.fecha else fecha_proceso
    trim_fac = obtener_trimestre(fecha_factura)
    trim_actual = obtener_trimestre(fecha_proceso)
    año_fac = fecha_factura.year
    num_trim_fac = (fecha_factura.month - 1) // 3 + 1
    año_actual = fecha_proceso.year
    num_trim_actual = (fecha_proceso.month - 1) // 3 + 1

    carpeta_anterior = f"{num_trim_fac} TRIMESTRE {año_fac}"
    carpeta_atrasada = f"{num_trim_actual} TRIMESTRE {año_actual}/ATRASADAS"

    # Nombre sin ATRASADA (para opción gracia)
    nombre_gracia = nombre_generado.replace("ATRASADA ", "")

    return {
        "archivo_original": resultado.archivo_generado or nombre_generado,
        "archivo_renombrado": nombre_gracia,
        "proveedor": resultado.proveedor.nombre if resultado.proveedor else resultado.remitente,
        "fecha_factura": fecha_factura.strftime("%Y-%m-%d"),
        "fecha_proceso": fecha_proceso.strftime("%Y-%m-%d"),
        "total": factura.total,
        "trimestre_factura": trim_fac,
        "trimestre_actual": trim_actual,
        "dia_proceso": fecha_proceso.day,
        "opciones": {
            "trimestre_anterior": {
                "carpeta": carpeta_anterior,
                "nombre": nombre_gracia,
            },
            "atrasada": {
                "carpeta": carpeta_atrasada,
                "nombre": f"ATRASADA {nombre_gracia}",
            },
        },
    }


def formatear_iban(iban: str) -> str:
    """Formatea IBAN con espacios cada 4 caracteres"""
    if not iban:
        return ""
    iban = iban.replace(" ", "").upper()
    return " ".join([iban[i:i+4] for i in range(0, len(iban), 4)])


# MaestroProveedores: importado de nucleo.maestro

# ============================================================================
# CONTROL DUPLICADOS
# ============================================================================

class ControlDuplicados:
    """Gestiona el control de emails y archivos procesados"""
    
    def __init__(self, ruta: str):
        self.ruta = ruta
        self.datos: Dict[str, Any] = {"emails": {}, "hashes": {}, "facturas": {}}
        self._cargar()
    
    def _cargar(self):
        """Carga el JSON de control"""
        if os.path.exists(self.ruta):
            try:
                with open(self.ruta, 'r', encoding='utf-8') as f:
                    self.datos = json.load(f)
            except json.JSONDecodeError:
                self.datos = {"emails": {}, "hashes": {}, "facturas": {}}

        for key in ["emails", "hashes", "facturas"]:
            if key not in self.datos:
                self.datos[key] = {}

    def get_ultima_ejecucion(self) -> Optional[str]:
        """Devuelve la fecha ISO de la última ejecución, o None si no hay."""
        return self.datos.get("ultima_ejecucion")

    def set_ultima_ejecucion(self, fecha_iso: str):
        """Guarda la fecha de la última ejecución."""
        self.datos["ultima_ejecucion"] = fecha_iso
    
    def guardar(self):
        """Guarda el JSON de control (escritura atómica con file lock)."""
        os.makedirs(os.path.dirname(self.ruta), exist_ok=True)
        ruta_lock = self.ruta + ".lock"
        ruta_tmp = self.ruta + ".tmp"

        # File lock para evitar corrupción si 2 instancias escriben a la vez
        lock_fd = None
        try:
            lock_fd = open(ruta_lock, "w")
            if os.name == "nt":
                import msvcrt
                msvcrt.locking(lock_fd.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl
                fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)

            with open(ruta_tmp, 'w', encoding='utf-8') as f:
                json.dump(self.datos, f, indent=2, ensure_ascii=False)
            os.replace(ruta_tmp, self.ruta)

        except (OSError, IOError) as e:
            logging.getLogger("gmail_module").warning("No se pudo obtener lock para %s: %s", self.ruta, e)
            # Fallback: escribir sin lock (mejor que perder datos)
            with open(ruta_tmp, 'w', encoding='utf-8') as f:
                json.dump(self.datos, f, indent=2, ensure_ascii=False)
            os.replace(ruta_tmp, self.ruta)
        finally:
            if lock_fd:
                lock_fd.close()
            try:
                os.remove(ruta_lock)
            except OSError:
                pass
    
    def email_procesado(self, email_id: str) -> bool:
        return email_id in self.datos["emails"]
    
    def hash_existe(self, hash_pdf: str) -> bool:
        return hash_pdf in self.datos["hashes"]
    
    def factura_existe(self, cif: str, referencia: str, nombre: str = "") -> bool:
        if not referencia:
            return False
        # v1.8: Usar nombre como fallback cuando CIF está vacío (ej: ODOO)
        identificador = cif if cif else nombre
        if not identificador:
            return False
        clave = f"{identificador}|{referencia}"
        return clave in self.datos["facturas"]
    
    def registrar(self, resultado: ResultadoProcesamiento):
        """Registra un procesamiento exitoso"""
        self.datos["emails"][resultado.email_id] = {
            "message_id": resultado.message_id,
            "fecha_proceso": datetime.now().isoformat(),
            "proveedor": resultado.proveedor.nombre if resultado.proveedor else None,
            "archivo": resultado.archivo_generado,
            "dropbox": resultado.dropbox_path
        }
        
        # v1.8: No sobreescribir hash/factura si ya existen (protección contra duplicados)
        if resultado.hash_pdf and resultado.hash_pdf not in self.datos["hashes"]:
            self.datos["hashes"][resultado.hash_pdf] = {
                "email_id": resultado.email_id,
                "archivo": resultado.archivo_generado
            }

        if resultado.proveedor and resultado.factura and resultado.factura.referencia:
            # v1.8: Usar nombre como fallback cuando CIF está vacío (ej: ODOO)
            identificador = resultado.proveedor.cif if resultado.proveedor.cif else resultado.proveedor.nombre
            if identificador:
                clave = f"{identificador}|{resultado.factura.referencia}"
                if clave not in self.datos["facturas"]:
                    self.datos["facturas"][clave] = {
                        "email_id": resultado.email_id,
                        "archivo": resultado.archivo_generado
                    }
    
    def registrar_email_visto(self, email_id: str, motivo: str = ""):
        """
        Registra que un email fue visto (aunque no tenga adjuntos procesables).
        Evita reprocesarlo en futuras ejecuciones.
        """
        self.datos["emails"][email_id] = {
            "fecha_proceso": datetime.now().isoformat(),
            "motivo": motivo
        }
    
    def registrar_y_guardar(self, resultado: ResultadoProcesamiento):
        """Registra un procesamiento Y guarda el JSON inmediatamente."""
        self.registrar(resultado)
        self.guardar()


# ============================================================================
# EXTRACTOR PDF
# ============================================================================

class ExtractorPDF:
    """Extrae datos de facturas en PDF usando nucleo/parser para el parseo"""

    # v1.7: Palabras que NO son referencias válidas (capturadas por error)
    # Usado también por _usar_extractor_dedicado
    REF_INVALIDAS = {
        'ERENCE', 'FERENCE', 'REFERENCE', 'REFERENCIA',
        'ERENCIA', 'ERENTE', 'RENCIA',
        'FACTURA', 'FECHA', 'TOTAL', 'IMPORTE', 'PEDIDO',
        'ALBARAN', 'CLIENTE', 'PROVEEDOR', 'NUMERO',
        'IVA', 'BASE', 'DATOS', 'PAGO', 'BANCO',
        'DOS', 'UNO', 'TRES',
    }

    def __init__(self, contenido: bytes):
        self.contenido = contenido
        self.texto = ""
        self.metodo = ""

    def extraer(self) -> FacturaExtraida:
        """Extrae datos del PDF usando pdfplumber y OCR como fallback"""
        resultado = FacturaExtraida()

        self.texto = self._extraer_texto_pdfplumber()
        self.metodo = "pdfplumber"

        if self.texto:
            resultado = self._parsear_texto()
            resultado.metodo = "pdfplumber"

        if OCR_DISPONIBLE and (not self.texto or not resultado.fecha or not resultado.total):
            texto_ocr = self._extraer_texto_ocr()
            if texto_ocr:
                self.texto = texto_ocr
                self.metodo = "ocr"
                resultado_ocr = self._parsear_texto()
                resultado_ocr.metodo = "ocr"

                if not resultado.fecha and resultado_ocr.fecha:
                    resultado.fecha = resultado_ocr.fecha
                if not resultado.total and resultado_ocr.total:
                    resultado.total = resultado_ocr.total
                if not resultado.referencia and resultado_ocr.referencia:
                    resultado.referencia = resultado_ocr.referencia

        resultado.exito = resultado.fecha is not None or resultado.total is not None
        return resultado

    def _extraer_texto_pdfplumber(self) -> str:
        """Extrae texto usando pdfplumber"""
        try:
            texto_completo = []
            with pdfplumber.open(io.BytesIO(self.contenido)) as pdf:
                for page in pdf.pages:
                    texto = page.extract_text()
                    if texto:
                        texto_completo.append(texto)
            return "\n".join(texto_completo)
        except Exception as e:
            logging.getLogger("gmail_module").debug("Error pdfplumber: %s", e)
            return ""

    def _extraer_texto_ocr(self) -> str:
        """Extrae texto usando OCR"""
        if not OCR_DISPONIBLE:
            return ""

        try:
            from pdf2image import convert_from_bytes

            images = convert_from_bytes(self.contenido)
            texto_completo = []

            for img in images:
                texto = pytesseract.image_to_string(img, lang='spa')
                if texto:
                    texto_completo.append(texto)

            return "\n".join(texto_completo)
        except Exception as e:
            logging.getLogger("gmail_module").debug("Error OCR: %s", e)
            return ""

    def _parsear_texto(self) -> FacturaExtraida:
        """Parsea el texto usando funciones centralizadas de nucleo/parser"""
        resultado = FacturaExtraida()

        # Fecha: nucleo/parser devuelve str "DD/MM/YYYY" → convertir a datetime
        fecha_str = extraer_fecha(self.texto)
        if fecha_str:
            try:
                resultado.fecha = datetime.strptime(fecha_str, "%d/%m/%Y")
                limite_futuro = datetime.now() + timedelta(days=CONFIG.MARGEN_FECHA_FUTURA_DIAS)
                if resultado.fecha > limite_futuro:
                    resultado.fecha_futura = True
            except ValueError:
                pass

        # Total
        resultado.total = extraer_total(self.texto)

        # Referencia — con validación anti-basura (v1.17)
        ref_candidata = extraer_referencia(self.texto) or ""
        if ref_candidata:
            ref_upper = ref_candidata.strip().upper()
            tiene_digito = any(c.isdigit() for c in ref_candidata)
            # Rechazar: palabras del diccionario REF_INVALIDAS, < 3 chars,
            # o solo letras sin dígitos (probablemente fragmento de nombre/dirección)
            if (ref_upper in self.REF_INVALIDAS
                    or len(ref_candidata.strip()) < 3
                    or not tiene_digito):
                resultado.referencia = ""
            else:
                resultado.referencia = ref_candidata.strip()
        else:
            resultado.referencia = ""

        # IBAN
        iban = extraer_iban(self.texto)
        if iban:
            resultado.iban_detectado = iban.replace(" ", "")

        return resultado


# ============================================================================
# GMAIL API
# ============================================================================

class GmailClient:
    """
    Cliente para interactuar con Gmail API.
    v1.6: Auto-reconexión ante errores de red (WinError 10054, ConnectionReset, etc.)
    """
    
    # Errores que indican pérdida de conexión y requieren reconexión
    ERRORES_RED = (
        ConnectionResetError,
        ConnectionAbortedError,
        ConnectionRefusedError,
        BrokenPipeError,
        TimeoutError,
        OSError,  # Incluye WinError 10054
    )
    
    def __init__(self, credentials_path: str, token_path: str):
        self.credentials_path = credentials_path
        self.token_path = token_path
        self.service = None
        self.label_ids = {}
        self._creds = None  # v1.6: guardar credenciales para reconexión
        self.logger = logging.getLogger("gmail_client")
    
    def conectar(self) -> bool:
        """Establece conexión con Gmail API"""
        creds = None
        
        if os.path.exists(self.token_path):
            creds = Credentials.from_authorized_user_file(self.token_path, CONFIG.GMAIL_SCOPES)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_path, CONFIG.GMAIL_SCOPES
                )
                creds = flow.run_local_server(port=0)
            
            with open(self.token_path, 'w') as token:
                token.write(creds.to_json())
        
        self._creds = creds  # v1.6: guardar para reconexión
        self.service = build('gmail', 'v1', credentials=creds)
        self._cargar_labels()
        return True
    
    def _reconectar(self):
        """
        v1.6: Reconecta con Gmail API usando credenciales existentes.
        Se llama automáticamente cuando se detecta un error de red.
        """
        self.logger.warning("Reconectando con Gmail API...")

        for intento in range(CONFIG.REINTENTOS):
            try:
                # Refrescar credenciales si expiraron
                if self._creds and self._creds.expired and self._creds.refresh_token:
                    self._creds.refresh(Request())
                    with open(self.token_path, 'w') as token:
                        token.write(self._creds.to_json())
                
                self.service = build('gmail', 'v1', credentials=self._creds)
                self.logger.info(f"Gmail reconectado ✓ (intento {intento + 1})")
                return
            except Exception as e:
                self.logger.warning(f"Reconexión fallida (intento {intento + 1}/{CONFIG.REINTENTOS}): {e}")
                if intento < CONFIG.REINTENTOS - 1:
                    time.sleep(CONFIG.ESPERA_RECONEXION * (intento + 1))
        
        raise ConnectionError("No se pudo reconectar con Gmail API tras múltiples intentos")
    
    def _api_call(self, operacion, *args, **kwargs):
        """
        v1.6: Wrapper para llamadas API con auto-reconexión.
        Si falla por error de red, reconecta y reintenta UNA vez.
        """
        try:
            return operacion(*args, **kwargs)
        except self.ERRORES_RED as e:
            # Detectar si es error de red genuino
            error_str = str(e).lower()
            es_error_red = any(x in error_str for x in [
                '10054', '10053', 'connection', 'reset', 'broken pipe',
                'timeout', 'timed out', 'forcibly closed', 'interrupción'
            ])
            
            if es_error_red:
                self.logger.warning(f"Error de red detectado: {e}. Reconectando...")
                time.sleep(CONFIG.ESPERA_RECONEXION)
                self._reconectar()
                # Reintentar UNA vez tras reconexión
                return operacion(*args, **kwargs)
            else:
                raise
    
    def _cargar_labels(self):
        """Carga los IDs de las etiquetas"""
        results = self.service.users().labels().list(userId='me').execute()
        labels = results.get('labels', [])
        
        for label in labels:
            self.label_ids[label['name']] = label['id']
    
    def obtener_emails_pendientes(self, max_results: int = 200, after_date: Optional[str] = None) -> List[Dict]:
        """
        Obtiene emails con etiqueta FACTURAS no procesados.
        v1.12: Paginación — recorre todas las páginas hasta max_results.
        v1.13: Filtro after_date (YYYY/MM/DD) para solo traer emails nuevos.
        """
        label_id = self.label_ids.get(CONFIG.LABEL_ORIGEN)
        if not label_id:
            raise ValueError(f"Etiqueta {CONFIG.LABEL_ORIGEN} no encontrada")

        all_messages = []
        page_token = None

        # v1.13: Construir query con filtro de fecha
        q = f"after:{after_date}" if after_date else None

        while len(all_messages) < max_results:
            kwargs = {
                'userId': 'me',
                'labelIds': [label_id],
                'maxResults': min(50, max_results - len(all_messages))
            }
            if q:
                kwargs['q'] = q
            if page_token:
                kwargs['pageToken'] = page_token

            results = self._api_call(
                lambda: self.service.users().messages().list(**kwargs).execute()
            )

            messages = results.get('messages', [])
            all_messages.extend(messages)

            page_token = results.get('nextPageToken')
            if not page_token:
                break

        emails = []
        for msg in all_messages:
            email_data = self._obtener_detalle_email(msg['id'])
            if email_data:
                emails.append(email_data)

        return emails
    
    def _obtener_detalle_email(self, email_id: str) -> Optional[Dict]:
        """Obtiene detalles de un email específico"""
        try:
            msg = self._api_call(
                lambda: self.service.users().messages().get(
                    userId='me',
                    id=email_id,
                    format='full'
                ).execute()
            )
            
            headers = {h['name']: h['value'] for h in msg['payload'].get('headers', [])}
            
            return {
                'id': email_id,
                'message_id': headers.get('Message-ID', ''),
                'from': headers.get('From', ''),
                'subject': headers.get('Subject', ''),
                'date': headers.get('Date', ''),
                'payload': msg['payload']
            }
        except Exception as e:
            self.logger.warning("Error obteniendo email %s: %s", email_id, e)
            return None

    def descargar_adjuntos(self, email_data: Dict) -> List[Tuple[str, bytes]]:
        """Descarga adjuntos PDF/JPG de un email"""
        adjuntos = []
        service = self  # referencia para _api_call dentro de closure
        
        def procesar_partes(partes):
            for parte in partes:
                filename = parte.get('filename', '')
                mime_type = parte.get('mimeType', '')
                
                if filename and (filename.lower().endswith('.pdf') or 
                                mime_type in ['application/pdf', 'image/jpeg', 'image/png']):
                    
                    body = parte.get('body', {})
                    attachment_id = body.get('attachmentId')
                    
                    if attachment_id:
                        att = service._api_call(
                            lambda: service.service.users().messages().attachments().get(
                                userId='me',
                                messageId=email_data['id'],
                                id=attachment_id
                            ).execute()
                        )
                        
                        data = base64.urlsafe_b64decode(att['data'])
                        adjuntos.append((filename, data))
                
                if 'parts' in parte:
                    procesar_partes(parte['parts'])
        
        payload = email_data.get('payload', {})
        if 'parts' in payload:
            procesar_partes(payload['parts'])
        
        return adjuntos
    
    def mover_a_procesados_y_marcar_leido(self, email_id: str):
        """
        Mueve email a etiqueta FACTURAS_PROCESADAS y lo marca como LEÍDO.
        v1.4: Añadido marcar como leído
        v1.6: Auto-reconexión via _api_call
        """
        label_origen = self.label_ids.get(CONFIG.LABEL_ORIGEN)
        label_destino = self.label_ids.get(CONFIG.LABEL_DESTINO)
        
        if not label_destino:
            label = self._api_call(
                lambda: self.service.users().labels().create(
                    userId='me',
                    body={'name': CONFIG.LABEL_DESTINO}
                ).execute()
            )
            label_destino = label['id']
            self.label_ids[CONFIG.LABEL_DESTINO] = label_destino
        
        # v1.4: Añadido 'UNREAD' a removeLabelIds para marcar como leído
        remove_labels = ['UNREAD']
        if label_origen:
            remove_labels.append(label_origen)
        
        self._api_call(
            lambda: self.service.users().messages().modify(
                userId='me',
                id=email_id,
                body={
                    'addLabelIds': [label_destino],
                    'removeLabelIds': remove_labels
                }
            ).execute()
        )


# ============================================================================
# DROPBOX LOCAL (v1.4 - copia a carpeta sincronizada)
# ============================================================================

class LocalDropboxClient:
    """
    Cliente que copia archivos a la carpeta local de Dropbox sincronizada.
    No requiere API ni token - usa la carpeta local en disco.
    
    Estructura destino:
    [DROPBOX_BASE]/FACTURAS 2026/FACTURAS RECIBIDAS/1 TRIMESTRE 2026/
    [DROPBOX_BASE]/FACTURAS 2026/FACTURAS RECIBIDAS/1 TRIMESTRE 2026/ATRASADAS/
    """
    
    def __init__(self, base_path: str):
        self.base_path = base_path
        if not os.path.exists(base_path):
            raise Exception(f"Carpeta Dropbox no encontrada: {base_path}")
    
    def subir_archivo(self, contenido: bytes, nombre_archivo: str, fecha_factura: datetime,
                      fecha_ejecucion: datetime, destino: str = None) -> Tuple[str, bool]:
        """
        Copia un archivo a la carpeta local de Dropbox con deduplicación inteligente.

        v1.5: Deduplicación multi-señal.
        v1.18: Ventana de gracia — destino controla carpeta.

        Returns:
            Tuple (ruta_archivo, ya_existia) — ya_existia=True si se saltó por duplicado
        """
        # v1.18: Usar destino si se proporciona
        if destino is None:
            destino = determinar_destino_factura(fecha_factura, fecha_ejecucion)

        # GRACIA → carpeta del trimestre de la FACTURA (no ejecución)
        # NORMAL → carpeta del trimestre de la factura (= ejecución)
        # ATRASADA → carpeta del trimestre de EJECUCIÓN + ATRASADAS/
        if destino == 'GRACIA':
            carpeta_trimestre = self.obtener_ruta_trimestre(fecha_factura)
            ruta_completa = os.path.join(self.base_path, carpeta_trimestre, nombre_archivo)
        elif destino == 'ATRASADA':
            carpeta_trimestre = self.obtener_ruta_trimestre(fecha_ejecucion)
            ruta_completa = os.path.join(self.base_path, carpeta_trimestre, "ATRASADAS", nombre_archivo)
        else:  # NORMAL
            carpeta_trimestre = self.obtener_ruta_trimestre(fecha_ejecucion)
            ruta_completa = os.path.join(self.base_path, carpeta_trimestre, nombre_archivo)
        
        # Crear carpeta si no existe
        carpeta = os.path.dirname(ruta_completa)
        os.makedirs(carpeta, exist_ok=True)
        
        # v1.5: Dedup — buscar contenido idéntico en TODA la carpeta destino
        hash_nuevo = hashlib.sha256(contenido).hexdigest()
        tamaño_nuevo = len(contenido)
        
        for archivo_existente in os.listdir(carpeta):
            ruta_existente = os.path.join(carpeta, archivo_existente)
            if not os.path.isfile(ruta_existente):
                continue
            # Pre-filtro rápido por tamaño (evita leer archivos de distinto tamaño)
            if os.path.getsize(ruta_existente) != tamaño_nuevo:
                continue
            # Tamaño coincide → comparar hash
            with open(ruta_existente, 'rb') as f:
                if hashlib.sha256(f.read()).hexdigest() == hash_nuevo:
                    # Contenido idéntico encontrado → no duplicar
                    return ruta_existente, True
        
        # No hay duplicado de contenido → resolver colisión de nombre si la hay
        if os.path.exists(ruta_completa):
            base, ext = os.path.splitext(ruta_completa)
            n = 2
            while os.path.exists(f"{base} {n}{ext}"):
                n += 1
            ruta_completa = f"{base} {n}{ext}"

        # Escribir archivo
        with open(ruta_completa, 'wb') as f:
            f.write(contenido)

        return ruta_completa, False
    
    def archivo_existe(self, ruta_relativa: str) -> bool:
        """Verifica si un archivo existe en la carpeta local"""
        ruta_completa = os.path.join(self.base_path, ruta_relativa.lstrip('/\\'))
        return os.path.exists(ruta_completa)
    
    def obtener_ruta_trimestre(self, fecha: datetime) -> str:
        """Genera la ruta relativa para un trimestre"""
        año = fecha.year
        trimestre = (fecha.month - 1) // 3 + 1
        nombre_trimestre = {1: "1 TRIMESTRE", 2: "2 TRIMESTRE", 3: "3 TRIMESTRE", 4: "4 TRIMESTRE"}
        return f"FACTURAS {año}/FACTURAS RECIBIDAS/{nombre_trimestre[trimestre]} {año}"


# ============================================================================
# EXCEL GENERATOR v1.4 - CON PESTAÑA SEPA
# ============================================================================

class ExcelGenerator:
    """
    Genera Excel de pagos con pestaña SEPA (v1.4).
    
    Pestaña 1: FACTURAS (todas las facturas)
    Pestaña 2: SEPA (solo TF, con desplegable IBAN ordenante)
    """
    
    COLUMNAS_FACTURAS = [
        "#", "ARCHIVO", "PROVEEDOR", "CIF", "FECHA_FACTURA", "REF",
        "TOTAL", "IBAN", "FORMA_PAGO", "ESTADO_PAGO", "MOV#", "OBS", "REMITENTE",
        "FECHA_PROCESO", "CUENTA"
    ]
    
    COLUMNAS_SEPA = [
        "#", "INCLUIR", "PROVEEDOR", "CIF", "IBAN_BENEFICIARIO", "IMPORTE",
        "REF_FACTURA", "CONCEPTO", "FECHA_EJECUCION", "IBAN_ORDENANTE", "ARCHIVO"
    ]
    
    ANCHOS_FACTURAS = {
        "#": 6, "ARCHIVO": 45, "PROVEEDOR": 30, "CIF": 12, "FECHA_FACTURA": 12,
        "REF": 15, "TOTAL": 12, "IBAN": 30, "FORMA_PAGO": 12, "ESTADO_PAGO": 12,
        "MOV#": 10, "OBS": 25, "REMITENTE": 35, "FECHA_PROCESO": 14, "CUENTA": 15
    }
    
    ANCHOS_SEPA = {
        "#": 6, "INCLUIR": 8, "PROVEEDOR": 35, "CIF": 12, "IBAN_BENEFICIARIO": 30,
        "IMPORTE": 12, "REF_FACTURA": 20, "CONCEPTO": 40, "FECHA_EJECUCION": 14,
        "IBAN_ORDENANTE": 30, "ARCHIVO": 40
    }
    
    def __init__(self, ruta: str):
        self.ruta = ruta
        self.wb = None
        self.ws_facturas = None
        self.ws_sepa = None
    
    def _crear_nuevo(self):
        """Crea un nuevo Excel con formato y dos pestañas"""
        self.wb = Workbook()
        
        # === PESTAÑA 1: FACTURAS ===
        self.ws_facturas = self.wb.active
        self.ws_facturas.title = "FACTURAS"
        self._crear_cabecera_facturas()
        
        # === PESTAÑA 2: SEPA ===
        self.ws_sepa = self.wb.create_sheet("SEPA")
        self._crear_cabecera_sepa()
    
    def _crear_cabecera_facturas(self):
        """Crea cabecera de pestaña FACTURAS"""
        font_header = Font(name='Aptos Display', size=10, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col, nombre in enumerate(self.COLUMNAS_FACTURAS, 1):
            cell = self.ws_facturas.cell(row=1, column=col, value=nombre)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            self.ws_facturas.column_dimensions[get_column_letter(col)].width = self.ANCHOS_FACTURAS.get(nombre, 15)
        
        self.ws_facturas.freeze_panes = 'A2'
    
    def _crear_cabecera_sepa(self):
        """Crea cabecera de pestaña SEPA con información del ordenante"""
        font_header = Font(name='Aptos Display', size=10, bold=True, color="FFFFFF")
        font_info = Font(name='Aptos Display', size=10, bold=True)
        fill_header = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        fill_info = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # === CABECERA INFORMATIVA (filas 1-5) ===
        info_rows = [
            ("DATOS DEL ORDENANTE", ""),
            ("Nombre:", CONFIG.NOMBRE_ORDENANTE),
            ("NIF-Sufijo:", CONFIG.NIF_SUFIJO),
            ("BIC:", CONFIG.BIC_ORDENANTE),
            ("IBANs disponibles:", f"TASCA: {CONFIG.IBAN_TASCA}  |  COMESTIBLES: {CONFIG.IBAN_COMESTIBLES}"),
        ]
        
        for row_num, (label, valor) in enumerate(info_rows, 1):
            cell_label = self.ws_sepa.cell(row=row_num, column=1, value=label)
            cell_label.font = font_info
            cell_label.fill = fill_info
            
            cell_valor = self.ws_sepa.cell(row=row_num, column=2, value=valor)
            cell_valor.font = Font(name='Aptos Display', size=10)
            cell_valor.fill = fill_info
            
            self.ws_sepa.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=5)
        
        # Fila vacía
        fila_cabecera = 7
        
        # === CABECERA DE COLUMNAS (fila 7) ===
        for col, nombre in enumerate(self.COLUMNAS_SEPA, 1):
            cell = self.ws_sepa.cell(row=fila_cabecera, column=col, value=nombre)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            self.ws_sepa.column_dimensions[get_column_letter(col)].width = self.ANCHOS_SEPA.get(nombre, 15)
        
        self.ws_sepa.freeze_panes = 'A8'
        
        # === VALIDACIÓN DE DATOS: Desplegable IBAN ordenante ===
        opciones_iban = f'"{CONFIG.IBAN_TASCA},{CONFIG.IBAN_COMESTIBLES}"'
        dv_iban = DataValidation(
            type="list",
            formula1=opciones_iban,
            allow_blank=True
        )
        dv_iban.error = "Selecciona un IBAN de la lista"
        dv_iban.errorTitle = "IBAN no válido"
        dv_iban.prompt = "Selecciona IBAN del ordenante"
        dv_iban.promptTitle = "IBAN Ordenante"
        
        self.ws_sepa.add_data_validation(dv_iban)
        dv_iban.add('J8:J500')
        
        # === VALIDACIÓN: Columna INCLUIR (checkbox simulado) ===
        dv_incluir = DataValidation(
            type="list",
            formula1='"✓,"',
            allow_blank=True
        )
        self.ws_sepa.add_data_validation(dv_incluir)
        dv_incluir.add('B8:B500')
    
    def abrir_o_crear(self):
        """Abre Excel existente o crea uno nuevo"""
        if os.path.exists(self.ruta):
            self.wb = load_workbook(self.ruta)
            self.ws_facturas = self.wb["FACTURAS"] if "FACTURAS" in self.wb.sheetnames else self.wb.active
            if "SEPA" not in self.wb.sheetnames:
                self.ws_sepa = self.wb.create_sheet("SEPA")
                self._crear_cabecera_sepa()
            else:
                self.ws_sepa = self.wb["SEPA"]
            # v1.8: Añadir columna CUENTA si falta
            self._migrar_si_necesario()
        else:
            self._crear_nuevo()

    def _migrar_si_necesario(self):
        """
        Migración acumulativa:
        - v1.7: Eliminar columna fantasma "FECHA PROCESO" (con espacio)
        - v1.8: Añadir columna CUENTA si no existe
        """
        if self.ws_facturas is None:
            return

        headers = [self.ws_facturas.cell(1, c).value
                   for c in range(1, self.ws_facturas.max_column + 1)]

        # v1.7 fix: Eliminar columna fantasma "FECHA PROCESO" (con espacio)
        # que causaba desplazamiento de todas las columnas al escribir filas
        if 'FECHA PROCESO' in headers:
            col_fantasma = headers.index('FECHA PROCESO') + 1  # 1-based
            self.ws_facturas.delete_cols(col_fantasma)
            # Recalcular headers tras borrar
            headers = [self.ws_facturas.cell(1, c).value
                       for c in range(1, self.ws_facturas.max_column + 1)]

        # v1.8: Añadir columna CUENTA si no existe
        if 'CUENTA' not in headers:
            font_header = Font(name='Aptos Display', size=10, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            col_nueva = self.ws_facturas.max_column + 1
            cell = self.ws_facturas.cell(row=1, column=col_nueva, value='CUENTA')
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            self.ws_facturas.column_dimensions[get_column_letter(col_nueva)].width = 15
    
    def añadir_fila(self, resultado: ResultadoProcesamiento):
        """Añade una fila a la pestaña FACTURAS y a SEPA si es TF"""
        fecha_hoy = datetime.now().strftime(CONFIG.FECHA_EXCEL)
        
        # Calcular número secuencial
        ultimo_num = 0
        for row in self.ws_facturas.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], int):
                ultimo_num = max(ultimo_num, row[0])
        
        num = ultimo_num + 1
        fila = self.ws_facturas.max_row + 1
        
        # Datos
        fecha_str = ""
        total_num = None

        if resultado.factura:
            if resultado.factura.fecha:
                fecha_str = resultado.factura.fecha.strftime(CONFIG.FECHA_EXCEL)
            if resultado.factura.total:
                total_num = float(resultado.factura.total)

        # v1.8: IBAN — solo escribir IBANs reales (2 letras + dígitos, >=15 chars)
        iban_valor = ""
        if resultado.proveedor and resultado.proveedor.iban:
            iban_clean = resultado.proveedor.iban.replace(" ", "")
            if len(iban_clean) >= 15 and iban_clean[:2].isalpha():
                iban_valor = formatear_iban(resultado.proveedor.iban)

        valores = [
            num,
            resultado.archivo_generado,
            resultado.proveedor.nombre if resultado.proveedor else "",
            resultado.proveedor.cif if resultado.proveedor else "",
            fecha_str,
            resultado.factura.referencia if resultado.factura else "",
            total_num,                                                      # v1.8: float, no string
            iban_valor,                                                     # v1.8: solo IBAN real
            resultado.proveedor.forma_pago if resultado.proveedor else "",
            "",  # ESTADO_PAGO
            "",  # MOV#
            resultado.motivo_revision if resultado.requiere_revision else "",
            resultado.remitente,
            datetime.now().date(),                                          # FECHA_PROCESO
            resultado.proveedor.cuenta if resultado.proveedor else "",      # v1.8: CUENTA
        ]
        
        # Estilos
        font_normal = Font(name='Aptos Display', size=10)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        fill_alt = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for col, valor in enumerate(valores, 1):
            cell = self.ws_facturas.cell(row=fila, column=col, value=valor)
            cell.font = font_normal
            cell.border = border
            if fila % 2 == 0:
                cell.fill = fill_alt
        
        # Si es TF, añadir también a pestaña SEPA
        if resultado.proveedor and resultado.proveedor.forma_pago == "TF":
            self._añadir_fila_sepa(resultado, num, fecha_hoy, total_num or 0)
    
    def _añadir_fila_sepa(self, resultado: ResultadoProcesamiento, num_factura: int, 
                          fecha_hoy: str, total: float):
        """Añade una fila a la pestaña SEPA"""
        # Encontrar última fila con datos (empezamos en fila 8)
        fila = 8
        for row in self.ws_sepa.iter_rows(min_row=8, max_col=1, values_only=True):
            if row[0]:
                fila += 1
            else:
                break
        
        # Obtener IBAN del beneficiario
        iban_beneficiario = ""
        if resultado.proveedor and resultado.proveedor.iban:
            iban_beneficiario = formatear_iban(resultado.proveedor.iban)
        else:
            iban_beneficiario = "⚠️ FALTA IBAN"
        
        # Generar concepto
        ref = resultado.factura.referencia if resultado.factura else ""
        concepto = f"PAGO FACTURA {ref}" if ref else "PAGO FACTURA"
        
        valores_sepa = [
            num_factura,                                    # #
            "",                                             # INCLUIR (vacío para marcar)
            resultado.proveedor.nombre,                     # PROVEEDOR
            resultado.proveedor.cif,                        # CIF
            iban_beneficiario,                              # IBAN_BENEFICIARIO
            total,                                          # IMPORTE
            ref,                                            # REF_FACTURA
            concepto,                                       # CONCEPTO
            fecha_hoy,                                      # FECHA_EJECUCION
            "",                                             # IBAN_ORDENANTE (desplegable)
            resultado.archivo_generado                      # ARCHIVO
        ]
        
        font_normal = Font(name='Aptos Display', size=10)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        fill_warning = PatternFill(start_color="FFEB3B", end_color="FFEB3B", fill_type="solid")
        
        for col, valor in enumerate(valores_sepa, 1):
            cell = self.ws_sepa.cell(row=fila, column=col, value=valor)
            cell.font = font_normal
            cell.border = border
            
            # Resaltar si falta IBAN
            if col == 5 and "FALTA" in str(valor):
                cell.fill = fill_warning
    
    def guardar(self):
        """Guarda el Excel"""
        os.makedirs(os.path.dirname(self.ruta), exist_ok=True)
        self.wb.save(self.ruta)


class ExcelFacturas:
    """
    Genera Excel de facturas simplificado (v1.6).
    
    Archivo: Facturas 1T26 Provisional.xlsx
    Pestaña: Facturas
    Columnas: NOMBRE | PROVEEDOR | Fec.Fac. | Factura | Total | Origen | OBS
    
    Se genera ADEMÁS del PAGOS_Gmail (no lo sustituye).
    Detecta duplicados por NOMBRE y marca con "DUPLICADO" en OBS.
    """
    
    COLUMNAS = ["NOMBRE", "PROVEEDOR", "Fec.Fac.", "Factura", "Total", "FECHA_PROCESO", "OBS"]
    
    ANCHOS = {
        "NOMBRE": 38,
        "PROVEEDOR": 30,
        "Fec.Fac.": 12,
        "Factura": 14,
        "Total": 13,
        "FECHA_PROCESO": 14,
        "OBS": 15,
    }
    
    def __init__(self, ruta: str):
        self.ruta = ruta
        self.wb = None
        self.ws = None
    
    def _crear_nuevo(self):
        """Crea un nuevo Excel con pestaña Facturas"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Facturas"
        
        font_header = Font(name='Aptos Light', size=11, bold=True)
        
        for col, nombre in enumerate(self.COLUMNAS, 1):
            cell = self.ws.cell(row=1, column=col, value=nombre)
            cell.font = font_header
            cell.alignment = Alignment(horizontal='center')
            self.ws.column_dimensions[get_column_letter(col)].width = self.ANCHOS.get(nombre, 13)
    
    def abrir_o_crear(self):
        """Abre Excel existente o crea uno nuevo"""
        if os.path.exists(self.ruta):
            self.wb = load_workbook(self.ruta)
            self.ws = self.wb["Facturas"] if "Facturas" in self.wb.sheetnames else self.wb.active
        else:
            self._crear_nuevo()
    
    def _obtener_nombres_existentes(self) -> dict:
        """Recoge NOMBRE (col A) + Total (col E) ya presentes → dict {nombre: total}"""
        nombres = {}
        for row in self.ws.iter_rows(min_row=2, max_col=5, values_only=True):
            if row[0]:
                nombres[str(row[0]).strip()] = row[4]
        return nombres
    
    def añadir_fila(self, resultado: ResultadoProcesamiento):
        """Añade una fila a la pestaña Facturas. Marca DUPLICADO si NOMBRE ya existe."""
        fila = self.ws.max_row + 1
        
        # Col A: NOMBRE (archivo sin extensión .pdf)
        nombre_archivo = resultado.archivo_generado
        if nombre_archivo.lower().endswith('.pdf'):
            nombre_archivo = nombre_archivo[:-4]
        
        # Detectar duplicado
        nombres_existentes = self._obtener_nombres_existentes()
        total_candidato = resultado.factura.total if resultado.factura else None
        es_duplicado = False
        if nombre_archivo.strip() in nombres_existentes:
            if total_candidato == nombres_existentes[nombre_archivo.strip()]:
                es_duplicado = True
        
        # Col B: PROVEEDOR
        proveedor = resultado.proveedor.nombre if resultado.proveedor else ""
        
        # Col C: Fec.Fac. (fecha factura como datetime)
        fecha_fac = None
        if resultado.factura and resultado.factura.fecha:
            fecha_fac = resultado.factura.fecha
        
        # Col D: Factura (referencia como texto forzado con ="xxx")
        ref = ""
        if resultado.factura and resultado.factura.referencia:
            ref = f'="{resultado.factura.referencia}"'
        
        # Col E: Total
        total = None
        if resultado.factura and resultado.factura.total:
            total = resultado.factura.total
        
        # Col F: FECHA_PROCESO (solo fecha, sin hora) v1.7
        fecha_proceso_col = datetime.now().date()
        
        # Col G: OBS
        obs_parts = []
        if es_duplicado:
            obs_parts.append("DUPLICADO")
        if resultado.factura and resultado.factura.es_proforma:
            obs_parts.append("PROFORMA")
        obs = " | ".join(obs_parts)
        
        font_normal = Font(name='Aptos Light', size=11)
        
        valores = [nombre_archivo, proveedor, fecha_fac, ref, total, fecha_proceso_col, obs]
        
        for col, valor in enumerate(valores, 1):
            cell = self.ws.cell(row=fila, column=col, value=valor)
            cell.font = font_normal
            
            # Formato fecha dd-mm-yy para Fec.Fac.
            if col == 3 and valor:
                cell.number_format = 'dd\\-mm\\-yy;@'
            # Formato fecha dd-mm-yy para FECHA_PROCESO
            elif col == 6:
                cell.number_format = 'dd\\-mm\\-yy'
    
    def guardar(self):
        """Guarda el Excel"""
        os.makedirs(os.path.dirname(self.ruta), exist_ok=True)
        self.wb.save(self.ruta)


# ============================================================================
# NOTIFICACIONES
# ============================================================================

class Notificador:
    """Envía notificaciones por email"""
    
    def __init__(self, gmail_service):
        self.service = gmail_service
    
    def enviar_resumen(
        self,
        procesados: List[ResultadoProcesamiento],
        errores: List[str],
        excel_path: str,
        dropbox_folder: str,
        log_path: str
    ):
        """Envía email de resumen HTML"""
        total = len(procesados)
        exitosos = sum(1 for p in procesados if not p.requiere_revision)
        revision = sum(1 for p in procesados if p.requiere_revision)
        alertas_rojas = sum(1 for p in procesados if p.alerta_roja)
        ibans_sugeridos = sum(1 for p in procesados if p.iban_sugerido)
        
        html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: 'Aptos Display', Arial, sans-serif; font-size: 14px; }}
                .header {{ background-color: #1F4E79; color: white; padding: 15px; }}
                .stats {{ margin: 20px 0; }}
                .stat-box {{ display: inline-block; padding: 10px 20px; margin: 5px; border-radius: 5px; }}
                .ok {{ background-color: #D4EDDA; color: #155724; }}
                .warn {{ background-color: #FFF3CD; color: #856404; }}
                .error {{ background-color: #F8D7DA; color: #721C24; }}
                .alerta-roja {{ background-color: #FF0000; color: white; font-weight: bold; padding: 15px; margin: 10px 0; border-radius: 5px; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th {{ background-color: #1F4E79; color: white; padding: 10px; text-align: left; }}
                td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                tr.roja {{ background-color: #FFCCCC; }}
                .paths {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin: 20px 0; }}
                .path {{ margin: 5px 0; font-family: monospace; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>📧 Gmail Module v{VERSION} - Resumen de Procesamiento</h2>
                <p>{datetime.now().strftime('%d/%m/%Y %H:%M')}</p>
            </div>
        """
        
        if alertas_rojas > 0:
            html += f"""
            <div class="alerta-roja">
                🔴 ALERTAS ROJAS: {alertas_rojas} facturas SIN TOTAL EXTRAÍDO<br>
                <small>Revisar manualmente ANTES de continuar. El total está en 0.00€</small>
            </div>
            """
        
        html += f"""
            <div class="stats">
                <div class="stat-box ok">✅ Procesados: {exitosos}</div>
                <div class="stat-box warn">⚠️ Requieren revisión: {revision}</div>
                <div class="stat-box error">❌ Errores: {len(errores)}</div>
            </div>
        """
        
        if ibans_sugeridos > 0:
            html += f"""
            <div style="background-color: #FFF3CD; padding: 15px; border-radius: 5px; margin: 10px 0;">
                💡 <strong>{ibans_sugeridos} IBANs detectados en PDFs</strong> que difieren del MAESTRO.<br>
                <small>Revisar archivo <code>⚠️_IBANS_SUGERIDOS_*.xlsx</code> antes de hacer pagos.</small>
            </div>
            """
        
        html += f"""
            <div class="paths">
                <h3>📁 Archivos</h3>
                <p class="path"><strong>Excel:</strong> {excel_path}</p>
                <p class="path"><strong>Dropbox:</strong> {dropbox_folder}</p>
                <p class="path"><strong>Log:</strong> {log_path}</p>
            </div>
        """
        
        items_revision = [p for p in procesados if p.requiere_revision]
        items_revision.sort(key=lambda x: (not x.alerta_roja, x.archivo_generado))
        
        if items_revision:
            html += """
            <h3>⚠️ Items que Requieren Revisión</h3>
            <table>
                <tr>
                    <th>Archivo</th>
                    <th>Remitente</th>
                    <th>Motivo</th>
                </tr>
            """
            for item in items_revision:
                clase_fila = 'class="roja"' if item.alerta_roja else ''
                html += f"""
                <tr {clase_fila}>
                    <td>{item.archivo_generado}</td>
                    <td>{item.remitente}</td>
                    <td>{item.motivo_revision}</td>
                </tr>
                """
            html += "</table>"
        
        if errores:
            html += """
            <h3>❌ Errores</h3>
            <ul>
            """
            for error in errores:
                html += f"<li>{error}</li>"
            html += "</ul>"
        
        html += """
        </body>
        </html>
        """
        
        asunto = f"Gmail Module v{VERSION} - {datetime.now().strftime('%d/%m/%Y')} - {total} facturas"
        if alertas_rojas > 0:
            asunto = f"🔴 ALERTAS ROJAS - {asunto}"
        
        message = MIMEMultipart('alternative')
        message['To'] = CONFIG.EMAIL_NOTIFICACION
        message['Subject'] = asunto
        message.attach(MIMEText(html, 'html'))
        
        raw = base64.urlsafe_b64encode(message.as_bytes()).decode()
        # v1.7: llamada directa (Notificador no tiene _api_call)
        self.service.users().messages().send(
            userId='me',
            body={'raw': raw}
        ).execute()


# ============================================================================
# BACKUP MANAGER
# ============================================================================

class BackupManager:
    """Gestiona backups de archivos críticos"""
    
    def __init__(self, backup_path: str, max_backups: int = 5):
        self.backup_path = backup_path
        self.max_backups = max_backups
        os.makedirs(backup_path, exist_ok=True)
    
    def crear_backup(self, archivo: str) -> Optional[str]:
        """Crea backup de un archivo"""
        if not os.path.exists(archivo):
            return None
        
        nombre = os.path.basename(archivo)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_backup = f"{os.path.splitext(nombre)[0]}_backup_{timestamp}{os.path.splitext(nombre)[1]}"
        ruta_backup = os.path.join(self.backup_path, nombre_backup)
        
        shutil.copy2(archivo, ruta_backup)
        self._limpiar_antiguos(nombre)
        
        return ruta_backup
    
    def _limpiar_antiguos(self, nombre_original: str):
        """Mantiene solo los últimos N backups"""
        base = os.path.splitext(nombre_original)[0]
        patron = f"{base}_backup_"
        
        backups = []
        for f in os.listdir(self.backup_path):
            if f.startswith(patron):
                ruta = os.path.join(self.backup_path, f)
                backups.append((ruta, os.path.getmtime(ruta)))
        
        backups.sort(key=lambda x: x[1], reverse=True)
        
        for ruta, _ in backups[self.max_backups:]:
            os.remove(ruta)


# ============================================================================
# PROCESADOR PRINCIPAL
# ============================================================================

class GmailProcessor:
    """Procesador principal del módulo Gmail"""
    
    def __init__(self, modo_test: bool = False, modo_produccion: bool = False):
        self.modo_test = modo_test
        self._modo_produccion = modo_produccion
        self.logger = configurar_logging(modo_test)

        self.maestro = MaestroProveedores(CONFIG.MAESTRO_PATH)
        self.control = ControlDuplicados(CONFIG.JSON_PATH)
        self.backup = BackupManager(CONFIG.BACKUPS_PATH)

        self.gmail = None
        self.dropbox = None

        self.resultados: List[ResultadoProcesamiento] = []
        self.errores: List[str] = []

        # v1.18: Contadores ventana de gracia
        self._contadores_gracia = 0
        self._contadores_pendientes = 0
    
    def ejecutar(self) -> bool:
        """Ejecuta el procesamiento completo"""
        self.logger.info("=" * 60)
        self.logger.info(f"GMAIL MODULE v{VERSION} - {'MODO TEST' if self.modo_test else 'PRODUCCIÓN'}")
        self.logger.info("=" * 60)
        
        try:
            self._conectar_servicios()

            # v1.13: Filtro por fecha — solo emails posteriores a la última ejecución
            after_date = None
            ultima = self.control.get_ultima_ejecucion()
            if ultima:
                try:
                    dt = datetime.fromisoformat(ultima)
                    # Restar 1 día para cubrir posibles desfases horarios
                    dt_filtro = dt - timedelta(days=1)
                    after_date = dt_filtro.strftime("%Y/%m/%d")
                    self.logger.info(f"Filtro fecha: after:{after_date} (última ejecución: {ultima[:10]})")
                except ValueError:
                    self.logger.warning(f"Fecha última ejecución inválida: {ultima}, procesando todo")

            emails = self.gmail.obtener_emails_pendientes(CONFIG.MAX_EMAILS, after_date=after_date)
            self.logger.info(f"Emails pendientes: {len(emails)}")

            if not emails:
                self.logger.info("No hay emails para procesar")
                return True

            self._crear_backups()

            fecha_proceso = datetime.now()
            for email in emails:
                self._procesar_email(email, fecha_proceso)

            # v1.13: Guardar fecha de esta ejecución como cursor
            self.control.set_ultima_ejecucion(fecha_proceso.isoformat())
            self.control.guardar()

            self._generar_proveedores_nuevos()
            self._enviar_notificacion()
            self._log_resumen()

            # Sync Google Drive (Estrategia B)
            if not self.modo_test:
                try:
                    from nucleo.sync_drive import sync_archivos
                    trimestre = obtener_trimestre(datetime.now())
                    archivos_drive = [
                        os.path.join(CONFIG.OUTPUT_PATH, f"PAGOS_Gmail_{trimestre}.xlsx"),
                        os.path.join(CONFIG.OUTPUT_PATH, f"Facturas {trimestre} Provisional.xlsx"),
                        CONFIG.MAESTRO_PATH,
                    ]
                    sync_archivos(archivos_drive, carpeta="Facturas")
                    self.logger.info("Sync Drive completado")
                except Exception as e:
                    self.logger.error("Error en sync Drive: %s", e)

            return True
            
        except Exception as e:
            self.logger.error(f"Error crítico: {e}", exc_info=True)
            self.errores.append(str(e))
            # Guardar lo que tengamos hasta ahora
            if not self.modo_test:
                try:
                    self.control.guardar()
                    self.logger.info("JSON de control guardado tras error crítico")
                except Exception as e2:
                    self.logger.error("No se pudo guardar JSON de control: %s", e2)
            return False

    def _conectar_servicios(self):
        """Conecta con Gmail y Dropbox Local"""
        self.logger.info("Conectando a Gmail...")
        credentials_path = os.path.join(CONFIG.GMAIL_PATH, "credentials.json")
        token_path = os.path.join(CONFIG.GMAIL_PATH, "token.json")
        
        self.gmail = GmailClient(credentials_path, token_path)
        self.gmail.conectar()
        self.logger.info("Gmail conectado ✓")
        
        # Dropbox — auto-selección Local vs API (v1.16)
        self.logger.info("Conectando a Dropbox...")
        try:
            from dropbox_selector import crear_cliente_dropbox
            self.dropbox = crear_cliente_dropbox(CONFIG)
            if self.dropbox:
                self.logger.info("Dropbox conectado ✓")
            else:
                self.logger.warning("Dropbox no disponible")
        except Exception as e:
            self.logger.warning(f"Dropbox no disponible: {e}")
            self.dropbox = None
    
    def _crear_backups(self):
        """Crea backups de archivos críticos"""
        self.logger.info("Creando backups...")
        self.backup.crear_backup(CONFIG.JSON_PATH)
        
        trimestre = obtener_trimestre(datetime.now())
        excel_path = os.path.join(CONFIG.OUTPUT_PATH, f"PAGOS_Gmail_{trimestre}.xlsx")
        if os.path.exists(excel_path):
            self.backup.crear_backup(excel_path)
        
        facturas_path = os.path.join(CONFIG.OUTPUT_PATH, f"Facturas {trimestre} Provisional.xlsx")
        if os.path.exists(facturas_path):
            self.backup.crear_backup(facturas_path)
        
        # v1.5: Verificar que los Excel no están bloqueados (abiertos en otra app)
        for path_check in [excel_path, facturas_path]:
            if os.path.exists(path_check):
                try:
                    with open(path_check, 'a'):
                        pass
                except PermissionError:
                    self.logger.error(f"❌ ARCHIVO BLOQUEADO: {path_check}")
                    self.logger.error("   Cierra el Excel y vuelve a ejecutar.")
                    raise SystemExit(1)
    
    def _procesar_email(self, email_data: Dict, fecha_proceso: datetime):
        """
        Procesa un email individual.
        
        v1.6 FLUJO ANTI-DUPLICADOS:
        1. Comprobar JSON PRIMERO (antes de todo)
        2. Mover a FACTURAS_PROCESADAS ANTES de procesar (así si falla, no se repite)
        3. Guardar JSON DESPUÉS de cada email (no al final)
        4. Registrar TODOS los emails vistos (con y sin adjuntos)
        """
        email_id = email_data['id']
        remitente = email_data.get('from', '')
        asunto = email_data.get('subject', '')
        
        # ── PASO 0: Comprobar si ya lo procesamos (JSON) ──
        if self.control.email_procesado(email_id):
            self.logger.info(f"  ↳ Email ya procesado (JSON), saltando: {asunto[:40]}...")
            # Por si acaso sigue en FACTURAS, moverlo
            if not self.modo_test:
                self._mover_email_seguro(email_id)
            return
        
        # ── PASO 1: Filtrar reenvíos propios ──
        email_remitente = remitente.lower()
        if "<" in email_remitente and ">" in email_remitente:
            email_remitente = email_remitente.split("<")[1].split(">")[0].strip()
        
        if email_remitente in CONFIG.EMAILS_IGNORAR:
            self.logger.info(f"Ignorando reenvío de {email_remitente}: {asunto[:40]}...")
            if not self.modo_test:
                self._mover_email_seguro(email_id)
                self.control.registrar_email_visto(email_id, "reenvío propio")
                self.control.guardar()
            return
        
        self.logger.info(f"Procesando: {asunto[:50]}...")
        
        # ── PASO 2: Mover a PROCESADAS ANTES de procesar ──
        # Así, si el script falla a mitad, el email NO reaparece la próxima semana
        if not self.modo_test:
            self._mover_email_seguro(email_id)
        
        resultado = ResultadoProcesamiento(
            email_id=email_id,
            message_id=email_data.get('message_id', ''),
            remitente=remitente,
            asunto=asunto
        )
        
        try:
            # ── PASO 3: Descargar y clasificar adjuntos ──
            adjuntos = self.gmail.descargar_adjuntos(email_data)
            if not adjuntos:
                self.logger.warning(f"  ↳ Sin adjuntos PDF/JPG")
                resultado.error = "Sin adjuntos"
                self.resultados.append(resultado)
                if not self.modo_test:
                    self.control.registrar_email_visto(email_id, "sin adjuntos")
                    self.control.guardar()
                return
            
            # ── PASO 4: Procesar TODOS los PDFs adjuntos (v1.18.1) ──
            pdfs = [(n, c) for n, c in adjuntos if n.lower().endswith('.pdf')]
            imagenes = [(n, c) for n, c in adjuntos if n.lower().endswith(('.jpg', '.jpeg', '.png'))]

            if pdfs:
                # Primer PDF → usa el resultado principal
                self._procesar_pdf(resultado, pdfs[0][0], pdfs[0][1], fecha_proceso)

                # PDFs adicionales → cada uno crea su propio resultado
                for i, (nombre_pdf, contenido_pdf) in enumerate(pdfs[1:], start=2):
                    self.logger.info(f"  ↳ PDF adicional #{i}: {nombre_pdf}")
                    resultado_extra = ResultadoProcesamiento(
                        email_id=f"{email_id}__pdf{i}",
                        message_id=email_data.get('message_id', ''),
                        remitente=remitente,
                        asunto=asunto,
                    )
                    self._procesar_pdf(resultado_extra, nombre_pdf, contenido_pdf, fecha_proceso)
                    self.resultados.append(resultado_extra)
                    # Registrar en JSON inmediatamente
                    if not self.modo_test:
                        if resultado_extra.es_duplicado:
                            self.control.registrar_email_visto(
                                f"{email_id}__pdf{i}",
                                resultado_extra.motivo_revision or "duplicado"
                            )
                        else:
                            self.control.registrar_y_guardar(resultado_extra)

                if len(pdfs) > 1:
                    self.logger.warning(
                        f"  ↳ ⚠️ {len(pdfs)} PDFs procesados en este email"
                    )

            # Imágenes solo si NO hubo ningún PDF y no es duplicado
            if not pdfs and not resultado.archivo_generado and not resultado.es_duplicado:
                for nombre_img, contenido_img in imagenes:
                    self._procesar_imagen(resultado, nombre_img, contenido_img, fecha_proceso)
                    break  # Solo una imagen (no procesar múltiples capturas)

            self.resultados.append(resultado)

            # ── PASO 5: Registrar en JSON y guardar INMEDIATAMENTE ──
            if not self.modo_test:
                if resultado.es_duplicado:
                    # v1.8: Duplicado → solo registrar email como visto (no sobreescribir facturas/hashes)
                    self.control.registrar_email_visto(email_id, resultado.motivo_revision or "duplicado")
                    self.control.guardar()
                else:
                    self.control.registrar_y_guardar(resultado)
            
        except Exception as e:
            self.logger.error(f"  ↳ Error: {e}")
            resultado.error = str(e)
            self.errores.append(f"{asunto}: {e}")
            self.resultados.append(resultado)
            # Registrar incluso si hubo error (ya movimos el email)
            if not self.modo_test:
                self.control.registrar_email_visto(email_id, f"error: {e}")
                self.control.guardar()
    
    def _mover_email_seguro(self, email_id: str):
        """
        Mueve email a PROCESADAS con retry.
        Si falla, solo logea warning (no bloquea el flujo).
        """
        for intento in range(CONFIG.REINTENTOS):
            try:
                self.gmail.mover_a_procesados_y_marcar_leido(email_id)
                return
            except Exception as e:
                if intento < CONFIG.REINTENTOS - 1:
                    self.logger.warning(f"  ↳ Retry mover email (intento {intento+1}): {e}")
                    time.sleep(2 ** intento)  # Backoff: 1s, 2s, 4s
                else:
                    self.logger.error(f"  ↳ No se pudo mover email a PROCESADAS: {e}")
    
    def _procesar_pdf(
        self,
        resultado: ResultadoProcesamiento,
        nombre_archivo: str,
        contenido: bytes,
        fecha_proceso: datetime
    ):
        """Procesa un archivo PDF"""
        ya_en_dropbox = False  # v1.18 fix: inicializar antes de cualquier uso
        hash_pdf = calcular_hash_archivo(contenido)
        if self.control.hash_existe(hash_pdf):
            self.logger.info(f"  ↳ PDF duplicado (hash), saltando")
            resultado.es_duplicado = True
            return
        
        resultado.hash_pdf = hash_pdf
        
        # v1.12: Extraer texto del PDF una sola vez para reutilizar
        texto_pdf_cache = ""
        try:
            with pdfplumber.open(io.BytesIO(contenido)) as pdf:
                for page in pdf.pages[:2]:
                    t = page.extract_text()
                    if t:
                        texto_pdf_cache += t + "\n"
        except Exception as e:
            self.logger.debug("Error extrayendo texto PDF para cache: %s", e)

        nombre_remitente = resultado.remitente.split("<")[0].strip() if "<" in resultado.remitente else resultado.remitente
        proveedor = self.maestro.identificar_proveedor(
            resultado.remitente,
            nombre_remitente,
            resultado.asunto
        )

        # v1.10: Si no se identificó por email/alias/fuzzy, intentar por CIF en PDF
        if not proveedor:
            proveedor = self.maestro.identificar_por_pdf(contenido, logger=self.logger, texto_pdf=texto_pdf_cache)

        resultado.proveedor = proveedor
        resultado.proveedor_identificado = proveedor is not None

        if not proveedor:
            resultado.requiere_revision = True
            resultado.motivo_revision = "Proveedor no identificado"
            self.logger.warning(f"  ↳ Proveedor no identificado")
        else:
            self.logger.info(f"  ↳ Proveedor: {proveedor.nombre}")

            # v1.11: Validación CIF — comprobar que el CIF del PDF coincide con el del proveedor
            if proveedor.cif:
                prov_pdf = self.maestro.identificar_por_pdf(contenido, texto_pdf=texto_pdf_cache)
                if prov_pdf and prov_pdf.cif:
                    cif_maestro = proveedor.cif.upper().replace(' ', '').replace('-', '')
                    cif_pdf = prov_pdf.cif.upper().replace(' ', '').replace('-', '')
                    if cif_maestro != cif_pdf:
                        self.logger.warning(
                            f"  ↳ ⚠️ CIF no coincide: identificado={proveedor.nombre} ({cif_maestro})"
                            f" vs PDF={prov_pdf.nombre} ({cif_pdf})"
                        )
                        # Reasignar al proveedor correcto del PDF
                        proveedor = prov_pdf
                        resultado.proveedor = proveedor
                        resultado.requiere_revision = True
                        resultado.motivo_revision = f"CIF PDF ({cif_pdf}) no coincide con identificación inicial"
                        self.logger.info(f"  ↳ Reasignado a: {proveedor.nombre}")

        # Extraer datos del PDF
        factura = None
        
        if proveedor and proveedor.tiene_extractor and proveedor.archivo_extractor:
            factura = self._usar_extractor_dedicado(proveedor, contenido)
            # El mensaje "Usando extractor" ya se muestra dentro de _usar_extractor_dedicado
        
        if not factura or not factura.exito:
            extractor = ExtractorPDF(contenido)
            factura = extractor.extraer()
        elif not factura.fecha or not factura.total:
            # v1.14: Extractor dedicado obtuvo datos parciales → complementar con genérico
            extractor = ExtractorPDF(contenido)
            factura_generica = extractor.extraer()
            if not factura.fecha and factura_generica.fecha:
                factura.fecha = factura_generica.fecha
                self.logger.debug(f"  ↳ Fecha completada por extractor genérico")
            if not factura.total and factura_generica.total:
                factura.total = factura_generica.total
                self.logger.debug(f"  ↳ Total completado por extractor genérico")
            if not factura.referencia and factura_generica.referencia:
                factura.referencia = factura_generica.referencia

        resultado.factura = factura
        
        if factura.fecha:
            self.logger.info(f"  ↳ Fecha: {factura.fecha.strftime('%d/%m/%Y')}")
            if factura.fecha_futura:
                resultado.requiere_revision = True
                resultado.motivo_revision = f"Fecha futura: {factura.fecha.strftime('%d/%m/%Y')}"
                self.logger.warning(f"  ↳ ⚠️ Fecha futura detectada")

            # v1.18.2: Fecha demasiado antigua
            dias_antiguedad = (fecha_proceso - factura.fecha).days
            if dias_antiguedad > FECHA_MAX_ANTIGUEDAD_DIAS:
                resultado.requiere_revision = True
                msg = f"Fecha muy antigua: {factura.fecha.strftime('%d/%m/%Y')} ({dias_antiguedad} días)"
                resultado.motivo_revision = (
                    (resultado.motivo_revision + " | " if resultado.motivo_revision else "") + msg
                )
                self.logger.warning(f"  ↳ ⚠️ {msg}")
        else:
            # v1.7: Sin fecha → no podemos saber si es ATRASADA → avisar
            resultado.requiere_revision = True
            aviso_atrasada = "⚠️ FECHA NO DETECTADA - verificar si es ATRASADA"
            if resultado.motivo_revision:
                resultado.motivo_revision = resultado.motivo_revision + " | " + aviso_atrasada
            else:
                resultado.motivo_revision = aviso_atrasada
            self.logger.warning(f"  ↳ ⚠️ Fecha no detectada - puede ser factura ATRASADA")
        
        if factura.total:
            self.logger.info(f"  ↳ Total: {factura.total:.2f}€")
        else:
            resultado.alerta_roja = True
            resultado.requiere_revision = True
            alerta = "🔴 ALERTA ROJA: Total no extraído - REVISIÓN OBLIGATORIA"
            if resultado.motivo_revision:
                resultado.motivo_revision = resultado.motivo_revision + " | " + alerta
            else:
                resultado.motivo_revision = alerta
            # v1.17: No asignar 0.00 — dejar None para que Excel muestre celda vacía
            self.logger.error(f"  ↳ 🔴 ALERTA ROJA: No se pudo extraer total del PDF")

        # v1.18.2: Validación rango de total
        if factura.total is not None:
            if 0 < abs(factura.total) < TOTAL_MIN_SOSPECHOSO:
                resultado.requiere_revision = True
                msg = f"Total sospechosamente bajo: {factura.total:.2f}€"
                resultado.motivo_revision = (
                    (resultado.motivo_revision + " | " if resultado.motivo_revision else "") + msg
                )
                self.logger.warning(f"  ↳ ⚠️ {msg}")

            elif abs(factura.total) > TOTAL_MAX_SOSPECHOSO:
                resultado.requiere_revision = True
                msg = f"Total sospechosamente alto: {factura.total:.2f}€"
                resultado.motivo_revision = (
                    (resultado.motivo_revision + " | " if resultado.motivo_revision else "") + msg
                )
                self.logger.warning(f"  ↳ ⚠️ {msg}")

            # v1.18.2: Detección de abonos (total negativo)
            if factura.total < 0:
                resultado.requiere_revision = True
                msg = f"POSIBLE ABONO (total negativo: {factura.total:.2f}€)"
                resultado.motivo_revision = (
                    (resultado.motivo_revision + " | " if resultado.motivo_revision else "") + msg
                )
                self.logger.warning(f"  ↳ ⚠️ {msg}")

        if factura.referencia:
            self.logger.info(f"  ↳ Ref: {factura.referencia}")
        
        if factura.iban_detectado and proveedor:
            iban_maestro = proveedor.iban.replace(" ", "").upper() if proveedor.iban else ""
            iban_pdf = factura.iban_detectado.replace(" ", "").upper()
            
            if iban_pdf != iban_maestro:
                resultado.iban_sugerido = factura.iban_detectado
                if iban_maestro:
                    self.logger.warning(f"  ↳ ⚠️ IBAN PDF ({iban_pdf[:8]}...) ≠ MAESTRO ({iban_maestro[:8]}...)")
                else:
                    self.logger.info(f"  ↳ 💡 IBAN detectado en PDF: {iban_pdf[:8]}... (no está en MAESTRO)")
        
        # v1.8: Detectar duplicado CIF+REF (o NOMBRE+REF si CIF vacío)
        es_duplicado_cif_ref = False
        if proveedor and factura.referencia:
            if self.control.factura_existe(proveedor.cif, factura.referencia, proveedor.nombre):
                es_duplicado_cif_ref = True
                resultado.es_duplicado = True
                self.logger.warning(f"  ↳ Factura CIF+REF duplicada — NO se guardará")
                resultado.requiere_revision = True
                resultado.motivo_revision = "Duplicado (CIF+REF) — descartado"

        # v1.18: Ventana de gracia — determinar destino antes de generar nombre
        fecha_factura_dt = factura.fecha if factura.fecha else fecha_proceso
        destino = determinar_destino_factura(fecha_factura_dt, fecha_proceso)

        nombre_generado = generar_nombre_archivo(
            proveedor, factura, fecha_proceso, resultado.remitente, destino=destino
        )
        resultado.archivo_generado = nombre_generado
        self.logger.info(f"  ↳ Archivo: {nombre_generado}")

        if destino == 'GRACIA':
            self.logger.info(
                f"  ↳ ✅ GRACIA: {obtener_trimestre(fecha_factura_dt)} subida a carpeta trimestre anterior "
                f"(día {fecha_proceso.day} de {fecha_proceso.strftime('%B')})"
            )
            self._contadores_gracia += 1
        elif destino == 'ATRASADA':
            self.logger.info(f"  ↳ ATRASADA: irá a subcarpeta ATRASADAS/")

        # v1.18: PENDIENTE_UBICACION — zona gris, requiere decisión
        if destino == 'PENDIENTE_UBICACION' and not es_duplicado_cif_ref:
            info_pendiente = construir_info_pendiente(resultado, factura, nombre_generado, fecha_proceso)

            if self.modo_test:
                self.logger.warning(f"  ↳ ⚠️ PENDIENTE_UBICACION (test, no se guarda en cola)")
            elif not self._modo_produccion:
                # Modo manual → preguntar en terminal
                decision = preguntar_destino_manual(info_pendiente)
                destino = decision
                # Regenerar nombre con el destino decidido
                nombre_generado = generar_nombre_archivo(
                    proveedor, factura, fecha_proceso, resultado.remitente, destino=destino
                )
                resultado.archivo_generado = nombre_generado
                self.logger.info(f"  ↳ Decisión manual: {destino} → {nombre_generado}")
            else:
                # Modo automático (--produccion) → guardar en cola
                guardar_en_cola_pendientes(info_pendiente, contenido, self.logger)
                self._contadores_pendientes += 1
                # Registrar en Excel con OBS PENDIENTE
                if not ya_en_dropbox:
                    resultado.motivo_revision = (
                        (resultado.motivo_revision + " | " if resultado.motivo_revision else "")
                        + "PENDIENTE_UBICACION"
                    )
                    trimestre = obtener_trimestre(fecha_proceso)
                    excel_path = os.path.join(CONFIG.OUTPUT_PATH, f"PAGOS_Gmail_{trimestre}.xlsx")
                    excel = ExcelGenerator(excel_path)
                    excel.abrir_o_crear()
                    excel.añadir_fila(resultado)
                    excel.guardar()

                    facturas_path = os.path.join(CONFIG.OUTPUT_PATH, f"Facturas {trimestre} Provisional.xlsx")
                    facturas_excel = ExcelFacturas(facturas_path)
                    facturas_excel.abrir_o_crear()
                    facturas_excel.añadir_fila(resultado)
                    facturas_excel.guardar()

                    self.logger.info(f"  ↳ Registrado en Excel con OBS=PENDIENTE_UBICACION")
                return  # No subir a Dropbox, PDF guardado en cola

        if es_duplicado_cif_ref:
            self.logger.info(f"  ↳ Duplicado CIF+REF: se omite Dropbox y Excel")
        else:
            # Subir a Dropbox (v1.4 local, v1.16 API, v1.18 ventana gracia)
            ya_en_dropbox = False
            if self.dropbox and not self.modo_test:
                ruta_dropbox, ya_en_dropbox = self.dropbox.subir_archivo(
                    contenido,
                    nombre_generado,
                    fecha_factura_dt,
                    fecha_proceso,
                    destino=destino
                )
                resultado.dropbox_path = ruta_dropbox
                if ya_en_dropbox:
                    self.logger.info(f"  ↳ Ya existe en Dropbox (contenido idéntico), saltando")
                else:
                    # Actualizar nombre si Dropbox añadió sufijo anti-colisión
                    nombre_real = os.path.basename(ruta_dropbox)
                    if nombre_real != nombre_generado:
                        resultado.archivo_generado = nombre_real
                        self.logger.info(f"  ↳ Renombrado a: {nombre_real} (colisión)")
                    self.logger.info(f"  ↳ Copiado a Dropbox")

            # Añadir al Excel PAGOS_Gmail (solo si no era duplicado en Dropbox)
            if not self.modo_test and not ya_en_dropbox:
                trimestre = obtener_trimestre(fecha_proceso)
                excel_path = os.path.join(CONFIG.OUTPUT_PATH, f"PAGOS_Gmail_{trimestre}.xlsx")

                excel = ExcelGenerator(excel_path)
                excel.abrir_o_crear()
                excel.añadir_fila(resultado)
                excel.guardar()

                # Añadir al Excel Facturas_XTYY_Provisional (v1.6)
                facturas_path = os.path.join(CONFIG.OUTPUT_PATH, f"Facturas {trimestre} Provisional.xlsx")
                facturas_excel = ExcelFacturas(facturas_path)
                facturas_excel.abrir_o_crear()
                facturas_excel.añadir_fila(resultado)
                facturas_excel.guardar()

                self.logger.info(f"  ↳ Añadido a Excel")

        # NOTA: El registro en JSON se hace en _procesar_email() con registrar_y_guardar()
    
    def _usar_extractor_dedicado(self, proveedor: Proveedor, contenido: bytes) -> Optional[FacturaExtraida]:
        """
        Usa el extractor dedicado del proveedor si existe.
        
        v1.4: Instancia correctamente las clases de PARSEO
        Los extractores de PARSEO son CLASES que heredan de ExtractorBase.
        """
        try:
            # Sanitizar nombre: solo basename, sin path traversal
            nombre_archivo = os.path.basename(proveedor.archivo_extractor)
            if not nombre_archivo.endswith('.py') or nombre_archivo != proveedor.archivo_extractor:
                self.logger.warning(f"  ↳ Nombre de extractor no válido: {proveedor.archivo_extractor}")
                return None

            ruta_extractor = os.path.join(CONFIG.EXTRACTORES_PATH, nombre_archivo)

            # Verificar containment: que la ruta resuelta esté dentro de EXTRACTORES_PATH
            ruta_real = os.path.realpath(ruta_extractor)
            carpeta_real = os.path.realpath(CONFIG.EXTRACTORES_PATH)
            if not ruta_real.startswith(carpeta_real + os.sep) and ruta_real != carpeta_real:
                self.logger.warning(f"  ↳ Extractor fuera de carpeta permitida: {ruta_real}")
                return None

            if not os.path.exists(ruta_extractor):
                self.logger.debug(f"  ↳ Extractor no encontrado: {ruta_extractor}")
                return None
            
            # Añadir carpeta PADRE al path para que Python encuentre el paquete 'extractores'
            # Los extractores hacen: from extractores.base import ExtractorBase
            # Por tanto necesitamos que 'Parseo' esté en el path, no 'extractores'
            carpeta_padre = os.path.dirname(CONFIG.EXTRACTORES_PATH)  # C:\..\Parseo
            if carpeta_padre not in sys.path:
                sys.path.insert(0, carpeta_padre)
            # También añadir extractores por si hay imports relativos
            if CONFIG.EXTRACTORES_PATH not in sys.path:
                sys.path.insert(0, CONFIG.EXTRACTORES_PATH)
            
            # Cargar módulo
            nombre_modulo = f"extractor_{nombre_archivo.replace('.py', '')}"
            spec = importlib.util.spec_from_file_location(nombre_modulo, ruta_extractor)
            modulo = importlib.util.module_from_spec(spec)
            sys.modules[nombre_modulo] = modulo
            spec.loader.exec_module(modulo)
            
            # Guardar PDF temporalmente
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(contenido)
                tmp_path = tmp.name
            
            try:
                datos = None
                
                # Buscar clase que herede de ExtractorBase
                clase_extractor = None
                for nombre in dir(modulo):
                    obj = getattr(modulo, nombre)
                    if isinstance(obj, type) and nombre.startswith('Extractor') and nombre != 'ExtractorBase':
                        clase_extractor = obj
                        break
                
                if clase_extractor:
                    instancia = clase_extractor()
                    
                    # Opción 1: Método extraer() que devuelve dict completo (como ceres.py)
                    if hasattr(instancia, 'extraer') and callable(instancia.extraer):
                        datos = instancia.extraer(tmp_path)
                        self.logger.info(f"  ↳ Usando extractor: {proveedor.archivo_extractor}")
                    
                    # Opción 2: Métodos individuales que reciben TEXTO (como borboton.py, bernal.py)
                    elif hasattr(instancia, 'extraer_total') or hasattr(instancia, 'extraer_fecha'):
                        self.logger.info(f"  ↳ Usando extractor: {proveedor.archivo_extractor}")
                        
                        # Extraer texto del PDF primero
                        texto_pdf = ""
                        try:
                            with pdfplumber.open(tmp_path) as pdf:
                                for page in pdf.pages:
                                    texto_pagina = page.extract_text()
                                    if texto_pagina:
                                        texto_pdf += texto_pagina + "\n"
                        except Exception as e:
                            self.logger.warning(f"  ↳ ⚠️ Error extrayendo texto en {proveedor.archivo_extractor}: {e}")
                        
                        # v1.6: Si pdfplumber no extrajo texto, intentar OCR
                        # (necesario para PDFs que son imágenes, como La Llildiria)
                        if not texto_pdf.strip() and OCR_DISPONIBLE:
                            try:
                                self.logger.debug(f"  ↳ PDF sin texto, intentando OCR...")
                                with pdfplumber.open(tmp_path) as pdf:
                                    for page in pdf.pages:
                                        img = page.to_image(resolution=300)
                                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as img_tmp:
                                            img.save(img_tmp.name)
                                            ocr_text = pytesseract.image_to_string(
                                                Image.open(img_tmp.name), lang='spa'
                                            )
                                            texto_pdf += ocr_text + "\n"
                                            os.unlink(img_tmp.name)
                            except Exception as e:
                                self.logger.warning(f"  ↳ ⚠️ Error OCR en {proveedor.archivo_extractor}: {e}")
                        
                        if texto_pdf:
                            datos = {}
                            
                            # Llamar métodos individuales pasando TEXTO
                            if hasattr(instancia, 'extraer_fecha'):
                                try:
                                    datos['fecha'] = instancia.extraer_fecha(texto_pdf)
                                except Exception as e:
                                    self.logger.warning(f"  ↳ ⚠️ extraer_fecha falló en {proveedor.archivo_extractor}: {e}")
                            
                            if hasattr(instancia, 'extraer_total'):
                                try:
                                    datos['total'] = instancia.extraer_total(texto_pdf)
                                except Exception as e:
                                    self.logger.warning(f"  ↳ ⚠️ extraer_total falló en {proveedor.archivo_extractor}: {e}")
                            
                            if hasattr(instancia, 'extraer_referencia'):
                                try:
                                    datos['referencia'] = instancia.extraer_referencia(texto_pdf)
                                except Exception as e:
                                    self.logger.warning(f"  ↳ ⚠️ extraer_referencia falló en {proveedor.archivo_extractor}: {e}")

                            # Detectar proforma
                            if hasattr(instancia, 'es_proforma'):
                                try:
                                    datos['es_proforma'] = instancia.es_proforma(texto_pdf)
                                except Exception as e:
                                    self.logger.warning(f"  ↳ ⚠️ es_proforma falló en {proveedor.archivo_extractor}: {e}")

                # Fallback: funciones a nivel módulo
                if not datos:
                    if hasattr(modulo, 'extraer_datos'):
                        datos = modulo.extraer_datos(tmp_path)
                    elif hasattr(modulo, 'extraer'):
                        datos = modulo.extraer(tmp_path)
                    elif hasattr(modulo, 'procesar'):
                        datos = modulo.procesar(tmp_path)
                
                if not datos:
                    return None
                
                # Convertir a FacturaExtraida
                resultado = FacturaExtraida()
                
                if isinstance(datos, dict):
                    if 'fecha' in datos and datos['fecha']:
                        if isinstance(datos['fecha'], datetime):
                            resultado.fecha = datos['fecha']
                        elif isinstance(datos['fecha'], str):
                            for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y']:
                                try:
                                    resultado.fecha = datetime.strptime(datos['fecha'], fmt)
                                    break
                                except ValueError:
                                    pass
                    
                    if 'total' in datos and datos['total']:
                        try:
                            total = datos['total']
                            if isinstance(total, str):
                                total = total.replace('€', '').replace(' ', '')
                                if ',' in total and '.' in total:
                                    total = total.replace('.', '').replace(',', '.')
                                elif ',' in total:
                                    total = total.replace(',', '.')
                            resultado.total = float(total)
                        except (ValueError, TypeError):
                            pass
                    
                    for campo in ['referencia', 'numero', 'ref', 'numero_factura']:
                        if campo in datos and datos[campo]:
                            ref_candidata = str(datos[campo]).strip()
                            # v1.17: validación anti-basura reforzada
                            if (ref_candidata.upper() not in ExtractorPDF.REF_INVALIDAS
                                    and len(ref_candidata) >= 3
                                    and sum(c.isdigit() for c in ref_candidata) >= 1):
                                resultado.referencia = ref_candidata
                            break
                    
                    resultado.exito = resultado.fecha is not None or resultado.total is not None
                    resultado.metodo = f"extractor:{proveedor.archivo_extractor}"
                    resultado.es_proforma = datos.get('es_proforma', False)

                return resultado
                
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)
        
        except Exception as e:
            self.logger.warning(f"  ↳ ⚠️ Extractor {proveedor.archivo_extractor} FALLÓ: {e}")
            self.logger.debug(f"  ↳ Traceback: {traceback.format_exc()}")
            return None
    
    def _procesar_imagen(
        self,
        resultado: ResultadoProcesamiento,
        nombre_archivo: str,
        contenido: bytes,
        fecha_proceso: datetime
    ):
        """Procesa un archivo de imagen"""
        resultado.requiere_revision = True
        resultado.motivo_revision = "Imagen - revisar manualmente"
        
        fecha_str = fecha_proceso.strftime("%m%d")
        trimestre = obtener_trimestre(fecha_proceso)
        nombre_generado = f"REVISAR {trimestre} {fecha_str} ({nombre_archivo}).pdf"
        resultado.archivo_generado = nombre_generado
        
        self.logger.warning(f"  ↳ Imagen, requiere revisión manual")
    
    def _generar_proveedores_nuevos(self):
        """Genera archivo con sugerencias de proveedores nuevos"""
        nuevos = [r for r in self.resultados if not r.proveedor_identificado and r.remitente]
        
        if nuevos:
            fecha = datetime.now().strftime("%Y%m%d")
            ruta = os.path.join(CONFIG.OUTPUT_PATH, f"PROVEEDORES_NUEVOS_{fecha}.txt")
            
            with open(ruta, 'w', encoding='utf-8') as f:
                f.write("PROVEEDORES NUEVOS DETECTADOS\n")
                f.write(f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n")
                f.write("=" * 60 + "\n\n")
                
                for r in nuevos:
                    f.write(f"Remitente: {r.remitente}\n")
                    f.write(f"Asunto: {r.asunto}\n")
                    f.write(f"Archivo: {r.archivo_generado}\n")
                    f.write("-" * 40 + "\n\n")
            
            self.logger.info(f"Archivo de proveedores nuevos: {ruta}")
    
    def _enviar_notificacion(self):
        """Envía email de notificación"""
        if self.modo_test:
            self.logger.info("Modo test - notificación no enviada")
            return
        
        try:
            trimestre = obtener_trimestre(datetime.now())
            excel_path = os.path.join(CONFIG.OUTPUT_PATH, f"PAGOS_Gmail_{trimestre}.xlsx")
            log_path = os.path.join(CONFIG.LOGS_PATH, f"{datetime.now().strftime('%Y-%m-%d')}.log")
            
            notificador = Notificador(self.gmail.service)
            notificador.enviar_resumen(
                self.resultados,
                self.errores,
                excel_path,
                CONFIG.DROPBOX_BASE,
                log_path
            )
            self.logger.info("Notificación enviada ✓")
        except Exception as e:
            self.logger.error(f"Error enviando notificación: {e}")
    
    def _log_resumen(self):
        """Muestra resumen del procesamiento"""
        total = len(self.resultados)
        exitosos = sum(1 for r in self.resultados if r.archivo_generado and not r.requiere_revision)
        revision = sum(1 for r in self.resultados if r.requiere_revision)
        
        self.logger.info("=" * 60)
        self.logger.info("RESUMEN")
        self.logger.info(f"  Total procesados: {total}")
        self.logger.info(f"  Exitosos: {exitosos}")
        self.logger.info(f"  Requieren revisión: {revision}")
        self.logger.info(f"  Errores: {len(self.errores)}")
        if self._contadores_gracia > 0:
            self.logger.info(f"  ✅ Ventana gracia: {self._contadores_gracia}")
        if self._contadores_pendientes > 0:
            self.logger.warning(f"  ⚠️ Pendientes ubicación: {self._contadores_pendientes}")
            self.logger.warning(f"     → Resolver en Streamlit o ejecutando gmail.py en modo manual")
        self.logger.info("=" * 60)

        # Exportar resumen como JSON para Streamlit
        self._exportar_resumen_json()

    def _exportar_resumen_json(self):
        """Exporta resumen de la ejecución como JSON para el puente de datos Streamlit."""
        try:
            total = len(self.resultados)
            exitosos = sum(1 for r in self.resultados if r.archivo_generado and not r.requiere_revision)
            revision = sum(1 for r in self.resultados if r.requiere_revision)

            proveedores_ok = []
            revision_list = []
            for r in self.resultados:
                nombre = r.proveedor.nombre if r.proveedor else r.remitente
                if r.archivo_generado and not r.requiere_revision:
                    if nombre and nombre not in proveedores_ok:
                        proveedores_ok.append(nombre)
                elif r.requiere_revision:
                    if nombre and nombre not in revision_list:
                        revision_list.append(nombre)

            resumen = {
                "fecha_ejecucion": datetime.now().isoformat(timespec="seconds"),
                "total_procesados": total,
                "exitosos": exitosos,
                "requieren_revision": revision,
                "errores": len(self.errores),
                "facturas_gracia": self._contadores_gracia,
                "facturas_pendientes": self._contadores_pendientes,
                "proveedores_ok": sorted(proveedores_ok),
                "revision": sorted(revision_list),
                "errores_detalle": self.errores[:20]
            }

            ruta_json = os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                "outputs", "logs_gmail", "gmail_resumen.json"
            )
            os.makedirs(os.path.dirname(ruta_json), exist_ok=True)
            with open(ruta_json, "w", encoding="utf-8") as f:
                json.dump(resumen, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Resumen JSON exportado: {ruta_json}")
        except Exception as e:
            self.logger.warning(f"No se pudo exportar resumen JSON: {e}")


# ============================================================================
# MAIN
# ============================================================================

def main():
    parser = argparse.ArgumentParser(description=f"Gmail Module v{VERSION} - Procesador de Facturas")
    parser.add_argument("--produccion", action="store_true", help="Ejecutar en modo producción")
    parser.add_argument("--test", action="store_true", help="Ejecutar en modo test")
    
    args = parser.parse_args()
    
    if not args.produccion and not args.test:
        print("Uso: python gmail.py --produccion  (o --test para pruebas)")
        print("\nModo test: procesa pero no modifica archivos ni envía emails")
        return
    
    modo_test = args.test
    processor = GmailProcessor(modo_test=modo_test, modo_produccion=args.produccion)
    
    exito = processor.ejecutar()
    sys.exit(0 if exito else 1)


if __name__ == "__main__":
    main()
